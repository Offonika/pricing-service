from __future__ import annotations

import json
from collections import Counter
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.core.config import Settings, get_settings
from app.models.procurement_order_formation import (
    ProcurementClassificationProposal,
    ProcurementLifecycleTransitionProposal,
    ProcurementOrderFormation,
    ProcurementOrderFormationEvent,
    ProcurementOrderFormationLine,
)
from app.services.assortment_lifecycle import (
    ASSORTMENT_STATUS_LABELS,
    ASSORTMENT_STATUS_LEGACY_LABELS,
    AssortmentLifecycleInput,
    AssortmentStatus,
    decide_assortment_status,
)
from app.services.assortment_lifecycle_classification_store import (
    ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE,
    ASSORTMENT_LIFECYCLE_RUN_TABLE,
)
from app.services.bitrix_procurement_order_formation_auth import (
    ProcurementOrderFormationSession,
)
from app.services.display_family_registry import matching_review_confirmations_by_code
from app.services.exporters.ut103_nomenclature_properties import (
    NomenclaturePropertyUpdateMessage,
    NomenclaturePropertyUpdateRow,
    PropertyUpdateExchangeResult,
)
from app.services.procurement_order_formation import (
    PROPERTY_UPDATE_SOURCE,
    STATUS_APPROVED_BY_PROPERTY_NAME,
    STATUS_CHANGED_AT_PROPERTY_NAME,
    STATUS_PROPERTY_NAME,
    STATUS_REASON_PROPERTY_NAME,
    STATUS_SOURCE_PROPERTY_NAME,
    approve_order,
    effective_assortment_status,
    get_order,
    line_blocker_details,
    line_blockers,
    normalize_guid,
    normalize_status,
    order_blockers,
    serialize_linked_process,
    serialize_order,
    serialize_proposal,
    status_label,
)
from app.services.procurement_order_registry import (
    LIFECYCLE_STATUS_LABELS as ORDER_LIFECYCLE_STATUS_LABELS,
)
from app.services.procurement_order_registry import (
    lifecycle_display_status,
)
from app.services.procurement_product_cards import bitrix_product_path
from app.services.procurement_supplier_profiles import (
    empty_supplier_profile,
    serialize_supplier_profile,
    supplier_profiles_by_ref,
)

DISPLAY_FOLDER = "дисплеи"
DISPLAY_RESPONSIBLE_NAME = "Омар"
LIFECYCLE_ORDER = ("fruit", "newborn", "new_item", "sales_start", "sale", "working")
# Подписи витрины: действующее название статуса. Прежнее живёт рядом в
# LIFECYCLE_LEGACY_LABELS и показывается на карточке отдельной строкой —
# решение пользователя 2026-08-19.
LIFECYCLE_LABELS = {
    "fruit": ASSORTMENT_STATUS_LABELS[AssortmentStatus.FRUIT],
    "newborn": ASSORTMENT_STATUS_LABELS[AssortmentStatus.NEWBORN],
    "newborn_need": ASSORTMENT_STATUS_LABELS[AssortmentStatus.NEWBORN_NEED],
    "new_item": ASSORTMENT_STATUS_LABELS[AssortmentStatus.NEW_ITEM],
    "sales_start": ASSORTMENT_STATUS_LABELS[AssortmentStatus.SALES_START],
    "sale": ASSORTMENT_STATUS_LABELS[AssortmentStatus.SALE],
    "working": ASSORTMENT_STATUS_LABELS[AssortmentStatus.WORKING],
    "review": "Разбор",
}
LIFECYCLE_LEGACY_LABELS = {
    "fruit": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.FRUIT],
    "newborn": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.NEWBORN],
    "newborn_need": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.NEWBORN_NEED],
    "new_item": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.NEW_ITEM],
    "sales_start": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.SALES_START],
    "sale": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.SALE],
    "working": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.WORKING],
    "review": "Review / разбор",
}
MANUAL_STATUS_ORDER = (
    "matrix",
    "on_demand",
    "replace_candidate",
    "nonliquid",
    "do_not_order",
    "pension",
    "review",
)
MANUAL_STATUS_LABELS = {
    "matrix": ASSORTMENT_STATUS_LABELS[AssortmentStatus.MATRIX],
    "on_demand": ASSORTMENT_STATUS_LABELS[AssortmentStatus.ON_DEMAND],
    "replace_candidate": ASSORTMENT_STATUS_LABELS[AssortmentStatus.REPLACE_CANDIDATE],
    "nonliquid": ASSORTMENT_STATUS_LABELS[AssortmentStatus.NONLIQUID],
    "do_not_order": ASSORTMENT_STATUS_LABELS[AssortmentStatus.DO_NOT_ORDER],
    "pension": ASSORTMENT_STATUS_LABELS[AssortmentStatus.PENSION],
    "review": "Разбор",
}
MANUAL_DECISION_NONBLOCKING_CODES = frozenset({"lifecycle_stage_not_exported"})
MANUAL_STATUS_LEGACY_LABELS = {
    "matrix": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.MATRIX],
    "on_demand": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.ON_DEMAND],
    "replace_candidate": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.REPLACE_CANDIDATE],
    "nonliquid": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.NONLIQUID],
    "do_not_order": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.DO_NOT_ORDER],
    "pension": ASSORTMENT_STATUS_LEGACY_LABELS[AssortmentStatus.PENSION],
    "review": "Review / разбор",
}
ORDER_STATUS_LABELS = {
    "draft": "На подтверждении",
    "review": "На проверке",
    "approved": "Проверен",
    "transmitting": "Передача в 1С",
    "transmitted": "Передан в 1С",
    "deferred": "Отложен",
    "superseded": "Заменён новым расчётом",
}
ORDER_CALCULATION_EXPORT_HEADERS = (
    "Предмет",
    "Категория",
    "Группа",
    "Номенклатура",
    "Артикул",
    "Поставщик",
    "Договор",
    "Склад",
    "Количество",
    "Закупочная цена",
    "Сумма",
    "Валюта",
    "Статус",
    "Номер 1С",
    "Дата 1С",
    "Контур",
    "Источник",
    "Открытый остаток",
    "Поступило",
    "Партия",
    "Дата заказа",
)
MANUAL_STATUS_RECOMMENDATIONS = {
    "matrix": "Проверить матрицу и минимальный запас",
    "on_demand": "Проверить явную клиентскую потребность",
    "replace_candidate": "Подтвердить товар для замены",
    "nonliquid": "Проверить остаток и последние продажи",
    "do_not_order": "Не включать в закупку",
    "pension": "Допродать остаток и закрыть карточку",
    "review": "Открыть ручной разбор",
}
DECISION_STATE_LABELS = {
    "ready": "Готово к подтверждению",
    "review": "Нужен разбор",
    "blocked": "Есть блокер",
    "control": "Контроль",
    "stale": "Расчёт устарел",
    "view": "Только просмотр",
}
DECISION_STATE_URGENCY = {
    "ready": "ready",
    "review": "warning",
    "blocked": "blocked",
    "control": "warning",
    "stale": "warning",
    "view": "neutral",
}
DECISION_STATE_PRIORITY = {
    "review": 0,
    "blocked": 1,
    "ready": 2,
}
ATTENTION_FACTS = (
    ("cargo_handoff_count_1c", "Передач в груз"),
    ("customer_order_count_1c", "Заказов покупателей"),
    ("customer_order_qty_1c", "Количество в заказах покупателей"),
    ("supplier_order_count_1c", "Заказов поставщику"),
    ("supplier_order_qty_1c", "Количество в заказах поставщику"),
)
APPROVAL_RESULT_KEYS = ("approved", "stale", "blocked", "conflict", "failed")


def sync_lifecycle_transition_proposals(
    db: Session,
    *,
    folder: str = DISPLAY_FOLDER,
    run_id: int | None = None,
    settings: Settings | None = None,
) -> dict[str, int]:
    settings = settings or get_settings()
    run = _load_run(db, folder=folder, run_id=run_id)
    if run is None:
        return {"created": 0, "updated": 0, "automatic": 0, "stale": 0, "run_id": 0}
    rows = _snapshot_rows(db, folder=folder, run_id=int(run["id"]))
    created = 0
    updated = 0
    automatic = 0
    active_keys: set[str] = set()
    for row in rows:
        candidate = _transition_candidate(row, run=run, settings=settings)
        if candidate is None:
            continue
        is_automatic = _is_automatic_lifecycle_transition(row, candidate)
        active_keys.add(candidate["idempotency_key"])
        if is_automatic:
            approved_at = datetime.now(UTC).replace(tzinfo=None)
            candidate.update(
                {
                    "status": "auto_applied",
                    "approved_at": approved_at,
                    "approved_by_actor": "system:onec-facts",
                    "approved_by_name": "Автоматически по фактам 1С",
                    "onec_status": "not_required",
                    "payload": {
                        **candidate["payload"],
                        "automatic": True,
                        "automatic_reason": "first_supplier_order_created",
                    },
                }
            )
        proposal = db.scalar(
            select(ProcurementLifecycleTransitionProposal).where(
                ProcurementLifecycleTransitionProposal.idempotency_key
                == candidate["idempotency_key"]
            )
        )
        applied_automatically = False
        if proposal is None:
            db.add(ProcurementLifecycleTransitionProposal(**candidate))
            created += 1
            automatic += int(is_automatic)
            applied_automatically = is_automatic
        elif proposal.status == "pending":
            for key in (
                "nomenclature_ref",
                "product_guid",
                "product_name",
                "folder",
                "reason",
                "facts",
                "blockers",
                "risk_codes",
                "facts_hash",
                "responsible_bitrix_user_id",
                "responsible_name",
                "payload",
            ):
                setattr(proposal, key, candidate[key])
            if is_automatic:
                proposal.status = candidate["status"]
                proposal.approved_at = candidate["approved_at"]
                proposal.approved_by_actor = candidate["approved_by_actor"]
                proposal.approved_by_name = candidate["approved_by_name"]
                proposal.onec_status = candidate["onec_status"]
                automatic += 1
                applied_automatically = True
            updated += 1
        elif is_automatic and proposal.status == "auto_applied":
            applied_automatically = True
        if applied_automatically:
            record_event(
                db,
                entity_type="lifecycle_transition",
                entity_id=candidate["nomenclature_code"],
                event_type="lifecycle_transition_auto_applied",
                idempotency_key=f"{candidate['idempotency_key']}:auto",
                actor="system:onec-facts",
                before={"status": candidate["current_status"]},
                after={"status": candidate["target_status"]},
                payload={
                    "run_id": candidate["run_id"],
                    "run_key": candidate["run_key"],
                    "reason": candidate["reason"],
                    "facts_hash": candidate["facts_hash"],
                },
            )

    pending_rows = db.scalars(
        select(ProcurementLifecycleTransitionProposal).where(
            ProcurementLifecycleTransitionProposal.folder.ilike(f"%{folder}%"),
            ProcurementLifecycleTransitionProposal.status == "pending",
        )
    ).all()
    stale_rows = [
        proposal
        for proposal in pending_rows
        if proposal.run_id != int(run["id"]) or proposal.idempotency_key not in active_keys
    ]
    for proposal in stale_rows:
        proposal.status = "stale"
        proposal.payload = {**(proposal.payload or {}), "stale_reason": "new_run_available"}
    db.commit()
    return {
        "created": created,
        "updated": updated,
        "automatic": automatic,
        "stale": len(stale_rows),
        "run_id": int(run["id"]),
    }


def build_dashboard(
    db: Session,
    *,
    folder: str = DISPLAY_FOLDER,
    settings: Settings | None = None,
) -> dict[str, Any]:
    settings = settings or get_settings()
    run = _load_run(db, folder=folder)
    latest_run_id = int(run["id"]) if run else 0
    rows = _snapshot_rows(db, folder=folder, run_id=latest_run_id or None)
    transition_query = select(ProcurementLifecycleTransitionProposal).where(
        ProcurementLifecycleTransitionProposal.folder.ilike(f"%{folder}%"),
        ProcurementLifecycleTransitionProposal.status == "pending",
    )
    if latest_run_id:
        transition_query = transition_query.where(
            ProcurementLifecycleTransitionProposal.run_id == latest_run_id
        )
    transitions = db.scalars(transition_query).all()
    status_counts: Counter[str] = Counter()
    review_counts: Counter[str] = Counter()
    for row in rows:
        code = _dashboard_status(str(row.get("status") or ""))
        if code in LIFECYCLE_ORDER:
            status_counts[code] += 1
        if bool(row.get("manual_review_required")) and code in LIFECYCLE_ORDER:
            review_counts[code] += 1
    action_counts: Counter[str] = Counter()
    ready_counts: Counter[str] = Counter()
    blocked_counts: Counter[str] = Counter()
    action_breakdowns: dict[str, Counter[str]] = {
        status_code: Counter() for status_code in LIFECYCLE_ORDER
    }
    for item in transitions:
        current_status = _dashboard_status(item.current_status)
        if current_status not in action_breakdowns:
            continue
        action_counts[current_status] += 1
        breakdown_key = (
            "review"
            if item.action_kind == "review" or not item.target_status
            else _dashboard_status(item.target_status)
        )
        action_breakdowns[current_status][breakdown_key] += 1
        decision_state = _decision_state(item, latest_run_id=latest_run_id)
        if decision_state == "blocked":
            blocked_counts[current_status] += 1
        elif decision_state == "ready":
            ready_counts[current_status] += 1
    cards: list[dict[str, Any]] = []
    for status_code in LIFECYCLE_ORDER:
        is_working = status_code == "working"
        action_count = action_counts[status_code]
        action_breakdown = {
            code: action_breakdowns[status_code][code]
            for code in (*LIFECYCLE_ORDER, "review")
            if action_breakdowns[status_code][code]
        }
        target_codes = [code for code in action_breakdown if code != "review"]
        target = target_codes[0] if len(action_breakdown) == 1 and len(target_codes) == 1 else None
        if action_count == 0:
            action_label = "Решений нет"
        elif len(action_breakdown) == 1 and "review" in action_breakdown:
            action_label = "На пересмотр"
        elif target:
            action_label = f"→ {_status_label(target)}"
        else:
            action_label = "Требует решения"
        cards.append(
            {
                "status": status_code,
                "label": LIFECYCLE_LABELS[status_code],
                "legacy_label": LIFECYCLE_LEGACY_LABELS.get(status_code, ""),
                "total_count": status_counts[status_code],
                "action_count": action_count,
                "action_kind": "review" if is_working else "transition",
                "action_label": action_label,
                "target_status": target,
                "action_breakdown": action_breakdown,
                "ready_count": ready_counts[status_code],
                "blocked_count": blocked_counts[status_code],
                "review_count": max(
                    review_counts[status_code],
                    action_breakdowns[status_code]["review"],
                ),
                "overdue_count": 0,
                "urgency": _card_urgency(
                    action_count,
                    review_counts[status_code],
                    blocked_counts[status_code],
                ),
            }
        )
    manual_counts = Counter(str(row.get("status") or "") for row in rows)
    manual_status_counts = {
        code: (review_counts["working"] if code == "review" else manual_counts[code])
        for code in MANUAL_STATUS_ORDER
    }
    decision_summary = Counter(
        _decision_state(item, latest_run_id=latest_run_id) for item in transitions
    )
    attention = [
        _dashboard_attention_item(item, latest_run_id=latest_run_id)
        for item in sorted(
            transitions,
            key=lambda value: (
                DECISION_STATE_PRIORITY.get(
                    _decision_state(value, latest_run_id=latest_run_id),
                    len(DECISION_STATE_PRIORITY),
                ),
                value.created_at or datetime.max,
            ),
        )
    ]
    manual_attention = _manual_attention_items(rows)
    return {
        "folder": folder,
        "responsible_user_id": settings.procurement_order_formation_display_responsible_user_id,
        "responsible_name": DISPLAY_RESPONSIBLE_NAME,
        "run_id": int(run["id"]) if run else None,
        "run_key": str(run["run_key"]) if run else None,
        "updated_at": run.get("finished_at") if run else None,
        "cards": cards,
        "decision_summary": {
            "ready_count": decision_summary["ready"],
            "review_count": decision_summary["review"],
            "blocked_count": decision_summary["blocked"],
        },
        "manual_status_counts": manual_status_counts,
        "attention": attention,
        "manual_attention": manual_attention,
    }


def list_lifecycle_transitions(
    db: Session,
    *,
    status: str,
    scope: str = "action",
    readiness: str = "all",
    search: str = "",
    proposal_id: int | None = None,
    page: int = 1,
    page_size: int = 50,
    folder: str = DISPLAY_FOLDER,
) -> dict[str, Any]:
    normalized_status = (
        "all" if status == "all" else _dashboard_status(normalize_status(status) or status)
    )
    if normalized_status != "all" and normalized_status not in LIFECYCLE_ORDER:
        raise ValueError("unsupported lifecycle status")
    if scope not in {"action", "all"}:
        raise ValueError("scope must be action or all")
    if readiness not in {"all", "ready", "review", "blocked", "stale"}:
        raise ValueError("unsupported readiness filter")
    page = max(page, 1)
    page_size = min(max(page_size, 1), 100)
    search_key = search.strip().casefold()
    latest_run = _load_run(db, folder=folder)
    latest_run_id = int(latest_run["id"]) if latest_run else 0

    if scope == "all":
        rows = _snapshot_rows(db, folder=folder, run_id=latest_run_id or None)
        serialized = [
            _serialize_snapshot_row(row, run=latest_run)
            for row in rows
            if normalized_status == "all"
            or _dashboard_status(str(row.get("status") or "")) == normalized_status
        ]
    else:
        proposal_status = "stale" if readiness == "stale" else "pending"
        proposal_filters = [
            ProcurementLifecycleTransitionProposal.folder.ilike(f"%{folder}%"),
            ProcurementLifecycleTransitionProposal.status == proposal_status,
        ]
        if normalized_status != "all":
            proposal_filters.append(
                ProcurementLifecycleTransitionProposal.current_status.in_(
                    _status_query_values(normalized_status)
                )
            )
        if proposal_status == "pending" and latest_run_id:
            proposal_filters.append(ProcurementLifecycleTransitionProposal.run_id == latest_run_id)
        if proposal_id is not None:
            proposal_filters.append(ProcurementLifecycleTransitionProposal.id == proposal_id)
        proposals = db.scalars(
            select(ProcurementLifecycleTransitionProposal)
            .where(*proposal_filters)
            .order_by(
                ProcurementLifecycleTransitionProposal.status,
                ProcurementLifecycleTransitionProposal.created_at,
            )
        ).all()
        serialized = [serialize_transition(item, latest_run_id=latest_run_id) for item in proposals]
    if search_key:
        serialized = [
            item
            for item in serialized
            if search_key in f"{item['nomenclature_code']} {item['product_name']}".casefold()
        ]
    if readiness != "all":
        serialized = [item for item in serialized if item["decision_state"] == readiness]
    total = len(serialized)
    start = (page - 1) * page_size
    items = serialized[start : start + page_size]
    return {
        "status": normalized_status,
        "scope": scope,
        "total": total,
        "page": page,
        "page_size": page_size,
        "ready_count": sum(1 for item in serialized if item["ready"]),
        "review_count": sum(1 for item in serialized if item["decision_state"] == "review"),
        "blocked_count": sum(1 for item in serialized if item["decision_state"] == "blocked"),
        "stale_count": sum(1 for item in serialized if item["stale"]),
        "items": items,
    }


def approve_lifecycle_transitions(
    db: Session,
    *,
    items: list[Mapping[str, Any]],
    idempotency_key: str,
    session: ProcurementOrderFormationSession,
    settings: Settings | None = None,
) -> dict[str, Any]:
    settings = settings or get_settings()
    if not 1 <= len(items) <= 100:
        raise ValueError("lifecycle approval batch must contain 1..100 rows")
    ensure_lifecycle_approver(session.user_id, settings=settings)
    existing_event = db.scalar(
        select(ProcurementOrderFormationEvent).where(
            ProcurementOrderFormationEvent.idempotency_key == idempotency_key
        )
    )
    if existing_event is not None:
        return dict(existing_event.payload or {})

    results: list[dict[str, Any]] = []
    approved: list[ProcurementLifecycleTransitionProposal] = []
    latest_run = _load_run(db, folder=DISPLAY_FOLDER)
    latest_run_id = int(latest_run["id"]) if latest_run else 0
    for raw in items:
        proposal_id = int(raw.get("proposal_id") or 0)
        proposal = db.get(ProcurementLifecycleTransitionProposal, proposal_id)
        if proposal is None:
            results.append(_approval_result(proposal_id, "failed", "Предложение не найдено"))
            continue
        if proposal.status != "pending":
            results.append(_approval_result(proposal_id, "conflict", "Предложение уже обработано"))
            continue
        if proposal.action_kind != "transition" or not proposal.target_status:
            results.append(
                _approval_result(proposal_id, "blocked", "Нужно отдельное ручное решение")
            )
            continue
        if (
            proposal.run_id != int(raw.get("expected_run_id") or 0)
            or proposal.run_id != latest_run_id
            or proposal.facts_hash != str(raw.get("facts_hash") or "")
        ):
            proposal.status = "stale"
            results.append(_approval_result(proposal_id, "stale", "Расчёт устарел"))
            continue
        expected_status = normalize_status(raw.get("expected_current_status"))
        if normalize_status(proposal.current_status) != expected_status:
            results.append(_approval_result(proposal_id, "conflict", "Текущий статус изменился"))
            continue
        if proposal.blockers:
            results.append(
                _approval_result(
                    proposal_id,
                    "blocked",
                    "; ".join(str(item) for item in proposal.blockers),
                )
            )
            continue
        if normalize_status(proposal.target_status) == "working":
            responsible_id = settings.procurement_order_formation_display_responsible_user_id
            if str(session.user_id) != str(responsible_id):
                results.append(
                    _approval_result(
                        proposal_id,
                        "blocked",
                        "Переход в Рабочий подтверждает ответственный за папку",
                    )
                )
                continue
        approved.append(proposal)

    mode = "internal"
    message_id: str | None = None
    xml_preview = ""
    written_path: Path | None = None
    if approved:
        approved_at = datetime.now(UTC).replace(tzinfo=None)
        for proposal in approved:
            proposal.status = "approved"
            proposal.approved_at = approved_at
            proposal.approved_by_actor = session.actor
            proposal.approved_by_bitrix_user_id = session.user_id
            proposal.approved_by_name = session.user_name or session.actor
            proposal.onec_message_id = None
            proposal.onec_status = "not_applicable"
            proposal.payload = {
                **(proposal.payload or {}),
                "storage": "pricing-service",
                "legacy_onec_export_disabled": True,
            }
            results.append(_approval_result(proposal.id, "approved", "Переход утверждён"))

    summary = {key: 0 for key in APPROVAL_RESULT_KEYS}
    for item in results:
        result = item["result"]
        if result in summary:
            summary[result] += 1
    response = {
        "mode": mode,
        "message_id": message_id,
        "xml_preview": xml_preview,
        "written_path": str(written_path) if written_path else None,
        "summary": summary,
        "items": sorted(results, key=lambda item: item["proposal_id"]),
    }
    record_event(
        db,
        entity_type="lifecycle_batch",
        entity_id=idempotency_key,
        event_type="lifecycle_transitions_approved",
        session=session,
        idempotency_key=idempotency_key,
        payload=response,
    )
    db.commit()
    return response


def decide_lifecycle_transition(
    db: Session,
    *,
    proposal_id: int,
    values: Mapping[str, Any],
    session: ProcurementOrderFormationSession,
    settings: Settings | None = None,
) -> dict[str, Any]:
    settings = settings or get_settings()
    ensure_lifecycle_approver(session.user_id, settings=settings)
    proposal = db.get(ProcurementLifecycleTransitionProposal, proposal_id)
    if proposal is None:
        raise LookupError("lifecycle proposal was not found")

    decision = normalize_status(values.get("decision"))
    if decision not in {"pension", "working"}:
        raise ValueError("manual lifecycle decision must be pension or working")
    reason = str(values.get("reason") or "").strip()
    if not reason:
        raise ValueError("manual lifecycle decision reason is required")
    replacement_sku_code = str(values.get("replacement_sku_code") or "").strip()
    no_replacement = bool(values.get("no_replacement"))
    if decision == "pension" and not replacement_sku_code and not no_replacement:
        raise ValueError("pension requires replacement_sku_code or no_replacement")

    existing_decision = dict((proposal.payload or {}).get("manual_decision") or {})
    if proposal.status != "pending":
        if existing_decision.get("decision") == decision and proposal.facts_hash == str(
            values.get("facts_hash") or ""
        ):
            approved_at = proposal.approved_at or proposal.updated_at or datetime.now(UTC)
            return {
                "proposal_id": proposal.id,
                "result": "approved",
                "message": "Решение уже сохранено",
                "decision": decision,
                "approved_at": approved_at,
            }
        raise ValueError("lifecycle proposal was already processed")

    latest_run = _load_run(db, folder=DISPLAY_FOLDER)
    latest_run_id = int(latest_run["id"]) if latest_run else 0
    if (
        proposal.run_id != int(values.get("expected_run_id") or 0)
        or proposal.run_id != latest_run_id
        or proposal.facts_hash != str(values.get("facts_hash") or "")
    ):
        proposal.status = "stale"
        db.commit()
        raise ValueError("lifecycle calculation is stale; refresh the queue")
    if proposal.action_kind == "transition" and proposal.target_status:
        raise ValueError("automatic transition must use batch approval")
    actionable_blockers = _manual_decision_blockers(proposal)
    if actionable_blockers:
        raise ValueError(
            "manual lifecycle decision has blockers: " + "; ".join(actionable_blockers)
        )

    approved_at = datetime.now(UTC).replace(tzinfo=None)
    before = serialize_transition(proposal, latest_run_id=latest_run_id)
    proposal.status = "approved"
    proposal.action_kind = "manual_decision"
    proposal.target_status = decision
    proposal.reason = reason
    proposal.approved_at = approved_at
    proposal.approved_by_actor = session.actor
    proposal.approved_by_bitrix_user_id = session.user_id
    proposal.approved_by_name = session.user_name or session.actor
    proposal.onec_status = "not_applicable"
    proposal.payload = {
        **(proposal.payload or {}),
        "storage": "pricing-service",
        "legacy_onec_export_disabled": True,
        "manual_decision": {
            "decision": decision,
            "reason": reason,
            "replacement_sku_code": replacement_sku_code or None,
            "no_replacement": no_replacement,
            "approved_at": approved_at.isoformat(),
            "approved_by": session.user_name or session.actor,
        },
    }
    response = {
        "proposal_id": proposal.id,
        "result": "approved",
        "message": (
            "Карточка переведена в «Допродаём»"
            if decision == "pension"
            else "Карточка оставлена в статусе «Рабочий»"
        ),
        "decision": decision,
        "approved_at": approved_at,
    }
    record_event(
        db,
        entity_type="lifecycle_transition",
        entity_id=proposal.id,
        event_type="lifecycle_manual_decision",
        session=session,
        idempotency_key=f"lifecycle-manual:{proposal.id}:{proposal.facts_hash}:{decision}",
        before=before,
        after={
            "status": proposal.status,
            "target_status": decision,
            "reason": reason,
            "replacement_sku_code": replacement_sku_code or None,
            "no_replacement": no_replacement,
        },
        payload=response,
    )
    db.commit()
    return response


def list_orders(
    db: Session,
    *,
    search: str = "",
    status: str = "",
    lifecycle_status: str = "",
    supplier: str = "",
    contour: str = "",
    onec_number: str = "",
    date_from: date | None = None,
    date_to: date | None = None,
    source: str = "",
    blockers: str = "all",
    page: int = 1,
    page_size: int = 50,
) -> dict[str, Any]:
    filtered = _filtered_orders(
        db,
        search=search,
        status=status,
        lifecycle_status=lifecycle_status,
        supplier=supplier,
        contour=contour,
        onec_number=onec_number,
        date_from=date_from,
        date_to=date_to,
        source=source,
        blockers=blockers,
    )
    summary = _orders_summary(filtered)
    page = max(page, 1)
    page_size = min(max(page_size, 1), 100)
    start = (page - 1) * page_size
    return {
        "total": len(filtered),
        "page": page,
        "page_size": page_size,
        "summary": summary,
        "items": [serialize_order_list_item(item) for item in filtered[start : start + page_size]],
    }


def build_order_assistant(
    db: Session,
    *,
    session: ProcurementOrderFormationSession | None = None,
    settings: Settings | None = None,
) -> dict[str, Any]:
    orders = list(db.scalars(_order_list_statement()).unique().all())
    orders = [
        order
        for order in orders
        if order.status in {"draft", "review", "error"}
        and order.onec_status not in {"pending", "transmitted"}
    ]
    orders.sort(key=lambda item: (item.order_date, item.updated_at), reverse=True)
    profiles = supplier_profiles_by_ref(db, (order.supplier_ref for order in orders))
    matching_confirmations = matching_review_confirmations_by_code(db)
    serialized_orders = []
    approver_ids = {
        str(value).strip()
        for value in (
            settings or get_settings()
        ).procurement_order_formation_classification_approver_user_ids
        if str(value).strip()
    }
    for order in orders:
        serialized = serialize_order(order)
        normalized_ref = str(order.supplier_ref or "").strip().lower()
        profile = profiles.get(normalized_ref)
        if profile is not None:
            serialized["supplier_profile"] = serialize_supplier_profile(profile)
        elif serialized.get("supplier_profile", {}).get("data_status") != "missing":
            serialized["supplier_profile"] = {
                **empty_supplier_profile(
                    supplier_ref=normalized_ref or None,
                    supplier_code=order.supplier_code,
                    supplier_name=order.supplier_name,
                ),
                **serialized["supplier_profile"],
            }
        else:
            serialized["supplier_profile"] = empty_supplier_profile(
                supplier_ref=normalized_ref or None,
                supplier_code=order.supplier_code,
                supplier_name=order.supplier_name,
            )
        serialized["supplier_profile"]["can_edit"] = bool(
            session and str(session.user_id) in approver_ids
        )
        for line in serialized["lines"]:
            confirmation = matching_confirmations.get(
                str(line.get("nomenclature_code") or "").strip()
            )
            recommendation = line.get("display_family_recommendation")
            if (
                confirmation
                and isinstance(recommendation, dict)
                and recommendation.get("registry_version_number")
                == confirmation.get("registry_version_number")
                and recommendation.get("registry_inventory_checksum")
                == confirmation.get("registry_inventory_checksum")
            ):
                recommendation["conflict_codes"] = [
                    code
                    for code in recommendation.get("conflict_codes") or []
                    if code not in {"accepted_matching_review", "manual_accepted_matching_review"}
                ]
                recommendation["matching_review_confirmed"] = True
                recommendation["matching_review_confirmed_at"] = confirmation.get("confirmed_at")
                recommendation["matching_review_confirmed_by"] = confirmation.get("confirmed_by")
                line["risk_codes"] = [
                    code
                    for code in line.get("risk_codes") or []
                    if code not in {"accepted_matching_review", "manual_accepted_matching_review"}
                ]
            proposal = line.get("latest_classification")
            if not isinstance(proposal, dict):
                continue
            can_decide = bool(
                session
                and proposal.get("status") == "proposed"
                and str(session.user_id) in approver_ids
                and str(proposal.get("requested_by_bitrix_user_id") or "") != str(session.user_id)
            )
            proposal["can_approve"] = can_decide
            proposal["can_reject"] = can_decide
        serialized_orders.append(serialized)
    lines = [line for order in serialized_orders for line in order["lines"] if not line["removed"]]
    ready_order_ids = {
        int(order["id"])
        for order in serialized_orders
        if not order["blockers"]
        and (active_lines := [line for line in order["lines"] if not line["removed"]])
        and all(_assistant_line_ready(line) for line in active_lines)
    }
    ready_lines = [
        line
        for order in serialized_orders
        if int(order["id"]) in ready_order_ids
        for line in order["lines"]
        if not line["removed"]
    ]
    updated_values = [order.updated_at for order in orders if order.updated_at is not None]
    return {
        "updated_at": max(updated_values) if updated_values else None,
        "summary": {
            "lines": len(lines),
            "ready_lines": len(ready_lines),
            "supplier_missing_lines": sum(
                1
                for order in orders
                for line in order.lines
                if not line.removed
                if not (order.supplier_ref or order.supplier_code)
            ),
            "price_changed_lines": sum(
                _assistant_decimal(line.get("price_change_pct")) not in {None, Decimal("0")}
                for line in lines
            ),
            "low_profitability_lines": sum(
                (value := _assistant_decimal(line.get("profitability_pct"))) is not None
                and value < Decimal("20")
                for line in lines
            ),
            "high_defect_lines": sum(
                (value := _assistant_decimal(line.get("supplier_defect_pct"))) is not None
                and value > Decimal("10")
                and (line.get("supplier_defect_attribution") == "supplier_exact")
                and int(line.get("supplier_defect_history_units") or 0) >= 100
                for line in lines
            ),
            "photo_missing_lines": sum(not line.get("photo_original_url") for line in lines),
            "orders": len(orders),
        },
        "orders": serialized_orders,
    }


def assemble_assistant_orders(
    db: Session,
    *,
    items: list[Mapping[str, Any]],
    idempotency_key: str,
    session: ProcurementOrderFormationSession,
) -> dict[str, Any]:
    results: list[dict[str, Any]] = []
    for item in items:
        order_id = int(item["order_id"])
        expected_version = int(item["expected_version"])
        event_key = f"assistant-assemble:{idempotency_key}:{order_id}"
        existing_event = db.scalar(
            select(ProcurementOrderFormationEvent).where(
                ProcurementOrderFormationEvent.idempotency_key == event_key
            )
        )
        if existing_event is not None:
            results.append({"order_id": order_id, "status": "approved", "message": "Уже собрано"})
            continue
        try:
            order = get_order(db, order_id)
        except LookupError:
            results.append(
                {"order_id": order_id, "status": "blocked", "message": "Заказ не найден"}
            )
            continue
        if order.status == "approved":
            results.append({"order_id": order_id, "status": "approved", "message": "Уже собрано"})
            continue
        if order.status not in {"draft", "review", "error"} or order.onec_status in {
            "pending",
            "transmitted",
        }:
            results.append(
                {
                    "order_id": order_id,
                    "status": "blocked",
                    "message": "Заказ уже вышел из очереди помощника",
                }
            )
            continue
        if order.version != expected_version:
            results.append(
                {
                    "order_id": order_id,
                    "status": "stale",
                    "message": "Версия изменилась — обновите помощник",
                }
            )
            continue
        serialized_lines = {int(line["id"]): line for line in serialize_order(order)["lines"]}
        missing_photo_lines = [
            line.line_number
            for line in order.lines
            if not line.removed
            and not serialized_lines.get(int(line.id or 0), {}).get("photo_original_url")
        ]
        missing_card_lines = [
            line.line_number
            for line in order.lines
            if not line.removed
            and not serialized_lines.get(int(line.id or 0), {}).get("product_card_url")
        ]
        blockers = order_blockers(order)
        if missing_photo_lines:
            blockers.append(
                "Нет исходного фото: строки " + ", ".join(map(str, missing_photo_lines))
            )
        if missing_card_lines:
            blockers.append(
                "Нет подтверждённой карточки товара: строки "
                + ", ".join(map(str, missing_card_lines))
            )
        if blockers:
            results.append(
                {
                    "order_id": order_id,
                    "status": "blocked",
                    "message": "; ".join(blockers),
                }
            )
            continue
        approved = approve_order(db, order_id, session)
        record_event(
            db,
            order_id=order_id,
            entity_type="order",
            entity_id=order_id,
            event_type="assistant_order_assembled",
            session=session,
            after=serialize_order(approved),
            idempotency_key=event_key,
        )
        db.commit()
        results.append(
            {
                "order_id": order_id,
                "status": "approved",
                "message": "Проект заказа собран",
            }
        )
    return {
        "approved": sum(item["status"] == "approved" for item in results),
        "blocked": sum(item["status"] == "blocked" for item in results),
        "stale": sum(item["status"] == "stale" for item in results),
        "items": results,
    }


def build_order_calculation_excel(
    db: Session,
    *,
    search: str = "",
    status: str = "",
    lifecycle_status: str = "",
    supplier: str = "",
    contour: str = "",
    onec_number: str = "",
    date_from: date | None = None,
    date_to: date | None = None,
    source: str = "",
    blockers: str = "all",
) -> bytes:
    orders = _filtered_orders(
        db,
        search=search,
        status=status,
        lifecycle_status=lifecycle_status,
        supplier=supplier,
        contour=contour,
        onec_number=onec_number,
        date_from=date_from,
        date_to=date_to,
        source=source,
        blockers=blockers,
    )
    active_lines = [line for order in orders for line in order.lines if not line.removed]
    nomenclature_codes = {line.nomenclature_code for line in active_lines if line.nomenclature_code}
    classification_by_code: dict[str, Mapping[str, Any]] = {}
    if nomenclature_codes:
        rows = db.execute(
            select(
                ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE.c.nomenclature_code,
                ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE.c.subject_1c,
                ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE.c.category_1c,
                ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE.c.folder,
                ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE.c.article,
            ).where(
                ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE.c.nomenclature_code.in_(
                    nomenclature_codes
                )
            )
        ).mappings()
        classification_by_code = {str(row["nomenclature_code"]): row for row in rows}

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Расчёт заказа"
    worksheet.freeze_panes = "A2"
    worksheet.append(ORDER_CALCULATION_EXPORT_HEADERS)

    for order in orders:
        for line in order.lines:
            if line.removed:
                continue
            classification = classification_by_code.get(str(line.nomenclature_code or ""), {})
            worksheet.append(
                (
                    classification.get("subject_1c") or "",
                    classification.get("category_1c") or "",
                    classification.get("folder") or "",
                    line.nomenclature_name,
                    classification.get("article") or "",
                    order.supplier_name,
                    order.contract_name,
                    order.warehouse_name,
                    line.final_quantity,
                    line.purchase_price,
                    line.amount,
                    line.currency,
                    ORDER_LIFECYCLE_STATUS_LABELS[
                        lifecycle_display_status(order, order_blockers(order))
                    ],
                    order.onec_document_number or "",
                    order.onec_document_date,
                    order.procurement_contour,
                    order.origin,
                    order.onec_open_quantity,
                    order.onec_received_quantity,
                    order.batch_id,
                    order.order_date,
                )
            )

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        row[8].number_format = "0.000"
        row[9].number_format = "#,##0.0000"
        row[10].number_format = "#,##0.00"
        row[14].number_format = "DD.MM.YYYY"
        row[17].number_format = "0.000"
        row[18].number_format = "0.000"
        row[20].number_format = "DD.MM.YYYY"
    for column_number, column_cells in enumerate(worksheet.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_number)].width = min(
            max(width + 2, 12), 45
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _filtered_orders(
    db: Session,
    *,
    search: str = "",
    status: str = "",
    lifecycle_status: str = "",
    supplier: str = "",
    contour: str = "",
    onec_number: str = "",
    date_from: date | None = None,
    date_to: date | None = None,
    source: str = "",
    blockers: str = "all",
) -> list[ProcurementOrderFormation]:
    statement = _order_list_statement()
    orders = list(db.scalars(statement).unique().all())
    search_key = search.strip().casefold()
    supplier_key = supplier.strip().casefold()
    contour_key = contour.strip().casefold()
    onec_number_key = onec_number.strip().casefold()
    source_key = source.strip().casefold()
    filtered: list[ProcurementOrderFormation] = []
    for order in orders:
        order_blocker_list = order_blockers(order)
        display_status = lifecycle_display_status(order, order_blocker_list)
        if not status and order.status == "superseded":
            continue
        if status and order.status != status:
            continue
        if lifecycle_status and display_status != lifecycle_status:
            continue
        if supplier_key and supplier_key not in order.supplier_name.casefold():
            continue
        if contour_key and order.procurement_contour.casefold() != contour_key:
            continue
        if (
            onec_number_key
            and onec_number_key not in str(order.onec_document_number or "").casefold()
        ):
            continue
        if source_key and order.origin.casefold() != source_key:
            continue
        if date_from and order.order_date < date_from:
            continue
        if date_to and order.order_date > date_to:
            continue
        if blockers == "without" and order_blocker_list:
            continue
        if blockers == "with" and not order_blocker_list:
            continue
        if search_key:
            haystack = " ".join(
                [
                    order.supplier_name,
                    order.contract_name,
                    str(order.onec_document_number or ""),
                    *(line.nomenclature_name for line in order.lines),
                    *(str(line.nomenclature_code or "") for line in order.lines),
                ]
            ).casefold()
            if search_key not in haystack:
                continue
        filtered.append(order)
    filtered.sort(key=lambda item: (item.order_date, item.updated_at), reverse=True)
    return filtered


def list_classification_proposals(
    db: Session,
    *,
    status: str = "",
    page: int = 1,
    page_size: int = 50,
    session: ProcurementOrderFormationSession | None = None,
    settings: Settings | None = None,
) -> dict[str, Any]:
    proposals = list(
        db.scalars(
            select(ProcurementClassificationProposal)
            .options(
                selectinload(ProcurementClassificationProposal.line).selectinload(
                    ProcurementOrderFormationLine.order
                )
            )
            .order_by(ProcurementClassificationProposal.requested_at.desc())
        ).all()
    )
    filtered = [item for item in proposals if not status or item.status == status]
    today = datetime.now(UTC).date()
    page = max(page, 1)
    page_size = min(max(page_size, 1), 100)
    start = (page - 1) * page_size
    # Кто вправе утвердить предложение. Без этого признака интерфейс показывал
    # активную кнопку «Принять» и автору предложения, а сервер отбивал её
    # правилом «второго сотрудника» — пользователь упирался в тупик.
    approver_ids = {
        str(value).strip()
        for value in (
            settings or get_settings()
        ).procurement_order_formation_classification_approver_user_ids
        if str(value).strip()
    }
    items = []
    for proposal in filtered[start : start + page_size]:
        line = proposal.line
        order = line.order
        serialized_proposal = serialize_proposal(proposal)
        self_proposed = bool(
            session and str(proposal.requested_by_bitrix_user_id or "") == str(session.user_id)
        )
        can_decide = bool(
            session
            and proposal.status == "proposed"
            and str(session.user_id) in approver_ids
            and not self_proposed
        )
        serialized_proposal["can_approve"] = can_decide
        serialized_proposal["can_reject"] = can_decide
        serialized_proposal["self_proposed"] = self_proposed
        items.append(
            {
                "proposal": serialized_proposal,
                "order_id": order.id,
                "order_version": order.version,
                "line_id": line.id,
                "line_version": line.version,
                "nomenclature_code": line.nomenclature_code,
                "nomenclature_ref": line.nomenclature_ref,
                "product_name": line.nomenclature_name,
                "supplier_name": order.supplier_name,
                "effective_status": effective_assortment_status(line),
            }
        )
    return {
        "total": len(filtered),
        "page": page,
        "page_size": page_size,
        "pending": sum(1 for item in proposals if item.status == "proposed"),
        "approved_today": sum(
            1
            for item in proposals
            if item.approved_at is not None and item.approved_at.date() == today
        ),
        "readback_conflicts": sum(1 for item in proposals if item.status == "conflict"),
        "items": items,
    }


def list_events(
    db: Session,
    *,
    order_id: int | None = None,
    event_type: str = "",
    page: int = 1,
    page_size: int = 50,
) -> dict[str, Any]:
    statement = select(ProcurementOrderFormationEvent)
    if order_id is not None:
        statement = statement.where(ProcurementOrderFormationEvent.order_id == order_id)
    if event_type:
        statement = statement.where(ProcurementOrderFormationEvent.event_type == event_type)
    events = list(
        db.scalars(statement.order_by(ProcurementOrderFormationEvent.created_at.desc())).all()
    )
    page = max(page, 1)
    page_size = min(max(page_size, 1), 100)
    start = (page - 1) * page_size
    return {
        "total": len(events),
        "page": page,
        "page_size": page_size,
        "items": [serialize_event(item) for item in events[start : start + page_size]],
    }


def record_event(
    db: Session,
    *,
    entity_type: str,
    entity_id: str | int,
    event_type: str,
    session: ProcurementOrderFormationSession | None = None,
    order_id: int | None = None,
    before: Mapping[str, Any] | None = None,
    after: Mapping[str, Any] | None = None,
    payload: Mapping[str, Any] | None = None,
    idempotency_key: str | None = None,
    actor: str = "system:procurement-order-formation",
) -> ProcurementOrderFormationEvent:
    if idempotency_key:
        existing = db.scalar(
            select(ProcurementOrderFormationEvent).where(
                ProcurementOrderFormationEvent.idempotency_key == idempotency_key
            )
        )
        if existing is not None:
            return existing
    event = ProcurementOrderFormationEvent(
        order_id=order_id,
        entity_type=entity_type,
        entity_id=str(entity_id),
        event_type=event_type,
        actor=session.actor if session else actor,
        bitrix_user_id=session.user_id if session else None,
        user_name=session.user_name if session else None,
        idempotency_key=idempotency_key,
        before=_jsonable(dict(before or {})),
        after=_jsonable(dict(after or {})),
        payload=_jsonable(dict(payload or {})),
    )
    db.add(event)
    db.flush()
    return event


def serialize_transition(
    proposal: ProcurementLifecycleTransitionProposal,
    *,
    latest_run_id: int,
) -> dict[str, Any]:
    decision_state = _decision_state(proposal, latest_run_id=latest_run_id)
    stale = decision_state == "stale"
    ready = decision_state == "ready"
    actionability = (
        "batch_approve" if ready else "manual_decision" if decision_state == "review" else "blocked"
    )
    blockers = _manual_decision_blockers(proposal)
    ignored_blockers = [
        str(item)
        for item in proposal.blockers or []
        if str(item) in MANUAL_DECISION_NONBLOCKING_CODES
    ]
    return {
        "proposal_id": proposal.id,
        "nomenclature_code": proposal.nomenclature_code,
        "nomenclature_ref": proposal.nomenclature_ref,
        "product_guid": proposal.product_guid,
        "product_name": proposal.product_name,
        "folder": proposal.folder,
        "action_kind": proposal.action_kind,
        "current_status": proposal.current_status,
        "current_status_label": _status_screen_label(proposal.current_status),
        "target_status": proposal.target_status,
        "target_status_label": _status_screen_label(proposal.target_status),
        "proposal_status": proposal.status,
        "reason": _pending_decision_reason(
            current_status=proposal.current_status,
            target_status=proposal.target_status,
            action_kind=proposal.action_kind,
            reason=proposal.reason,
        ),
        "facts": dict(proposal.facts or {}),
        "blockers": blockers,
        "risk_codes": sorted({*list(proposal.risk_codes or []), *ignored_blockers}),
        "run_id": proposal.run_id,
        "run_key": proposal.run_key,
        "facts_hash": proposal.facts_hash,
        "responsible_bitrix_user_id": proposal.responsible_bitrix_user_id,
        "responsible_name": proposal.responsible_name,
        "decision_state": decision_state,
        "actionability": actionability,
        "suggested_manual_status": "pension" if actionability == "manual_decision" else None,
        "ready": ready,
        "selectable": ready,
        "stale": stale,
        "created_at": proposal.created_at,
    }


def serialize_order_list_item(order: ProcurementOrderFormation) -> dict[str, Any]:
    active_lines = [line for line in order.lines if not line.removed]
    blockers = order_blockers(order)
    display_status = lifecycle_display_status(order, blockers)
    total_quantity = sum((line.final_quantity for line in active_lines), Decimal("0"))
    return {
        "id": order.id,
        "stable_key": order.stable_key,
        "status": order.status,
        "lifecycle_status": display_status,
        "lifecycle_status_label": ORDER_LIFECYCLE_STATUS_LABELS[display_status],
        "origin": order.origin,
        "version": order.version,
        "supplier_name": order.supplier_name,
        "contract_name": order.contract_name,
        "warehouse_name": order.warehouse_name,
        "currency": order.currency,
        "route": order.route,
        "batch_id": order.batch_id,
        "order_date": order.order_date,
        "responsible_name": order.responsible_name,
        "source_run_id": order.source_run_id,
        "onec_status": order.onec_status,
        "onec_document_number": order.onec_document_number,
        "onec_document_ref": order.onec_document_ref,
        "onec_document_date": order.onec_document_date,
        "onec_error": order.onec_error,
        "procurement_contour": order.procurement_contour,
        "bitrix_item_url": order.bitrix_item_url,
        "linked_process": serialize_linked_process(order),
        "expected_receipt_date": order.expected_receipt_date,
        "supplier_dispatch_date": order.supplier_dispatch_date,
        "cargo_dropoff_date": order.cargo_dropoff_date,
        "ordered_quantity": order.onec_ordered_quantity or total_quantity,
        "open_quantity": order.onec_open_quantity,
        "received_quantity": order.onec_received_quantity,
        "last_onec_sync_at": order.last_onec_sync_at,
        "sync_conflict": order.sync_conflict,
        "line_count": len(active_lines),
        "total_quantity": total_quantity,
        "total_amount": sum((line.amount for line in active_lines), Decimal("0")),
        "blockers": blockers,
        "blocked_products": _blocked_products(active_lines),
        "updated_at": order.updated_at,
    }


def serialize_event(event: ProcurementOrderFormationEvent) -> dict[str, Any]:
    return {
        "id": event.id,
        "order_id": event.order_id,
        "entity_type": event.entity_type,
        "entity_id": event.entity_id,
        "event_type": event.event_type,
        "actor": event.actor,
        "bitrix_user_id": event.bitrix_user_id,
        "user_name": event.user_name,
        "before": dict(event.before or {}),
        "after": dict(event.after or {}),
        "payload": dict(event.payload or {}),
        "product": _event_product(event),
        "created_at": event.created_at,
    }


def _blocked_products(lines: Iterable[ProcurementOrderFormationLine]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line in lines:
        blockers = line_blockers(line)
        if not blockers:
            continue
        rows.append(
            {
                "line_id": line.id,
                "line_number": line.line_number,
                "bitrix_product_id": line.bitrix_product_id,
                "xml_id": line.bitrix_product_xml_id,
                "nomenclature_code": line.nomenclature_code,
                "name": line.nomenclature_name,
                "blocker_count": len(blockers),
                "blockers": line_blocker_details(line),
                "bitrix_url": bitrix_product_path(
                    line.bitrix_product_id,
                    catalog_id=get_settings().procurement_product_card_catalog_id,
                ),
            }
        )
    return rows


def _event_product(event: ProcurementOrderFormationEvent) -> dict[str, Any] | None:
    payloads = [dict(event.after or {}), dict(event.before or {}), dict(event.payload or {})]
    line_id = str(event.entity_id) if event.entity_type == "order_line" else ""
    for payload in payloads:
        lines = payload.get("lines")
        if isinstance(lines, list):
            for line in lines:
                if not isinstance(line, Mapping):
                    continue
                if line_id and str(line.get("id") or "") != line_id:
                    continue
                product_id = str(line.get("bitrix_product_id") or "").strip()
                if product_id:
                    return {
                        "bitrix_product_id": product_id,
                        "xml_id": str(line.get("bitrix_product_xml_id") or ""),
                        "nomenclature_code": line.get("nomenclature_code"),
                        "name": str(line.get("nomenclature_name") or ""),
                        "bitrix_url": bitrix_product_path(
                            product_id,
                            catalog_id=get_settings().procurement_product_card_catalog_id,
                        ),
                    }
        product_id = str(payload.get("bitrix_product_id") or "").strip()
        if product_id:
            return {
                "bitrix_product_id": product_id,
                "xml_id": str(payload.get("bitrix_product_xml_id") or ""),
                "nomenclature_code": payload.get("nomenclature_code"),
                "name": str(payload.get("nomenclature_name") or ""),
                "bitrix_url": bitrix_product_path(
                    product_id,
                    catalog_id=get_settings().procurement_product_card_catalog_id,
                ),
            }
    return None


def record_lifecycle_property_update_exchange_result(
    db: Session,
    result: PropertyUpdateExchangeResult,
) -> list[ProcurementLifecycleTransitionProposal]:
    proposals = list(
        db.scalars(
            select(ProcurementLifecycleTransitionProposal).where(
                ProcurementLifecycleTransitionProposal.onec_message_id == result.message_id
            )
        ).all()
    )
    if not proposals:
        return []
    conflict = any(
        "conflict" in f"{item.result} {item.message}".casefold()
        or "конфликт" in f"{item.result} {item.message}".casefold()
        for item in result.item_results
    )
    for proposal in proposals:
        if result.ok:
            proposal.status = "applied"
            proposal.onec_status = "success"
            proposal.onec_error = None
        elif conflict:
            proposal.status = "conflict"
            proposal.onec_status = "conflict"
            proposal.onec_error = result.errors or "1C current value conflict"
        else:
            proposal.status = "failed"
            proposal.onec_status = "error"
            proposal.onec_error = result.errors or "1C property update failed"
    db.commit()
    return proposals


def ensure_lifecycle_approver(user_id: str, *, settings: Settings) -> None:
    allowed = {
        str(item).strip()
        for item in settings.procurement_order_formation_lifecycle_approver_user_ids
        if str(item).strip()
    }
    if str(user_id).strip() not in allowed:
        raise PermissionError("user cannot approve lifecycle transitions")


def _load_run(
    db: Session,
    *,
    folder: str,
    run_id: int | None = None,
) -> Mapping[str, Any] | None:
    statement = select(ASSORTMENT_LIFECYCLE_RUN_TABLE)
    if run_id is not None:
        statement = statement.where(ASSORTMENT_LIFECYCLE_RUN_TABLE.c.id == run_id)
    else:
        statement = statement.where(
            ASSORTMENT_LIFECYCLE_RUN_TABLE.c.folder.ilike(f"%{folder}%"),
            ASSORTMENT_LIFECYCLE_RUN_TABLE.c.source_status == "ready",
        ).order_by(ASSORTMENT_LIFECYCLE_RUN_TABLE.c.finished_at.desc())
    return db.execute(statement.limit(1)).mappings().first()


def _snapshot_rows(
    db: Session,
    *,
    folder: str,
    run_id: int | None,
) -> list[Mapping[str, Any]]:
    statement = select(ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE).where(
        ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE.c.folder.ilike(f"%{folder}%")
    )
    if run_id is not None:
        statement = statement.where(
            ASSORTMENT_LIFECYCLE_CLASSIFICATION_TABLE.c.last_run_id == run_id
        )
    return list(db.execute(statement).mappings().all())


def _transition_candidate(
    row: Mapping[str, Any],
    *,
    run: Mapping[str, Any],
    settings: Settings,
) -> dict[str, Any] | None:
    source = dict(row.get("source_record") or {})
    row_status = normalize_status(row.get("status")) or str(row.get("status") or "")
    manual_status = normalize_status(
        source.get("manual_status")
        or source.get("ManualStatus")
        or source.get("current_status")
        or source.get("assortment_status")
    )
    action_kind = "transition"
    current_status = manual_status or row_status
    target_status = normalize_status(row.get("recommended_status"))
    reason = str(row.get("reason_text") or "")

    if row_status == "newborn_need":
        action_kind = "review"
        current_status = "newborn"
        target_status = None
    elif row_status == "working" and (
        bool(row.get("manual_review_required")) or bool(row.get("blockers"))
    ):
        action_kind = "review"
        current_status = "working"
        target_status = None
    elif target_status is None and manual_status in LIFECYCLE_ORDER:
        fact_target = _fact_target_from_source(source, str(row.get("nomenclature_code") or ""))
        if fact_target and fact_target != manual_status:
            target_status = fact_target
    elif target_status is None and "fact_status_decision_requires_1c_approval" in list(
        row.get("export_blockers") or []
    ):
        fact_decision = dict(source.get("fact_status_decision") or {})
        current_status = normalize_status(fact_decision.get("source_status")) or "fruit"
        target_status = row_status

    if action_kind == "transition":
        if not target_status or target_status == current_status:
            return None
        if target_status not in LIFECYCLE_ORDER:
            return None
    elif current_status != "working" and row_status != "newborn_need":
        return None

    product_ref = str(row.get("product_ref") or "").strip() or None
    product_guid = _safe_guid(product_ref)
    blockers = [
        str(item)
        for item in list(row.get("blockers") or []) + list(row.get("export_blockers") or [])
        if str(item) not in {"ut103_export_blocked", "fact_status_decision_requires_1c_approval"}
    ]
    if action_kind == "transition" and not product_guid:
        blockers.append("catalog_guid_missing")
    facts = {
        "reason_codes": list(row.get("reason_codes") or []),
        "source": row.get("source"),
        "classified_at": _jsonable(row.get("classified_at")),
        "evidence": dict((source.get("fact_status_decision") or {}).get("evidence") or {}),
    }
    run_id = int(run["id"])
    idempotency_key = (
        f"proc-lifecycle:{row.get('nomenclature_code')}:{run_id}:"
        f"{current_status}:{target_status or 'review'}:{action_kind}"
    )
    return {
        "nomenclature_code": str(row.get("nomenclature_code") or ""),
        "nomenclature_ref": product_ref,
        "product_guid": product_guid,
        "product_name": str(row.get("name") or ""),
        "folder": str(row.get("folder") or ""),
        "action_kind": action_kind,
        "current_status": current_status,
        "target_status": target_status,
        "status": "pending",
        "reason": reason,
        "facts": facts,
        "blockers": sorted(set(blockers)),
        "risk_codes": list(row.get("reason_codes") or []),
        "run_id": run_id,
        "run_key": str(run.get("run_key") or run_id),
        "facts_hash": str(row.get("source_hash") or ""),
        "idempotency_key": idempotency_key,
        "responsible_bitrix_user_id": (
            settings.procurement_order_formation_display_responsible_user_id
        ),
        "responsible_name": DISPLAY_RESPONSIBLE_NAME,
        "payload": {"snapshot_id": row.get("id")},
    }


def _is_automatic_lifecycle_transition(
    row: Mapping[str, Any],
    candidate: Mapping[str, Any],
) -> bool:
    if candidate.get("current_status") != "fruit" or candidate.get("target_status") != "newborn":
        return False
    reason_codes = {str(item) for item in row.get("reason_codes") or []}
    source = dict(row.get("source_record") or {})
    evidence = dict((source.get("fact_status_decision") or {}).get("evidence") or {})
    try:
        supplier_order_count = int(evidence.get("supplier_order_count_1c") or 0)
    except (TypeError, ValueError):
        supplier_order_count = 0
    return (
        "supplier_order_without_cargo" in reason_codes
        or bool(source.get("first_supplier_order_at"))
        or supplier_order_count > 0
    )


def _fact_target_from_source(source: Mapping[str, Any], nomenclature_code: str) -> str | None:
    try:
        decision = decide_assortment_status(
            AssortmentLifecycleInput(
                nomenclature_code=nomenclature_code,
                created_at=_parse_date(source.get("created_at")),
                first_supplier_order_at=_parse_date(source.get("first_supplier_order_at")),
                supplier_order_cargo_handoff_dates=_parse_dates(
                    source.get("supplier_order_cargo_handoff_dates")
                ),
                receipt_dates=_parse_dates(source.get("receipt_dates")),
                has_need_signal=bool(source.get("has_need_signal")),
            )
        )
    except (TypeError, ValueError):
        return None
    target = decision.recommended_status or decision.status
    return target.value


def _build_transition_message(
    proposals: Iterable[ProcurementLifecycleTransitionProposal],
    *,
    message_id: str,
    mode: str,
    approved_by: str,
) -> NomenclaturePropertyUpdateMessage:
    rows: list[NomenclaturePropertyUpdateRow] = []
    changed_at = datetime.now(UTC).date()
    for proposal in proposals:
        reason = proposal.reason
        base_key = proposal.idempotency_key
        rows.extend(
            (
                NomenclaturePropertyUpdateRow(
                    idempotency_key=f"{base_key}:status",
                    nomenclature_code=proposal.nomenclature_code,
                    property_name=STATUS_PROPERTY_NAME,
                    value_type="property_value",
                    new_value_name=_onec_status_value(proposal.target_status),
                    new_value_tag=str(proposal.target_status or ""),
                    expected_current_value_name=_onec_status_value(proposal.current_status),
                    expected_current_value_tag=proposal.current_status,
                    reason=reason,
                    approved_by=approved_by,
                ),
                NomenclaturePropertyUpdateRow(
                    idempotency_key=f"{base_key}:reason",
                    nomenclature_code=proposal.nomenclature_code,
                    property_name=STATUS_REASON_PROPERTY_NAME,
                    value_type="string",
                    new_value=reason,
                    reason=reason,
                    approved_by=approved_by,
                ),
                NomenclaturePropertyUpdateRow(
                    idempotency_key=f"{base_key}:changed-at",
                    nomenclature_code=proposal.nomenclature_code,
                    property_name=STATUS_CHANGED_AT_PROPERTY_NAME,
                    value_type="date",
                    new_value=changed_at,
                    reason=reason,
                    approved_by=approved_by,
                ),
                NomenclaturePropertyUpdateRow(
                    idempotency_key=f"{base_key}:source",
                    nomenclature_code=proposal.nomenclature_code,
                    property_name=STATUS_SOURCE_PROPERTY_NAME,
                    value_type="string",
                    new_value=PROPERTY_UPDATE_SOURCE,
                    reason=reason,
                    approved_by=approved_by,
                ),
                NomenclaturePropertyUpdateRow(
                    idempotency_key=f"{base_key}:approved-by",
                    nomenclature_code=proposal.nomenclature_code,
                    property_name=STATUS_APPROVED_BY_PROPERTY_NAME,
                    value_type="string",
                    new_value=approved_by,
                    reason=reason,
                    approved_by=approved_by,
                ),
            )
        )
    return NomenclaturePropertyUpdateMessage(
        message_id=message_id,
        rows=tuple(rows),
        mode=mode,
        approved_by=approved_by,
        source=PROPERTY_UPDATE_SOURCE,
    )


def _serialize_snapshot_row(
    row: Mapping[str, Any],
    *,
    run: Mapping[str, Any] | None,
) -> dict[str, Any]:
    status_code = _dashboard_status(str(row.get("status") or ""))
    run_id = int(row.get("last_run_id") or (run.get("id") if run else 0) or 0)
    return {
        "proposal_id": None,
        "nomenclature_code": str(row.get("nomenclature_code") or ""),
        "nomenclature_ref": str(row.get("product_ref") or "") or None,
        "product_guid": _safe_guid(row.get("product_ref")),
        "product_name": str(row.get("name") or ""),
        "folder": str(row.get("folder") or ""),
        "action_kind": "view",
        "current_status": status_code,
        "current_status_label": _status_screen_label(status_code),
        "target_status": None,
        "target_status_label": None,
        "proposal_status": "snapshot",
        "reason": str(row.get("reason_text") or ""),
        "facts": {"reason_codes": list(row.get("reason_codes") or [])},
        "blockers": list(row.get("blockers") or []),
        "risk_codes": list(row.get("reason_codes") or []),
        "run_id": run_id,
        "run_key": str(run.get("run_key") if run else run_id),
        "facts_hash": str(row.get("source_hash") or ""),
        "responsible_bitrix_user_id": None,
        "responsible_name": None,
        "decision_state": "view",
        "ready": False,
        "selectable": False,
        "stale": False,
        "created_at": row.get("classified_at"),
    }


def _order_list_statement():
    return select(ProcurementOrderFormation).options(
        selectinload(ProcurementOrderFormation.lines).selectinload(
            ProcurementOrderFormationLine.classification_proposals
        )
    )


def _orders_summary(orders: Iterable[ProcurementOrderFormation]) -> dict[str, Any]:
    orders_list = list(orders)
    active_lines = [line for order in orders_list for line in order.lines if not line.removed]
    statuses = Counter(
        lifecycle_display_status(order, order_blockers(order)) for order in orders_list
    )
    return {
        "orders": len(orders_list),
        "lines": len(active_lines),
        "quantity": sum((line.final_quantity for line in active_lines), Decimal("0")),
        "amount": sum((line.amount for line in active_lines), Decimal("0")),
        "by_status": dict(statuses),
    }


def _approval_result(proposal_id: int, result: str, message: str) -> dict[str, Any]:
    return {"proposal_id": proposal_id, "result": result, "message": message}


def _status_query_values(status: str) -> tuple[str, ...]:
    return ("newborn", "newborn_need") if status == "newborn" else (status,)


def _dashboard_status(status: str) -> str:
    return "newborn" if status == "newborn_need" else status


def _status_label(status: str | None) -> str:
    if status is None:
        return ""
    normalized = normalize_status(status) or status
    return (
        LIFECYCLE_LABELS.get(normalized)
        or MANUAL_STATUS_LABELS.get(normalized)
        or status_label(normalized)
        or str(status)
    )


def _onec_status_value(status: str | None) -> str:
    """Название статуса для свойства номенклатуры в 1С.

    Экранные подписи переименованы (решение 2026-08-19), а в учётной системе
    значения остались прежними, поэтому в обмен уходит старое название.
    """

    return _status_legacy_label(status) or _status_label(status)


def _status_legacy_label(status: str | None) -> str:
    """Прежнее название статуса или пусто, если его не переименовывали."""

    if status is None:
        return ""
    normalized = normalize_status(status) or status
    legacy = LIFECYCLE_LEGACY_LABELS.get(normalized) or MANUAL_STATUS_LEGACY_LABELS.get(normalized)
    return legacy if legacy and legacy != _status_label(normalized) else ""


def _status_screen_label(status: str | None) -> str:
    """Действующее название плюс прежнее в скобках — для строк списков."""

    label = _status_label(status)
    legacy = _status_legacy_label(status)
    return f"{label} ({legacy})" if legacy else label


def _decision_state(
    proposal: ProcurementLifecycleTransitionProposal,
    *,
    latest_run_id: int,
) -> str:
    if proposal.status == "stale" or proposal.run_id != latest_run_id:
        return "stale"
    if _manual_decision_blockers(proposal):
        return "blocked"
    if proposal.action_kind != "transition" or not proposal.target_status:
        return "review"
    return "ready"


def _manual_decision_blockers(proposal: ProcurementLifecycleTransitionProposal) -> list[str]:
    blockers = [str(item) for item in proposal.blockers or []]
    if proposal.action_kind != "transition" or not proposal.target_status:
        return [item for item in blockers if item not in MANUAL_DECISION_NONBLOCKING_CODES]
    return blockers


def _attention_action_label(proposal: ProcurementLifecycleTransitionProposal) -> str:
    if proposal.action_kind != "transition" or not proposal.target_status:
        return "Открыть разбор"
    return f"{_status_screen_label(proposal.current_status)} → {_status_screen_label(proposal.target_status)}"


def _attention_fact_summary(proposal: ProcurementLifecycleTransitionProposal) -> str:
    facts = dict(proposal.facts or {})
    evidence = facts.get("evidence")
    if not isinstance(evidence, Mapping):
        return "Факты 1С сохранены в расчёте"
    for key, label in ATTENTION_FACTS:
        value = evidence.get(key)
        if value is None or value == "" or value is False or value == 0:
            continue
        return f"{label}: {value}"
    return "Факты 1С сохранены в расчёте"


def _dashboard_attention_item(
    proposal: ProcurementLifecycleTransitionProposal,
    *,
    latest_run_id: int,
) -> dict[str, Any]:
    decision_state = _decision_state(proposal, latest_run_id=latest_run_id)
    state_label = DECISION_STATE_LABELS[decision_state]
    action_label = _attention_action_label(proposal)
    fact_summary = _attention_fact_summary(proposal)
    return {
        "proposal_id": proposal.id,
        "nomenclature_code": proposal.nomenclature_code,
        "product_name": proposal.product_name,
        "current_status": proposal.current_status,
        "current_status_label": _status_screen_label(proposal.current_status),
        "kind": "lifecycle",
        "filter_status": _dashboard_status(proposal.current_status),
        "action_label": action_label,
        "fact_summary": fact_summary,
        "decision_state": decision_state,
        "decision_state_label": state_label,
        "reason": fact_summary,
        "recommendation": action_label,
        "deadline_label": state_label,
        "urgency": DECISION_STATE_URGENCY[decision_state],
        "responsible_name": proposal.responsible_name or DISPLAY_RESPONSIBLE_NAME,
        "overdue": decision_state == "stale",
    }


def _manual_attention_items(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows:
        row_status = str(row.get("status") or "")
        dashboard_status = _dashboard_status(row_status)
        if row_status in MANUAL_STATUS_ORDER and row_status != "review":
            filter_status = row_status
            current_status = row_status
            current_status_label = _status_screen_label(row_status)
        elif dashboard_status == "working" and bool(row.get("manual_review_required")):
            filter_status = "review"
            current_status = "working"
            current_status_label = _status_screen_label("working")
        else:
            continue
        blockers = list(row.get("blockers") or []) + list(row.get("export_blockers") or [])
        has_blockers = bool(blockers)
        decision_state = "blocked" if has_blockers else "control"
        reason = str(row.get("reason_text") or "Требуется ручная проверка")
        # Страховка от новых статусов: витрина не должна падать целиком из-за
        # отсутствующей подсказки.
        action_label = MANUAL_STATUS_RECOMMENDATIONS.get(filter_status, "Проверить решение")
        items.append(
            {
                "proposal_id": None,
                "nomenclature_code": str(row.get("nomenclature_code") or ""),
                "product_name": str(row.get("name") or ""),
                "current_status": current_status,
                "current_status_label": current_status_label,
                "kind": "manual",
                "filter_status": filter_status,
                "action_label": action_label,
                "fact_summary": reason,
                "decision_state": decision_state,
                "decision_state_label": DECISION_STATE_LABELS[decision_state],
                "reason": reason,
                "recommendation": action_label,
                "deadline_label": DECISION_STATE_LABELS[decision_state],
                "urgency": DECISION_STATE_URGENCY[decision_state],
                "responsible_name": DISPLAY_RESPONSIBLE_NAME,
                "overdue": False,
            }
        )
    order = {status: index for index, status in enumerate(MANUAL_STATUS_ORDER)}
    return sorted(
        items,
        key=lambda item: (
            order.get(item["filter_status"], len(order)),
            item["product_name"].casefold(),
        ),
    )


def _pending_decision_reason(
    *,
    current_status: str,
    target_status: str | None,
    action_kind: str,
    reason: str,
) -> str:
    text = str(reason or "").strip()
    if action_kind != "transition" or not target_status:
        return text or "Требуется ручной разбор"
    prefix = f"Рекомендуется переход {_status_screen_label(current_status)} → {_status_screen_label(target_status)}."
    lowered = text.casefold()
    fact_marker = "по твердым фактам 1с:"
    marker_index = lowered.find(fact_marker)
    if marker_index >= 0:
        detail = text[marker_index + len(fact_marker) :]
        detail = detail.split(". 1С не менялась", 1)[0].strip(" .")
        detail = detail.replace("есть cargo/передачи:", "передач в груз —")
        detail = detail.replace("cargo/передачи:", "передачи в груз —")
        return f"{prefix} Факты 1С: {detail}." if detail else prefix
    if lowered.startswith("статус ") and ". " in text:
        text = text.split(". ", 1)[1].strip()
    if not text:
        return prefix
    return f"{prefix} {text}"


def _card_urgency(action_count: int, review_count: int, blocked_count: int = 0) -> str:
    if blocked_count > 0:
        return "blocked"
    if action_count > 0:
        return "action"
    if review_count > 0:
        return "warning"
    return "normal"


def _safe_guid(value: Any) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        normalized = normalize_guid(text)
    except ValueError:
        return None
    if len(normalized) != 36 or normalized.count("-") != 4:
        return None
    return normalized


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def _parse_dates(value: Any) -> tuple[date, ...]:
    if not isinstance(value, (list, tuple)):
        return ()
    return tuple(item for item in (_parse_date(raw) for raw in value) if item is not None)


def _jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, Mapping):
        return {str(key): _jsonable(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_jsonable(item) for item in value]
    try:
        json.dumps(value)
    except TypeError:
        return str(value)
    return value


def _assistant_line_ready(line: Mapping[str, Any]) -> bool:
    return bool(
        not line.get("removed")
        and not line.get("blockers")
        and line.get("product_card_url")
        and line.get("photo_original_url")
    )


def _assistant_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (ValueError, ArithmeticError):
        return None
