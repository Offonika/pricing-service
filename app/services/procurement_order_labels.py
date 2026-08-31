from __future__ import annotations

import hashlib
import hmac
import json
import re
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font
from PIL import Image, ImageDraw
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.infrastructure.db import build_onec_engine_from_settings
from app.services.procurement_labels import _draw_barcode, fetch_onec_supplier_order_lines
from app.services.procurement_order_formation import get_order, serialize_order_label_source

LABEL_SIZES_MM = {"50x40": (50, 40), "40x30": (40, 30)}
LABEL_DPI = 300


class LabelSourceChangedError(ValueError):
    pass


def _positive_integer(value: Any) -> int:
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("Количество этикеток должно быть целым числом") from exc
    if number <= 0 or number != number.to_integral_value():
        raise ValueError("Количество этикеток должно быть положительным целым числом")
    return int(number)


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw[:10])
    except ValueError:
        return None


def _read_onec_order_rows(onec_number: str) -> list[dict[str, Any]]:
    engine = build_onec_engine_from_settings()
    try:
        return fetch_onec_supplier_order_lines(engine, onec_number)
    finally:
        engine.dispose()


def _preview_checksum(preview: dict[str, Any]) -> str:
    payload = {
        "order_id": preview["order_id"],
        "onec_number": preview["onec_number"],
        "onec_date": preview["onec_date"].isoformat(),
        "label_size": preview["label_size"],
        "rows": preview["rows"],
    }
    serialized = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(serialized).hexdigest()


def build_preview_from_rows(
    *,
    order_id: int,
    onec_number: str,
    onec_date: date | datetime | str | None,
    label_size: str,
    rows: Iterable[dict[str, Any]],
    max_page_count: int = 1000,
) -> dict[str, Any]:
    if label_size not in LABEL_SIZES_MM:
        raise ValueError("Поддерживаются размеры 50x40 и 40x30")
    source_rows = list(rows)
    canonical_number = next(
        (
            str(row.get("order_number") or "").strip()
            for row in source_rows
            if row.get("order_number")
        ),
        onec_number,
    )
    canonical_date = next(
        (_as_date(row.get("order_date")) for row in source_rows if _as_date(row.get("order_date"))),
        None,
    ) or _as_date(onec_date)
    result_rows: list[dict[str, Any]] = []
    blockers: list[str] = []
    for raw in source_rows:
        line_no = int(raw.get("line_no") or 0)
        try:
            quantity = _positive_integer(raw.get("quantity"))
        except ValueError as exc:
            blockers.append(f"строка {line_no}: {exc}")
            quantity = 0
        barcode = str(raw.get("barcode") or "").strip()
        if not barcode:
            blockers.append(f"строка {line_no}: не найден штрихкод 1С")
        result_rows.append(
            {
                "line_no": line_no,
                "onec_item_code": str(raw.get("onec_item_code") or "").strip(),
                "item_name": str(raw.get("item_name") or "").strip(),
                "article_1c": str(raw.get("article_1c") or "").strip(),
                "barcode": barcode,
                "quantity": quantity,
            }
        )
    if not result_rows:
        blockers.append(f"Не найдены строки заказа поставщику {canonical_number} в 1С")
    if canonical_date is None:
        blockers.append(f"Не найдена дата заказа поставщику {canonical_number} в 1С")
    total_labels = sum(row["quantity"] for row in result_rows)
    separators = max(0, len(result_rows) - 1)
    total_pages = total_labels + separators
    if total_pages > max_page_count:
        blockers.append(
            f"Слишком много страниц: {total_pages}. Максимум для одного файла — {max_page_count}"
        )
    preview = {
        "order_id": order_id,
        "onec_number": canonical_number,
        "onec_date": canonical_date,
        "label_size": label_size,
        "max_page_count": max_page_count,
        "position_count": len(result_rows),
        "product_label_count": total_labels,
        "separator_count": separators,
        "total_page_count": total_pages,
        "ready": not blockers,
        "blockers": blockers,
        "rows": result_rows,
    }
    if canonical_date is not None:
        preview["source_checksum"] = _preview_checksum(preview)
    else:
        preview["source_checksum"] = ""
    return preview


def build_order_label_preview(
    db: Session, order_id: int, *, label_size: str, settings: Settings
) -> dict[str, Any]:
    order = get_order(db, order_id)
    label_source = serialize_order_label_source(order)
    if label_source is None:
        raise ValueError(
            "Укажите номер существующего заказа 1С или создайте черновик из приложения"
        )
    onec_number = label_source["onec_number"]
    rows = _read_onec_order_rows(onec_number)
    return build_preview_from_rows(
        order_id=order_id,
        onec_number=onec_number,
        onec_date=label_source.get("onec_date"),
        label_size=label_size,
        rows=rows,
        max_page_count=settings.procurement_order_formation_label_max_pages,
    )


def link_order_label_source(
    db: Session,
    order_id: int,
    *,
    onec_number: str,
    label_size: str,
    settings: Settings,
) -> tuple[dict[str, Any], dict[str, Any]]:
    order = get_order(db, order_id)
    if str(order.onec_document_number or "").strip():
        raise ValueError("Для заказа уже получен номер 1С из штатного обмена")
    requested_number = onec_number.strip()
    rows = _read_onec_order_rows(requested_number)
    if not rows:
        raise ValueError(f"Заказ поставщику {requested_number} не найден в 1С")
    preview = build_preview_from_rows(
        order_id=order.id,
        onec_number=requested_number,
        onec_date=None,
        label_size=label_size,
        rows=rows,
        max_page_count=settings.procurement_order_formation_label_max_pages,
    )
    if not preview["onec_date"]:
        raise ValueError(f"Не найдена дата заказа поставщику {preview['onec_number']} в 1С")
    order.label_onec_document_number = preview["onec_number"]
    order.label_onec_document_date = preview["onec_date"]
    order.label_source_linked_at = datetime.now(UTC).replace(tzinfo=None)
    db.flush()
    label_source = serialize_order_label_source(order)
    if label_source is None:
        raise RuntimeError("Не удалось сохранить источник этикеток")
    return label_source, preview


def ensure_label_source_checksum(preview: dict[str, Any], expected_checksum: str) -> None:
    if not hmac.compare_digest(preview["source_checksum"], expected_checksum):
        raise LabelSourceChangedError(
            "Заказ 1С изменился после проверки; повторите проверку количества"
        )


def _order_marker(preview: dict[str, Any]) -> str:
    document_date = _as_date(preview.get("onec_date"))
    if document_date is None:
        raise ValueError("Не найдена дата заказа 1С")
    onec_number = str(preview.get("onec_number") or "").strip()
    match = re.search(r"(\d+)$", onec_number)
    suffix = match.group(1)[-4:].zfill(4) if match else onec_number[-4:]
    return f"{document_date:%d.%m.%y}/{suffix}"


def _label_image(row: dict[str, Any], label_size: str, order_marker: str) -> Image.Image:
    width_mm, height_mm = LABEL_SIZES_MM[label_size]
    width = round(width_mm / 25.4 * LABEL_DPI)
    height = round(height_mm / 25.4 * LABEL_DPI)
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    margin = max(12, round(width * 0.04))
    from app.services.procurement_labels import _font, _trim_text

    code_font = _font(max(22, round(height * 0.075)), bold=True)
    text_font = _font(max(18, round(height * 0.06)))
    barcode_font = _font(max(20, round(height * 0.065)))
    marker_font = _font(max(19, round(height * 0.058)), bold=True)
    usable_width = width - margin * 2
    code_line = row["onec_item_code"]
    if row.get("article_1c"):
        code_line += f"   {row['article_1c']}"
    draw.text(
        (margin, margin),
        _trim_text(draw, code_line, code_font, usable_width),
        fill="black",
        font=code_font,
    )
    draw.text(
        (margin, margin + round(height * 0.10)),
        _trim_text(draw, row["item_name"], text_font, usable_width),
        fill="black",
        font=text_font,
    )
    barcode_top = round(height * 0.40)
    barcode_height = round(height * 0.30)
    _draw_barcode(
        draw,
        barcode=row["barcode"],
        left=margin,
        top=barcode_top,
        width=usable_width,
        height=barcode_height,
    )
    barcode_text = row["barcode"]
    text_width = draw.textlength(barcode_text, font=barcode_font)
    draw.text(
        ((width - text_width) / 2, barcode_top + barcode_height + 8),
        barcode_text,
        fill="black",
        font=barcode_font,
    )
    marker_width = draw.textlength(order_marker, font=marker_font)
    draw.text(
        (width - margin - marker_width, height - margin - round(height * 0.07)),
        order_marker,
        fill="black",
        font=marker_font,
    )
    return image


def _pdf_bytes(
    pages: list[tuple[bytes | None, int, int]], width_pt: float, height_pt: float
) -> bytes:
    objects: list[bytes] = [b"", b""]
    page_refs: list[int] = []
    image_refs: dict[bytes, int] = {}
    for jpeg, width_px, height_px in pages:
        image_ref = None
        if jpeg is not None:
            image_ref = image_refs.get(jpeg)
            if image_ref is None:
                image_ref = len(objects) + 1
                image_refs[jpeg] = image_ref
                objects.append(
                    (
                        f"<< /Type /XObject /Subtype /Image /Width {width_px} /Height {height_px} "
                        f"/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode /Length {len(jpeg)} >>\nstream\n"
                    ).encode()
                    + jpeg
                    + b"\nendstream"
                )
        content = (
            b""
            if image_ref is None
            else f"q {width_pt:.4f} 0 0 {height_pt:.4f} 0 0 cm /Im0 Do Q".encode()
        )
        content_ref = len(objects) + 1
        objects.append(
            f"<< /Length {len(content)} >>\nstream\n".encode() + content + b"\nendstream"
        )
        page_ref = len(objects) + 1
        resources = "<< >>" if image_ref is None else f"<< /XObject << /Im0 {image_ref} 0 R >> >>"
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width_pt:.4f} {height_pt:.4f}] /Resources {resources} /Contents {content_ref} 0 R >>".encode()
        )
        page_refs.append(page_ref)
    objects[0] = b"<< /Type /Catalog /Pages 2 0 R >>"
    objects[1] = (
        f"<< /Type /Pages /Count {len(page_refs)} /Kids [{' '.join(f'{ref} 0 R' for ref in page_refs)}] >>".encode()
    )
    output = BytesIO()
    output.write(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for number, payload in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f"{number} 0 obj\n".encode() + payload + b"\nendobj\n")
    xref = output.tell()
    output.write(f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode())
    for offset in offsets[1:]:
        output.write(f"{offset:010d} 00000 n \n".encode())
    output.write(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode()
    )
    return output.getvalue()


def build_pdf(preview: dict[str, Any]) -> bytes:
    if not preview["ready"]:
        raise ValueError("Нельзя сформировать этикетки: исправьте ошибки preview")
    pages: list[tuple[bytes | None, int, int]] = []
    order_marker = _order_marker(preview)
    for index, row in enumerate(preview["rows"]):
        image = _label_image(row, preview["label_size"], order_marker)
        jpeg_io = BytesIO()
        image.save(jpeg_io, format="JPEG", quality=92, dpi=(LABEL_DPI, LABEL_DPI))
        jpeg = jpeg_io.getvalue()
        pages.extend([(jpeg, image.width, image.height)] * row["quantity"])
        if index + 1 < len(preview["rows"]):
            pages.append((None, image.width, image.height))
    width_mm, height_mm = LABEL_SIZES_MM[preview["label_size"]]
    return _pdf_bytes(pages, width_mm / 25.4 * 72, height_mm / 25.4 * 72)


def build_xlsx(preview: dict[str, Any]) -> bytes:
    if not preview["ready"]:
        raise ValueError("Нельзя сформировать этикетки: исправьте ошибки preview")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Этикетки"
    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 62 if preview["label_size"] == "50x40" else 48
    current = 1
    order_marker = _order_marker(preview)
    for position, row in enumerate(preview["rows"]):
        barcode_image = _label_image(row, preview["label_size"], order_marker)
        image_io = BytesIO()
        barcode_image.save(image_io, format="PNG", optimize=True)
        image_bytes = image_io.getvalue()
        for copy_no in range(1, row["quantity"] + 1):
            sheet.cell(
                current,
                2,
                f"{row['onec_item_code']}    Заказ {preview['onec_number']}    {order_marker}",
            ).font = Font(bold=True, size=12)
            sheet.cell(current + 1, 2, row["item_name"]).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
            image = ExcelImage(BytesIO(image_bytes))
            image.width = 360 if preview["label_size"] == "50x40" else 300
            image.height = 288 if preview["label_size"] == "50x40" else 225
            sheet.add_image(image, f"B{current + 2}")
            sheet.row_dimensions[current].height = 20
            sheet.row_dimensions[current + 1].height = 34
            sheet.row_dimensions[current + 2].height = image.height * 0.75
            sheet.cell(current + 3, 1, f"{position + 1}.{copy_no}").font = Font(
                size=8, color="808080"
            )
            current += 5
        if position + 1 < len(preview["rows"]):
            # Пустой разделитель занимает высоту полноценной этикетки, а не
            # обычной строки Excel: на рулонной печати это физический пропуск.
            sheet.row_dimensions[current].height = image.height * 0.75 + 54
            current += 5
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
