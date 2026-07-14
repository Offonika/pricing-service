from __future__ import annotations

import hashlib
from collections import Counter
from datetime import UTC, date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session, joinedload

from app.models.weekly_kpi_report import WeeklyKpiReportMetricSnapshot, WeeklyKpiReportSnapshot

READY_LIFECYCLE = "published"
READY_ELIGIBILITY = "eligible"
READY_ARTIFACT = "ready"
MONEY_QUANT = Decimal("0.01")
VALUE_QUANT = Decimal("0.0001")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")


def last_completed_week_end(today: date | None = None) -> date:
    anchor = today or date.today()
    return anchor - timedelta(days=(anchor.weekday() + 1) % 7)


def _utcnow() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def week_bounds_from_end(week_end: date) -> tuple[date, date]:
    return week_end - timedelta(days=6), week_end


def _quantize(value: Decimal | int | float | None, quant: Decimal = VALUE_QUANT) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        source = value
    else:
        source = Decimal(str(value))
    return source.quantize(quant, rounding=ROUND_HALF_UP)


def _artifact_relative_path(report: WeeklyKpiReportSnapshot) -> Path:
    safe_employee_key = report.employee_key.replace("/", "-")
    filename = (
        f"weekly-kpi-{safe_employee_key}-{report.week_start.isoformat()}-"
        f"to-{report.week_end.isoformat()}-r{report.revision}.xlsx"
    )
    return Path(report.week_end.isoformat()) / safe_employee_key / filename


def _format_metric_value(value: Decimal | None, unit: str | None) -> str:
    if value is None:
        return ""
    unit_normalized = (unit or "").strip().lower()
    if unit_normalized in {"rub", "rur", "money"}:
        return f"{_quantize(value, MONEY_QUANT):,.2f}".replace(",", " ")
    if unit_normalized in {"pct", "percent", "%"}:
        return f"{_quantize(value, Decimal('0.01'))}%"
    return f"{_quantize(value):,.4f}".replace(",", " ")


def _base_query(session: Session):
    return session.query(WeeklyKpiReportSnapshot).options(
        joinedload(WeeklyKpiReportSnapshot.metrics)
    )


def _ready_query(session: Session):
    return _base_query(session).filter(
        WeeklyKpiReportSnapshot.lifecycle_status == READY_LIFECYCLE,
        WeeklyKpiReportSnapshot.eligibility_status == READY_ELIGIBILITY,
        WeeklyKpiReportSnapshot.artifact_status == READY_ARTIFACT,
    )


def build_weekly_kpi_report_health(
    session: Session,
    *,
    week_end: date,
) -> dict[str, Any]:
    rows = (
        _base_query(session)
        .filter(WeeklyKpiReportSnapshot.week_end == week_end)
        .order_by(WeeklyKpiReportSnapshot.report_key, WeeklyKpiReportSnapshot.revision.desc())
        .all()
    )
    lifecycle_counts = Counter(item.lifecycle_status for item in rows)
    eligibility_counts = Counter(item.eligibility_status for item in rows)
    artifact_counts = Counter(item.artifact_status for item in rows)
    latest_generated_at = max((item.generated_at for item in rows), default=None)
    ready_count = sum(
        1
        for item in rows
        if item.lifecycle_status == READY_LIFECYCLE
        and item.eligibility_status == READY_ELIGIBILITY
        and item.artifact_status == READY_ARTIFACT
    )
    if ready_count:
        status = "ready"
    elif rows:
        status = "draft"
    else:
        status = "missing"
    return {
        "as_of": week_end,
        "freshness_status": status,
        "source_status": "ready" if rows else "empty",
        "week_end": week_end,
        "status": status,
        "report_count": len(rows),
        "ready_count": ready_count,
        "lifecycle_counts": dict(lifecycle_counts),
        "eligibility_counts": dict(eligibility_counts),
        "artifact_counts": dict(artifact_counts),
        "latest_generated_at": latest_generated_at,
    }


def list_ready_weekly_kpi_reports(
    session: Session,
    *,
    week_end: date,
    employee_key: str | None = None,
    bitrix_user_id: str | None = None,
    limit: int | None = None,
) -> list[WeeklyKpiReportSnapshot]:
    query = _ready_query(session).filter(WeeklyKpiReportSnapshot.week_end == week_end)
    if employee_key:
        query = query.filter(WeeklyKpiReportSnapshot.employee_key == employee_key)
    if bitrix_user_id:
        query = query.filter(
            (WeeklyKpiReportSnapshot.bitrix_user_id == bitrix_user_id)
            | (WeeklyKpiReportSnapshot.bitrix_box_user_id == bitrix_user_id)
        )
    query = query.order_by(WeeklyKpiReportSnapshot.employee_name, WeeklyKpiReportSnapshot.id)
    if limit is not None:
        query = query.limit(limit)
    return query.all()


def get_ready_weekly_kpi_report(
    session: Session,
    *,
    report_id: int,
) -> WeeklyKpiReportSnapshot | None:
    return _ready_query(session).filter(WeeklyKpiReportSnapshot.id == report_id).first()


def publish_weekly_kpi_reports(
    session: Session,
    *,
    week_end: date,
    report_keys: Iterable[str] | None = None,
) -> dict[str, Any]:
    query = _base_query(session).filter(
        WeeklyKpiReportSnapshot.week_end == week_end,
        WeeklyKpiReportSnapshot.lifecycle_status == "draft",
        WeeklyKpiReportSnapshot.eligibility_status == "eligible",
    )
    if report_keys:
        query = query.filter(WeeklyKpiReportSnapshot.report_key.in_(list(report_keys)))
    rows = query.all()
    published_at = _utcnow()
    for row in rows:
        row.lifecycle_status = "published"
        row.artifact_status = "pending"
        row.published_at = published_at
    return {
        "week_end": week_end.isoformat(),
        "published_count": len(rows),
        "report_ids": [item.id for item in rows],
    }


def _write_summary_sheet(workbook: Workbook, report: WeeklyKpiReportSnapshot) -> None:
    sheet = workbook.active
    sheet.title = "summary"
    sheet.freeze_panes = "A2"
    sheet.append(["Поле", "Значение"])
    header = (
        report.summary_payload.get("header") if isinstance(report.summary_payload, dict) else {}
    )
    rows = [
        ("Сотрудник", report.employee_name),
        ("Ключ сотрудника", report.employee_key),
        ("Роль", report.role_code or ""),
        ("Должность", report.position_name or ""),
        ("Неделя с", report.week_start.isoformat()),
        ("Неделя по", report.week_end.isoformat()),
        ("Revision", report.revision),
        ("Signal", report.overall_signal or ""),
        ("Header title", (header or {}).get("title", "")),
        ("Header subtitle", (header or {}).get("subtitle", "")),
        ("Eligibility", report.eligibility_status),
        ("Eligibility reason", report.eligibility_reason or ""),
    ]
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 90


def _write_metrics_sheet(
    workbook: Workbook,
    metrics: list[WeeklyKpiReportMetricSnapshot],
) -> None:
    sheet = workbook.create_sheet("kpi_metrics")
    sheet.freeze_panes = "A2"
    headers = [
        "metric_code",
        "metric_name",
        "unit",
        "fact_value",
        "plan_value",
        "achievement_pct",
        "bonus_preview_amount",
        "previous_fact_value",
        "delta_abs",
        "delta_pct",
        "signal",
        "source_system",
        "source_entity",
        "comment",
    ]
    sheet.append(headers)
    for metric in metrics:
        sheet.append(
            [
                metric.metric_code,
                metric.metric_name,
                metric.unit,
                float(metric.fact_value or 0),
                float(metric.plan_value) if metric.plan_value is not None else None,
                float(metric.achievement_pct) if metric.achievement_pct is not None else None,
                (
                    float(metric.bonus_preview_amount)
                    if metric.bonus_preview_amount is not None
                    else None
                ),
                (
                    float(metric.previous_fact_value)
                    if metric.previous_fact_value is not None
                    else None
                ),
                float(metric.delta_abs) if metric.delta_abs is not None else None,
                float(metric.delta_pct) if metric.delta_pct is not None else None,
                metric.signal,
                metric.source_system,
                metric.source_entity,
                metric.comment,
            ]
        )
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.auto_filter.ref = f"A1:N{max(sheet.max_row, 1)}"
    for column in ("A", "B", "C", "K", "L", "M", "N"):
        sheet.column_dimensions[column].width = 22


def _write_meta_sheet(workbook: Workbook, report: WeeklyKpiReportSnapshot) -> None:
    sheet = workbook.create_sheet("_meta")
    sheet.append(["field", "value"])
    rows = [
        ("report_id", report.id),
        ("report_key", report.report_key),
        ("revision", report.revision),
        ("lifecycle_status", report.lifecycle_status),
        ("artifact_status", report.artifact_status),
        ("generated_at", report.generated_at.isoformat() if report.generated_at else ""),
        ("published_at", report.published_at.isoformat() if report.published_at else ""),
        ("bitrix_user_id", report.bitrix_user_id or ""),
        ("bitrix_box_user_id", report.bitrix_box_user_id or ""),
        ("source_as_of", report.source_as_of.isoformat() if report.source_as_of else ""),
    ]
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 48


def export_weekly_kpi_report_artifact(
    report: WeeklyKpiReportSnapshot,
    *,
    output_dir: Path,
) -> tuple[Path, str]:
    workbook = Workbook()
    _write_summary_sheet(workbook, report)
    _write_metrics_sheet(workbook, report.metrics)
    _write_meta_sheet(workbook, report)

    relative_path = _artifact_relative_path(report)
    full_path = output_dir / relative_path
    full_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(full_path)
    digest = hashlib.sha256(full_path.read_bytes()).hexdigest()
    return full_path, digest


def build_pending_weekly_kpi_artifacts(
    session: Session,
    *,
    output_dir: Path,
    week_end: date | None = None,
    report_ids: Iterable[int] | None = None,
) -> dict[str, Any]:
    query = _base_query(session).filter(
        WeeklyKpiReportSnapshot.lifecycle_status == "published",
        WeeklyKpiReportSnapshot.eligibility_status == "eligible",
        WeeklyKpiReportSnapshot.artifact_status.in_(["pending", "failed"]),
    )
    if week_end is not None:
        query = query.filter(WeeklyKpiReportSnapshot.week_end == week_end)
    if report_ids:
        query = query.filter(WeeklyKpiReportSnapshot.id.in_(list(report_ids)))

    built: list[dict[str, Any]] = []
    rows = query.order_by(
        WeeklyKpiReportSnapshot.week_end, WeeklyKpiReportSnapshot.employee_name
    ).all()
    for row in rows:
        try:
            artifact_path, digest = export_weekly_kpi_report_artifact(row, output_dir=output_dir)
            row.artifact_path = str(artifact_path)
            row.artifact_sha256 = digest
            row.artifact_status = "ready"
            built.append(
                {"report_id": row.id, "artifact_path": str(artifact_path), "sha256": digest}
            )
        except Exception:
            row.artifact_status = "failed"
            raise

    return {
        "built_count": len(built),
        "reports": built,
    }


def build_weekly_kpi_report_manifest(
    report: WeeklyKpiReportSnapshot,
    *,
    include_metrics: bool,
) -> dict[str, Any]:
    payload = {
        "report_id": report.id,
        "report_key": report.report_key,
        "revision": report.revision,
        "overall_signal": report.overall_signal,
        "summary_payload": report.summary_payload or {},
        "employee": {
            "employee_key": report.employee_key,
            "employee_name": report.employee_name,
            "role_code": report.role_code,
            "position_code": report.position_code,
            "position_name": report.position_name,
            "bitrix_user_id": report.bitrix_user_id,
            "bitrix_box_user_id": report.bitrix_box_user_id,
        },
        "period": {
            "week_start": report.week_start,
            "week_end": report.week_end,
            "source_as_of": report.source_as_of,
        },
        "artifact_url": f"/api/management/weekly-kpi-reports/{report.id}/artifact",
    }
    if include_metrics:
        payload["metrics"] = [
            {
                "metric_code": metric.metric_code,
                "metric_name": metric.metric_name,
                "unit": metric.unit,
                "fact_value": metric.fact_value,
                "plan_value": metric.plan_value,
                "achievement_pct": metric.achievement_pct,
                "bonus_preview_amount": metric.bonus_preview_amount,
                "previous_fact_value": metric.previous_fact_value,
                "delta_abs": metric.delta_abs,
                "delta_pct": metric.delta_pct,
                "signal": metric.signal,
                "source_system": metric.source_system,
                "source_entity": metric.source_entity,
                "comment": metric.comment,
            }
            for metric in report.metrics
        ]
    return payload
