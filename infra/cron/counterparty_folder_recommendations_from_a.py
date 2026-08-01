#!/usr/bin/env python3
"""Pull counterparty folder recommendations from server A and export CSV artifacts."""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import json
import os
import urllib.error
import urllib.parse
import urllib.request
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:  # Deployed on Openclaw as a flat scripts directory.
    from weekly_kpi_reports_from_a import (  # type: ignore
        DEFAULT_LOCAL_ENV_FILE,
        DEFAULT_LOCAL_SOURCE_URL,
        _build_fetcher,
        _load_env,
    )
except ImportError:  # Local tests/imports from pricing-service repo.
    from infra.cron.weekly_kpi_reports_from_a import (
        DEFAULT_LOCAL_ENV_FILE,
        DEFAULT_LOCAL_SOURCE_URL,
        _build_fetcher,
        _load_env,
    )


DEFAULT_STATE_PATH = (
    "/home/deploy/.openclaw/workspace/.data/counterparty-folder-recommendations/state.json"
)
DEFAULT_ARTIFACT_DIR = (
    "/home/deploy/.openclaw/workspace/.data/counterparty-folder-recommendations/artifacts"
)
REPORT_ENDPOINT = "/api/management/counterparty-folder-recommendations"
STATUS_MOVE_RECOMMENDED = "move_recommended"
STATUS_OK = "ok"
STATUS_NO_OVERDUE = "no_overdue"
STATUS_NEEDS_REVIEW = "needs_review"
STATUS_VALUES = (
    STATUS_MOVE_RECOMMENDED,
    STATUS_OK,
    STATUS_NO_OVERDUE,
    STATUS_NEEDS_REVIEW,
)
QUEUE_ACTIONABLE = "actionable"
QUEUE_ALL = "all"
DELIVERY_LEGACY = "legacy"
DELIVERY_DAILY_DELTA = "daily_delta"
DELIVERY_WEEKLY_SUMMARY = "weekly_summary"
PAYMENT_TERM_SOURCE_FALLBACK = "fallback_7_days_read_only"
REVIEW_REASON_LABELS = {
    "missing_origin_document": "не найден исходный документ долга",
    "origin_document_not_found": "исходный документ долга не найден в 1С",
    "origin_document_department_missing": "у документа долга не заполнено подразделение",
    "department_folder_missing": "у подразделения долга не настроена папка",
    "current_counterparty_folder_missing": "у контрагента не найдена текущая папка",
    "folder_mismatch_payment_term_missing": (
        "папка отличается, но срок оплаты не заполнен; " "в отчете применен расчетный срок 7 дней"
    ),
    "spb_cross_folder_manual_review": "межпапочный СПБ-кейс, нужна ручная проверка",
    "excluded_employee_folder": "исключено: контрагент или папка сотрудников",
    "excluded_wholesale_counterparty": "исключено: оптовый клиент или оптовый отдел",
    "excluded_supplier_folder": "исключено: контрагент находится в папке поставщиков",
    "excluded_site_payment_on_pickup": "исключено: выдача без оплаты, отвечает сайт",
    "excluded_maklab_spb_prosvet": "исключено: Маклаб СПБ ПРОСВЕТ, не трогаем",
    "below_min_balance_threshold": "скрыто из ежедневного списка: сумма долга ниже порога",
    "excluded_china_supplier_group": "исключено: группа доступа Поставщики Китай",
    "open_structure_document_not_found": (
        "не удалось определить открытые реализации по структуре 1С"
    ),
    "origin_document_needs_order_payment_check": (
        "долг сайта: проверить заказ, оплату картой и фактический источник долга"
    ),
    "origin_document_structure_unconfirmed": (
        "источник долга требует проверки структуры документа в 1С"
    ),
    "origin_document_structure_confirmed_manual_review": (
        "структура 1С подтверждает открытый остаток; кандидат требует ручной проверки"
    ),
    "origin_document_closed_by_structure": (
        "выбранный документ закрыт по структуре 1С; нужно найти фактический источник долга"
    ),
    "document_comment_history_required": ("нужна проверка комментария или истории документа в 1С"),
}
DOCUMENT_STRUCTURE_STATUS_LABELS = {
    "confirmed_open": "открытый остаток подтвержден структурой 1С",
    "closed_by_structure": "выбранный документ закрыт по структуре 1С",
    "ambiguous": "структура документа не дала однозначный ответ",
    "not_found": "документ не найден в 1С",
}
STATEMENT_RULE_LABELS = {
    "statement_direct_payment_match": "РТУ закрыта оплатой рядом в ведомости",
    "statement_multi_sale_payment_match": "оплата закрыла несколько соседних РТУ",
    "statement_bottom_up_balance_cutoff": "выбрано снизу вверх по общей сумме долга",
    "statement_unmatched_open_sale": "открытая РТУ по ведомостной логике",
    "statement_structure_confirmed_open": "открыто по структуре 1С и ведомостной логике",
    "statement_structure_closed": "закрыто по структуре 1С",
}
BITRIX_WEBHOOK_ENV_KEYS = (
    "COUNTERPARTY_FOLDER_RECOMMENDATIONS_BITRIX_WEBHOOK_BASE",
    "CONTRACTOR_PROJECT_REPORT_BITRIX_WEBHOOK_BASE",
    "BITRIX_BOX_WEBHOOK_BASE",
    "BITRIX24_BOX_WEBHOOK_URL",
    "BITRIX_WEBHOOK_BASE",
    "BITRIX24_WEBHOOK_URL",
)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pull counterparty folder recommendations and export a dry-run CSV."
    )
    parser.add_argument(
        "--date",
        dest="snapshot_date",
        help="Receivables snapshot date in YYYY-MM-DD format; default is today.",
    )
    parser.add_argument(
        "--delivery-mode",
        choices=(DELIVERY_LEGACY, DELIVERY_DAILY_DELTA, DELIVERY_WEEKLY_SUMMARY),
        default=DELIVERY_LEGACY,
        help="Legacy report, actionable delta, or weekly aggregate.",
    )
    parser.add_argument(
        "--status",
        choices=STATUS_VALUES,
        default=STATUS_MOVE_RECOMMENDED,
        help="Recommendation status filter; default is move_recommended.",
    )
    parser.add_argument("--limit", type=int, help="Optional max row count.")
    parser.add_argument(
        "--dry-run", action="store_true", help="Fetch and summarize without writing state/artifact."
    )
    parser.add_argument("--force", action="store_true", help="Export even if this revision exists.")
    parser.add_argument(
        "--notify-bitrix-task-id",
        type=int,
        help="Post a deduplicated summary comment to this Bitrix task when the report has rows.",
    )
    parser.add_argument(
        "--notify-empty",
        action="store_true",
        help="Also post Bitrix comments for empty reports. Default: skip empty reports.",
    )
    parser.add_argument("--json", action="store_true", help="Emit machine-readable summary.")
    return parser.parse_args()


def _parse_date(value: str | None) -> date:
    if value:
        return date.fromisoformat(value)
    return date.today()


def _utcnow() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _load_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"version": 2, "reports": {}, "active_signals": {}}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        return {"version": 2, "reports": {}, "active_signals": {}}
    payload.setdefault("reports", {})
    payload.setdefault("active_signals", {})
    payload.setdefault("weekly_window", {})
    payload["version"] = 2
    return payload


def _signal_content_hash(row: dict[str, Any]) -> str:
    relevant = {
        key: row.get(key)
        for key in (
            "signal_key",
            "counterparty_ref",
            "current_folder_ref",
            "recommended_folder_ref",
            "debt_document_ref",
            "current_balance",
            "review_reason",
            "effective_overdue_days",
        )
    }
    raw = json.dumps(relevant, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _daily_delta_report(
    report: dict[str, Any], state: dict[str, Any], *, snapshot_date: date
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    rows = [row for row in report.get("payload", []) if isinstance(row, dict)]
    current = {
        str(row.get("signal_key")): {
            "hash": _signal_content_hash(row),
            "last_seen": snapshot_date.isoformat(),
            "row": row,
        }
        for row in rows
        if row.get("signal_key")
    }
    previous = state.get("active_signals")
    if not isinstance(previous, dict):
        previous = {}
    changed_keys = [
        key
        for key, value in current.items()
        if key not in previous or (previous.get(key) or {}).get("hash") != value["hash"]
    ]
    closed_keys = sorted(set(previous) - set(current))
    delta_rows = [current[key]["row"] for key in changed_keys]
    delta = dict(report)
    delta["payload"] = delta_rows
    delta["report_revision"] = hashlib.sha256(
        json.dumps(
            {"date": snapshot_date.isoformat(), "changed": changed_keys, "closed": closed_keys},
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()[:12]
    delta["summary"] = {
        **dict(report.get("summary") or {}),
        "total_count": len(delta_rows),
        "new_or_changed_count": len(delta_rows),
        "closed_count": len(closed_keys),
        "remaining_actionable_count": len(current),
    }
    weekly = state.get("weekly_window")
    if not isinstance(weekly, dict):
        weekly = {}
    weekly.setdefault("start", snapshot_date.isoformat())
    weekly["new_or_changed"] = int(weekly.get("new_or_changed") or 0) + len(delta_rows)
    weekly["closed"] = int(weekly.get("closed") or 0) + len(closed_keys)
    weekly["last_date"] = snapshot_date.isoformat()
    updates = {"active_signals": current, "weekly_window": weekly}
    metrics = {
        "new_or_changed_count": len(delta_rows),
        "closed_count": len(closed_keys),
        "remaining_actionable_count": len(current),
    }
    return delta, updates, metrics


def _publication_suppression_reason(report: dict[str, Any], state: dict[str, Any]) -> str | None:
    source_status = _safe(report.get("source_status"))
    if source_status != "cache_ready":
        return f"source_not_cache_ready:{source_status or 'missing'}"
    current = _summary_int(report, "source_snapshot_count")
    previous = int(state.get("last_source_snapshot_count") or 0)
    difference = abs(current - previous)
    if previous and difference >= 100 and difference / previous > 0.5:
        return f"row_count_anomaly:{previous}->{current}"
    return None


def _save_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(state, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _safe(value: Any) -> str:
    return str(value or "").strip()


def _safe_path_chunk(value: Any) -> str:
    safe = _safe(value).replace("/", "-").replace("\\", "-")
    return safe or "unknown"


def _first_present(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def _format_dt(value: Any) -> str:
    if not value:
        return ""
    return str(value).replace("T", " ")[:19]


def _csv_number(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        return str(Decimal(str(value)).quantize(Decimal("0.01")))
    except (InvalidOperation, ValueError):
        return str(value)


def _summary_int(report: dict[str, Any], key: str) -> int:
    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    try:
        return int(summary.get(key) or 0)
    except (TypeError, ValueError):
        return 0


def _state_key(snapshot_date: date, *, status: str, report_revision: str) -> str:
    return f"{snapshot_date.isoformat()}|{status}|{report_revision}"


def _report_state_key(
    snapshot_date: date,
    *,
    status: str,
    report_revision: str,
    delivery_mode: str,
) -> str:
    scope = status if delivery_mode == DELIVERY_LEGACY else delivery_mode
    return _state_key(snapshot_date, status=scope, report_revision=report_revision)


def _finish_delivery(
    action: dict[str, Any],
    *,
    state: dict[str, Any],
    state_path: Path,
    snapshot_date: date,
    delivery_mode: str,
    dry_run: bool,
) -> dict[str, Any]:
    if (
        delivery_mode == DELIVERY_WEEKLY_SUMMARY
        and not dry_run
        and action.get("delivery_action") == "deliver"
    ):
        state["weekly_window"] = {
            "start": snapshot_date.isoformat(),
            "new_or_changed": 0,
            "closed": 0,
            "last_date": snapshot_date.isoformat(),
        }
        _save_state(state_path, state)
    return action


def _delivery_state_key(
    snapshot_date: date,
    *,
    status: str,
    report_revision: str,
    target: str,
) -> str:
    return f"{snapshot_date.isoformat()}|{status}|{report_revision}|{target}"


def _status_label(status: str) -> str:
    return {
        STATUS_MOVE_RECOMMENDED: "готовые рекомендации к переносу",
        STATUS_NEEDS_REVIEW: "ручная проверка",
        STATUS_OK: "совпадает по правилу",
        STATUS_NO_OVERDUE: "без просрочки",
    }.get(status, status)


def _review_reason_label(reason: Any) -> str:
    reason_key = _safe(reason)
    if not reason_key:
        return ""
    return REVIEW_REASON_LABELS.get(reason_key, reason_key)


def _preview_title(row: dict[str, Any], index: int) -> str:
    name = _safe(row.get("counterparty_name")) or _safe(row.get("counterparty_ref")) or "контрагент"
    code = _safe(row.get("counterparty_code"))
    if code:
        return f"{index}. {name} (код клиента: {code})"
    return f"{index}. {name}"


def _payment_term_source_label(source: Any) -> str:
    source_key = _safe(source)
    if source_key == PAYMENT_TERM_SOURCE_FALLBACK:
        return "расчетно 7 дней только для отчета"
    return source_key


def _document_structure_status_label(status: Any) -> str:
    status_key = _safe(status)
    if not status_key:
        return ""
    return DOCUMENT_STRUCTURE_STATUS_LABELS.get(status_key, status_key)


def _statement_rule_label(rule: Any) -> str:
    rule_key = _safe(rule)
    if not rule_key:
        return ""
    return STATEMENT_RULE_LABELS.get(rule_key, rule_key)


def _format_linked_documents(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    labels: list[str] = []
    for item in value:
        if not isinstance(item, dict):
            continue
        doc_type = _safe(item.get("document_type"))
        number = _safe(item.get("document_number"))
        doc_date = _format_dt(item.get("document_date"))
        amount = _csv_number(item.get("amount"))
        basis_kind = _safe(item.get("basis_kind"))
        label = " ".join(chunk for chunk in (doc_type, number) if chunk)
        if doc_date:
            label = f"{label} от {doc_date}" if label else doc_date
        if amount:
            label = f"{label} ({amount} ₽)" if label else f"{amount} ₽"
        if basis_kind:
            label = f"{label}, основание: {basis_kind}" if label else f"основание: {basis_kind}"
        if label:
            labels.append(label)
    return " | ".join(labels)


def _format_open_debt_documents(value: Any, *, limit: int | None = None) -> str:
    if not isinstance(value, list):
        return ""
    labels: list[str] = []
    for item in value:
        if not isinstance(item, dict):
            continue
        number = _safe(item.get("document_number"))
        doc_date = _format_dt(item.get("document_date"))
        amount = _csv_number(item.get("open_amount"))
        department = _safe(item.get("debt_department_name"))
        responsible = _safe(
            _first_present(
                item.get("document_responsible_name"),
                item.get("manager_name"),
            )
        )
        rule = _statement_rule_label(item.get("statement_selection_rule"))
        balance_after = _csv_number(item.get("statement_balance_after"))
        label = " ".join(chunk for chunk in ("Реализация", number) if chunk)
        if doc_date:
            label = f"{label} от {doc_date}" if label else doc_date
        if amount:
            label = f"{label}, остаток {amount} ₽" if label else f"остаток {amount} ₽"
        if department:
            label = f"{label}, подразделение: {department}" if label else department
        if responsible:
            label = (
                f"{label}, ответственный РТУ: {responsible}"
                if label
                else f"ответственный РТУ: {responsible}"
            )
        if balance_after:
            label = (
                f"{label}, конечный остаток: {balance_after} ₽"
                if label
                else f"конечный остаток: {balance_after} ₽"
            )
        if rule:
            label = f"{label}, правило: {rule}" if label else f"правило: {rule}"
        if label:
            labels.append(label)
        if limit is not None and len(labels) >= limit:
            break
    if limit is not None and len(value) > limit:
        labels.append(f"еще {len(value) - limit} в CSV")
    return " | ".join(labels)


def _bitrix_webhook_base(env: dict[str, str]) -> str:
    for key in BITRIX_WEBHOOK_ENV_KEYS:
        value = _safe(env.get(key)).rstrip("/")
        if value:
            return value
    raise RuntimeError(
        "Missing Bitrix webhook base: set one of " + ", ".join(BITRIX_WEBHOOK_ENV_KEYS)
    )


def _bitrix_call(
    base_url: str,
    method: str,
    payload: dict[str, Any],
    *,
    timeout: int,
) -> dict[str, Any]:
    request = urllib.request.Request(
        base_url.rstrip("/") + f"/{method}.json",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        method="POST",
    )
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    try:
        with opener.open(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        try:
            data = json.loads(raw)
        except json.JSONDecodeError as decode_error:
            raise RuntimeError(
                f"Bitrix API HTTP {exc.code} for {method}: {raw[:300]}"
            ) from decode_error
        raise RuntimeError(
            f"Bitrix API error for {method}: {data.get('error')} "
            f"{data.get('error_description', '')}"
        ) from exc

    data = json.loads(raw)
    if "error" in data:
        raise RuntimeError(
            f"Bitrix API error for {method}: {data.get('error')} "
            f"{data.get('error_description', '')}"
        )
    return data


def _bitrix_form_call(base_url: str, method: str, params: list[tuple[str, str]]) -> dict[str, Any]:
    request = urllib.request.Request(
        base_url.rstrip("/") + f"/{method}.json",
        data=urllib.parse.urlencode(params, doseq=True).encode("utf-8"),
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    with opener.open(request, timeout=60) as response:
        data = json.loads(response.read().decode("utf-8"))
    if "error" in data:
        raise RuntimeError(
            f"Bitrix API error for {method}: {data.get('error')} "
            f"{data.get('error_description', '')}"
        )
    return data


def _to_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def post_bitrix_task_comment(
    base_url: str,
    *,
    task_id: int,
    message: str,
    timeout: int,
) -> int:
    data = _bitrix_call(
        base_url,
        "task.commentitem.add",
        {"TASKID": task_id, "FIELDS": {"POST_MESSAGE": message}},
        timeout=timeout,
    )
    return int(data.get("result") or 0)


def upload_bitrix_disk_file(
    base_url: str,
    *,
    folder_id: int,
    file_path: Path,
) -> int:
    encoded = base64.b64encode(file_path.read_bytes()).decode("ascii")
    data = _bitrix_form_call(
        base_url,
        "disk.folder.uploadfile",
        [
            ("id", str(folder_id)),
            ("data[NAME]", file_path.name),
            ("fileContent[0]", file_path.name),
            ("fileContent[1]", encoded),
            ("generateUniqueName", "true"),
        ],
    )
    result = data.get("result") or {}
    file_object_id = _to_int(result.get("ID") if isinstance(result, dict) else result)
    if file_object_id is None:
        raise RuntimeError("Bitrix disk.folder.uploadfile returned empty object id")
    return file_object_id


def attach_bitrix_file_to_task(
    base_url: str,
    *,
    task_id: int,
    file_object_id: int,
) -> int:
    data = _bitrix_form_call(
        base_url,
        "tasks.task.files.attach",
        [("taskId", str(task_id)), ("fileId", str(file_object_id))],
    )
    result = data.get("result") or {}
    attachment_id = _to_int(result.get("attachmentId") if isinstance(result, dict) else result)
    if attachment_id is None:
        raise RuntimeError("Bitrix tasks.task.files.attach returned empty attachment id")
    return attachment_id


def deliver_bitrix_task_attachment(
    base_url: str,
    *,
    folder_id: int,
    task_id: int,
    file_path: Path,
) -> dict[str, int]:
    file_object_id = upload_bitrix_disk_file(
        base_url,
        folder_id=folder_id,
        file_path=file_path,
    )
    attachment_id = attach_bitrix_file_to_task(
        base_url,
        task_id=task_id,
        file_object_id=file_object_id,
    )
    return {
        "bitrix_file_object_id": file_object_id,
        "bitrix_attachment_id": attachment_id,
    }


def export_recommendations_csv(report: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    rows = report.get("payload") if isinstance(report.get("payload"), list) else []

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(["Отчет", "Контроль папок контрагентов по просроченной дебиторке"])
        writer.writerow(["Дата снапшота", report.get("as_of") or report.get("snapshot_date")])
        writer.writerow(["Ревизия", report.get("report_revision")])
        writer.writerow(["Всего в выгрузке", summary.get("total_count", len(rows))])
        writer.writerow(["К переносу", summary.get("move_recommended_count", 0)])
        writer.writerow(["На ручную проверку", summary.get("needs_review_count", 0)])
        writer.writerow([])
        writer.writerow(
            [
                "Контрагент",
                "Код клиента",
                "Текущая папка",
                "Рекомендуемая папка",
                "Подразделение долга",
                "Сумма",
                "Открытые документы по ведомостной логике 1С",
                "Основной открытый документ",
                "Ответственный РТУ",
                "Правило выбора источника",
                "Конечный остаток ведомости",
                "Сегмент ведомости",
                "Документ витрины дебиторки",
                "Дата долга",
                "Просрочка дней",
                "Глубина кредита",
                "Дата просрочки",
                "Источник срока оплаты",
                "Статус проверки структуры",
                "Остаток по структуре",
                "Сумма реализации по структуре",
                "Закрывающие документы по структуре",
                "Связанный заказ",
                "Статус",
                "Причина проверки",
                "Причина проверки код",
                "Менеджер долга",
                "Текущий менеджер",
                "Контрагент ref",
                "Документ ref",
            ]
        )
        for item in rows:
            if not isinstance(item, dict):
                continue
            debt_document_label = " ".join(
                chunk
                for chunk in (
                    _safe(item.get("debt_document_number")),
                    _safe(item.get("debt_document_ref")),
                )
                if chunk
            )
            legacy_document_label = " ".join(
                chunk
                for chunk in (
                    _safe(item.get("origin_document_number")),
                    _safe(item.get("origin_document_ref")),
                )
                if chunk
            )
            writer.writerow(
                [
                    item.get("counterparty_name") or item.get("counterparty_ref"),
                    item.get("counterparty_code"),
                    item.get("current_folder_name"),
                    item.get("recommended_folder_name"),
                    item.get("debt_department_name"),
                    _csv_number(item.get("current_balance")),
                    _format_open_debt_documents(item.get("open_debt_documents")),
                    debt_document_label,
                    _first_present(
                        item.get("debt_document_responsible_name"),
                        item.get("origin_manager_name"),
                    ),
                    _statement_rule_label(
                        (item.get("open_debt_documents") or [{}])[0].get("statement_selection_rule")
                        if isinstance(item.get("open_debt_documents"), list)
                        and item.get("open_debt_documents")
                        else ""
                    ),
                    _csv_number(item.get("statement_balance_after")),
                    "–".join(
                        chunk
                        for chunk in (
                            _safe(item.get("statement_segment_start_row")),
                            _safe(item.get("statement_segment_end_row")),
                        )
                        if chunk
                    ),
                    legacy_document_label,
                    _format_dt(
                        _first_present(
                            item.get("debt_document_date"),
                            item.get("origin_document_date"),
                        )
                    ),
                    _first_present(item.get("effective_overdue_days"), item.get("overdue_days")),
                    _first_present(
                        item.get("effective_credit_depth_days"),
                        item.get("credit_depth_days"),
                    ),
                    _format_dt(
                        _first_present(item.get("effective_due_date"), item.get("due_date"))
                    ),
                    _payment_term_source_label(
                        _first_present(
                            item.get("effective_payment_term_source"),
                            item.get("payment_term_source"),
                        )
                    ),
                    _document_structure_status_label(item.get("document_structure_status")),
                    _csv_number(item.get("document_structure_open_amount")),
                    _csv_number(item.get("document_structure_sale_amount")),
                    _format_linked_documents(item.get("document_structure_linked_documents")),
                    " от ".join(
                        chunk
                        for chunk in (
                            _safe(item.get("document_structure_order_number")),
                            _format_dt(item.get("document_structure_order_date")),
                        )
                        if chunk
                    ),
                    item.get("status"),
                    _review_reason_label(item.get("review_reason")),
                    item.get("review_reason"),
                    item.get("origin_manager_name"),
                    item.get("current_manager_name"),
                    item.get("counterparty_ref"),
                    item.get("debt_document_ref") or item.get("origin_document_ref"),
                ]
            )

    return output_path


def _excel_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(Decimal(str(value)).quantize(Decimal("0.01")))
    except (InvalidOperation, ValueError):
        return None


def _primary_statement_rule(item: dict[str, Any]) -> str:
    open_debt_documents = item.get("open_debt_documents")
    if isinstance(open_debt_documents, list) and open_debt_documents:
        first = open_debt_documents[0]
        if isinstance(first, dict):
            return _statement_rule_label(first.get("statement_selection_rule"))
    return _statement_rule_label(item.get("statement_selection_rule"))


def _statement_segment_label(item: dict[str, Any]) -> str:
    return "–".join(
        chunk
        for chunk in (
            _safe(item.get("statement_segment_start_row")),
            _safe(item.get("statement_segment_end_row")),
        )
        if chunk
    )


def export_recommendations_xlsx(report: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    data_sheet = workbook.create_sheet("Проверка")

    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    rows = report.get("payload") if isinstance(report.get("payload"), list) else []

    summary_rows = [
        ("Отчет", "Контроль папок контрагентов по просроченной дебиторке"),
        ("Дата снапшота", report.get("as_of") or report.get("snapshot_date")),
        ("Ревизия", report.get("report_revision")),
        ("Всего в выгрузке", summary.get("total_count", len(rows))),
        ("К переносу", summary.get("move_recommended_count", 0)),
        ("На ручную проверку", summary.get("needs_review_count", 0)),
        ("Скрыто мелких долгов", summary.get("below_min_balance_count", 0)),
    ]
    for row in summary_rows:
        summary_sheet.append(list(row))

    headers = [
        "Контрагент",
        "Код клиента",
        "Текущая папка",
        "Рекомендуемая папка",
        "Подразделение долга",
        "Сумма",
        "Открытые документы по ведомостной логике 1С",
        "Основной открытый документ",
        "Ответственный РТУ",
        "Правило выбора источника",
        "Конечный остаток ведомости",
        "Сегмент ведомости",
        "Документ витрины дебиторки",
        "Дата долга",
        "Просрочка дней",
        "Глубина кредита",
        "Дата просрочки",
        "Источник срока оплаты",
        "Статус проверки структуры",
        "Остаток по структуре",
        "Сумма реализации по структуре",
        "Закрывающие документы по структуре",
        "Связанный заказ",
        "Статус",
        "Причина проверки",
        "Причина проверки код",
        "Менеджер долга",
        "Текущий менеджер",
        "Контрагент ref",
        "Документ ref",
    ]
    data_sheet.append(headers)
    for item in rows:
        if not isinstance(item, dict):
            continue
        debt_document_label = " ".join(
            chunk
            for chunk in (
                _safe(item.get("debt_document_number")),
                _safe(item.get("debt_document_ref")),
            )
            if chunk
        )
        legacy_document_label = " ".join(
            chunk
            for chunk in (
                _safe(item.get("origin_document_number")),
                _safe(item.get("origin_document_ref")),
            )
            if chunk
        )
        data_sheet.append(
            [
                item.get("counterparty_name") or item.get("counterparty_ref"),
                item.get("counterparty_code"),
                item.get("current_folder_name"),
                item.get("recommended_folder_name"),
                item.get("debt_department_name"),
                _excel_number(item.get("current_balance")),
                _format_open_debt_documents(item.get("open_debt_documents")),
                debt_document_label,
                _first_present(
                    item.get("debt_document_responsible_name"),
                    item.get("origin_manager_name"),
                ),
                _primary_statement_rule(item),
                _excel_number(item.get("statement_balance_after")),
                _statement_segment_label(item),
                legacy_document_label,
                _format_dt(
                    _first_present(
                        item.get("debt_document_date"),
                        item.get("origin_document_date"),
                    )
                ),
                _first_present(item.get("effective_overdue_days"), item.get("overdue_days")),
                _first_present(
                    item.get("effective_credit_depth_days"),
                    item.get("credit_depth_days"),
                ),
                _format_dt(_first_present(item.get("effective_due_date"), item.get("due_date"))),
                _payment_term_source_label(
                    _first_present(
                        item.get("effective_payment_term_source"),
                        item.get("payment_term_source"),
                    )
                ),
                _document_structure_status_label(item.get("document_structure_status")),
                _excel_number(item.get("document_structure_open_amount")),
                _excel_number(item.get("document_structure_sale_amount")),
                _format_linked_documents(item.get("document_structure_linked_documents")),
                " от ".join(
                    chunk
                    for chunk in (
                        _safe(item.get("document_structure_order_number")),
                        _format_dt(item.get("document_structure_order_date")),
                    )
                    if chunk
                ),
                item.get("status"),
                _review_reason_label(item.get("review_reason")),
                item.get("review_reason"),
                item.get("origin_manager_name"),
                item.get("current_manager_name"),
                item.get("counterparty_ref"),
                item.get("debt_document_ref") or item.get("origin_document_ref"),
            ]
        )

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    stripe_fill = PatternFill(fill_type="solid", fgColor="F7FBFF")
    header_font = Font(bold=True, color="FFFFFF")
    border_side = Side(style="thin", color="D9E2F3")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for sheet in (summary_sheet, data_sheet):
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_index in range(2, sheet.max_row + 1):
            use_stripe = row_index % 2 == 0
            for column_index in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                cell.border = cell_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if use_stripe:
                    cell.fill = stripe_fill
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="top")
        if sheet.max_row > 1:
            sheet.freeze_panes = "A2"
        sheet.sheet_view.zoomScale = 90
        for column_index in range(1, sheet.max_column + 1):
            width = 10
            for row_index in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row_index, column=column_index).value
                if value is None:
                    continue
                width = max(width, len(str(value)) + 2)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(width, 48)

    if data_sheet.max_row > 1:
        data_sheet.auto_filter.ref = (
            f"A1:{get_column_letter(data_sheet.max_column)}{data_sheet.max_row}"
        )

    workbook.save(output_path)
    return output_path


def render_bitrix_comment(
    report: dict[str, Any],
    *,
    artifact_path: str,
    xlsx_path: str | None = None,
    xlsx_attached: bool = False,
    status: str,
) -> str:
    rows = report.get("payload") if isinstance(report.get("payload"), list) else []
    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    as_of = _safe(report.get("as_of") or report.get("snapshot_date"))
    report_revision = _safe(report.get("report_revision"))
    total_count = _summary_int(report, "total_count")
    move_count = _summary_int(report, "move_recommended_count")
    review_count = _summary_int(report, "needs_review_count")
    below_min_count = _summary_int(report, "below_min_balance_count")
    min_balance = summary.get("min_recommendation_balance")

    if status == DELIVERY_WEEKLY_SUMMARY:
        reasons = summary.get("review_reason_counts") or {}
        lines = [
            "Добрый день.",
            "",
            f"📊 Недельная сводка по закреплению клиентов за {as_of}",
            f"• Новые/изменённые: {summary.get('weekly_new_or_changed', 0)}",
            f"• Закрытые: {summary.get('weekly_closed', 0)}",
            f"• Осталось actionable: {summary.get('actionable_count', 0)}",
            f"• Бизнес-проверка: {summary.get('business_review_count', 0)}",
            f"• Ошибки данных: {summary.get('data_quality_count', 0)}",
        ]
        if isinstance(reasons, dict) and reasons:
            lines.extend(["", "Основные причины качества данных:"])
            for reason, count in sorted(reasons.items(), key=lambda item: -int(item[1]))[:5]:
                lines.append(f"• {_review_reason_label(reason)}: {count}")
        return "\n".join(lines)

    if status == DELIVERY_DAILY_DELTA:
        lines = [
            "Добрый день.",
            "",
            f"📌 Новые и изменённые подтверждённые сигналы за {as_of}",
            f"• Новые/изменённые: {summary.get('new_or_changed_count', total_count)}",
            f"• Закрытые: {summary.get('closed_count', 0)}",
            f"• Осталось actionable: {summary.get('remaining_actionable_count', 0)}",
            "• Автоматический перенос папок не выполнялся.",
            f"• 📎 Excel: {'прикреплен к задаче' if xlsx_attached else 'не требуется'}",
            f"• CSV audit: {artifact_path}",
        ]
    else:
        lines = [
            "Добрый день.",
            "",
            f"📌 Отчет по контролю папок контрагентов за {as_of}",
            f"Тип отчета: {_status_label(status)}",
            "",
            "📊 Сводка",
            f"• Всего строк: {total_count}",
            f"• ⏸️ Готовых рекомендаций к переносу: {move_count}",
            f"• 🔎 На ручную проверку: {review_count}",
            f"• 🧹 Скрыто мелких долгов ниже {min_balance or 500} ₽: {below_min_count}",
            f"• Ревизия: {report_revision}",
            f"• 📎 Excel: {'прикреплен к задаче' if xlsx_attached else 'не прикреплен'}",
            f"• CSV fallback: {artifact_path}",
        ]
    if xlsx_path:
        lines.append(f"• XLSX fallback: {xlsx_path}")
    lines.append("")
    if status == STATUS_MOVE_RECOMMENDED:
        lines.extend(
            [
                "⏸️ Готовые рекомендации временно выключены.",
                "Источник долга сначала должен быть подтвержден структурой документа 1С.",
                "Автоматический перенос в 1С не выполнялся.",
            ]
        )
    elif status == STATUS_NEEDS_REVIEW:
        lines.extend(
            [
                "🔎 Это не рекомендации к переносу.",
                "Нужно вручную проверить источник долга, документ или подразделение.",
            ]
        )
        review_reason_counts = summary.get("review_reason_counts")
        if isinstance(review_reason_counts, dict) and review_reason_counts:
            lines.extend(["", "⚠️ Причины ручной проверки"])
            for reason, count in sorted(review_reason_counts.items()):
                lines.append(f"• {_review_reason_label(reason)}: {count}")
    elif status != DELIVERY_DAILY_DELTA:
        lines.append("ℹ️ Отчет служебный, без автоматических изменений в 1С.")

    preview_limit = 20 if status == DELIVERY_DAILY_DELTA else 10
    preview_rows = [row for row in rows if isinstance(row, dict)][:preview_limit]
    if preview_rows:
        lines.extend(["", "🧾 Первые строки"])
        for index, row in enumerate(preview_rows, start=1):
            current_folder = _safe(row.get("current_folder_name"))
            recommended_folder = _safe(row.get("recommended_folder_name"))
            department = _safe(row.get("debt_department_name"))
            document_number = _safe(row.get("debt_document_number"))
            document_date = _format_dt(row.get("debt_document_date"))
            document_responsible = _safe(
                _first_present(
                    row.get("debt_document_responsible_name"),
                    row.get("origin_manager_name"),
                )
            )
            reason = _review_reason_label(row.get("review_reason"))
            lines.extend(["", _preview_title(row, index)])
            lines.append(f"💰 Сумма: {_csv_number(row.get('current_balance'))} ₽")
            if current_folder:
                folder_line = f"📁 Текущая папка: {current_folder}"
                if recommended_folder:
                    folder_line += f" → {recommended_folder}"
                lines.append(folder_line)
            elif recommended_folder:
                lines.append(f"📁 Рекомендуемая папка: {recommended_folder}")
            if department:
                lines.append(f"🏬 Подразделение долга: {department}")
            open_documents = _format_open_debt_documents(
                row.get("open_debt_documents"),
                limit=3,
            )
            if open_documents:
                lines.append("🧾 Открытые документы по ведомостной логике 1С: " f"{open_documents}")
            elif document_number or document_date:
                document_parts = [chunk for chunk in (document_number, document_date) if chunk]
                lines.append(
                    "🧾 Открытый документ по ведомостной логике 1С: "
                    f"{' от '.join(document_parts)}"
                )
            if document_responsible:
                lines.append(f"👤 Ответственный РТУ: {document_responsible}")
            open_debt_documents = row.get("open_debt_documents")
            if isinstance(open_debt_documents, list) and open_debt_documents:
                first_open_doc = open_debt_documents[0]
                if isinstance(first_open_doc, dict):
                    statement_rule = _statement_rule_label(
                        first_open_doc.get("statement_selection_rule")
                    )
                    if statement_rule:
                        lines.append(f"🧭 Правило выбора: {statement_rule}")
            balance_after = _csv_number(row.get("statement_balance_after"))
            segment_label = _statement_segment_label(row)
            if balance_after:
                lines.append(f"📈 Конечный остаток ведомости: {balance_after} ₽")
            if segment_label:
                lines.append(f"📍 Сегмент ведомости: строки {segment_label}")
            structure_status = _document_structure_status_label(
                row.get("document_structure_status")
            )
            if structure_status:
                structure_line = f"🔗 Структура 1С: {structure_status}"
                open_amount = _csv_number(row.get("document_structure_open_amount"))
                if open_amount:
                    structure_line += f", остаток {open_amount} ₽"
                lines.append(structure_line)
            order_label = " от ".join(
                chunk
                for chunk in (
                    _safe(row.get("document_structure_order_number")),
                    _format_dt(row.get("document_structure_order_date")),
                )
                if chunk
            )
            if order_label:
                lines.append(f"🧷 Связанный заказ: {order_label}")
            linked_documents = _format_linked_documents(
                row.get("document_structure_linked_documents")
            )
            if linked_documents:
                lines.append(f"💳 Закрывающие документы: {linked_documents}")
            if reason:
                lines.append(f"⚠️ Причина: {reason}")
        if len(rows) > len(preview_rows):
            lines.extend(["", f"Еще строк в CSV: {len(rows) - len(preview_rows)}."])

    return "\n".join(lines)


def _maybe_deliver_bitrix(
    *,
    action: dict[str, Any],
    report: dict[str, Any],
    state: dict[str, Any],
    state_path: Path,
    snapshot_date: date,
    status: str,
    report_revision: str,
    artifact_path: str,
    xlsx_path: str | None,
    task_id: int | None,
    notify_empty: bool,
    dry_run: bool,
    force: bool,
    deliver_comment: Callable[[int, str], int] | None,
    deliver_attachment: Callable[[int, Path], dict[str, int]] | None = None,
) -> dict[str, Any]:
    if task_id is None:
        return action

    total_count = _summary_int(report, "total_count")
    target = f"bitrix-task:{task_id}"
    delivery_key = _delivery_state_key(
        snapshot_date,
        status=status,
        report_revision=report_revision,
        target=target,
    )
    action["delivery_channel"] = "bitrix_task_comment"
    action["delivery_target"] = target

    if total_count <= 0 and not notify_empty:
        action["delivery_action"] = "skip_no_rows"
        action["delivered"] = 0
        return action

    deliveries = state.setdefault("deliveries", {})
    current = deliveries.get(delivery_key)
    if isinstance(current, dict) and current.get("delivery_status") == "delivered" and not force:
        action["delivery_action"] = "noop"
        action["delivery_reason"] = "already_delivered"
        action["bitrix_comment_id"] = current.get("bitrix_comment_id")
        action["delivered"] = 0
        return action

    if dry_run or deliver_comment is None:
        message = render_bitrix_comment(
            report,
            artifact_path=artifact_path,
            xlsx_path=xlsx_path,
            xlsx_attached=False,
            status=status,
        )
        action["delivery_action"] = "dry_run"
        action["delivery_message"] = message
        if xlsx_path:
            action["xlsx_path"] = xlsx_path
        action["delivered"] = 0
        return action

    xlsx_attached = False
    attachment_result: dict[str, int] = {}
    if xlsx_path and deliver_attachment is not None:
        attachment_result = deliver_attachment(task_id, Path(xlsx_path))
        xlsx_attached = True
        action["delivery_attachment_action"] = "attach"
        action.update(attachment_result)
    elif xlsx_path:
        action["delivery_attachment_action"] = "skip_missing_disk_folder"

    message = render_bitrix_comment(
        report,
        artifact_path=artifact_path,
        xlsx_path=xlsx_path,
        xlsx_attached=xlsx_attached,
        status=status,
    )
    comment_id = deliver_comment(task_id, message)
    deliveries[delivery_key] = {
        "delivery_status": "delivered",
        "date": snapshot_date.isoformat(),
        "report_revision": report_revision,
        "status_filter": status,
        "artifact_path": artifact_path,
        "xlsx_path": xlsx_path,
        "target": target,
        "bitrix_comment_id": comment_id,
        **attachment_result,
        "delivered_at": _utcnow().isoformat(),
    }
    _save_state(state_path, state)
    action["delivery_action"] = "deliver"
    action["bitrix_comment_id"] = comment_id
    action["delivered"] = 1
    return action


def sync_counterparty_folder_recommendations(
    *,
    fetch_json: Callable[[str, dict[str, str]], Any],
    snapshot_date: date,
    state_path: Path,
    artifact_dir: Path,
    status: str = STATUS_MOVE_RECOMMENDED,
    limit: int | None = None,
    dry_run: bool = False,
    force: bool = False,
    bitrix_task_id: int | None = None,
    notify_empty: bool = False,
    deliver_comment: Callable[[int, str], int] | None = None,
    deliver_attachment: Callable[[int, Path], dict[str, int]] | None = None,
    delivery_mode: str = DELIVERY_LEGACY,
) -> dict[str, Any]:
    params = {"date": snapshot_date.isoformat()}
    if delivery_mode == DELIVERY_LEGACY:
        params["status"] = status
    else:
        params["queue"] = QUEUE_ACTIONABLE if delivery_mode == DELIVERY_DAILY_DELTA else QUEUE_ALL
    if limit is not None:
        params["limit"] = str(limit)

    try:
        report = fetch_json(REPORT_ENDPOINT, params)
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, ValueError) as error:
        return {
            "status": "error",
            "date": snapshot_date.isoformat(),
            "action": "failed",
            "error": str(error),
            "exported": 0,
            "noop": 0,
            "failed": 1,
            "publication_suppressed": True,
            "suppression_reason": "source_timeout_or_error",
        }

    if not isinstance(report, dict):
        return {
            "status": "error",
            "date": snapshot_date.isoformat(),
            "action": "failed",
            "error": "source returned non-object payload",
            "exported": 0,
            "noop": 0,
            "failed": 1,
        }

    report_revision = _safe(report.get("report_revision"))
    if not report_revision:
        return {
            "status": "error",
            "date": snapshot_date.isoformat(),
            "action": "failed",
            "error": "source payload has no report_revision",
            "exported": 0,
            "noop": 0,
            "failed": 1,
        }

    state = _load_state(state_path)
    delivery_report = report
    state_updates: dict[str, Any] = {}
    delta_metrics: dict[str, Any] = {}
    delivery_status = status
    if delivery_mode == DELIVERY_DAILY_DELTA:
        delivery_report, state_updates, delta_metrics = _daily_delta_report(
            report, state, snapshot_date=snapshot_date
        )
        delivery_status = DELIVERY_DAILY_DELTA
    elif delivery_mode == DELIVERY_WEEKLY_SUMMARY:
        weekly = state.get("weekly_window") if isinstance(state.get("weekly_window"), dict) else {}
        delivery_report = {
            **report,
            "payload": [],
            "report_revision": f"week-{snapshot_date.isoformat()}",
            "summary": {
                **dict(report.get("summary") or {}),
                "total_count": 1,
                "weekly_new_or_changed": int((weekly or {}).get("new_or_changed") or 0),
                "weekly_closed": int((weekly or {}).get("closed") or 0),
            },
        }
        delivery_status = DELIVERY_WEEKLY_SUMMARY
        notify_empty = True
    suppression_reason = (
        _publication_suppression_reason(report, state) if delivery_mode != DELIVERY_LEGACY else None
    )
    key = _report_state_key(
        snapshot_date,
        status=status,
        report_revision=report_revision,
        delivery_mode=delivery_mode,
    )
    current = (state.get("reports") or {}).get(key)
    if isinstance(current, dict) and current.get("export_status") == "exported" and not force:
        current_xlsx_path = _safe(current.get("xlsx_path")) or None
        action = {
            "status": "ok",
            "date": snapshot_date.isoformat(),
            "report_revision": report_revision,
            "action": "noop",
            "reason": "already_exported",
            "artifact_path": current.get("artifact_path"),
            "xlsx_path": current_xlsx_path,
            "status_filter": status,
            "total_count": _summary_int(report, "total_count"),
            "move_recommended_count": _summary_int(report, "move_recommended_count"),
            "needs_review_count": _summary_int(report, "needs_review_count"),
            "below_min_balance_count": _summary_int(report, "below_min_balance_count"),
            "exported": 0,
            "noop": 1,
            "failed": 0,
        }
        if suppression_reason:
            action.update(
                {
                    "delivery_action": "suppressed",
                    "publication_suppressed": True,
                    "suppression_reason": suppression_reason,
                }
            )
            return action
        delivered_action = _maybe_deliver_bitrix(
            action=action,
            report=delivery_report,
            state=state,
            state_path=state_path,
            snapshot_date=snapshot_date,
            status=delivery_status,
            report_revision=str(delivery_report.get("report_revision") or report_revision),
            artifact_path=str(current.get("artifact_path") or ""),
            xlsx_path=(
                current_xlsx_path
                if delivery_mode != DELIVERY_DAILY_DELTA
                or _summary_int(delivery_report, "total_count") > 20
                else None
            ),
            task_id=bitrix_task_id,
            notify_empty=notify_empty,
            dry_run=dry_run,
            force=force,
            deliver_comment=deliver_comment,
            deliver_attachment=deliver_attachment,
        )
        return _finish_delivery(
            delivered_action,
            state=state,
            state_path=state_path,
            snapshot_date=snapshot_date,
            delivery_mode=delivery_mode,
            dry_run=dry_run,
        )

    artifact_scope = status if delivery_mode == DELIVERY_LEGACY else delivery_mode
    artifact_path = (
        artifact_dir
        / snapshot_date.isoformat()
        / f"counterparty-folder-{_safe_path_chunk(artifact_scope)}-{report_revision}.csv"
    )
    xlsx_path = artifact_path.with_suffix(".xlsx")
    action = {
        "status": "ok",
        "date": snapshot_date.isoformat(),
        "report_revision": report_revision,
        "action": "dry_run" if dry_run else "export",
        "artifact_path": str(artifact_path),
        "xlsx_path": str(xlsx_path),
        "status_filter": status,
        "limit": limit,
        "source_snapshot_count": _summary_int(report, "source_snapshot_count"),
        "total_count": _summary_int(report, "total_count"),
        "move_recommended_count": _summary_int(report, "move_recommended_count"),
        "needs_review_count": _summary_int(report, "needs_review_count"),
        "below_min_balance_count": _summary_int(report, "below_min_balance_count"),
        "exported": 0 if dry_run else 1,
        "noop": 0,
        "failed": 0,
    }
    if dry_run:
        if suppression_reason:
            action.update(
                {
                    "delivery_action": "suppressed",
                    "publication_suppressed": True,
                    "suppression_reason": suppression_reason,
                }
            )
            return action
        delivered_action = _maybe_deliver_bitrix(
            action=action,
            report=delivery_report,
            state=state,
            state_path=state_path,
            snapshot_date=snapshot_date,
            status=delivery_status,
            report_revision=str(delivery_report.get("report_revision") or report_revision),
            artifact_path=str(artifact_path),
            xlsx_path=(
                str(xlsx_path)
                if delivery_mode != DELIVERY_DAILY_DELTA
                or _summary_int(delivery_report, "total_count") > 20
                else None
            ),
            task_id=bitrix_task_id,
            notify_empty=notify_empty,
            dry_run=dry_run,
            force=force,
            deliver_comment=deliver_comment,
            deliver_attachment=deliver_attachment,
        )
        return _finish_delivery(
            delivered_action,
            state=state,
            state_path=state_path,
            snapshot_date=snapshot_date,
            delivery_mode=delivery_mode,
            dry_run=dry_run,
        )

    export_recommendations_csv(report, artifact_path)
    export_recommendations_xlsx(report, xlsx_path)
    state.setdefault("reports", {})[key] = {
        "export_status": "exported",
        "date": snapshot_date.isoformat(),
        "report_revision": report_revision,
        "status_filter": status,
        "limit": limit,
        "artifact_path": str(artifact_path),
        "xlsx_path": str(xlsx_path),
        "exported_at": _utcnow().isoformat(),
    }
    if not suppression_reason:
        state.update(state_updates)
        state["last_source_snapshot_count"] = _summary_int(report, "source_snapshot_count")
    _save_state(state_path, state)
    action.update(delta_metrics)
    if suppression_reason:
        action.update(
            {
                "delivery_action": "suppressed",
                "publication_suppressed": True,
                "suppression_reason": suppression_reason,
            }
        )
        return action
    delivered_action = _maybe_deliver_bitrix(
        action=action,
        report=delivery_report,
        state=state,
        state_path=state_path,
        snapshot_date=snapshot_date,
        status=delivery_status,
        report_revision=str(delivery_report.get("report_revision") or report_revision),
        artifact_path=str(artifact_path),
        xlsx_path=(
            str(xlsx_path)
            if delivery_mode != DELIVERY_DAILY_DELTA
            or _summary_int(delivery_report, "total_count") > 20
            else None
        ),
        task_id=bitrix_task_id,
        notify_empty=notify_empty,
        dry_run=dry_run,
        force=force,
        deliver_comment=deliver_comment,
        deliver_attachment=deliver_attachment,
    )
    return _finish_delivery(
        delivered_action,
        state=state,
        state_path=state_path,
        snapshot_date=snapshot_date,
        delivery_mode=delivery_mode,
        dry_run=dry_run,
    )


def render_summary(summary: dict[str, Any]) -> str:
    lines = [
        f"counterparty_folder_recommendations_from_a: {summary.get('status', 'unknown')}",
        f"Дата: {summary.get('date', '-')}",
        f"Действие: {summary.get('action', '-')}",
        (
            "exported: {exported}; noop: {noop}; failed: {failed}; "
            "к переносу: {move}; ручная проверка: {review}; "
            "скрыто мелких: {below_min}; всего строк: {total}"
        ).format(
            exported=summary.get("exported", 0),
            noop=summary.get("noop", 0),
            failed=summary.get("failed", 0),
            move=summary.get("move_recommended_count", 0),
            review=summary.get("needs_review_count", 0),
            below_min=summary.get("below_min_balance_count", 0),
            total=summary.get("total_count", 0),
        ),
    ]
    if summary.get("report_revision"):
        lines.append(f"Ревизия: {summary['report_revision']}")
    if summary.get("artifact_path"):
        lines.append(f"Файл: {summary['artifact_path']}")
    if summary.get("xlsx_path"):
        lines.append(f"Excel: {summary['xlsx_path']}")
    if summary.get("reason"):
        lines.append(f"Причина: {summary['reason']}")
    if summary.get("delivery_action"):
        lines.append(f"Доставка: {summary['delivery_action']}")
    if summary.get("suppression_reason"):
        lines.append(f"Публикация подавлена: {summary['suppression_reason']}")
    if summary.get("delivery_attachment_action"):
        lines.append(f"Вложение: {summary['delivery_attachment_action']}")
    if summary.get("bitrix_comment_id"):
        lines.append(f"Комментарий Bitrix: {summary['bitrix_comment_id']}")
    if summary.get("bitrix_attachment_id"):
        lines.append(f"Вложение Bitrix: {summary['bitrix_attachment_id']}")
    if summary.get("error"):
        lines.append(f"Ошибка: {summary['error']}")
    return "\n".join(lines)


def main() -> None:
    args = _parse_args()
    env = _load_env(
        os.getenv("MANAGEMENT_B_ENV_FILE")
        or os.getenv("OPENCLAW_ENV_FILE")
        or os.getenv("PRICING_ENV_FILE")
        or DEFAULT_LOCAL_ENV_FILE
    )
    source_url = (
        env.get("MANAGEMENT_SOURCE_URL")
        or env.get("RETURN_SCHEME_SOURCE_URL")
        or DEFAULT_LOCAL_SOURCE_URL
    )
    source_token = (
        env.get("MANAGEMENT_SOURCE_TOKEN")
        or env.get("RETURN_SCHEME_SOURCE_TOKEN")
        or env.get("MANAGEMENT_INTERNAL_API_TOKEN")
        or env.get("RETURN_SCHEME_INTERNAL_API_TOKEN")
    )
    if not source_token:
        raise SystemExit(
            "Missing required env: MANAGEMENT_SOURCE_TOKEN|RETURN_SCHEME_SOURCE_TOKEN|"
            "MANAGEMENT_INTERNAL_API_TOKEN|RETURN_SCHEME_INTERNAL_API_TOKEN"
        )

    timeout = int(
        env.get("COUNTERPARTY_FOLDER_RECOMMENDATIONS_SOURCE_TIMEOUT_SECONDS")
        or env.get("MANAGEMENT_ADAPTER_TIMEOUT_SECONDS", "180")
    )
    retries = int(
        env.get("COUNTERPARTY_FOLDER_RECOMMENDATIONS_SOURCE_RETRIES")
        or env.get("MANAGEMENT_ADAPTER_RETRIES", "0")
    )
    retry_delay = float(env.get("MANAGEMENT_ADAPTER_RETRY_DELAY_SECONDS", "1.0"))
    fetch_json = _build_fetcher(
        source_url=source_url,
        token=source_token,
        timeout=timeout,
        retries=retries,
        retry_delay=retry_delay,
    )
    bitrix_base_url = (
        _bitrix_webhook_base(env) if args.notify_bitrix_task_id and not args.dry_run else ""
    )
    disk_folder_id = _to_int(env.get("COUNTERPARTY_FOLDER_RECOMMENDATIONS_B24_DISK_FOLDER_ID"))

    summary = sync_counterparty_folder_recommendations(
        fetch_json=fetch_json,
        snapshot_date=_parse_date(args.snapshot_date),
        state_path=Path(
            env.get("COUNTERPARTY_FOLDER_RECOMMENDATIONS_STATE_PATH", DEFAULT_STATE_PATH)
        ),
        artifact_dir=Path(
            env.get("COUNTERPARTY_FOLDER_RECOMMENDATIONS_ARTIFACT_DIR", DEFAULT_ARTIFACT_DIR)
        ),
        status=args.status,
        limit=args.limit,
        dry_run=args.dry_run,
        force=args.force,
        bitrix_task_id=args.notify_bitrix_task_id,
        notify_empty=args.notify_empty,
        deliver_comment=(
            (
                lambda task_id, message: post_bitrix_task_comment(
                    bitrix_base_url,
                    task_id=task_id,
                    message=message,
                    timeout=int(
                        env.get("COUNTERPARTY_FOLDER_RECOMMENDATIONS_BITRIX_TIMEOUT_SEC", "30")
                    ),
                )
            )
            if args.notify_bitrix_task_id and not args.dry_run
            else None
        ),
        deliver_attachment=(
            (
                lambda task_id, file_path: deliver_bitrix_task_attachment(
                    bitrix_base_url,
                    folder_id=disk_folder_id,
                    task_id=task_id,
                    file_path=file_path,
                )
            )
            if args.notify_bitrix_task_id and disk_folder_id is not None and not args.dry_run
            else None
        ),
        delivery_mode=args.delivery_mode,
    )

    if args.json:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return
    print(render_summary(summary))


if __name__ == "__main__":
    main()
