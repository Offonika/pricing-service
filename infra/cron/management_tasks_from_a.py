#!/usr/bin/env python3
"""Consume management task payloads and upsert tasks in Bitrix24."""

from __future__ import annotations

import argparse
import base64
import errno
import hashlib
import json
import os
import socket
import time
import urllib.error
import urllib.parse
import urllib.request
from collections.abc import Mapping
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_LOCAL_SOURCE_URL = "http://127.0.0.1:18080"
DEFAULT_LOCAL_ENV_FILE = "/opt/MM/pricing-service/.env"
DELIVERY_INTENT_TASK = "task"
DELIVERY_CHANNEL_BITRIX_TASK = "bitrix_task"
BITRIX_CONTOUR_CLOUD = "cloud"
BITRIX_CONTOUR_BOX = "box"
BITRIX_CONTOUR_VALUES = {BITRIX_CONTOUR_CLOUD, BITRIX_CONTOUR_BOX}
MIGRATION_STATE_CLOUD_PRIMARY = "cloud_primary"
MIGRATION_STATE_BOX_SHADOW = "box_shadow"
MIGRATION_STATE_BOX_PRIMARY = "box_primary"
MIGRATION_STATE_CLOUD_FALLBACK = "cloud_fallback"
MIGRATION_STATE_VALUES = {
    MIGRATION_STATE_CLOUD_PRIMARY,
    MIGRATION_STATE_BOX_SHADOW,
    MIGRATION_STATE_BOX_PRIMARY,
    MIGRATION_STATE_CLOUD_FALLBACK,
}

RULE_RECEIVABLE_NEW_DAILY = "receivable_new_daily"
RULE_RECEIVABLE_NEW_DAILY_BATCH = "receivable_new_daily_batch"
RULE_RECEIVABLE_OVERDUE = "receivable_overdue"
RULE_RECEIVABLE_OVERDUE_BATCH = "receivable_overdue_batch"
RULE_RECEIVABLE_FINANCE_DAILY_BATCH = "receivable_finance_daily_batch"
RULE_RECEIVABLE_EMPLOYEE = "receivable_employee"
RULE_RECEIVABLE_EMPLOYEE_BATCH = "receivable_employee_batch"
RULE_RECEIVABLE_FIRED_MANAGER = "receivable_fired_manager"
RULE_RECEIVABLE_FIRED_MANAGER_BATCH = "receivable_fired_manager_batch"
RULE_RECEIVABLE_ADJUSTMENT = "receivable_adjustment_candidate"
RULE_RECEIVABLE_ADJUSTMENT_LARGE = "receivable_adjustment_candidate_large"
RULE_RECEIVABLE_ADJUSTMENT_BATCH = "receivable_adjustment_candidate_batch"
SEVERITY_ORDER = {"warning": 1, "high": 2, "critical": 3}
BATCH_REPORT_HEADERS = [
    "№",
    "Контрагент",
    "Сумма долга, ₽",
    "Контрагент ref",
    "Исходный dedupe key",
]
OVERDUE_BATCH_REPORT_HEADERS = [
    "№",
    "Контрагент",
    "Сумма долга, ₽",
    "Срок оплаты",
    "Дней просрочки",
    "Источник срока",
    "Возраст долга",
    "Запрет отгрузки",
    "Контрагент ref",
    "Исходный dedupe key",
]
CASE_BATCH_REPORT_HEADERS = [
    "Категория",
    "№",
    "Контрагент",
    "Сумма долга, ₽",
    "Возраст долга",
    "Активность",
    "Текущий ответственный",
    "Исходный документ",
    "Контрагент ref",
    "Исходный dedupe key",
]
BATCH_RULE_CODES = {
    RULE_RECEIVABLE_NEW_DAILY_BATCH,
    RULE_RECEIVABLE_OVERDUE_BATCH,
    RULE_RECEIVABLE_FINANCE_DAILY_BATCH,
    RULE_RECEIVABLE_EMPLOYEE_BATCH,
    RULE_RECEIVABLE_FIRED_MANAGER_BATCH,
    RULE_RECEIVABLE_ADJUSTMENT_BATCH,
}

FINANCE_DAILY_BATCH_SOURCE_RULES = {
    RULE_RECEIVABLE_NEW_DAILY_BATCH,
    RULE_RECEIVABLE_EMPLOYEE_BATCH,
    RULE_RECEIVABLE_FIRED_MANAGER_BATCH,
    RULE_RECEIVABLE_ADJUSTMENT_BATCH,
}
SUPPRESSED_RECEIVABLE_RULE_CODES = {
    RULE_RECEIVABLE_ADJUSTMENT,
    RULE_RECEIVABLE_ADJUSTMENT_BATCH,
    RULE_RECEIVABLE_EMPLOYEE,
    RULE_RECEIVABLE_EMPLOYEE_BATCH,
    RULE_RECEIVABLE_FIRED_MANAGER,
    RULE_RECEIVABLE_FIRED_MANAGER_BATCH,
}


class PartialTaskDeliveryError(RuntimeError):
    """Task side effect already happened; do not fallback to another contour."""


class FallbackableTaskSyncError(RuntimeError):
    """Task sync failed before durable side effects in the current contour."""


class UnsafeToFallbackTaskSyncError(RuntimeError):
    """Task sync result is uncertain; cross-contour fallback may create duplicates."""


def _load_env(path: str | None) -> dict[str, str]:
    env = os.environ.copy()
    if not path:
        return env
    env_path = Path(path)
    if not env_path.exists():
        return env
    for raw in env_path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pull management task payloads from server A and upsert them in Bitrix24."
    )
    parser.add_argument("--date", dest="anchor_date", help="Anchor date in YYYY-MM-DD format")
    parser.add_argument(
        "--dry-run", action="store_true", help="Render actions without side effects"
    )
    parser.add_argument("--json", action="store_true", help="Emit machine-readable summary")
    return parser.parse_args()


def _env_flag(value: str | None) -> bool:
    if value is None:
        return False
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _http_json(
    url: str,
    *,
    method: str = "GET",
    headers: dict[str, str] | None = None,
    payload: dict[str, Any] | None = None,
    timeout: int = 60,
) -> dict[str, Any]:
    data = None
    request_headers = dict(headers or {})
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        request_headers["Content-Type"] = "application/json"
    request = urllib.request.Request(url, data=data, headers=request_headers, method=method)
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    with opener.open(request, timeout=timeout) as response:
        body = response.read().decode("utf-8")
        if not body:
            return {}
        return json.loads(body)


def _b24_call(base_url: str, method: str, params: list[tuple[str, str]]) -> dict[str, Any]:
    url = f"{base_url.rstrip('/')}/{method}.json"
    data = None
    if params:
        data = urllib.parse.urlencode(params, doseq=True).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST" if data is not None else "GET",
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = json.loads(response.read().decode("utf-8"))
    if payload.get("error"):
        raise RuntimeError(
            f"Bitrix24 {method}: {payload['error']} {payload.get('error_description', '')}"
        )
    return payload


def _is_fallbackable_create_error(error: Exception) -> bool:
    if isinstance(error, urllib.error.HTTPError):
        return False
    if not isinstance(error, urllib.error.URLError):
        return False

    reason = error.reason
    if isinstance(reason, (ConnectionRefusedError, socket.gaierror)):
        return True
    if isinstance(reason, OSError):
        return reason.errno in {
            errno.ECONNREFUSED,
            errno.ENETUNREACH,
            errno.EHOSTUNREACH,
        }
    return False


def _load_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"tasks": {}}
    return json.loads(path.read_text(encoding="utf-8"))


def _save_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(state, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _parse_date(value: str | None) -> date:
    if value:
        return date.fromisoformat(value)
    return datetime.now().date()


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(str(value))
    except (TypeError, ValueError):
        return None


def _parse_csv_ints(raw: str | None) -> list[int]:
    if not raw:
        return []
    items: list[int] = []
    for chunk in raw.split(","):
        parsed = _to_int(chunk.strip())
        if parsed is not None:
            items.append(parsed)
    return items


def _parse_csv_strings(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [chunk.strip() for chunk in raw.split(",") if chunk.strip()]


def _parse_bool(value: Any, *, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "on"}:
        return True
    if normalized in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _parse_code_overrides(raw: str | None) -> dict[str, int]:
    if not raw:
        return {}
    overrides: dict[str, int] = {}
    for chunk in raw.split(","):
        item = chunk.strip()
        if not item or ":" not in item:
            continue
        code, value = item.split(":", 1)
        parsed = _to_int(value.strip())
        if code.strip() and parsed is not None:
            overrides[code.strip()] = parsed
    return overrides


def _load_team_role_map(path: str | None) -> dict[str, dict[str, Any]]:
    if not path:
        return {}
    team_path = Path(path)
    if not team_path.exists():
        return {}
    payload = yaml.safe_load(team_path.read_text(encoding="utf-8")) or {}
    role_map: dict[str, dict[str, Any]] = {}
    for employee in payload.get("employees", []):
        role_code = str(employee.get("role_code") or "").strip()
        legacy_bitrix_id = _to_int(employee.get("bitrix24_id"))
        cloud_bitrix_id = _to_int(
            employee.get("bitrix_cloud_user_id")
            or employee.get("bitrix24_cloud_user_id")
            or employee.get("bitrix24_cloud_id")
            or employee.get("bitrix_cloud_id")
            or employee.get("bitrix24_id")
        )
        box_bitrix_id = _to_int(
            employee.get("bitrix_box_user_id")
            or employee.get("bitrix24_box_user_id")
            or employee.get("bitrix24_box_id")
            or employee.get("bitrix_box_id")
        )
        if role_code and any(
            value is not None for value in (legacy_bitrix_id, cloud_bitrix_id, box_bitrix_id)
        ):
            role_map[role_code] = {
                "legacy": legacy_bitrix_id,
                BITRIX_CONTOUR_CLOUD: cloud_bitrix_id,
                BITRIX_CONTOUR_BOX: box_bitrix_id,
                "bitrix_box_assistant_enabled": _parse_bool(
                    employee.get("bitrix_box_assistant_enabled"),
                    default=False,
                ),
                "allowed_delivery_channels": list(
                    employee.get("allowed_delivery_channels")
                    or employee.get("delivery_channels")
                    or [DELIVERY_CHANNEL_BITRIX_TASK]
                ),
            }
    return role_map


ROLE_ALIASES = {
    "finance": "cfo",
    "finance_pool": "cfo",
    "retail_supervisor": "coo",
    "sales_manager": "cco",
}


def _delivery_state_key(
    dedupe_key: str,
    *,
    delivery_intent: str,
    contour: str,
    preserve_legacy_cloud: bool = True,
) -> str:
    if (
        preserve_legacy_cloud
        and delivery_intent == DELIVERY_INTENT_TASK
        and contour == BITRIX_CONTOUR_CLOUD
    ):
        return dedupe_key
    return f"{delivery_intent}|{contour}|{dedupe_key}"


def _normalize_contour(value: str | None) -> str:
    normalized = str(value or "").strip().lower()
    if normalized in BITRIX_CONTOUR_VALUES:
        return normalized
    return BITRIX_CONTOUR_CLOUD


def _normalize_migration_state(value: str | None) -> str:
    normalized = str(value or "").strip().lower()
    if normalized in MIGRATION_STATE_VALUES:
        return normalized
    return MIGRATION_STATE_CLOUD_PRIMARY


def _resolve_override_value(
    code: str,
    *,
    overrides: Mapping[str, Any],
    contour: str,
) -> int | None:
    value = overrides.get(code)
    if value is None:
        return None
    if isinstance(value, Mapping):
        contour_value = _to_int(value.get(contour))
        if contour_value is not None:
            return contour_value
        if contour == BITRIX_CONTOUR_BOX:
            return None
        return _to_int(value.get("default"))
    return _to_int(value)


def _resolve_team_role_id(
    role_entry: Any,
    *,
    contour: str,
) -> int | None:
    if isinstance(role_entry, int):
        if contour == BITRIX_CONTOUR_BOX:
            return None
        return role_entry
    if isinstance(role_entry, Mapping):
        contour_value = _to_int(role_entry.get(contour))
        if contour_value is not None:
            return contour_value
        if contour == BITRIX_CONTOUR_BOX:
            return None
        legacy_value = _to_int(role_entry.get("legacy"))
        if legacy_value is not None:
            return legacy_value
    return None


def _resolve_role_id(
    code: str,
    *,
    overrides: Mapping[str, Any],
    team_roles: Mapping[str, Any],
    contour: str,
    default_id: int | None = None,
) -> int | None:
    override_id = _resolve_override_value(code, overrides=overrides, contour=contour)
    if override_id is not None:
        return override_id
    alias_code = ROLE_ALIASES.get(code, code)
    override_id = _resolve_override_value(alias_code, overrides=overrides, contour=contour)
    if override_id is not None:
        return override_id
    if alias_code in team_roles:
        team_role_id = _resolve_team_role_id(team_roles[alias_code], contour=contour)
        if team_role_id is not None:
            return team_role_id
    return default_id


def _resolve_created_by_id(
    *,
    payload: dict[str, Any],
    owner_overrides: Mapping[str, Any],
    team_roles: Mapping[str, Any],
    contour: str,
    default_created_by_id: int | None,
) -> int | None:
    created_by_code = str(payload.get("created_by_code") or "").strip()
    if not created_by_code:
        return default_created_by_id
    return _resolve_role_id(
        created_by_code,
        overrides=owner_overrides,
        team_roles=team_roles,
        contour=contour,
        default_id=default_created_by_id,
    )


def _build_fetcher(
    *,
    source_url: str,
    token: str,
    timeout: int,
    retries: int,
    retry_delay: float,
) -> Callable[[str, dict[str, str]], dict[str, Any]]:
    base = source_url.rstrip("/")
    headers = {"Authorization": f"Bearer {token}"}

    def _fetch(path: str, params: dict[str, str]) -> dict[str, Any]:
        query = urllib.parse.urlencode(params)
        url = f"{base}{path}"
        if query:
            url = f"{url}?{query}"

        attempts = max(1, retries + 1)
        last_error: Exception | None = None
        for attempt in range(attempts):
            try:
                return _http_json(url, headers=headers, timeout=timeout)
            except (
                urllib.error.HTTPError,
                urllib.error.URLError,
                TimeoutError,
                ValueError,
            ) as error:
                last_error = error
                if attempt + 1 >= attempts:
                    break
                time.sleep(retry_delay)
        assert last_error is not None
        raise last_error

    return _fetch


def _fingerprint_payload(payload: dict[str, Any]) -> str:
    references = payload.get("references") or []
    references_signature = hashlib.sha256(
        json.dumps(
            references,
            ensure_ascii=False,
            sort_keys=True,
            default=str,
        ).encode("utf-8")
    ).hexdigest()
    normalized = json.dumps(
        {
            "title": payload.get("title"),
            "description": _build_description(payload),
            "summary": payload.get("summary"),
            "severity": payload.get("severity"),
            "owner_code": payload.get("owner_code"),
            "watcher_codes": payload.get("watcher_codes") or [],
            "reaction_deadline_at": payload.get("reaction_deadline_at"),
            "due_at": payload.get("due_at"),
            "metrics": payload.get("metrics") or {},
            "references_count": len(references),
            "references_signature": references_signature,
            "tags": payload.get("tags") or [],
        },
        ensure_ascii=False,
        sort_keys=True,
        default=str,
    )
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def _format_timestamp(value: str | None) -> str:
    if not value:
        return "не задано"
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return value
    return parsed.strftime("%Y-%m-%d %H:%M")


def _format_date_label(value: str | None) -> str:
    if not value:
        return "не задана"
    try:
        parsed = date.fromisoformat(value)
    except ValueError:
        return value
    return parsed.strftime("%d.%m.%Y")


def _extract_batch_date(payload: dict[str, Any]) -> str | None:
    dedupe_key = str(payload.get("dedupe_key") or "")
    parts = dedupe_key.split("|")
    if len(parts) >= 2:
        return parts[1]
    entity_ref = str(payload.get("entity_ref") or "")
    if entity_ref.count(":") >= 2:
        return entity_ref.rsplit(":", 1)[-1]
    return None


def _build_receivable_batch_description(payload: dict[str, Any]) -> str:
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_OVERDUE_BATCH:
        return _build_receivable_overdue_batch_description(payload)
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_FINANCE_DAILY_BATCH:
        return _build_receivable_finance_daily_batch_description(payload)
    if str(payload.get("rule_code") or "") in {
        RULE_RECEIVABLE_EMPLOYEE_BATCH,
        RULE_RECEIVABLE_FIRED_MANAGER_BATCH,
        RULE_RECEIVABLE_ADJUSTMENT_BATCH,
    }:
        return _build_receivable_case_batch_description(payload)

    metrics = payload.get("metrics") or {}
    references = payload.get("references") or []
    counterparty_count = _to_int(metrics.get("counterparty_count")) or len(references)
    total_balance = _format_amount(_to_decimal(metrics.get("current_balance_total")))
    report_date = _format_date_label(_extract_batch_date(payload))

    lines = [
        "Что произошло",
        f"- За {report_date} выявлено {counterparty_count} новых контрагентов с дебиторкой на {total_balance}.",
        "",
        "Как отобрано",
        '- В задачу включены только кейсы сегмента "новая дебиторка за день" (`new_daily`).',
        f"- Дата отбора: {report_date}.",
        "- Один контрагент = один кейс, но в Bitrix они объединены в одну дневную задачу, чтобы не создавать шум.",
        "",
        "Что нужно сделать",
        "- Проверить причину возникновения долга по каждому контрагенту.",
        "- Зафиксировать решение: оплата, контроль менеджера, ограничение продаж или эскалация.",
        f"- До {_format_timestamp(payload.get('due_at'))} закрыть разбор верхних сумм риска.",
    ]

    if references:
        lines.extend(["", "Топ-10 по сумме"])
        ranked = sorted(
            references,
            key=lambda item: _to_decimal(item.get("current_balance")),
            reverse=True,
        )
        for index, item in enumerate(ranked[:10], start=1):
            counterparty_name = item.get("counterparty_name") or item.get("counterparty_ref")
            amount = _format_amount(_to_decimal(item.get("current_balance")))
            lines.append(f"{index}. {counterparty_name} — {amount}")

    lines.extend(
        [
            "",
            "Служебно",
            f"- Реакция до: {_format_timestamp(payload.get('reaction_deadline_at'))}",
            f"- Идентификатор пакета: {payload.get('dedupe_key')}",
        ]
    )
    return "\n".join(lines)


def _build_receivable_batch_report_filename(payload: dict[str, Any]) -> str:
    report_date = _extract_batch_date(payload) or datetime.now().date().isoformat()
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_OVERDUE_BATCH:
        return f"receivable-overdue-{report_date}.xlsx"
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_FINANCE_DAILY_BATCH:
        return f"receivable-finance-daily-{report_date}.xlsx"
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_EMPLOYEE_BATCH:
        return f"receivable-employee-{report_date}.xlsx"
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_FIRED_MANAGER_BATCH:
        return f"receivable-fired-manager-{report_date}.xlsx"
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_ADJUSTMENT_BATCH:
        return f"receivable-adjustment-{report_date}.xlsx"
    return f"receivable-new-daily-{report_date}.xlsx"


def _export_receivable_batch_xlsx(payload: dict[str, Any], output_path: Path) -> Path:
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_OVERDUE_BATCH:
        return _export_receivable_overdue_batch_xlsx(payload, output_path)
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_FINANCE_DAILY_BATCH:
        return _export_receivable_finance_daily_batch_xlsx(payload, output_path)
    if str(payload.get("rule_code") or "") in {
        RULE_RECEIVABLE_EMPLOYEE_BATCH,
        RULE_RECEIVABLE_FIRED_MANAGER_BATCH,
        RULE_RECEIVABLE_ADJUSTMENT_BATCH,
    }:
        return _export_receivable_case_batch_xlsx(payload, output_path)

    rows: list[list[Any]] = [BATCH_REPORT_HEADERS]
    references = payload.get("references") or []
    ranked = sorted(
        references,
        key=lambda item: _to_decimal(item.get("current_balance")),
        reverse=True,
    )
    for index, item in enumerate(ranked, start=1):
        rows.append(
            [
                index,
                item.get("counterparty_name") or item.get("counterparty_ref"),
                float(_to_decimal(item.get("current_balance"))),
                item.get("counterparty_ref") or "",
                item.get("original_task_key") or "",
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_xlsx(output_path, sheet_name="new_daily_receivables", rows=rows)
    return output_path


def _payment_term_source_label(value: str | None) -> str:
    mapping = {
        "planned_payment_date": "Согласованная дата оплаты",
        "credit_depth_days": "Глубина кредита",
        "missing": "Срок не зафиксирован",
    }
    return mapping.get(str(value or ""), str(value or "Не задано"))


def _format_bucket_breakdown(metrics: dict[str, Any]) -> str:
    bucket_counts = metrics.get("aged_bucket_counts") or {}
    bucket_totals = metrics.get("aged_bucket_totals") or {}
    ordered = []
    for bucket in ("0-7", "8-30", "31-60", "61-90", "90+"):
        count = _to_int(bucket_counts.get(bucket)) or 0
        total = _to_decimal(bucket_totals.get(bucket))
        if count <= 0:
            continue
        ordered.append(f"{bucket}: {count} на {_format_amount(total)}")
    if not ordered:
        return "нет данных"
    return "; ".join(ordered)


def _build_receivable_overdue_batch_description(payload: dict[str, Any]) -> str:
    metrics = payload.get("metrics") or {}
    references = payload.get("references") or []
    counterparty_count = _to_int(metrics.get("counterparty_count")) or len(references)
    total_balance = _format_amount(_to_decimal(metrics.get("current_balance_total")))
    report_date = _format_date_label(_extract_batch_date(payload))

    lines = [
        "Что произошло",
        f"- На {report_date} выявлено {counterparty_count} просроченных долгов на {total_balance}.",
        "",
        "Как отобрано",
        "- В задачу включены только кейсы, где истёк согласованный срок оплаты или глубина кредита.",
        f"- Разбивка по возрасту долга: {_format_bucket_breakdown(metrics)}.",
        "",
        "Что нужно сделать",
        "- Проверить просроченные контрагенты и подтвердить план взыскания.",
        "- По верхним суммам риска зафиксировать статус: оплата, эскалация или ограничение продаж.",
        f"- До {_format_timestamp(payload.get('due_at'))} закрыть разбор верхних сумм риска.",
    ]

    if references:
        lines.extend(["", "Топ-20 по сумме"])
        ranked = sorted(
            references,
            key=lambda item: _to_decimal(item.get("current_balance")),
            reverse=True,
        )
        for index, item in enumerate(ranked[:20], start=1):
            counterparty_name = item.get("counterparty_name") or item.get("counterparty_ref")
            amount = _format_amount(_to_decimal(item.get("current_balance")))
            overdue_days = _to_int(item.get("overdue_days")) or 0
            due_date = _format_date_label(item.get("due_date"))
            lines.append(
                f"{index}. {counterparty_name} — {amount}; срок {due_date}; просрочка {overdue_days} дн."
            )

    lines.extend(
        [
            "",
            "Служебно",
            f"- Реакция до: {_format_timestamp(payload.get('reaction_deadline_at'))}",
            f"- Идентификатор пакета: {payload.get('dedupe_key')}",
        ]
    )
    return "\n".join(lines)


def _export_receivable_overdue_batch_xlsx(payload: dict[str, Any], output_path: Path) -> Path:
    rows: list[list[Any]] = [OVERDUE_BATCH_REPORT_HEADERS]
    references = payload.get("references") or []
    ranked = sorted(
        references,
        key=lambda item: _to_decimal(item.get("current_balance")),
        reverse=True,
    )
    for index, item in enumerate(ranked, start=1):
        rows.append(
            [
                index,
                item.get("counterparty_name") or item.get("counterparty_ref"),
                float(_to_decimal(item.get("current_balance"))),
                item.get("due_date") or "",
                _to_int(item.get("overdue_days")) or 0,
                _payment_term_source_label(item.get("payment_term_source")),
                item.get("aged_bucket") or "",
                "Да" if item.get("shipment_ban") else "Нет",
                item.get("counterparty_ref") or "",
                item.get("original_task_key") or "",
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_xlsx(output_path, sheet_name="overdue_receivables", rows=rows)
    return output_path


def _finance_daily_section_label(rule_code: str) -> str:
    mapping = {
        RULE_RECEIVABLE_NEW_DAILY_BATCH: "Новая дебиторка",
        RULE_RECEIVABLE_EMPLOYEE_BATCH: "Долги сотрудников",
        RULE_RECEIVABLE_FIRED_MANAGER_BATCH: "Уволенные менеджеры",
        RULE_RECEIVABLE_ADJUSTMENT_BATCH: "Корректировка",
    }
    return mapping.get(rule_code, rule_code)


def _build_receivable_finance_daily_batch_description(payload: dict[str, Any]) -> str:
    metrics = payload.get("metrics") or {}
    report_date = _format_date_label(_extract_batch_date(payload))
    section_metrics = metrics.get("sections") or {}
    total_count = _to_int(metrics.get("counterparty_count")) or 0
    total_balance = _format_amount(_to_decimal(metrics.get("current_balance_total")))

    lines = [
        "Что произошло",
        f"- За {report_date} сформирован единый дневной финансовый пакет по дебиторке.",
        f"- Всего кейсов: {total_count} на {total_balance}.",
        "",
        "Что внутри",
    ]

    for rule_code in (
        RULE_RECEIVABLE_NEW_DAILY_BATCH,
        RULE_RECEIVABLE_EMPLOYEE_BATCH,
        RULE_RECEIVABLE_FIRED_MANAGER_BATCH,
        RULE_RECEIVABLE_ADJUSTMENT_BATCH,
    ):
        section = section_metrics.get(rule_code) or {}
        count = _to_int(section.get("counterparty_count")) or 0
        balance = _format_amount(_to_decimal(section.get("current_balance_total")))
        lines.append(f"- {_finance_daily_section_label(rule_code)}: {count} на {balance}.")

    lines.extend(
        [
            "",
            "Как работать",
            "- Открыть один пакет вместо нескольких отдельных daily-задач.",
            "- Сначала разобрать новую дебиторку и долги сотрудников, затем кейсы на корректировку и уволенных менеджеров.",
            f"- До {_format_timestamp(payload.get('due_at'))} зафиксировать статусы по верхним суммам риска.",
            "",
            "Служебно",
            f"- Реакция до: {_format_timestamp(payload.get('reaction_deadline_at'))}",
            f"- Идентификатор пакета: {payload.get('dedupe_key')}",
        ]
    )
    return "\n".join(lines)


def _export_receivable_finance_daily_batch_xlsx(payload: dict[str, Any], output_path: Path) -> Path:
    rows: list[list[Any]] = [CASE_BATCH_REPORT_HEADERS]
    references = payload.get("references") or []
    section_order = {
        RULE_RECEIVABLE_NEW_DAILY_BATCH: 0,
        RULE_RECEIVABLE_EMPLOYEE_BATCH: 1,
        RULE_RECEIVABLE_FIRED_MANAGER_BATCH: 2,
        RULE_RECEIVABLE_ADJUSTMENT_BATCH: 3,
    }
    ranked = sorted(
        references,
        key=lambda item: (
            section_order.get(str(item.get("batch_rule_code") or ""), 99),
            -float(_to_decimal(item.get("current_balance"))),
            str(item.get("counterparty_name") or item.get("counterparty_ref") or ""),
        ),
    )
    for index, item in enumerate(ranked, start=1):
        rows.append(
            [
                _finance_daily_section_label(str(item.get("batch_rule_code") or "")),
                index,
                item.get("counterparty_name") or item.get("counterparty_ref"),
                float(_to_decimal(item.get("current_balance"))),
                item.get("aged_bucket") or "",
                item.get("activity_segment") or "",
                item.get("current_manager_name") or item.get("current_manager_ref") or "",
                item.get("origin_document_number") or item.get("origin_document_ref") or "",
                item.get("counterparty_ref") or "",
                item.get("original_task_key") or "",
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_xlsx(output_path, sheet_name="finance_daily", rows=rows)
    return output_path


def _case_batch_copy(payload: dict[str, Any]) -> tuple[str, str, str]:
    rule_code = str(payload.get("rule_code") or "")
    mapping = {
        RULE_RECEIVABLE_EMPLOYEE_BATCH: (
            "долгов сотрудников",
            "В задачу включены только кейсы, где должником является сотрудник компании.",
            "Передать кейсы в финансы и HR для отдельного разбора.",
        ),
        RULE_RECEIVABLE_FIRED_MANAGER_BATCH: (
            "долгов на уволенных менеджерах",
            "В задачу включены только кейсы, где текущий или исходный ответственный менеджер уволен.",
            "Передать кейсы в финансовый пул и назначить новый план взыскания.",
        ),
        RULE_RECEIVABLE_ADJUSTMENT_BATCH: (
            "кейсов на корректировку",
            "В задачу включены старые или неактивные долги, требующие решения по корректировке.",
            "Проверить кейсы и определить: взыскание, корректировка или заморозка.",
        ),
    }
    return mapping[rule_code]


def _build_receivable_case_batch_description(payload: dict[str, Any]) -> str:
    metrics = payload.get("metrics") or {}
    references = payload.get("references") or []
    counterparty_count = _to_int(metrics.get("counterparty_count")) or len(references)
    total_balance = _format_amount(_to_decimal(metrics.get("current_balance_total")))
    report_date = _format_date_label(_extract_batch_date(payload))
    label, selection_line, action_line = _case_batch_copy(payload)

    lines = [
        "Что произошло",
        f"- На {report_date} выявлено {counterparty_count} {label} на {total_balance}.",
        "",
        "Как отобрано",
        f"- {selection_line}",
        "",
        "Что нужно сделать",
        f"- {action_line}",
        f"- До {_format_timestamp(payload.get('due_at'))} закрыть разбор верхних сумм риска.",
    ]

    if references:
        lines.extend(["", "Топ-20 по сумме"])
        ranked = sorted(
            references,
            key=lambda item: _to_decimal(item.get("current_balance")),
            reverse=True,
        )
        for index, item in enumerate(ranked[:20], start=1):
            counterparty_name = item.get("counterparty_name") or item.get("counterparty_ref")
            amount = _format_amount(_to_decimal(item.get("current_balance")))
            aged_bucket = item.get("aged_bucket") or "не задан"
            activity = item.get("activity_segment") or "не задан"
            lines.append(
                f"{index}. {counterparty_name} — {amount}; возраст {aged_bucket}; активность {activity}."
            )

    lines.extend(
        [
            "",
            "Служебно",
            f"- Реакция до: {_format_timestamp(payload.get('reaction_deadline_at'))}",
            f"- Идентификатор пакета: {payload.get('dedupe_key')}",
        ]
    )
    return "\n".join(lines)


def _export_receivable_case_batch_xlsx(payload: dict[str, Any], output_path: Path) -> Path:
    rows: list[list[Any]] = [CASE_BATCH_REPORT_HEADERS]
    references = payload.get("references") or []
    ranked = sorted(
        references,
        key=lambda item: _to_decimal(item.get("current_balance")),
        reverse=True,
    )
    for index, item in enumerate(ranked, start=1):
        rows.append(
            [
                _finance_daily_section_label(str(payload.get("rule_code") or "")),
                index,
                item.get("counterparty_name") or item.get("counterparty_ref"),
                float(_to_decimal(item.get("current_balance"))),
                item.get("aged_bucket") or "",
                item.get("activity_segment") or "",
                item.get("current_manager_name") or item.get("current_manager_ref") or "",
                item.get("origin_document_number") or item.get("origin_document_ref") or "",
                item.get("counterparty_ref") or "",
                item.get("original_task_key") or "",
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_xlsx(output_path, sheet_name="receivable_cases", rows=rows)
    return output_path


def _upload_b24_disk_file(*, webhook_url: str, folder_id: int, file_path: Path) -> int:
    encoded = base64.b64encode(file_path.read_bytes()).decode("ascii")
    response = _b24_call(
        webhook_url,
        "disk.folder.uploadfile",
        [
            ("id", str(folder_id)),
            ("data[NAME]", file_path.name),
            ("fileContent[0]", file_path.name),
            ("fileContent[1]", encoded),
            ("generateUniqueName", "true"),
        ],
    )
    result = response.get("result") or {}
    object_id = _to_int(result.get("ID"))
    if object_id is None:
        raise RuntimeError("Bitrix24 disk.folder.uploadfile returned empty object id")
    return object_id


def _write_xlsx(output_path: Path, *, sheet_name: str, rows: list[list[Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]

    if not rows:
        rows = [[]]

    for row in rows:
        sheet.append(row)

    max_row = sheet.max_row
    max_column = sheet.max_column
    if max_row == 0 or max_column == 0:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    stripe_fill = PatternFill(fill_type="solid", fgColor="F7FBFF")
    header_font = Font(bold=True, color="FFFFFF")
    border_side = Side(style="thin", color="D9E2F3")
    cell_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )

    amount_headers = {"Сумма долга, ₽"}
    centered_headers = {"№", "Дней просрочки", "Возраст долга", "Запрет отгрузки"}
    wide_headers = {
        "Категория": 20,
        "Контрагент": 34,
        "Текущий ответственный": 26,
        "Исходный документ": 22,
        "Контрагент ref": 22,
        "Исходный dedupe key": 42,
        "Срок оплаты": 20,
        "Источник срока": 24,
        "Активность": 16,
    }

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(2, max_row + 1):
        use_stripe = row_index % 2 == 0
        for column_index in range(1, max_column + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            header_value = str(sheet.cell(row=1, column=column_index).value or "")
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if use_stripe:
                cell.fill = stripe_fill
            if header_value in amount_headers and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif header_value in centered_headers:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    if max_row > 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"

    sheet.sheet_view.zoomScale = 90
    sheet.row_dimensions[1].height = 24

    for column_index in range(1, max_column + 1):
        header_value = str(sheet.cell(row=1, column=column_index).value or "")
        width = wide_headers.get(header_value, 10)
        for row_index in range(1, max_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            if isinstance(value, float):
                rendered = f"{value:,.2f}"
            else:
                rendered = str(value)
            width = max(width, len(rendered) + 2)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(width, 48)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _attach_b24_file_to_task(*, webhook_url: str, task_id: int, file_object_id: int) -> int:
    response = _b24_call(
        webhook_url,
        "tasks.task.files.attach",
        [("taskId", str(task_id)), ("fileId", str(file_object_id))],
    )
    attachment_id = _to_int((response.get("result") or {}).get("attachmentId"))
    if attachment_id is None:
        raise RuntimeError("Bitrix24 tasks.task.files.attach returned empty attachment id")
    return attachment_id


def _build_description(payload: dict[str, Any]) -> str:
    if str(payload.get("rule_code") or "") == RULE_RECEIVABLE_ADJUSTMENT_LARGE:
        metrics = payload.get("metrics") or {}
        current_balance = _format_amount(_to_decimal(metrics.get("current_balance")))
        current_manager = next(
            (
                ref.get("current_manager_name") or ref.get("current_manager_ref")
                for ref in (payload.get("references") or [])
                if ref.get("kind") == "current_manager"
            ),
            "не назначен",
        )
        aged_bucket = metrics.get("aged_bucket") or "не задан"
        activity_segment = metrics.get("activity_segment") or "не задана"
        lines = [
            "Что произошло",
            (
                f"- Выявлен крупный кейс на корректировку: "
                f"{payload.get('entity_name') or payload.get('entity_ref')} на {current_balance}."
            ),
            "",
            "Как отобрано",
            "- В задачу попадают только adjustment_candidates с суммой 10 000 ₽ и выше.",
            "- Один контрагент = одна живая задача до полного разбора кейса.",
            "",
            "Что нужно сделать",
            "- Сначала попытаться взыскать долг.",
            "- Если взыскание нереально, подготовить решение на списание.",
            (
                "- Разбор ведёт руководитель розничной сети; "
                f"менеджер в кейсе сейчас: {current_manager}."
            ),
            f"- Возраст долга: {aged_bucket}; активность: {activity_segment}.",
            f"- До {_format_timestamp(payload.get('due_at'))} зафиксировать итог решения по кейсу.",
            "",
            "Служебно",
            f"- Реакция до: {_format_timestamp(payload.get('reaction_deadline_at'))}",
            f"- Идентификатор кейса: {payload.get('dedupe_key')}",
        ]
        return "\n".join(lines)

    if str(payload.get("rule_code") or "") in BATCH_RULE_CODES:
        return _build_receivable_batch_description(payload)

    lines = [
        f"Источник: {payload.get('source_type')}",
        f"Правило: {payload.get('rule_code')}",
        f"Серьёзность: {payload.get('severity')}",
        f"Сущность: {payload.get('entity_name') or payload.get('entity_ref')}",
        f"Сводка: {payload.get('summary')}",
        f"Реакция до: {_format_timestamp(payload.get('reaction_deadline_at'))}",
        f"Закрыть до: {_format_timestamp(payload.get('due_at'))}",
        f"Dedupe key: {payload.get('dedupe_key')}",
    ]

    metrics = payload.get("metrics") or {}
    if metrics:
        lines.append("")
        lines.append("Метрики:")
        for key, value in sorted(metrics.items()):
            lines.append(f"- {key}: {value}")

    references = payload.get("references") or []
    if references:
        lines.append("")
        lines.append("Ссылки:")
        for item in references:
            rendered = ", ".join(
                f"{key}={value}" for key, value in item.items() if value not in (None, "", [], {})
            )
            lines.append(f"- {rendered}")

    tags = payload.get("tags") or []
    if tags:
        lines.append("")
        lines.append(f"Теги: {', '.join(tags)}")

    return "\n".join(lines)


def _task_fields(
    payload: dict[str, Any],
    *,
    assignee_id: int,
    observer_ids: list[int],
    created_by_id: int | None = None,
) -> list[tuple[str, str]]:
    params = [
        (
            "fields[TITLE]",
            str(payload.get("title") or payload.get("dedupe_key") or "Management task"),
        ),
        ("fields[DESCRIPTION]", _build_description(payload)),
        ("fields[RESPONSIBLE_ID]", str(assignee_id)),
        ("fields[DEADLINE]", str(payload.get("due_at") or "")),
    ]
    if _parse_bool(payload.get("allow_assignee_change_deadline"), default=False):
        params.append(("fields[ALLOW_CHANGE_DEADLINE]", "Y"))
    if created_by_id is not None:
        params.append(("fields[CREATED_BY]", str(created_by_id)))
    for observer_id in observer_ids:
        params.append(("fields[AUDITORS][]", str(observer_id)))
    return params


def _create_b24_task(
    *,
    webhook_url: str,
    payload: dict[str, Any],
    assignee_id: int,
    observer_ids: list[int],
    created_by_id: int | None = None,
) -> int:
    try:
        response = _b24_call(
            webhook_url,
            "tasks.task.add",
            _task_fields(
                payload,
                assignee_id=assignee_id,
                observer_ids=observer_ids,
                created_by_id=created_by_id,
            ),
        )
    except Exception as error:
        if _is_fallbackable_create_error(error):
            raise FallbackableTaskSyncError(
                "Bitrix24 tasks.task.add failed before durable side effect"
            ) from error
        raise
    result = response.get("result")
    task_id = _to_int(result)
    if task_id is None and isinstance(result, dict):
        task_id = _to_int(result.get("task", {}).get("id"))
        if task_id is None:
            task_id = _to_int(result.get("id"))
    if task_id is None:
        raise RuntimeError("Bitrix24 tasks.task.add returned empty result")
    return task_id


def _update_b24_task(
    *,
    webhook_url: str,
    task_id: int,
    payload: dict[str, Any],
    assignee_id: int,
    observer_ids: list[int],
    created_by_id: int | None = None,
) -> None:
    params = [("taskId", str(task_id))]
    params.extend(
        _task_fields(
            payload,
            assignee_id=assignee_id,
            observer_ids=observer_ids,
            created_by_id=created_by_id,
        )
    )
    _b24_call(webhook_url, "tasks.task.update", params)


def _build_sync_summary(anchor_date: date, payloads: list[dict[str, Any]]) -> dict[str, Any]:
    by_rule: dict[str, int] = {}
    for item in payloads:
        rule_code = str(item.get("rule_code") or "unknown")
        by_rule[rule_code] = by_rule.get(rule_code, 0) + 1
    return {
        "anchor_date": anchor_date.isoformat(),
        "payload_count": len(payloads),
        "by_rule": dict(sorted(by_rule.items())),
        "by_contour": {},
        "created": 0,
        "updated": 0,
        "noop": 0,
        "actions": [],
    }


def _record_contour_action(summary: dict[str, Any], contour: str, action: str) -> None:
    contour_stats = summary.setdefault("by_contour", {}).setdefault(
        contour,
        {"created": 0, "updated": 0, "noop": 0},
    )
    if action in {"create", "recreate", "fallback_create"}:
        contour_stats["created"] += 1
    elif action in {"update", "fallback_update"}:
        contour_stats["updated"] += 1
    elif action == "noop":
        contour_stats["noop"] += 1


def _to_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _format_amount(value: Decimal) -> str:
    quantized = value.quantize(Decimal("1"))
    return f"{quantized:,.0f} ₽".replace(",", " ")


def _collapse_receivable_new_daily(
    payloads: list[dict[str, Any]], *, anchor_date: date
) -> list[dict[str, Any]]:
    batch_candidates = [
        item for item in payloads if str(item.get("rule_code") or "") == RULE_RECEIVABLE_NEW_DAILY
    ]
    if len(batch_candidates) <= 1:
        return payloads

    other_payloads = [
        item for item in payloads if str(item.get("rule_code") or "") != RULE_RECEIVABLE_NEW_DAILY
    ]
    ranked = sorted(
        batch_candidates,
        key=lambda item: (
            _to_decimal((item.get("metrics") or {}).get("current_balance")),
            str(item.get("entity_name") or item.get("entity_ref") or ""),
        ),
        reverse=True,
    )
    total_balance = sum(
        (_to_decimal((item.get("metrics") or {}).get("current_balance")) for item in ranked),
        start=Decimal("0"),
    )
    owner_code = str(ranked[0].get("owner_code") or "finance")
    watcher_codes = sorted(
        {str(code) for item in ranked for code in (item.get("watcher_codes") or []) if str(code)}
    )
    severity = max(
        (str(item.get("severity") or "warning") for item in ranked),
        key=lambda value: SEVERITY_ORDER.get(value, 0),
    )
    reaction_deadline_candidates = [
        str(item.get("reaction_deadline_at") or "")
        for item in ranked
        if item.get("reaction_deadline_at")
    ]
    due_at_candidates = [str(item.get("due_at") or "") for item in ranked if item.get("due_at")]
    reaction_deadline_at = (
        min(reaction_deadline_candidates) if reaction_deadline_candidates else None
    )
    due_at = max(due_at_candidates) if due_at_candidates else None
    top_items = [
        f"{item.get('entity_name') or item.get('entity_ref')}: "
        f"{_format_amount(_to_decimal((item.get('metrics') or {}).get('current_balance')))}"
        for item in ranked[:10]
    ]
    references = []
    for item in ranked:
        references.append(
            {
                "kind": "receivable_case_batch_item",
                "counterparty_ref": item.get("entity_ref"),
                "counterparty_name": item.get("entity_name"),
                "current_balance": str((item.get("metrics") or {}).get("current_balance") or ""),
                "original_task_key": item.get("dedupe_key"),
            }
        )
    batch_payload = {
        "rule_code": RULE_RECEIVABLE_NEW_DAILY_BATCH,
        "source_type": "receivable_case_batch",
        "entity_ref": f"receivables:new_daily:{anchor_date.isoformat()}",
        "entity_name": f"Новая дебиторка за {anchor_date.isoformat()}",
        "severity": severity,
        "owner_code": owner_code,
        "watcher_codes": watcher_codes,
        "title": f"Дебиторка: новая дебиторка за {anchor_date.isoformat()}",
        "summary": (
            f"Новых контрагентов с дебиторкой: {len(ranked)} на {_format_amount(total_balance)}. "
            f"Топ: {'; '.join(top_items)}."
        ),
        "reaction_deadline_at": reaction_deadline_at,
        "due_at": due_at,
        "dedupe_key": f"{RULE_RECEIVABLE_NEW_DAILY_BATCH}|{anchor_date.isoformat()}",
        "tags": ["management", "receivables", RULE_RECEIVABLE_NEW_DAILY, "batch"],
        "metrics": {
            "counterparty_count": len(ranked),
            "current_balance_total": str(total_balance.quantize(Decimal("0.01"))),
        },
        "references": references,
    }
    return other_payloads + [batch_payload]


def _collapse_receivable_overdue(
    payloads: list[dict[str, Any]], *, anchor_date: date, overdue_batch_weekday: int
) -> list[dict[str, Any]]:
    batch_candidates = [
        item for item in payloads if str(item.get("rule_code") or "") == RULE_RECEIVABLE_OVERDUE
    ]
    if not batch_candidates:
        return payloads

    if anchor_date.isoweekday() != overdue_batch_weekday:
        return [
            item for item in payloads if str(item.get("rule_code") or "") != RULE_RECEIVABLE_OVERDUE
        ]

    other_payloads = [
        item for item in payloads if str(item.get("rule_code") or "") != RULE_RECEIVABLE_OVERDUE
    ]
    ranked = sorted(
        batch_candidates,
        key=lambda item: (
            _to_decimal((item.get("metrics") or {}).get("current_balance")),
            str(item.get("entity_name") or item.get("entity_ref") or ""),
        ),
        reverse=True,
    )
    total_balance = sum(
        (_to_decimal((item.get("metrics") or {}).get("current_balance")) for item in ranked),
        start=Decimal("0"),
    )
    watcher_codes = sorted(
        {str(code) for item in ranked for code in (item.get("watcher_codes") or []) if str(code)}
        | {"finance"}
    )
    severity = max(
        (str(item.get("severity") or "warning") for item in ranked),
        key=lambda value: SEVERITY_ORDER.get(value, 0),
    )
    reaction_deadline_candidates = [
        str(item.get("reaction_deadline_at") or "")
        for item in ranked
        if item.get("reaction_deadline_at")
    ]
    due_at_candidates = [str(item.get("due_at") or "") for item in ranked if item.get("due_at")]
    reaction_deadline_at = (
        min(reaction_deadline_candidates) if reaction_deadline_candidates else None
    )
    due_at = max(due_at_candidates) if due_at_candidates else None

    aged_bucket_counts: dict[str, int] = {}
    aged_bucket_totals: dict[str, str] = {}
    running_totals: dict[str, Decimal] = {}
    for item in ranked:
        metrics = item.get("metrics") or {}
        bucket = str(metrics.get("aged_bucket") or "unknown")
        aged_bucket_counts[bucket] = aged_bucket_counts.get(bucket, 0) + 1
        running_totals[bucket] = running_totals.get(bucket, Decimal("0")) + _to_decimal(
            metrics.get("current_balance")
        )
    for bucket, amount in running_totals.items():
        aged_bucket_totals[bucket] = str(amount.quantize(Decimal("0.01")))

    top_items = [
        f"{item.get('entity_name') or item.get('entity_ref')}: "
        f"{_format_amount(_to_decimal((item.get('metrics') or {}).get('current_balance')))}"
        for item in ranked[:20]
    ]
    references = []
    for item in ranked:
        metrics = item.get("metrics") or {}
        references.append(
            {
                "kind": "receivable_case_batch_item",
                "counterparty_ref": item.get("entity_ref"),
                "counterparty_name": item.get("entity_name"),
                "current_balance": str(metrics.get("current_balance") or ""),
                "due_date": metrics.get("due_date"),
                "overdue_days": metrics.get("overdue_days"),
                "payment_term_source": metrics.get("payment_term_source"),
                "aged_bucket": metrics.get("aged_bucket"),
                "shipment_ban": bool(metrics.get("shipment_ban")),
                "original_task_key": item.get("dedupe_key"),
            }
        )
    batch_payload = {
        "rule_code": RULE_RECEIVABLE_OVERDUE_BATCH,
        "source_type": "receivable_case_batch",
        "entity_ref": f"receivables:overdue:{anchor_date.isoformat()}",
        "entity_name": f"Просроченная дебиторка за {anchor_date.isoformat()}",
        "severity": severity,
        "owner_code": "cco",
        "watcher_codes": watcher_codes,
        "title": f"Дебиторка: просроченная дебиторка на {anchor_date.isoformat()}",
        "summary": (
            f"Просроченных контрагентов: {len(ranked)} на {_format_amount(total_balance)}. "
            f"Разбивка: {_format_bucket_breakdown({'aged_bucket_counts': aged_bucket_counts, 'aged_bucket_totals': aged_bucket_totals})}. "
            f"Топ: {'; '.join(top_items[:10])}."
        ),
        "reaction_deadline_at": reaction_deadline_at,
        "due_at": due_at,
        "dedupe_key": f"{RULE_RECEIVABLE_OVERDUE_BATCH}|{anchor_date.isoformat()}",
        "tags": ["management", "receivables", RULE_RECEIVABLE_OVERDUE, "batch"],
        "metrics": {
            "counterparty_count": len(ranked),
            "current_balance_total": str(total_balance.quantize(Decimal("0.01"))),
            "aged_bucket_counts": aged_bucket_counts,
            "aged_bucket_totals": aged_bucket_totals,
        },
        "references": references,
    }
    return other_payloads + [batch_payload]


def _collapse_receivable_case_batch(
    payloads: list[dict[str, Any]],
    *,
    anchor_date: date,
    source_rule_code: str,
    batch_rule_code: str,
    owner_code: str,
    title_prefix: str,
    source_type: str,
    watcher_codes_seed: set[str] | None = None,
) -> list[dict[str, Any]]:
    batch_candidates = [
        item for item in payloads if str(item.get("rule_code") or "") == source_rule_code
    ]
    if not batch_candidates:
        return payloads

    other_payloads = [
        item for item in payloads if str(item.get("rule_code") or "") != source_rule_code
    ]
    ranked = sorted(
        batch_candidates,
        key=lambda item: (
            _to_decimal((item.get("metrics") or {}).get("current_balance")),
            str(item.get("entity_name") or item.get("entity_ref") or ""),
        ),
        reverse=True,
    )
    total_balance = sum(
        (_to_decimal((item.get("metrics") or {}).get("current_balance")) for item in ranked),
        start=Decimal("0"),
    )
    watchers = {
        str(code) for item in ranked for code in (item.get("watcher_codes") or []) if str(code)
    }
    if watcher_codes_seed:
        watchers |= watcher_codes_seed
    severity = max(
        (str(item.get("severity") or "warning") for item in ranked),
        key=lambda value: SEVERITY_ORDER.get(value, 0),
    )
    reaction_deadline_candidates = [
        str(item.get("reaction_deadline_at") or "")
        for item in ranked
        if item.get("reaction_deadline_at")
    ]
    due_at_candidates = [str(item.get("due_at") or "") for item in ranked if item.get("due_at")]
    reaction_deadline_at = (
        min(reaction_deadline_candidates) if reaction_deadline_candidates else None
    )
    due_at = max(due_at_candidates) if due_at_candidates else None

    top_items = [
        f"{item.get('entity_name') or item.get('entity_ref')}: "
        f"{_format_amount(_to_decimal((item.get('metrics') or {}).get('current_balance')))}"
        for item in ranked[:10]
    ]
    references = []
    for item in ranked:
        metrics = item.get("metrics") or {}
        refs = item.get("references") or []
        origin_ref = next((ref for ref in refs if ref.get("kind") == "origin_document"), {})
        references.append(
            {
                "kind": "receivable_case_batch_item",
                "counterparty_ref": item.get("entity_ref"),
                "counterparty_name": item.get("entity_name"),
                "current_balance": str(metrics.get("current_balance") or ""),
                "aged_bucket": metrics.get("aged_bucket"),
                "activity_segment": metrics.get("activity_segment"),
                "current_manager_ref": next(
                    (
                        ref.get("current_manager_ref")
                        for ref in refs
                        if ref.get("kind") == "current_manager"
                    ),
                    None,
                ),
                "current_manager_name": next(
                    (
                        ref.get("current_manager_name")
                        for ref in refs
                        if ref.get("kind") == "current_manager"
                    ),
                    None,
                ),
                "origin_document_ref": origin_ref.get("document_ref"),
                "origin_document_number": origin_ref.get("document_number"),
                "original_task_key": item.get("dedupe_key"),
            }
        )
    batch_payload = {
        "rule_code": batch_rule_code,
        "source_type": source_type,
        "entity_ref": f"receivables:{batch_rule_code}:{anchor_date.isoformat()}",
        "entity_name": f"{title_prefix} за {anchor_date.isoformat()}",
        "severity": severity,
        "owner_code": owner_code,
        "watcher_codes": sorted(watchers),
        "title": f"Дебиторка: {title_prefix.lower()} за {anchor_date.isoformat()}",
        "summary": (
            f"Контрагентов: {len(ranked)} на {_format_amount(total_balance)}. "
            f"Топ: {'; '.join(top_items)}."
        ),
        "reaction_deadline_at": reaction_deadline_at,
        "due_at": due_at,
        "dedupe_key": f"{batch_rule_code}|{anchor_date.isoformat()}",
        "tags": ["management", "receivables", source_rule_code, "batch"],
        "metrics": {
            "counterparty_count": len(ranked),
            "current_balance_total": str(total_balance.quantize(Decimal("0.01"))),
        },
        "references": references,
    }
    return other_payloads + [batch_payload]


def _collapse_receivable_finance_daily_batch(
    payloads: list[dict[str, Any]], *, anchor_date: date
) -> list[dict[str, Any]]:
    batch_candidates = [
        item
        for item in payloads
        if str(item.get("rule_code") or "") in FINANCE_DAILY_BATCH_SOURCE_RULES
    ]
    if not batch_candidates:
        return payloads

    other_payloads = [
        item
        for item in payloads
        if str(item.get("rule_code") or "") not in FINANCE_DAILY_BATCH_SOURCE_RULES
    ]
    ranked = sorted(
        batch_candidates,
        key=lambda item: (
            _to_decimal((item.get("metrics") or {}).get("current_balance_total")),
            str(item.get("title") or item.get("entity_name") or item.get("entity_ref") or ""),
        ),
        reverse=True,
    )
    total_balance = sum(
        (_to_decimal((item.get("metrics") or {}).get("current_balance_total")) for item in ranked),
        start=Decimal("0"),
    )
    total_count = sum(
        _to_int((item.get("metrics") or {}).get("counterparty_count")) or 0 for item in ranked
    )
    watchers = {
        str(code) for item in ranked for code in (item.get("watcher_codes") or []) if str(code)
    }
    severity = max(
        (str(item.get("severity") or "warning") for item in ranked),
        key=lambda value: SEVERITY_ORDER.get(value, 0),
    )
    reaction_deadline_candidates = [
        str(item.get("reaction_deadline_at") or "")
        for item in ranked
        if item.get("reaction_deadline_at")
    ]
    due_at_candidates = [str(item.get("due_at") or "") for item in ranked if item.get("due_at")]
    reaction_deadline_at = (
        min(reaction_deadline_candidates) if reaction_deadline_candidates else None
    )
    due_at = max(due_at_candidates) if due_at_candidates else None

    section_metrics: dict[str, dict[str, str | int]] = {}
    references: list[dict[str, Any]] = []
    for item in ranked:
        rule_code = str(item.get("rule_code") or "")
        metrics = item.get("metrics") or {}
        section_metrics[rule_code] = {
            "counterparty_count": _to_int(metrics.get("counterparty_count")) or 0,
            "current_balance_total": str(
                _to_decimal(metrics.get("current_balance_total")).quantize(Decimal("0.01"))
            ),
        }
        for ref in item.get("references") or []:
            references.append({**ref, "batch_rule_code": rule_code})

    batch_payload = {
        "rule_code": RULE_RECEIVABLE_FINANCE_DAILY_BATCH,
        "source_type": "receivable_case_batch",
        "entity_ref": f"receivables:finance_daily:{anchor_date.isoformat()}",
        "entity_name": f"Ежедневный финансовый пакет дебиторки за {anchor_date.isoformat()}",
        "severity": severity,
        "owner_code": "finance",
        "watcher_codes": sorted(watchers),
        "title": f"Дебиторка: единый финансовый пакет за {anchor_date.isoformat()}",
        "summary": (
            f"Единый daily-пакет по дебиторке: {total_count} кейсов на "
            f"{_format_amount(total_balance)}."
        ),
        "reaction_deadline_at": reaction_deadline_at,
        "due_at": due_at,
        "dedupe_key": f"{RULE_RECEIVABLE_FINANCE_DAILY_BATCH}|{anchor_date.isoformat()}",
        "tags": ["management", "receivables", "finance_daily", "batch"],
        "metrics": {
            "counterparty_count": total_count,
            "current_balance_total": str(total_balance.quantize(Decimal("0.01"))),
            "sections": section_metrics,
        },
        "references": references,
    }
    return other_payloads + [batch_payload]


def _filter_suppressed_payloads(payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        item
        for item in payloads
        if str(item.get("rule_code") or "") not in SUPPRESSED_RECEIVABLE_RULE_CODES
    ]


def _build_state_entry(
    *,
    task_id: int,
    fingerprint: str,
    payload: dict[str, Any],
    responsible_id: int,
    observer_ids: list[int],
    created_by_id: int | None,
    delivery_intent: str,
    contour: str,
    migration_state: str,
) -> dict[str, Any]:
    return {
        "task_id": task_id,
        "fingerprint": fingerprint,
        "title": payload.get("title"),
        "rule_code": payload.get("rule_code"),
        "responsible_id": responsible_id,
        "observer_ids": observer_ids,
        "created_by_id": created_by_id,
        "delivery_intent": delivery_intent,
        "delivery_channel": DELIVERY_CHANNEL_BITRIX_TASK,
        "bitrix_contour": contour,
        "migration_state": migration_state,
        "updated_at": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
    }


def _resolve_observer_ids(
    *,
    payload: dict[str, Any],
    responsible_id: int,
    watcher_overrides: Mapping[str, Any],
    team_roles: Mapping[str, Any],
    default_observer_ids: list[int],
    contour: str,
) -> list[int]:
    observer_ids: list[int] = []
    seen_observers: set[int] = set()
    for watcher_code in payload.get("watcher_codes") or []:
        watcher_id = _resolve_role_id(
            str(watcher_code),
            overrides=watcher_overrides,
            team_roles=team_roles,
            contour=contour,
            default_id=None,
        )
        if watcher_id is None or watcher_id == responsible_id or watcher_id in seen_observers:
            continue
        observer_ids.append(watcher_id)
        seen_observers.add(watcher_id)
    if bool(payload.get("suppress_default_observers")):
        return observer_ids
    for observer_id in default_observer_ids:
        if observer_id != responsible_id and observer_id not in seen_observers:
            observer_ids.append(observer_id)
            seen_observers.add(observer_id)
    return observer_ids


def _legacy_state_aliases(
    *,
    dedupe_key: str,
    contour: str,
    delivery_intent: str,
) -> list[str]:
    aliases = [_delivery_state_key(dedupe_key, delivery_intent=delivery_intent, contour=contour)]
    if contour == BITRIX_CONTOUR_CLOUD and delivery_intent == DELIVERY_INTENT_TASK:
        aliases.append(dedupe_key)
    return aliases


def _get_current_task_state(
    task_state: Mapping[str, Any],
    *,
    dedupe_key: str,
    contour: str,
    delivery_intent: str,
) -> tuple[str, str | None, dict[str, Any] | None]:
    primary_key = _delivery_state_key(
        dedupe_key,
        delivery_intent=delivery_intent,
        contour=contour,
    )
    if primary_key in task_state:
        current = task_state.get(primary_key)
        return primary_key, None, current if isinstance(current, dict) else None

    if contour == BITRIX_CONTOUR_CLOUD and delivery_intent == DELIVERY_INTENT_TASK:
        legacy_key = dedupe_key
        current = task_state.get(legacy_key)
        return (
            primary_key,
            legacy_key if isinstance(current, dict) else None,
            current if isinstance(current, dict) else None,
        )

    return primary_key, None, None


def _build_delivery_targets(
    *,
    webhook_url: str | None,
    owner_overrides: Mapping[str, Any],
    watcher_overrides: Mapping[str, Any],
    default_responsible_id: int | None,
    default_observer_ids: list[int],
    default_created_by_id: int | None,
    disk_folder_id: int | None,
    delivery_targets: list[dict[str, Any]] | None,
) -> list[dict[str, Any]]:
    if delivery_targets:
        return delivery_targets
    if not webhook_url:
        raise RuntimeError("Missing Bitrix24 webhook URL for management task delivery")
    return [
        {
            "contour": BITRIX_CONTOUR_CLOUD,
            "mode": "primary",
            "webhook_url": webhook_url,
            "owner_overrides": dict(owner_overrides),
            "watcher_overrides": dict(watcher_overrides),
            "default_responsible_id": default_responsible_id,
            "default_observer_ids": list(default_observer_ids),
            "default_created_by_id": default_created_by_id,
            "disk_folder_id": disk_folder_id,
        }
    ]


def _build_contour_target_from_env(
    env: Mapping[str, str],
    *,
    contour: str,
    team_roles: Mapping[str, Any],
) -> dict[str, Any] | None:
    normalized_contour = _normalize_contour(contour)
    contour_prefix = f"MANAGEMENT_B24_{normalized_contour.upper()}_"

    if normalized_contour == BITRIX_CONTOUR_CLOUD:
        webhook_url = (
            env.get(f"{contour_prefix}WEBHOOK_URL")
            or env.get("MANAGEMENT_B24_WEBHOOK_URL")
            or env.get("BITRIX24_WEBHOOK_URL")
        )
        disk_folder_id = _to_int(
            env.get(f"{contour_prefix}DISK_FOLDER_ID") or env.get("MANAGEMENT_B24_DISK_FOLDER_ID")
        )
        if disk_folder_id is None:
            disk_folder_id = 3
    else:
        webhook_url = (
            env.get(f"{contour_prefix}WEBHOOK_URL")
            or env.get("MANAGEMENT_B24_BOX_WEBHOOK_URL")
            or env.get("BITRIX24_BOX_WEBHOOK_URL")
        )
        disk_folder_id = _to_int(
            env.get(f"{contour_prefix}DISK_FOLDER_ID")
            or env.get("MANAGEMENT_B24_BOX_DISK_FOLDER_ID")
        )

    if not webhook_url:
        return None

    if normalized_contour == BITRIX_CONTOUR_CLOUD:
        owner_overrides = _parse_code_overrides(
            env.get(f"{contour_prefix}OWNER_OVERRIDES") or env.get("MANAGEMENT_B24_OWNER_OVERRIDES")
        )
        watcher_overrides = _parse_code_overrides(
            env.get(f"{contour_prefix}WATCHER_OVERRIDES")
            or env.get("MANAGEMENT_B24_WATCHER_OVERRIDES")
        )
        default_responsible_id = _to_int(
            env.get(f"{contour_prefix}DEFAULT_RESPONSIBLE_ID")
            or env.get("MANAGEMENT_B24_DEFAULT_RESPONSIBLE_ID")
        )
        default_observer_ids = _parse_csv_ints(
            env.get(f"{contour_prefix}DEFAULT_OBSERVER_IDS")
            or env.get("MANAGEMENT_B24_DEFAULT_OBSERVER_IDS")
        )
        default_created_by_id = _to_int(
            env.get(f"{contour_prefix}CREATED_BY_ID") or env.get("MANAGEMENT_B24_CREATED_BY_ID")
        )
    else:
        owner_overrides = _parse_code_overrides(env.get(f"{contour_prefix}OWNER_OVERRIDES"))
        watcher_overrides = _parse_code_overrides(env.get(f"{contour_prefix}WATCHER_OVERRIDES"))
        default_responsible_id = _to_int(
            env.get(f"{contour_prefix}DEFAULT_RESPONSIBLE_ID")
            or env.get("BITRIX24_BOX_RESPONSIBLE_ID")
        )
        default_observer_ids = _parse_csv_ints(env.get(f"{contour_prefix}DEFAULT_OBSERVER_IDS"))
        default_created_by_id = _to_int(env.get(f"{contour_prefix}CREATED_BY_ID"))
    if default_responsible_id is None:
        default_responsible_id = _resolve_team_role_id(
            team_roles.get("cfo"), contour=normalized_contour
        )
    return {
        "contour": normalized_contour,
        "mode": "primary",
        "webhook_url": webhook_url,
        "owner_overrides": owner_overrides,
        "watcher_overrides": watcher_overrides,
        "default_responsible_id": default_responsible_id,
        "default_observer_ids": default_observer_ids,
        "default_created_by_id": default_created_by_id,
        "disk_folder_id": disk_folder_id,
    }


def _build_env_delivery_targets(
    env: Mapping[str, str],
    *,
    team_roles: Mapping[str, Any],
) -> tuple[str, list[dict[str, Any]]]:
    migration_state = _normalize_migration_state(env.get("MANAGEMENT_B24_TASK_MIGRATION_STATE"))
    cloud_target = _build_contour_target_from_env(
        env,
        contour=BITRIX_CONTOUR_CLOUD,
        team_roles=team_roles,
    )
    box_target = _build_contour_target_from_env(
        env,
        contour=BITRIX_CONTOUR_BOX,
        team_roles=team_roles,
    )

    if migration_state == MIGRATION_STATE_CLOUD_PRIMARY:
        if not cloud_target:
            raise SystemExit(
                "Missing required env for cloud_primary: "
                "MANAGEMENT_B24_CLOUD_WEBHOOK_URL|MANAGEMENT_B24_WEBHOOK_URL|BITRIX24_WEBHOOK_URL"
            )
        return migration_state, [cloud_target]

    if migration_state == MIGRATION_STATE_BOX_SHADOW:
        if not cloud_target or not box_target:
            raise SystemExit(
                "Missing required env for box_shadow: cloud and box webhooks must both be configured"
            )
        box_target["mode"] = "shadow"
        return migration_state, [cloud_target, box_target]

    if migration_state == MIGRATION_STATE_BOX_PRIMARY:
        if not box_target:
            raise SystemExit(
                "Missing required env for box_primary: MANAGEMENT_B24_BOX_WEBHOOK_URL|BITRIX24_BOX_WEBHOOK_URL"
            )
        return migration_state, [box_target]

    if not box_target or not cloud_target:
        raise SystemExit(
            "Missing required env for cloud_fallback: box and cloud webhooks must both be configured"
        )
    box_target["fallback"] = {
        **cloud_target,
        "mode": "fallback",
    }
    return migration_state, [box_target]


def _sync_payload_to_target(
    *,
    payload: dict[str, Any],
    target: dict[str, Any],
    task_state: dict[str, Any],
    summary: dict[str, Any],
    anchor_date: date,
    dry_run: bool,
    report_dir: Path,
    delivery_intent: str,
    migration_state: str,
    create_task: Callable[..., int],
    update_task: Callable[..., None],
    persist_state: Callable[[], None] | None = None,
) -> bool:
    dedupe_key = str(payload.get("dedupe_key") or "")
    contour = _normalize_contour(target.get("contour"))
    primary_key, legacy_key, current = _get_current_task_state(
        task_state,
        dedupe_key=dedupe_key,
        contour=contour,
        delivery_intent=delivery_intent,
    )

    owner_overrides = target.get("owner_overrides") or {}
    watcher_overrides = target.get("watcher_overrides") or {}
    team_roles = target.get("team_roles") or {}
    webhook_url = str(target.get("webhook_url") or "").strip()
    default_responsible_id = _to_int(target.get("default_responsible_id"))
    default_observer_ids = list(target.get("default_observer_ids") or [])
    default_created_by_id = _to_int(target.get("default_created_by_id"))
    disk_folder_id = _to_int(target.get("disk_folder_id"))

    responsible_id = _resolve_role_id(
        str(payload.get("owner_code") or ""),
        overrides=owner_overrides,
        team_roles=team_roles,
        contour=contour,
        default_id=default_responsible_id,
    )
    if responsible_id is None:
        raise RuntimeError(
            f"Cannot resolve responsible user for owner_code={payload.get('owner_code')} and contour={contour}"
        )

    observer_ids = _resolve_observer_ids(
        payload=payload,
        responsible_id=responsible_id,
        watcher_overrides=watcher_overrides,
        team_roles=team_roles,
        default_observer_ids=default_observer_ids,
        contour=contour,
    )
    created_by_id = _resolve_created_by_id(
        payload=payload,
        owner_overrides=owner_overrides,
        team_roles=team_roles,
        contour=contour,
        default_created_by_id=default_created_by_id,
    )

    fingerprint = _fingerprint_payload(payload)
    needs_batch_attachment = (
        disk_folder_id is not None
        and str(payload.get("rule_code") or "") in BATCH_RULE_CODES
        and (not current or not current.get("attachment_id"))
    )

    action_record = {
        "dedupe_key": dedupe_key,
        "state_key": primary_key,
        "title": payload.get("title"),
        "responsible_id": responsible_id,
        "observer_ids": observer_ids,
        "created_by_id": created_by_id,
        "delivery_intent": delivery_intent,
        "delivery_channel": DELIVERY_CHANNEL_BITRIX_TASK,
        "bitrix_contour": contour,
        "delivery_mode": str(target.get("mode") or "primary"),
    }

    def _attach_batch_report(task_id: int) -> None:
        report_path = (
            report_dir
            / contour
            / anchor_date.isoformat()
            / _build_receivable_batch_report_filename(payload)
        )
        _export_receivable_batch_xlsx(payload, report_path)
        file_object_id = _upload_b24_disk_file(
            webhook_url=webhook_url,
            folder_id=disk_folder_id,
            file_path=report_path,
        )
        attachment_id = _attach_b24_file_to_task(
            webhook_url=webhook_url,
            task_id=task_id,
            file_object_id=file_object_id,
        )
        task_state[primary_key]["report_path"] = str(report_path)
        task_state[primary_key]["disk_object_id"] = file_object_id
        task_state[primary_key]["attachment_id"] = attachment_id

    def _persist_partial_state() -> None:
        if not dry_run and persist_state is not None:
            persist_state()

    if current and current.get("fingerprint") == fingerprint and current.get("task_id"):
        if needs_batch_attachment:
            task_id = int(current["task_id"])
            action_record["action"] = "attach_report"
            action_record["task_id"] = task_id
            task_state[primary_key] = {
                **current,
                **_build_state_entry(
                    task_id=task_id,
                    fingerprint=fingerprint,
                    payload=payload,
                    responsible_id=responsible_id,
                    observer_ids=observer_ids,
                    created_by_id=created_by_id,
                    delivery_intent=delivery_intent,
                    contour=contour,
                    migration_state=migration_state,
                ),
                "attachment_id": current.get("attachment_id"),
                "disk_object_id": current.get("disk_object_id"),
                "report_path": current.get("report_path"),
            }
            if dry_run:
                summary["actions"].append(action_record)
                _record_contour_action(summary, contour, "update")
                summary["updated"] += 1
                return False
            if legacy_key and legacy_key != primary_key:
                task_state.pop(legacy_key, None)
            _persist_partial_state()
            try:
                _attach_batch_report(task_id)
            except Exception as error:
                raise PartialTaskDeliveryError(
                    f"Batch attachment failed after task sync for dedupe_key={dedupe_key} "
                    f"and contour={contour}"
                ) from error
            summary["actions"].append(action_record)
            _record_contour_action(summary, contour, "update")
            summary["updated"] += 1
            return True

        summary["noop"] += 1
        action_record["action"] = "noop"
        action_record["task_id"] = current.get("task_id")
        summary["actions"].append(action_record)
        _record_contour_action(summary, contour, "noop")
        if legacy_key and legacy_key != primary_key:
            task_state[primary_key] = {
                **current,
                "delivery_intent": delivery_intent,
                "delivery_channel": DELIVERY_CHANNEL_BITRIX_TASK,
                "bitrix_contour": contour,
                "migration_state": migration_state,
            }
            task_state.pop(legacy_key, None)
            return True
        return False

    if dry_run:
        action = "update" if current and current.get("task_id") else "create"
        action_record["action"] = action
        action_record["task_id"] = current.get("task_id") if current else None
        if action == "update":
            summary["updated"] += 1
        else:
            summary["created"] += 1
        summary["actions"].append(action_record)
        _record_contour_action(summary, contour, action)
        return False

    if current and current.get("task_id"):
        task_id = int(current["task_id"])
        try:
            update_task(
                webhook_url=webhook_url,
                task_id=task_id,
                payload=payload,
                assignee_id=responsible_id,
                observer_ids=observer_ids,
                created_by_id=created_by_id,
            )
        except Exception as error:
            raise PartialTaskDeliveryError(
                f"Bitrix task update failed for dedupe_key={dedupe_key} and contour={contour}"
            ) from error
        else:
            summary["updated"] += 1
            action_record["action"] = "update"
            _record_contour_action(summary, contour, "update")
        action_record["task_id"] = task_id
    else:
        try:
            task_id = create_task(
                webhook_url=webhook_url,
                payload=payload,
                assignee_id=responsible_id,
                observer_ids=observer_ids,
                created_by_id=created_by_id,
            )
        except FallbackableTaskSyncError:
            raise
        except Exception as error:
            raise UnsafeToFallbackTaskSyncError(
                f"Bitrix task create result is uncertain for dedupe_key={dedupe_key} and contour={contour}"
            ) from error
        summary["created"] += 1
        action_record["action"] = "create"
        action_record["task_id"] = task_id
        _record_contour_action(summary, contour, "create")

    task_state[primary_key] = _build_state_entry(
        task_id=task_id,
        fingerprint=fingerprint,
        payload=payload,
        responsible_id=responsible_id,
        observer_ids=observer_ids,
        created_by_id=created_by_id,
        delivery_intent=delivery_intent,
        contour=contour,
        migration_state=migration_state,
    )
    if legacy_key and legacy_key != primary_key:
        task_state.pop(legacy_key, None)

    if disk_folder_id is not None and str(payload.get("rule_code") or "") in BATCH_RULE_CODES:
        _persist_partial_state()
        try:
            _attach_batch_report(task_id)
        except Exception as error:
            raise PartialTaskDeliveryError(
                f"Batch attachment failed after task sync for dedupe_key={dedupe_key} "
                f"and contour={contour}"
            ) from error

    summary["actions"].append(action_record)
    return True


def sync_management_tasks(
    *,
    fetch_json: Callable[[str, dict[str, str]], dict[str, Any]],
    webhook_url: str | None,
    anchor_date: date,
    state_path: Path,
    owner_overrides: Mapping[str, Any],
    watcher_overrides: Mapping[str, Any],
    team_roles: Mapping[str, Any],
    default_responsible_id: int | None,
    default_observer_ids: list[int],
    default_created_by_id: int | None,
    report_dir: Path,
    disk_folder_id: int | None,
    overdue_batch_weekday: int = 1,
    dry_run: bool = False,
    create_task: Callable[..., int] = _create_b24_task,
    update_task: Callable[..., None] = _update_b24_task,
    delivery_targets: list[dict[str, Any]] | None = None,
    delivery_intent: str = DELIVERY_INTENT_TASK,
    migration_state: str = MIGRATION_STATE_CLOUD_PRIMARY,
) -> dict[str, Any]:
    response = fetch_json("/api/management/task-payloads", {"date": anchor_date.isoformat()})
    payloads = list(response.get("payload", []))
    payloads = _filter_suppressed_payloads(payloads)
    payloads = _collapse_receivable_new_daily(payloads, anchor_date=anchor_date)
    payloads = _collapse_receivable_overdue(
        payloads,
        anchor_date=anchor_date,
        overdue_batch_weekday=overdue_batch_weekday,
    )
    payloads = _collapse_receivable_case_batch(
        payloads,
        anchor_date=anchor_date,
        source_rule_code=RULE_RECEIVABLE_EMPLOYEE,
        batch_rule_code=RULE_RECEIVABLE_EMPLOYEE_BATCH,
        owner_code="finance",
        title_prefix="Долги сотрудников",
        source_type="receivable_case_batch",
        watcher_codes_seed={"hr"},
    )
    payloads = _collapse_receivable_case_batch(
        payloads,
        anchor_date=anchor_date,
        source_rule_code=RULE_RECEIVABLE_FIRED_MANAGER,
        batch_rule_code=RULE_RECEIVABLE_FIRED_MANAGER_BATCH,
        owner_code="finance_pool",
        title_prefix="Долги на уволенных менеджерах",
        source_type="receivable_case_batch",
    )
    payloads = _collapse_receivable_case_batch(
        payloads,
        anchor_date=anchor_date,
        source_rule_code=RULE_RECEIVABLE_ADJUSTMENT,
        batch_rule_code=RULE_RECEIVABLE_ADJUSTMENT_BATCH,
        owner_code="finance",
        title_prefix="Кейсы на корректировку",
        source_type="receivable_case_batch",
    )
    payloads = _filter_suppressed_payloads(payloads)
    summary = _build_sync_summary(anchor_date, payloads)
    summary["delivery_intent"] = delivery_intent
    summary["migration_state"] = _normalize_migration_state(migration_state)

    effective_targets = _build_delivery_targets(
        webhook_url=webhook_url,
        owner_overrides=owner_overrides,
        watcher_overrides=watcher_overrides,
        default_responsible_id=default_responsible_id,
        default_observer_ids=default_observer_ids,
        default_created_by_id=default_created_by_id,
        disk_folder_id=disk_folder_id,
        delivery_targets=delivery_targets,
    )

    state = _load_state(state_path)
    task_state = state.setdefault("tasks", {})
    legacy_receivable_rules = {
        RULE_RECEIVABLE_NEW_DAILY,
        RULE_RECEIVABLE_EMPLOYEE,
        RULE_RECEIVABLE_FIRED_MANAGER,
        RULE_RECEIVABLE_ADJUSTMENT,
    }
    stale_keys = [
        key
        for key, item in list(task_state.items())
        if str(item.get("rule_code") or "") in legacy_receivable_rules
    ]
    for key in stale_keys:
        task_state.pop(key, None)
    changed = bool(stale_keys)

    for payload in payloads:
        dedupe_key = str(payload.get("dedupe_key") or "")
        if not dedupe_key:
            continue
        for target in effective_targets:
            target_with_context = {
                **target,
                "team_roles": team_roles,
            }
            try:
                target_changed = _sync_payload_to_target(
                    payload=payload,
                    target=target_with_context,
                    task_state=task_state,
                    summary=summary,
                    anchor_date=anchor_date,
                    dry_run=dry_run,
                    report_dir=report_dir,
                    delivery_intent=delivery_intent,
                    migration_state=summary["migration_state"],
                    create_task=create_task,
                    update_task=update_task,
                    persist_state=lambda: _save_state(state_path, state),
                )
            except PartialTaskDeliveryError:
                raise
            except UnsafeToFallbackTaskSyncError:
                raise
            except FallbackableTaskSyncError as error:
                fallback = target.get("fallback")
                if not fallback or dry_run:
                    raise
                fallback_with_context = {
                    **fallback,
                    "team_roles": team_roles,
                }
                target_changed = _sync_payload_to_target(
                    payload=payload,
                    target=fallback_with_context,
                    task_state=task_state,
                    summary=summary,
                    anchor_date=anchor_date,
                    dry_run=dry_run,
                    report_dir=report_dir,
                    delivery_intent=delivery_intent,
                    migration_state=summary["migration_state"],
                    create_task=create_task,
                    update_task=update_task,
                    persist_state=lambda: _save_state(state_path, state),
                )
                if summary["actions"]:
                    summary["actions"][-1][
                        "action"
                    ] = f"fallback_{summary['actions'][-1]['action']}"
                    summary["actions"][-1]["fallback_from_contour"] = target.get("contour")
                    summary["actions"][-1]["fallback_reason"] = str(error)
            changed = changed or target_changed
            if target_changed and not dry_run:
                _save_state(state_path, state)

    if changed and not dry_run:
        _save_state(state_path, state)

    return summary


def render_summary(summary: dict[str, Any]) -> str:
    lines = [
        f"Дата: {summary['anchor_date']}",
        f"Payload'ов: {summary['payload_count']}",
        f"Создано: {summary['created']}",
        f"Обновлено: {summary['updated']}",
        f"Без изменений: {summary['noop']}",
    ]
    if summary.get("disabled_reason"):
        lines.append(f"Отключено: {summary['disabled_reason']}")
    if summary.get("migration_state"):
        lines.append(f"Миграция Bitrix: {summary['migration_state']}")
    if summary["by_rule"]:
        rules = ", ".join(f"{key}={value}" for key, value in summary["by_rule"].items())
        lines.append(f"По правилам: {rules}")
    if summary.get("by_contour"):
        contour_parts = []
        for contour, counters in sorted(summary["by_contour"].items()):
            contour_parts.append(
                f"{contour}: create={counters.get('created', 0)}, "
                f"update={counters.get('updated', 0)}, noop={counters.get('noop', 0)}"
            )
        lines.append("По контурам: " + "; ".join(contour_parts))
    for item in summary["actions"][:10]:
        lines.append(
            f"- {item['action']}[{item.get('bitrix_contour', 'cloud')}]: "
            f"{item.get('title') or item['dedupe_key']} "
            f"(resp={item['responsible_id']}, task={item.get('task_id', '-')})"
        )
    return "\n".join(lines)


def main() -> None:
    args = _parse_args()
    env = _load_env(
        os.getenv("MANAGEMENT_B_ENV_FILE")
        or os.getenv("OPENCLAW_ENV_FILE")
        or os.getenv("PRICING_ENV_FILE")
        or DEFAULT_LOCAL_ENV_FILE
    )
    source_url = (
        env.get("MANAGEMENT_SOURCE_URL")
        or env.get("RETURN_SCHEME_SOURCE_URL")
        or DEFAULT_LOCAL_SOURCE_URL
    )
    source_token = (
        env.get("MANAGEMENT_SOURCE_TOKEN")
        or env.get("RETURN_SCHEME_SOURCE_TOKEN")
        or env.get("MANAGEMENT_INTERNAL_API_TOKEN")
        or env.get("RETURN_SCHEME_INTERNAL_API_TOKEN")
    )
    if not source_token:
        raise SystemExit(
            "Missing required env: MANAGEMENT_SOURCE_TOKEN|RETURN_SCHEME_SOURCE_TOKEN|"
            "MANAGEMENT_INTERNAL_API_TOKEN|RETURN_SCHEME_INTERNAL_API_TOKEN"
        )

    fetch_json = _build_fetcher(
        source_url=source_url,
        token=source_token,
        timeout=int(env.get("MANAGEMENT_ADAPTER_TIMEOUT_SECONDS", "20")),
        retries=int(env.get("MANAGEMENT_ADAPTER_RETRIES", "2")),
        retry_delay=float(env.get("MANAGEMENT_ADAPTER_RETRY_DELAY_SECONDS", "1.0")),
    )
    team_roles = _load_team_role_map(
        env.get(
            "MANAGEMENT_B24_TEAM_PATH",
            "/home/deploy/.openclaw/workspace/org/mastermobile/team.yaml",
        )
    )
    migration_state, delivery_targets = _build_env_delivery_targets(env, team_roles=team_roles)
    overdue_batch_weekday = _to_int(env.get("MANAGEMENT_RECEIVABLE_OVERDUE_WEEKDAY")) or 1
    if overdue_batch_weekday < 1 or overdue_batch_weekday > 7:
        overdue_batch_weekday = 1
    state_path = Path(
        env.get(
            "MANAGEMENT_TASKS_STATE_PATH",
            "/home/deploy/.openclaw/workspace/.data/management-tasks/state.json",
        )
    )
    report_dir = Path(
        env.get(
            "MANAGEMENT_TASKS_REPORT_DIR",
            "/home/deploy/.openclaw/workspace/.data/management-tasks/reports",
        )
    )
    anchor_date = _parse_date(args.anchor_date)

    if _env_flag(env.get("MANAGEMENT_TASKS_DISABLED")):
        summary = {
            "anchor_date": anchor_date.isoformat(),
            "payload_count": 0,
            "created": 0,
            "updated": 0,
            "noop": 0,
            "by_rule": {},
            "by_contour": {},
            "actions": [],
            "delivery_intent": DELIVERY_INTENT_TASK,
            "migration_state": _normalize_migration_state(migration_state),
            "disabled_reason": "MANAGEMENT_TASKS_DISABLED=true",
        }
        if args.json:
            print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
        else:
            print(render_summary(summary))
        return

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url=None,
        anchor_date=anchor_date,
        state_path=state_path,
        owner_overrides={},
        watcher_overrides={},
        team_roles=team_roles,
        default_responsible_id=None,
        default_observer_ids=[],
        default_created_by_id=None,
        report_dir=report_dir,
        disk_folder_id=None,
        overdue_batch_weekday=overdue_batch_weekday,
        dry_run=args.dry_run,
        delivery_targets=delivery_targets,
        migration_state=migration_state,
    )
    if args.json:
        print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    else:
        print(render_summary(summary))


if __name__ == "__main__":
    main()
