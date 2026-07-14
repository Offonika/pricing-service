#!/usr/bin/env python3
"""Pull new-daily receivables from server A and deliver a morning XLSX to Telegram."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_LOCAL_SOURCE_URL = "http://127.0.0.1:18080"
DEFAULT_LOCAL_ENV_FILE = "/opt/MM/pricing-service/.env"
DEFAULT_STATE_PATH = "/home/deploy/.openclaw/workspace/.data/new-daily-receivables/state.json"
DEFAULT_ARTIFACT_DIR = "/home/deploy/.openclaw/workspace/.data/new-daily-receivables/artifacts"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REPORT_KEY_PREFIX = "receivable-new-daily"
PENDING_STATE_SUFFIX = "|pending"
RETRYABLE_HTTP_STATUS_CODES = {408, 425, 429, 500, 502, 503, 504}


def _load_env(path: str | None) -> dict[str, str]:
    env = os.environ.copy()
    if not path:
        return env
    env_path = Path(path)
    if not env_path.exists():
        return env
    for raw in env_path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pull new-daily receivables from server A and send morning XLSX to Telegram."
    )
    parser.add_argument("--date", dest="anchor_date", help="Anchor date in YYYY-MM-DD format")
    parser.add_argument(
        "--dry-run", action="store_true", help="Render actions without side effects"
    )
    parser.add_argument("--json", action="store_true", help="Emit machine-readable summary")
    return parser.parse_args()


def _parse_date(value: str | None) -> date:
    if value:
        return date.fromisoformat(value)
    return date.today()


def _utcnow() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _http_json(
    url: str,
    *,
    headers: dict[str, str] | None = None,
    timeout: int,
) -> Any:
    request = urllib.request.Request(url, headers=headers or {})
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    with opener.open(request, timeout=timeout) as response:
        body = response.read().decode("utf-8")
        if not body:
            return {}
        return json.loads(body)


def _is_retryable_source_error(error: Exception) -> bool:
    if isinstance(error, urllib.error.HTTPError):
        return error.code in RETRYABLE_HTTP_STATUS_CODES
    return isinstance(error, (urllib.error.URLError, TimeoutError))


def _build_fetcher(
    *,
    source_url: str,
    token: str,
    timeout: int,
    retries: int,
    retry_delay: float,
) -> Callable[[str, dict[str, str]], Any]:
    base = source_url.rstrip("/")
    headers = {"Authorization": f"Bearer {token}"}

    def _fetch(path: str, params: dict[str, str]) -> Any:
        query = urllib.parse.urlencode(params)
        url = f"{base}{path}"
        if query:
            url = f"{url}?{query}"

        attempts = max(1, retries + 1)
        last_error: Exception | None = None
        for attempt in range(attempts):
            try:
                return _http_json(url, headers=headers, timeout=timeout)
            except (
                urllib.error.HTTPError,
                urllib.error.URLError,
                TimeoutError,
                ValueError,
            ) as error:
                last_error = error
                if attempt + 1 >= attempts:
                    break
                time.sleep(retry_delay)
        assert last_error is not None
        raise last_error

    return _fetch


def _payload_items(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, dict):
        items = payload.get("payload", [])
        if isinstance(items, list):
            return [item for item in items if isinstance(item, dict)]
    return []


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _format_money(value: Decimal) -> str:
    return f"{value:,.0f} ₽".replace(",", " ")


def _format_dt(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    rendered = str(value).strip()
    if not rendered:
        return ""
    try:
        return datetime.fromisoformat(rendered).strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return rendered


def _to_excel_temporal(value: Any) -> date | datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    rendered = str(value).strip()
    if not rendered:
        return None
    try:
        parsed = datetime.fromisoformat(rendered)
    except ValueError:
        return None
    if parsed.time() == datetime.min.time():
        return parsed.date()
    return parsed


def _parse_origin_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    rendered = str(value).strip()
    if not rendered:
        return None
    try:
        return datetime.fromisoformat(rendered).date()
    except ValueError:
        return None


def _age_days(anchor_date: date, item: dict[str, Any]) -> int | None:
    origin_date = _parse_origin_date(item.get("origin_document_date"))
    if origin_date is None:
        return None
    return (anchor_date - origin_date).days


def _normalize_items(anchor_date: date, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ranked = sorted(
        items,
        key=lambda item: (
            _to_decimal(item.get("current_balance")),
            str(item.get("counterparty_name") or item.get("counterparty_ref") or ""),
        ),
        reverse=True,
    )
    normalized: list[dict[str, Any]] = []
    for item in ranked:
        normalized.append(
            {
                "counterparty_ref": str(item.get("counterparty_ref") or ""),
                "counterparty_name": str(item.get("counterparty_name") or ""),
                "current_balance": str(
                    _to_decimal(item.get("current_balance")).quantize(Decimal("0.01"))
                ),
                "origin_document_ref": str(item.get("origin_document_ref") or ""),
                "origin_document_number": str(item.get("origin_document_number") or ""),
                "origin_document_date": str(item.get("origin_document_date") or ""),
                "current_manager_name": str(item.get("current_manager_name") or ""),
                "planned_payment_date": str(item.get("planned_payment_date") or ""),
                "due_date": str(item.get("due_date") or ""),
                "overdue_days": item.get("overdue_days"),
                "age_days": _age_days(anchor_date, item),
            }
        )
    return normalized


def _build_revision(anchor_date: date, items: list[dict[str, Any]]) -> str:
    normalized = {
        "anchor_date": anchor_date.isoformat(),
        "items": _normalize_items(anchor_date, items),
    }
    payload = json.dumps(normalized, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()[:12]


def _load_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"reports": {}}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        return {"reports": {}}
    payload.setdefault("reports", {})
    return payload


def _save_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(state, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _state_key(report_key: str, revision: str) -> str:
    return f"{report_key}|r{revision}"


def _pending_state_key(report_key: str) -> str:
    return f"{report_key}{PENDING_STATE_SUFFIX}"


def _has_prior_delivered_revision(
    state: dict[str, Any],
    *,
    report_key: str,
    revision: str,
) -> bool:
    reports = state.get("reports") or {}
    if not isinstance(reports, dict):
        return False
    for item in reports.values():
        if not isinstance(item, dict):
            continue
        if item.get("delivery_status") != "delivered":
            continue
        if str(item.get("report_key") or "") != report_key:
            continue
        if str(item.get("revision") or "") == revision:
            continue
        return True
    return False


def _has_pending_source_state(state: dict[str, Any], *, report_key: str) -> bool:
    reports = state.get("reports") or {}
    pending = reports.get(_pending_state_key(report_key))
    return isinstance(pending, dict) and pending.get("delivery_status") == "pending_source"


def _clear_pending_source_state(state: dict[str, Any], *, report_key: str) -> None:
    reports = state.get("reports") or {}
    if isinstance(reports, dict):
        reports.pop(_pending_state_key(report_key), None)


def _mark_pending_source_state(
    state: dict[str, Any],
    *,
    report_key: str,
    anchor_date: date,
    freshness_status: str,
    source_status: str,
    latest_snapshot_date: str | None,
) -> None:
    state.setdefault("reports", {})[_pending_state_key(report_key)] = {
        "report_key": report_key,
        "anchor_date": anchor_date.isoformat(),
        "delivery_status": "pending_source",
        "freshness_status": freshness_status,
        "source_status": source_status,
        "latest_snapshot_date": latest_snapshot_date,
        "updated_at": _utcnow().isoformat(),
    }


def _receivables_component_ready_for_date(
    fetch_json: Callable[[str, dict[str, str]], Any],
    *,
    anchor_date: date,
) -> tuple[bool, str | None]:
    payload = fetch_json("/api/management/health", {"date": anchor_date.isoformat()})
    components = payload.get("components", []) if isinstance(payload, dict) else []
    for item in components:
        if not isinstance(item, dict):
            continue
        if str(item.get("component") or "") != "receivables":
            continue
        latest_snapshot_date = str(item.get("latest_snapshot_date") or "").strip() or None
        source_status = str(item.get("source_status") or "").strip().lower()
        return source_status == "ready" and latest_snapshot_date == anchor_date.isoformat(), (
            latest_snapshot_date
        )
    return False, None


def _collect_pending_anchor_dates(
    state: dict[str, Any],
    *,
    today: date,
    lookback_days: int,
) -> list[date]:
    reports = state.get("reports") or {}
    if not isinstance(reports, dict):
        return []
    dates: set[date] = set()
    for item in reports.values():
        if not isinstance(item, dict):
            continue
        if item.get("delivery_status") != "pending_source":
            continue
        anchor_raw = str(item.get("anchor_date") or "").strip()
        if not anchor_raw:
            continue
        try:
            anchor_date = date.fromisoformat(anchor_raw)
        except ValueError:
            continue
        lag_days = (today - anchor_date).days
        if 0 <= lag_days <= lookback_days:
            dates.add(anchor_date)
    return sorted(dates)


def _collect_recent_anchor_dates(*, today: date, lookback_days: int) -> list[date]:
    safe_lookback_days = max(0, lookback_days)
    return sorted({today - timedelta(days=offset) for offset in range(safe_lookback_days + 1)})


def _safe_path_chunk(value: str) -> str:
    safe = value.replace("/", "-").replace("\\", "-").strip()
    return safe or "unknown"


def _build_report_path(
    *, artifact_dir: Path, anchor_date: date, report_key: str, revision: str
) -> Path:
    return (
        artifact_dir
        / anchor_date.isoformat()
        / _safe_path_chunk(report_key)
        / _safe_path_chunk(revision)
        / f"Новая_дебиторка_{anchor_date.isoformat()}.xlsx"
    )


def _export_new_daily_receivables_xlsx(
    *,
    anchor_date: date,
    items: list[dict[str, Any]],
    output_path: Path,
) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    sheet = workbook.create_sheet("Новая дебиторка")

    headers = [
        "№",
        "Контрагент",
        "Сумма долга, ₽",
        "Возраст долга, дн",
        "Исходный документ",
        "Дата возникновения",
        "Текущий менеджер",
        "План оплаты",
        "Due date",
        "Просрочка, дн",
    ]
    ranked = sorted(
        items,
        key=lambda item: (
            _to_decimal(item.get("current_balance")),
            str(item.get("counterparty_name") or item.get("counterparty_ref") or ""),
        ),
        reverse=True,
    )

    total_balance = sum(
        (_to_decimal(item.get("current_balance")) for item in ranked),
        start=Decimal("0"),
    )
    overdue_count = sum(1 for item in ranked if int(item.get("overdue_days") or 0) > 0)
    planned_count = sum(1 for item in ranked if item.get("planned_payment_date"))
    max_balance = max(
        (_to_decimal(item.get("current_balance")) for item in ranked), default=Decimal("0")
    )

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    subheader_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    title_fill = PatternFill(fill_type="solid", fgColor="EAF3E2")
    stripe_fill = PatternFill(fill_type="solid", fgColor="F7FBFF")
    overdue_fill = PatternFill(fill_type="solid", fgColor="FDE9E7")
    high_balance_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    total_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    header_font = Font(bold=True, color="FFFFFF")
    border_side = Side(style="thin", color="D9E2F3")
    cell_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )

    summary_sheet.merge_cells("A1:B1")
    summary_sheet["A1"] = "Утренняя новая дебиторка"
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet["A1"].fill = title_fill
    summary_sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    summary_rows = [
        ("Дата отчета", anchor_date),
        ("Правило", "Новая дебиторка: долг не закрыт более 3 дней"),
        ("Контрагентов", len(ranked)),
        ("Общая сумма, ₽", float(total_balance)),
        ("Средний долг, ₽", float(total_balance / len(ranked)) if ranked else 0.0),
        ("Просрочено", overdue_count),
        ("Есть план оплаты", planned_count),
        ("Максимальный долг, ₽", float(max_balance)),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=2):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=value)

    top_start_row = len(summary_rows) + 4
    summary_sheet.cell(row=top_start_row, column=1, value="Топ контрагентов по сумме")
    summary_sheet.cell(row=top_start_row, column=1).font = Font(bold=True)
    summary_sheet.cell(row=top_start_row + 1, column=1, value="Контрагент")
    summary_sheet.cell(row=top_start_row + 1, column=2, value="Сумма долга, ₽")
    for cell in summary_sheet[top_start_row + 1]:
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for offset, item in enumerate(ranked[:5], start=top_start_row + 2):
        summary_sheet.cell(
            row=offset,
            column=1,
            value=item.get("counterparty_name") or item.get("counterparty_ref") or "",
        )
        summary_sheet.cell(
            row=offset,
            column=2,
            value=float(_to_decimal(item.get("current_balance"))),
        )

    for row_index in range(2, top_start_row):
        label_cell = summary_sheet.cell(row=row_index, column=1)
        value_cell = summary_sheet.cell(row=row_index, column=2)
        label_cell.font = Font(bold=True)
        label_cell.fill = subheader_fill
        label_cell.border = cell_border
        label_cell.alignment = Alignment(vertical="center")
        value_cell.border = cell_border
        value_cell.alignment = Alignment(vertical="center")
        if row_index in {5, 6, 9}:
            value_cell.number_format = "#,##0.00"
        elif isinstance(value_cell.value, date):
            value_cell.number_format = "dd.mm.yyyy"

    for row_index in range(top_start_row + 2, top_start_row + 2 + min(len(ranked), 5)):
        for column_index in (1, 2):
            cell = summary_sheet.cell(row=row_index, column=column_index)
            cell.border = cell_border
            if row_index % 2 == 0:
                cell.fill = stripe_fill
        summary_sheet.cell(row=row_index, column=2).number_format = "#,##0.00"

    summary_sheet.freeze_panes = "A2"
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 18
    summary_last_row = top_start_row + 1 + max(1, min(len(ranked), 5))
    summary_sheet.auto_filter.ref = f"A{top_start_row + 1}:B{summary_last_row}"
    summary_sheet.sheet_view.zoomScale = 95

    sheet.merge_cells("A1:J1")
    sheet["A1"] = f"Новая дебиторка на {anchor_date.strftime('%d.%m.%Y')}"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells("A2:J2")
    sheet["A2"] = "В отчёт попадают долги, которые остаются открытыми более 3 дней."
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")

    sheet.append(headers)
    header_row = 3
    data_start_row = header_row + 1

    for cell in sheet[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index, item in enumerate(ranked, start=1):
        row_index = data_start_row + index - 1
        sheet.append(
            [
                index,
                item.get("counterparty_name") or item.get("counterparty_ref") or "",
                float(_to_decimal(item.get("current_balance"))),
                _age_days(anchor_date, item),
                item.get("origin_document_number") or item.get("origin_document_ref") or "",
                _to_excel_temporal(item.get("origin_document_date"))
                or _format_dt(item.get("origin_document_date")),
                item.get("current_manager_name") or item.get("current_manager_ref") or "",
                _to_excel_temporal(item.get("planned_payment_date"))
                or _format_dt(item.get("planned_payment_date")),
                _to_excel_temporal(item.get("due_date")) or _format_dt(item.get("due_date")),
                item.get("overdue_days"),
            ]
        )
        balance = _to_decimal(item.get("current_balance"))
        is_overdue = int(item.get("overdue_days") or 0) > 0
        is_high_balance = balance >= Decimal("100000")
        for column_index in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = stripe_fill
            if is_high_balance and column_index in (2, 3):
                cell.fill = high_balance_fill
            if is_overdue and column_index in (9, 10):
                cell.fill = overdue_fill

        sheet.cell(row=row_index, column=3).number_format = "#,##0.00"
        for column_index in (6, 8, 9):
            cell = sheet.cell(row=row_index, column=column_index)
            if isinstance(cell.value, datetime):
                cell.number_format = "dd.mm.yyyy hh:mm"
            elif isinstance(cell.value, date):
                cell.number_format = "dd.mm.yyyy"
        sheet.cell(row=row_index, column=1).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        sheet.cell(row=row_index, column=3).alignment = Alignment(
            horizontal="right", vertical="top"
        )
        sheet.cell(row=row_index, column=4).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        sheet.cell(row=row_index, column=10).alignment = Alignment(
            horizontal="center", vertical="top"
        )

    total_row = data_start_row + len(ranked)
    sheet.cell(row=total_row, column=1, value="Итого")
    sheet.cell(
        row=total_row,
        column=3,
        value=f"=SUM(C{data_start_row}:C{max(data_start_row, total_row - 1)})",
    )
    sheet.cell(
        row=total_row,
        column=4,
        value=f"=MAX(D{data_start_row}:D{max(data_start_row, total_row - 1)})",
    )
    sheet.cell(
        row=total_row,
        column=10,
        value=f"=SUM(J{data_start_row}:J{max(data_start_row, total_row - 1)})",
    )
    for column_index in range(1, len(headers) + 1):
        cell = sheet.cell(row=total_row, column=column_index)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = cell_border
    sheet.cell(row=total_row, column=3).number_format = "#,##0.00"
    sheet.cell(row=total_row, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    widths = {
        1: 6,
        2: 34,
        3: 16,
        4: 16,
        5: 24,
        6: 20,
        7: 24,
        8: 20,
        9: 20,
        10: 14,
    }
    for column_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_idx)].width = width
    sheet.freeze_panes = "A4"
    last_data_row = max(data_start_row, total_row - 1)
    sheet.auto_filter.ref = f"A{header_row}:J{last_data_row}"
    sheet.sheet_view.zoomScale = 90

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _parse_chat_ids(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [chunk.strip() for chunk in raw.split(",") if chunk.strip()]


def _resolve_chat_ids(env: dict[str, str]) -> list[str]:
    return _parse_chat_ids(
        env.get("MANAGEMENT_NEW_DAILY_TELEGRAM_CHAT_ID")
        or env.get("WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_CHAT_ID")
        or env.get("WEEKLY_MANAGER_SALES_B_TELEGRAM_CHAT_ID")
    )


def _send_telegram_document(
    *,
    token: str,
    chat_id: str,
    message: str,
    report_path: Path,
    timeout: int = 60,
) -> None:
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    boundary = f"----newdailyreceivables{int(time.time() * 1000)}"
    file_bytes = report_path.read_bytes()

    parts = []
    for name, value in (("chat_id", chat_id), ("caption", message)):
        parts.append(f"--{boundary}\r\n".encode())
        parts.append(f'Content-Disposition: form-data; name="{name}"\r\n\r\n{value}\r\n'.encode())

    parts.append(f"--{boundary}\r\n".encode())
    parts.append(
        (
            f'Content-Disposition: form-data; name="document"; filename="{report_path.name}"\r\n'
            f"Content-Type: {XLSX_MEDIA_TYPE}\r\n\r\n"
        ).encode()
    )
    parts.append(file_bytes)
    parts.append(b"\r\n")
    parts.append(f"--{boundary}--\r\n".encode())
    body = b"".join(parts)

    request = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        response.read()


def _build_caption(*, anchor_date: date, items: list[dict[str, Any]], is_correction: bool) -> str:
    total_balance = sum(
        (_to_decimal(item.get("current_balance")) for item in items), start=Decimal("0")
    )
    top_items = []
    for item in sorted(
        items,
        key=lambda row: _to_decimal(row.get("current_balance")),
        reverse=True,
    )[:5]:
        name = str(item.get("counterparty_name") or item.get("counterparty_ref") or "без названия")
        top_items.append(f"{name}: {_format_money(_to_decimal(item.get('current_balance')))}")

    lines = []
    if is_correction:
        lines.append("Исправленная версия утреннего отчета.")
    lines.append(f"Новая дебиторка > 3 дней на {anchor_date.strftime('%d.%m.%Y')}.")
    lines.append(f"Контрагентов: {len(items)} на {_format_money(total_balance)}.")
    if top_items:
        lines.append("Топ: " + "; ".join(top_items) + ".")
    caption = "\n".join(lines)
    return caption[:1024]


def sync_new_daily_receivables_report(
    *,
    fetch_json: Callable[[str, dict[str, str]], Any],
    deliver_report: Callable[..., dict[str, Any]],
    anchor_date: date,
    state_path: Path,
    artifact_dir: Path,
    dry_run: bool = False,
    pending_retry_attempts: int = 0,
    pending_retry_delay_seconds: float = 0.0,
) -> dict[str, Any]:
    report_key = f"{REPORT_KEY_PREFIX}|{anchor_date.isoformat()}"
    state = _load_state(state_path)

    payload: Any = None
    items: list[dict[str, Any]] = []
    freshness_status = "unknown"
    source_status = "unknown"
    latest_snapshot_date: str | None = None

    attempts = max(0, pending_retry_attempts)
    for attempt in range(attempts + 1):
        try:
            payload = fetch_json("/api/receivables/new-daily", {"date": anchor_date.isoformat()})
        except Exception as error:
            if _is_retryable_source_error(error):
                if attempt >= attempts:
                    if not dry_run:
                        _mark_pending_source_state(
                            state,
                            report_key=report_key,
                            anchor_date=anchor_date,
                            freshness_status="error",
                            source_status="error",
                            latest_snapshot_date=None,
                        )
                        _save_state(state_path, state)
                    return {
                        "status": "pending",
                        "anchor_date": anchor_date.isoformat(),
                        "freshness_status": "error",
                        "source_status": "error",
                        "latest_snapshot_date": None,
                        "error": str(error),
                        "fetched": 0,
                        "delivered": 0,
                        "noop": 0,
                        "failed": 0,
                        "sent_documents": 0,
                        "actions": [
                            {
                                "action": "pending_source_error",
                                "report_key": report_key,
                            }
                        ],
                    }
                if pending_retry_delay_seconds > 0:
                    time.sleep(pending_retry_delay_seconds)
                continue
            return {
                "status": "error",
                "anchor_date": anchor_date.isoformat(),
                "error": str(error),
                "fetched": 0,
                "delivered": 0,
                "noop": 0,
                "failed": 1,
                "sent_documents": 0,
                "actions": [],
            }

        items = _payload_items(payload)
        freshness_status = str(payload.get("freshness_status") or "unknown")
        source_status = str(payload.get("source_status") or "unknown")

        if items:
            break
        if freshness_status == "fresh" and source_status == "ready":
            break

        try:
            source_ready, latest_snapshot_date = _receivables_component_ready_for_date(
                fetch_json,
                anchor_date=anchor_date,
            )
        except Exception:
            source_ready = False
            latest_snapshot_date = None

        if source_ready:
            break
        if attempt >= attempts:
            if not dry_run:
                _mark_pending_source_state(
                    state,
                    report_key=report_key,
                    anchor_date=anchor_date,
                    freshness_status=freshness_status,
                    source_status=source_status,
                    latest_snapshot_date=latest_snapshot_date,
                )
                _save_state(state_path, state)
            return {
                "status": "pending",
                "anchor_date": anchor_date.isoformat(),
                "freshness_status": freshness_status,
                "source_status": source_status,
                "latest_snapshot_date": latest_snapshot_date,
                "fetched": 1,
                "delivered": 0,
                "noop": 0,
                "failed": 0,
                "sent_documents": 0,
                "actions": [{"action": "pending_source", "report_key": report_key}],
            }
        if pending_retry_delay_seconds > 0:
            time.sleep(pending_retry_delay_seconds)

    if not items:
        if not dry_run:
            _clear_pending_source_state(state, report_key=report_key)
            _save_state(state_path, state)
        return {
            "status": "ok",
            "anchor_date": anchor_date.isoformat(),
            "freshness_status": freshness_status,
            "source_status": source_status,
            "fetched": 1,
            "delivered": 0,
            "noop": 1,
            "failed": 0,
            "sent_documents": 0,
            "actions": [{"action": "noop_empty", "report_key": report_key}],
        }

    revision = _build_revision(anchor_date, items)
    current_state_key = _state_key(report_key, revision)
    current = (state.get("reports") or {}).get(current_state_key)
    if isinstance(current, dict) and current.get("delivery_status") == "delivered":
        return {
            "status": "ok",
            "anchor_date": anchor_date.isoformat(),
            "freshness_status": freshness_status,
            "source_status": source_status,
            "fetched": 1,
            "delivered": 0,
            "noop": 1,
            "failed": 0,
            "sent_documents": 0,
            "actions": [{"action": "noop", "report_key": report_key, "revision": revision}],
        }

    is_correction = _has_prior_delivered_revision(
        state,
        report_key=report_key,
        revision=revision,
    ) or _has_pending_source_state(state, report_key=report_key)
    report_path = _build_report_path(
        artifact_dir=artifact_dir,
        anchor_date=anchor_date,
        report_key=report_key,
        revision=revision,
    )
    action_record = {
        "action": "deliver" if not dry_run else "dry_run",
        "report_key": report_key,
        "revision": revision,
        "item_count": len(items),
        "is_correction": is_correction,
    }

    if dry_run:
        return {
            "status": "ok",
            "anchor_date": anchor_date.isoformat(),
            "freshness_status": freshness_status,
            "source_status": source_status,
            "fetched": 1,
            "delivered": 1,
            "noop": 0,
            "failed": 0,
            "sent_documents": 0,
            "actions": [action_record],
        }

    try:
        _export_new_daily_receivables_xlsx(
            anchor_date=anchor_date,
            items=items,
            output_path=report_path,
        )
        delivery_result = deliver_report(
            report_path=report_path,
            caption=_build_caption(
                anchor_date=anchor_date, items=items, is_correction=is_correction
            ),
            items=items,
            is_correction=is_correction,
        )
        sent_documents = int(delivery_result.get("sent_count") or 0)
        state.setdefault("reports", {})[current_state_key] = {
            "report_key": report_key,
            "revision": revision,
            "anchor_date": anchor_date.isoformat(),
            "delivery_status": "delivered",
            "artifact_path": str(report_path),
            "sent_documents": sent_documents,
            "is_correction": is_correction,
            "delivered_at": _utcnow().isoformat(),
        }
        _clear_pending_source_state(state, report_key=report_key)
        _save_state(state_path, state)
        action_record["artifact_path"] = str(report_path)
        action_record["sent_documents"] = sent_documents
        return {
            "status": "ok",
            "anchor_date": anchor_date.isoformat(),
            "freshness_status": freshness_status,
            "source_status": source_status,
            "fetched": 1,
            "delivered": 1,
            "noop": 0,
            "failed": 0,
            "sent_documents": sent_documents,
            "actions": [action_record],
        }
    except Exception as error:
        state.setdefault("reports", {})[current_state_key] = {
            "report_key": report_key,
            "revision": revision,
            "anchor_date": anchor_date.isoformat(),
            "delivery_status": "failed",
            "error": str(error),
            "updated_at": _utcnow().isoformat(),
        }
        _save_state(state_path, state)
        action_record["action"] = "failed"
        action_record["error"] = str(error)
        return {
            "status": "error",
            "anchor_date": anchor_date.isoformat(),
            "freshness_status": freshness_status,
            "source_status": source_status,
            "error": str(error),
            "fetched": 1,
            "delivered": 0,
            "noop": 0,
            "failed": 1,
            "sent_documents": 0,
            "actions": [action_record],
        }


def render_summary(summary: dict[str, Any]) -> str:
    lines = [
        f"new_daily_receivables_from_a: {summary.get('status', 'unknown')}",
        f"Дата: {summary.get('anchor_date', '-')}",
        f"Источник: {summary.get('freshness_status', 'unknown')}/{summary.get('source_status', 'unknown')}",
        (
            "Получено: {fetched}; delivered: {delivered}; noop: {noop}; "
            "failed: {failed}; sent_documents: {sent_documents}"
        ).format(
            fetched=summary.get("fetched", 0),
            delivered=summary.get("delivered", 0),
            noop=summary.get("noop", 0),
            failed=summary.get("failed", 0),
            sent_documents=summary.get("sent_documents", 0),
        ),
    ]
    error = summary.get("error")
    if error:
        lines.append(f"Ошибка: {error}")
    for item in summary.get("actions") or []:
        lines.append(
            f"- {item.get('action')}: {item.get('report_key')} "
            f"(revision={item.get('revision')}, correction={item.get('is_correction', False)})"
        )
    return "\n".join(lines)


def main() -> None:
    args = _parse_args()
    env = _load_env(
        os.getenv("MANAGEMENT_NEW_DAILY_B_ENV_FILE")
        or os.getenv("OPENCLAW_ENV_FILE")
        or os.getenv("PRICING_ENV_FILE")
        or DEFAULT_LOCAL_ENV_FILE
    )
    source_url = (
        env.get("MANAGEMENT_SOURCE_URL")
        or env.get("RETURN_SCHEME_SOURCE_URL")
        or DEFAULT_LOCAL_SOURCE_URL
    )
    source_token = (
        env.get("MANAGEMENT_SOURCE_TOKEN")
        or env.get("RETURN_SCHEME_SOURCE_TOKEN")
        or env.get("MANAGEMENT_INTERNAL_API_TOKEN")
        or env.get("RETURN_SCHEME_INTERNAL_API_TOKEN")
    )
    if not source_token:
        raise SystemExit(
            "Missing required env: MANAGEMENT_SOURCE_TOKEN|RETURN_SCHEME_SOURCE_TOKEN|"
            "MANAGEMENT_INTERNAL_API_TOKEN|RETURN_SCHEME_INTERNAL_API_TOKEN"
        )

    timeout = int(env.get("MANAGEMENT_ADAPTER_TIMEOUT_SECONDS", "20"))
    retries = int(env.get("MANAGEMENT_ADAPTER_RETRIES", "2"))
    retry_delay = float(env.get("MANAGEMENT_ADAPTER_RETRY_DELAY_SECONDS", "1.0"))
    fetch_json = _build_fetcher(
        source_url=source_url,
        token=source_token,
        timeout=timeout,
        retries=retries,
        retry_delay=retry_delay,
    )

    telegram_token = (
        env.get("MANAGEMENT_NEW_DAILY_TELEGRAM_TOKEN")
        or env.get("WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_TOKEN")
        or env.get("WEEKLY_MANAGER_SALES_B_TELEGRAM_TOKEN")
        or env.get("TELEGRAM_TOKEN_MM")
    )
    chat_ids = _resolve_chat_ids(env)
    if not args.dry_run and (not telegram_token or not chat_ids):
        raise SystemExit(
            "Missing required env: MANAGEMENT_NEW_DAILY_TELEGRAM_TOKEN|"
            "WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_TOKEN|WEEKLY_MANAGER_SALES_B_TELEGRAM_TOKEN|"
            "TELEGRAM_TOKEN_MM and MANAGEMENT_NEW_DAILY_TELEGRAM_CHAT_ID|"
            "WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_CHAT_ID|WEEKLY_MANAGER_SALES_B_TELEGRAM_CHAT_ID"
        )

    state_path = Path(env.get("MANAGEMENT_NEW_DAILY_STATE_PATH", DEFAULT_STATE_PATH))
    artifact_dir = Path(env.get("MANAGEMENT_NEW_DAILY_REPORT_DIR", DEFAULT_ARTIFACT_DIR))
    anchor_date = _parse_date(args.anchor_date)
    pending_retry_attempts = int(env.get("MANAGEMENT_NEW_DAILY_PENDING_RETRY_ATTEMPTS", "3"))
    pending_retry_delay_seconds = float(
        env.get("MANAGEMENT_NEW_DAILY_PENDING_RETRY_DELAY_SECONDS", "60")
    )
    pending_lookback_days = int(env.get("MANAGEMENT_NEW_DAILY_PENDING_LOOKBACK_DAYS", "3"))
    delivery_lookback_days = int(env.get("MANAGEMENT_NEW_DAILY_DELIVERY_LOOKBACK_DAYS", "3"))

    def _deliver(
        *,
        report_path: Path,
        caption: str,
        items: list[dict[str, Any]],
        is_correction: bool,
    ) -> dict[str, Any]:
        del items, is_correction
        assert telegram_token is not None
        sent_count = 0
        for chat_id in chat_ids:
            _send_telegram_document(
                token=telegram_token,
                chat_id=chat_id,
                message=caption,
                report_path=report_path,
            )
            sent_count += 1
        return {"sent_count": sent_count}

    state = _load_state(state_path)
    pending_dates = (
        []
        if args.anchor_date
        else _collect_pending_anchor_dates(
            state,
            today=anchor_date,
            lookback_days=pending_lookback_days,
        )
    )
    recovery_dates = (
        []
        if args.anchor_date
        else _collect_recent_anchor_dates(
            today=anchor_date,
            lookback_days=delivery_lookback_days,
        )
    )
    dates_to_process = sorted({*pending_dates, *recovery_dates, anchor_date})
    summaries = [
        sync_new_daily_receivables_report(
            fetch_json=fetch_json,
            deliver_report=_deliver,
            anchor_date=run_date,
            state_path=state_path,
            artifact_dir=artifact_dir,
            dry_run=args.dry_run,
            pending_retry_attempts=pending_retry_attempts if run_date == anchor_date else 0,
            pending_retry_delay_seconds=(
                pending_retry_delay_seconds if run_date == anchor_date else 0.0
            ),
        )
        for run_date in dates_to_process
    ]
    if args.json:
        if len(summaries) == 1:
            print(json.dumps(summaries[0], ensure_ascii=False, indent=2))
        else:
            print(json.dumps({"summaries": summaries}, ensure_ascii=False, indent=2))
        return
    print("\n\n".join(render_summary(summary) for summary in summaries))


if __name__ == "__main__":
    main()
