#!/usr/bin/env python3
"""Pull monthly customer price-type recommendations from server A and deliver XLSX to Telegram."""

from __future__ import annotations

import argparse
import json
import os
import time
import urllib.error
import urllib.request
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill

try:  # Deployed on Openclaw as a flat scripts directory.
    from weekly_kpi_reports_from_a import (  # type: ignore
        DEFAULT_LOCAL_ENV_FILE,
        DEFAULT_LOCAL_SOURCE_URL,
        _build_fetcher,
        _load_env,
    )
except ImportError:  # Local tests/imports from pricing-service repo.
    from infra.cron.weekly_kpi_reports_from_a import (
        DEFAULT_LOCAL_ENV_FILE,
        DEFAULT_LOCAL_SOURCE_URL,
        _build_fetcher,
        _load_env,
    )


DEFAULT_STATE_PATH = (
    "/home/deploy/.openclaw/workspace/.data/retail-price-type-recommendations/state.json"
)
DEFAULT_ARTIFACT_DIR = (
    "/home/deploy/.openclaw/workspace/.data/retail-price-type-recommendations/artifacts"
)
REPORT_ENDPOINT = "/api/management/retail-customer-price-type-recommendations"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pull retail customer price-type recommendations and send XLSX to Telegram."
    )
    parser.add_argument(
        "--month",
        help="Closed month in YYYY-MM format; default is previous calendar month.",
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Render actions without side effects"
    )
    parser.add_argument("--force", action="store_true", help="Send even if this month is delivered")
    parser.add_argument("--json", action="store_true", help="Emit machine-readable summary")
    return parser.parse_args()


def previous_month(today: date | None = None) -> str:
    anchor = today or date.today()
    first_day = anchor.replace(day=1)
    closed = first_day - timedelta(days=1)
    return closed.strftime("%Y-%m")


def _utcnow() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _load_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"months": {}}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        return {"months": {}}
    payload.setdefault("months", {})
    return payload


def _save_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(state, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _safe(value: Any) -> str:
    return str(value or "").strip()


def _format_dt(value: Any) -> str:
    if not value:
        return ""
    rendered = str(value)
    return rendered.replace("T", " ")[:19]


def _excel_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return None


def _summary_value(report: dict[str, Any], key: str) -> int:
    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    try:
        return int(summary.get(key) or 0)
    except (TypeError, ValueError):
        return 0


def _parse_chat_ids(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [chunk.strip() for chunk in raw.split(",") if chunk.strip()]


def _resolve_chat_ids(env: dict[str, str]) -> list[str]:
    return _parse_chat_ids(
        env.get("RETAIL_PRICE_TYPE_B_TELEGRAM_CHAT_ID")
        or env.get("RETAIL_PRICE_TYPE_TELEGRAM_CHAT_ID")
        or env.get("RETAIL_PRICE_TYPE_ASSISTANT_TELEGRAM_CHAT_ID")
        or env.get("WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_CHAT_ID")
        or env.get("WEEKLY_MANAGER_SALES_B_TELEGRAM_CHAT_ID")
    )


def _send_telegram_document(
    *,
    token: str,
    chat_id: str,
    message: str,
    report_path: Path,
    timeout: int = 60,
) -> None:
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    boundary = f"----retailpricetypes{int(time.time() * 1000)}"
    file_bytes = report_path.read_bytes()

    parts = []
    for name, value in (("chat_id", chat_id), ("caption", message)):
        parts.append(f"--{boundary}\r\n".encode())
        parts.append(f'Content-Disposition: form-data; name="{name}"\r\n\r\n{value}\r\n'.encode())

    parts.append(f"--{boundary}\r\n".encode())
    parts.append(
        (
            f'Content-Disposition: form-data; name="document"; filename="{report_path.name}"\r\n'
            f"Content-Type: {XLSX_MEDIA_TYPE}\r\n\r\n"
        ).encode()
    )
    parts.append(file_bytes)
    parts.append(b"\r\n")
    parts.append(f"--{boundary}--\r\n".encode())
    body = b"".join(parts)

    request = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        response.read()


def render_telegram_message(report: dict[str, Any]) -> str:
    month = _safe(report.get("month")) or "-"
    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    actionable = _summary_value(report, "actionable_count")
    lines = [
        f"Ежемесячный отчет по типам цен клиентов за {month}.",
        (
            "К изменению: "
            f"{actionable}; поставить серебро {_summary_value(report, 'set_silver_count')}; "
            f"поставить золото {_summary_value(report, 'set_gold_count')}; "
            f"понизить до серебра {_summary_value(report, 'downgrade_to_silver_count')}; "
            f"перевести на бронзу {_summary_value(report, 'downgrade_to_bronze_count')}."
        ),
        "Правило: новый тип цен ставится с 1 числа по чистым продажам прошлого месяца.",
        "Excel во вложении.",
    ]
    rules = summary.get("rules") if isinstance(summary, dict) else None
    if isinstance(rules, dict):
        lines.append(
            "Пороги: "
            f"серебро {rules.get('silver', '300 000..1 200 000 ₽')}; "
            f"золото {rules.get('gold', 'от 1 200 000 ₽')}."
        )
    return "\n".join(lines)


def export_recommendations_xlsx(report: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Типы цен"

    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    ws.append(["Отчет", "Рекомендации по типам цен клиентов"])
    ws.append(["Месяц", report.get("month")])
    ws.append(["К изменению", summary.get("actionable_count", 0)])
    ws.append(["Поставить серебро", summary.get("set_silver_count", 0)])
    ws.append(["Поставить золото", summary.get("set_gold_count", 0)])
    ws.append(["Понизить до серебра", summary.get("downgrade_to_silver_count", 0)])
    ws.append(["Перевести на бронзу", summary.get("downgrade_to_bronze_count", 0)])
    ws.append([])

    headers = [
        "Действие",
        "Контрагент",
        "Текущий тип цен",
        "Рекомендуемый тип цен",
        "Продажи (чистые)",
        "Продажи прошлый месяц",
        "Изменение продаж",
        "Изменение, %",
        "Возвраты",
        "Документов",
        "Последняя продажа",
        "Текущий тип видели",
        "Комментарий правила",
        "Код 1С",
    ]
    header_row = ws.max_row + 1
    ws.append(headers)
    for item in report.get("payload") or []:
        if not isinstance(item, dict):
            continue
        ws.append(
            [
                item.get("action_label"),
                item.get("counterparty_name"),
                item.get("current_price_type") or item.get("current_level_label"),
                item.get("recommended_price_type"),
                _excel_number(item.get("net_sales_amount", item.get("purchase_amount"))),
                _excel_number(
                    item.get("previous_net_sales_amount", item.get("previous_purchase_amount"))
                ),
                _excel_number(
                    item.get("net_sales_delta_amount", item.get("purchase_delta_amount"))
                ),
                _excel_number(item.get("net_sales_delta_pct", item.get("purchase_delta_pct"))),
                _excel_number(item.get("return_amount")),
                item.get("document_count"),
                _format_dt(item.get("last_sale_at")),
                _format_dt(item.get("current_price_seen_at")),
                item.get("rule_note"),
                item.get("counterparty_code"),
            ]
        )

    title_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = title_fill
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    money_columns = {"E", "F", "G", "I"}
    for column in money_columns:
        for cell in ws[column][header_row:]:
            cell.number_format = '#,##0.00 "₽"'
    for cell in ws["H"][header_row:]:
        cell.number_format = "0.0%"
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:N{ws.max_row}"
    if ws.max_row > header_row:
        first_data_row = header_row + 1
        last_data_row = ws.max_row
        positive_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
        negative_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
        promotion_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
        downgrade_fill = PatternFill(fill_type="solid", fgColor="FCE5CD")
        for column in ("G", "H"):
            value_range = f"{column}{first_data_row}:{column}{last_data_row}"
            ws.conditional_formatting.add(
                value_range,
                CellIsRule(operator="greaterThan", formula=["0"], fill=positive_fill),
            )
            ws.conditional_formatting.add(
                value_range,
                CellIsRule(operator="lessThan", formula=["0"], fill=negative_fill),
            )
        table_range = f"A{first_data_row}:N{last_data_row}"
        ws.conditional_formatting.add(
            table_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("Поставить",$A{first_data_row}))'],
                fill=promotion_fill,
            ),
        )
        ws.conditional_formatting.add(
            table_range,
            FormulaRule(
                formula=[
                    f'OR(ISNUMBER(SEARCH("Понизить",$A{first_data_row})),'
                    f'ISNUMBER(SEARCH("Перевести",$A{first_data_row})))'
                ],
                fill=downgrade_fill,
            ),
        )
    widths = {
        "A": 22,
        "B": 36,
        "C": 20,
        "D": 22,
        "E": 16,
        "F": 18,
        "G": 18,
        "H": 14,
        "I": 14,
        "J": 12,
        "K": 20,
        "L": 20,
        "M": 58,
        "N": 22,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(output_path)
    return output_path


def sync_retail_price_type_recommendations(
    *,
    fetch_json: Callable[[str, dict[str, str]], Any],
    deliver_report: Callable[..., dict[str, Any]],
    month: str,
    state_path: Path,
    artifact_dir: Path,
    delivery_target: str,
    dry_run: bool = False,
    force: bool = False,
) -> dict[str, Any]:
    state = _load_state(state_path)
    current = (state.get("months") or {}).get(month)
    if isinstance(current, dict) and current.get("delivery_status") == "delivered" and not force:
        return {
            "status": "ok",
            "month": month,
            "action": "noop",
            "reason": "already_delivered",
            "delivered": 0,
            "noop": 1,
            "failed": 0,
        }

    try:
        report = fetch_json(
            REPORT_ENDPOINT,
            {
                "month": month,
                "actionable_only": "true",
                "buyers_group_only": "true",
                "buyer_group_name": "ПОКУПАТЕЛИ",
            },
        )
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, ValueError) as error:
        return {
            "status": "error",
            "month": month,
            "action": "failed",
            "error": str(error),
            "delivered": 0,
            "noop": 0,
            "failed": 1,
        }

    if not isinstance(report, dict):
        return {
            "status": "error",
            "month": month,
            "action": "failed",
            "error": "source returned non-object payload",
            "delivered": 0,
            "noop": 0,
            "failed": 1,
        }

    source_status = str(report.get("source_status") or "unknown")
    if source_status == "empty":
        return {
            "status": "ok",
            "month": month,
            "action": "noop",
            "reason": "empty_source",
            "delivered": 0,
            "noop": 1,
            "failed": 0,
        }

    artifact_path = artifact_dir / month / f"retail-price-types-{month}.xlsx"
    export_recommendations_xlsx(report, artifact_path)
    message = render_telegram_message(report)
    action = {
        "status": "ok",
        "month": month,
        "action": "dry_run" if dry_run else "deliver",
        "artifact_path": str(artifact_path),
        "delivery_channel": "telegram",
        "delivery_target": delivery_target,
        "actionable_count": _summary_value(report, "actionable_count"),
        "delivered": 0 if dry_run else 1,
        "noop": 0,
        "failed": 0,
    }
    if dry_run:
        return action

    try:
        delivery_result = deliver_report(
            message=message,
            artifact_path=artifact_path,
            report_key=f"retail-price-types|{month}",
            revision=1,
        )
    except Exception as error:
        state.setdefault("months", {})[month] = {
            "delivery_status": "failed",
            "error": str(error),
            "updated_at": _utcnow().isoformat(),
        }
        _save_state(state_path, state)
        return {
            **action,
            "status": "error",
            "action": "failed",
            "error": str(error),
            "delivered": 0,
            "failed": 1,
        }

    state.setdefault("months", {})[month] = {
        "delivery_status": "delivered",
        "delivery_channel": "telegram",
        "delivery_target": delivery_target,
        "artifact_path": str(artifact_path),
        "telegram_sent_count": delivery_result.get("sent_count"),
        "telegram_chat_ids": delivery_result.get("chat_ids"),
        "delivered_at": _utcnow().isoformat(),
    }
    _save_state(state_path, state)
    action.update(
        {
            "telegram_sent_count": delivery_result.get("sent_count"),
            "telegram_chat_ids": delivery_result.get("chat_ids"),
        }
    )
    return action


def render_summary(summary: dict[str, Any]) -> str:
    lines = [
        f"retail_price_type_recommendations_from_a: {summary.get('status', 'unknown')}",
        f"Месяц: {summary.get('month', '-')}",
        f"Действие: {summary.get('action', '-')}",
        ("delivered: {delivered}; noop: {noop}; failed: {failed}; к изменению: {count}").format(
            delivered=summary.get("delivered", 0),
            noop=summary.get("noop", 0),
            failed=summary.get("failed", 0),
            count=summary.get("actionable_count", 0),
        ),
    ]
    if summary.get("artifact_path"):
        lines.append(f"Файл: {summary['artifact_path']}")
    if summary.get("telegram_sent_count") is not None:
        lines.append(f"Telegram отправок: {summary['telegram_sent_count']}")
    if summary.get("reason"):
        lines.append(f"Причина: {summary['reason']}")
    if summary.get("error"):
        lines.append(f"Ошибка: {summary['error']}")
    return "\n".join(lines)


def main() -> None:
    args = _parse_args()
    env = _load_env(
        os.getenv("MANAGEMENT_B_ENV_FILE")
        or os.getenv("OPENCLAW_ENV_FILE")
        or os.getenv("PRICING_ENV_FILE")
        or DEFAULT_LOCAL_ENV_FILE
    )
    source_url = (
        env.get("MANAGEMENT_SOURCE_URL")
        or env.get("RETURN_SCHEME_SOURCE_URL")
        or DEFAULT_LOCAL_SOURCE_URL
    )
    source_token = (
        env.get("MANAGEMENT_SOURCE_TOKEN")
        or env.get("RETURN_SCHEME_SOURCE_TOKEN")
        or env.get("MANAGEMENT_INTERNAL_API_TOKEN")
        or env.get("RETURN_SCHEME_INTERNAL_API_TOKEN")
    )
    if not source_token:
        raise SystemExit(
            "Missing required env: MANAGEMENT_SOURCE_TOKEN|RETURN_SCHEME_SOURCE_TOKEN|"
            "MANAGEMENT_INTERNAL_API_TOKEN|RETURN_SCHEME_INTERNAL_API_TOKEN"
        )

    timeout = int(env.get("MANAGEMENT_ADAPTER_TIMEOUT_SECONDS", "20"))
    retries = int(env.get("MANAGEMENT_ADAPTER_RETRIES", "2"))
    retry_delay = float(env.get("MANAGEMENT_ADAPTER_RETRY_DELAY_SECONDS", "1.0"))
    fetch_json = _build_fetcher(
        source_url=source_url,
        token=source_token,
        timeout=timeout,
        retries=retries,
        retry_delay=retry_delay,
    )

    telegram_token = (
        env.get("RETAIL_PRICE_TYPE_B_TELEGRAM_TOKEN")
        or env.get("RETAIL_PRICE_TYPE_TELEGRAM_TOKEN")
        or env.get("RETAIL_PRICE_TYPE_ASSISTANT_TELEGRAM_TOKEN")
        or env.get("WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_TOKEN")
        or env.get("WEEKLY_MANAGER_SALES_B_TELEGRAM_TOKEN")
        or env.get("TELEGRAM_TOKEN_MM")
    )
    chat_ids = _resolve_chat_ids(env)
    if not args.dry_run and (not telegram_token or not chat_ids):
        raise SystemExit(
            "Missing required env: RETAIL_PRICE_TYPE_B_TELEGRAM_TOKEN|"
            "RETAIL_PRICE_TYPE_TELEGRAM_TOKEN|RETAIL_PRICE_TYPE_ASSISTANT_TELEGRAM_TOKEN|"
            "WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_TOKEN|WEEKLY_MANAGER_SALES_B_TELEGRAM_TOKEN|"
            "TELEGRAM_TOKEN_MM and RETAIL_PRICE_TYPE_B_TELEGRAM_CHAT_ID|"
            "RETAIL_PRICE_TYPE_TELEGRAM_CHAT_ID|RETAIL_PRICE_TYPE_ASSISTANT_TELEGRAM_CHAT_ID|"
            "WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_CHAT_ID|WEEKLY_MANAGER_SALES_B_TELEGRAM_CHAT_ID"
        )

    state_path = Path(env.get("RETAIL_PRICE_TYPE_STATE_PATH", DEFAULT_STATE_PATH))
    artifact_dir = Path(env.get("RETAIL_PRICE_TYPE_REPORT_DIR", DEFAULT_ARTIFACT_DIR))
    month = args.month or previous_month()

    def _deliver(
        *,
        message: str,
        artifact_path: Path,
        report_key: str,
        revision: int,
    ) -> dict[str, Any]:
        del report_key, revision
        assert telegram_token is not None
        sent_count = 0
        for chat_id in chat_ids:
            _send_telegram_document(
                token=telegram_token,
                chat_id=chat_id,
                message=message,
                report_path=artifact_path,
            )
            sent_count += 1
        return {"sent_count": sent_count, "chat_ids": chat_ids}

    summary = sync_retail_price_type_recommendations(
        fetch_json=fetch_json,
        deliver_report=_deliver,
        month=month,
        state_path=state_path,
        artifact_dir=artifact_dir,
        delivery_target=",".join(chat_ids) if chat_ids else "telegram",
        dry_run=args.dry_run,
        force=args.force,
    )

    if args.json:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return
    print(render_summary(summary))


if __name__ == "__main__":
    main()
