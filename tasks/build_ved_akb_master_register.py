"""Build a VED AKB master register from 1C SKU, product data, and prior VED files."""

from __future__ import annotations

import argparse
import csv
import os
from collections import defaultdict
from collections.abc import Iterable
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import or_, select, text
from sqlalchemy.orm import Session, selectinload

from app.infrastructure.db.engines import build_engine
from app.models import Product
from app.services.sku import generate_sku_for_product

ORDER_NUMBER = "РБГУ0000377"
DEFAULT_BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_ENV_FILE = DEFAULT_BASE_DIR / ".env"
DEFAULT_OUTPUT_DIR = DEFAULT_BASE_DIR / ".local" / "procurement-labels" / "РБГУ0000377-v1"
DEFAULT_CURRENT_XLSX = DEFAULT_OUTPUT_DIR / "ved-akb-rbgu0000377-sku-ds-discrepancies.xlsx"
DEFAULT_OUTPUT_XLSX = DEFAULT_OUTPUT_DIR / "мастер-таблица-АКБ-РБГУ0000377-для-декларации.xlsx"
DEFAULT_OLD_PACK_DIR = (
    DEFAULT_BASE_DIR.parent / ".local" / "bitrix-task-context" / "task-143" / "ved-akb-doc-pack"
)
DEFAULT_OLD_SKU_CSV = DEFAULT_OLD_PACK_DIR / "sku-data-1c-sku-verified.csv"
DEFAULT_OLD_TECH_CSV = DEFAULT_OLD_PACK_DIR / "sku-data-1c-verified.csv"
DEFAULT_OLD_NORMALIZATION_CSV = DEFAULT_OLD_PACK_DIR / "onec-sku-normalization-proposal.csv"

REQUIRED_FIELDS = (
    "sku_1c",
    "trade_name_family",
    "compatibility",
    "capacity_mah",
    "voltage_v",
    "energy_wh",
    "dim",
    "bms",
    "connector",
    "tnved",
    "un_code",
    "gtin_ean13",
    "gost_r_ds_number",
    "gost_r_ds_covers_sku",
)

MASTER_COLUMNS = (
    "line_no",
    "onec_order",
    "onec_item_code",
    "article_1c",
    "sku_1c",
    "item_name_1c",
    "qty",
    "unit",
    "barcode_1c",
    "barcode_source",
    "generated_sku",
    "generation_status",
    "generation_reasons",
    "planned_sku_db",
    "fact_sku_db",
    "external_sku_candidate",
    "old_onec_sku",
    "old_sku_relation",
    "trade_name_family",
    "compatibility",
    "capacity_mah",
    "voltage_v",
    "energy_wh",
    "dim",
    "bms",
    "connector",
    "tnved",
    "un_code",
    "current_eaeu_ds_number",
    "current_eaeu_ds_type",
    "gost_r_ds_number",
    "gost_r_ds_covers_sku",
    "gtin_ean13",
    "old_normalization_status",
    "old_normalization_decision",
    "missing_fields",
    "review_status",
    "next_action",
)

SHEET_TITLES = {
    "master": "Мастер-реестр",
    "missing": "Что проверить",
    "conflicts": "Конфликты SKU",
    "broker": "Для брокера",
    "onec": "Кандидаты 1С",
}

COLUMN_TITLES = {
    "line_no": "№ строки",
    "onec_order": "Заказ 1С",
    "onec_item_code": "Код номенклатуры 1С",
    "article_1c": "Артикул 1С (числовой)",
    "sku_1c": "SKU 1С (отдельное поле)",
    "item_name_1c": "Наименование 1С",
    "qty": "Количество",
    "unit": "Ед.",
    "barcode_1c": "Штрихкод 1С",
    "barcode_source": "Источник штрихкода",
    "generated_sku": "SKU генератора (кандидат)",
    "generation_status": "Статус генерации SKU",
    "generation_reasons": "Комментарий генератора SKU",
    "planned_sku_db": "Плановый SKU в базе",
    "fact_sku_db": "Фактический SKU в базе",
    "external_sku_candidate": "Старый F5/OEM SKU (кандидат)",
    "old_onec_sku": "Старый SKU 1С из пакета",
    "old_sku_relation": "Комментарий по старому SKU",
    "trade_name_family": "TradeName / семейная формула для ДС",
    "compatibility": "Совместимость",
    "capacity_mah": "Ёмкость, mAh",
    "voltage_v": "Напряжение, V",
    "energy_wh": "Энергия, Wh",
    "dim": "Размеры, DIM",
    "bms": "BMS / плата защиты",
    "connector": "Коннектор",
    "tnved": "ТН ВЭД",
    "un_code": "UN Code",
    "current_eaeu_ds_number": "Текущая ДС ЕАЭС",
    "current_eaeu_ds_type": "Тип текущей ДС ЕАЭС",
    "gost_r_ds_number": "Новая ДС ГОСТ Р",
    "new_gost_r_ds_number": "Новая ДС ГОСТ Р",
    "gost_r_ds_covers_sku": "Покрытие новой ДС по SKU",
    "gtin_ean13": "GTIN/EAN-13",
    "old_normalization_status": "Статус старой нормализации",
    "old_normalization_decision": "Решение старой нормализации",
    "missing_fields": "Что не заполнено / ошибки",
    "review_status": "Статус проверки",
    "next_action": "Что сделать",
    "property_name": "Свойство 1С",
    "candidate_value": "Кандидат значения",
    "reason": "Почему / источник",
    "question_for_broker": "Вопрос брокеру",
}

REVIEW_STATUS_TITLES = {
    "missing_sku_1c": "Нет SKU 1С",
    "sku_policy_conflict": "Конфликт SKU",
    "missing_gost_r_ds": "Нет новой ДС ГОСТ Р / нет покрытия SKU",
    "missing_fields": "Не заполнены свойства",
    "ready_for_broker_review": "Готово к проверке брокером",
}

GENERATION_STATUS_TITLES = {
    "generated": "Сгенерирован",
    "manual_review": "Нужна ручная проверка",
    "conflict": "Конфликт",
}

BARCODE_SOURCE_TITLES = {
    "1c_internal": "Штрихкод из 1С",
    "catalog_barcode": "Штрихкод из каталога",
    "catalog_gtin": "GTIN из каталога",
}

OLD_NORMALIZATION_STATUS_TITLES = {
    "ready": "Готово к применению после проверки",
    "blocked": "Блокер / не менять без подтверждения",
    "matched": "Совпало",
}

OLD_NORMALIZATION_DECISION_TITLES = {
    "Use declaration/VED SKU in 1C SKU field": (
        "Использовать SKU из декларации/ВЭД в отдельном поле SKU 1С"
    ),
    "Do not change until supplier confirms": "Не менять до подтверждения поставщика",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build the RBGU0000377 AKB master register with separate 1C article and SKU."
    )
    parser.add_argument("--order-number", default=ORDER_NUMBER)
    parser.add_argument("--current-xlsx", type=Path, default=DEFAULT_CURRENT_XLSX)
    parser.add_argument("--old-sku-csv", type=Path, default=DEFAULT_OLD_SKU_CSV)
    parser.add_argument("--old-tech-csv", type=Path, default=DEFAULT_OLD_TECH_CSV)
    parser.add_argument("--old-normalization-csv", type=Path, default=DEFAULT_OLD_NORMALIZATION_CSV)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_XLSX)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    _load_env_file(DEFAULT_ENV_FILE)
    database_url = os.environ.get("DATABASE_URL")
    onec_database_url = os.environ.get("ONEC_DATABASE_URL")
    if not database_url:
        raise SystemExit("DATABASE_URL is not configured")
    if not onec_database_url:
        raise SystemExit("ONEC_DATABASE_URL is not configured")

    current_rows = _load_current_ds_rows(args.current_xlsx)
    old_sku_rows = _load_csv_by_code(args.old_sku_csv)
    old_tech_rows = _load_csv_by_code(args.old_tech_csv)
    old_normalization_rows = _load_csv_by_code(args.old_normalization_csv)

    onec_engine = build_engine(onec_database_url, pool_pre_ping=True)
    onec_lines = _fetch_onec_order_lines(onec_engine, args.order_number)

    app_engine = build_engine(database_url, pool_pre_ping=True)
    with Session(app_engine) as session:
        products_by_code = _load_products_by_code(
            session,
            codes=[_clean(row.get("onec_item_code")) for row in onec_lines],
            articles=[_clean(row.get("article_1c")) for row in onec_lines],
            skus=[_clean(row.get("sku_1c")) for row in onec_lines],
        )
        master_rows = _build_master_rows(
            session=session,
            order_number=args.order_number,
            onec_lines=onec_lines,
            current_rows=current_rows,
            old_sku_rows=old_sku_rows,
            old_tech_rows=old_tech_rows,
            old_normalization_rows=old_normalization_rows,
            products_by_code=products_by_code,
        )

    workbook = _build_workbook(master_rows)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(
        {
            "output": str(args.output),
            "rows": len(master_rows),
            "missing_rows": sum(1 for row in master_rows if row["missing_fields"]),
            "sku_conflicts": sum(
                1 for row in master_rows if row["review_status"] == "sku_policy_conflict"
            ),
        }
    )


def _load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        name, value = line.split("=", 1)
        name = name.strip()
        if name in {"DATABASE_URL", "ONEC_DATABASE_URL"}:
            os.environ.setdefault(name, value.strip().strip('"').strip("'"))


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return _format_decimal(value)
    cleaned = str(value).strip()
    if cleaned.endswith(".0") and cleaned[:-2].isdigit():
        return cleaned[:-2]
    return cleaned


def _format_decimal(value: Decimal) -> str:
    return format(value.normalize(), "f").rstrip("0").rstrip(".") or "0"


def _load_current_ds_rows(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook["Для ДС ГОСТ Р"]
    headers = [_clean(cell.value) for cell in worksheet[1]]
    rows: dict[str, dict[str, Any]] = {}
    for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(raw_row):
            continue
        row = dict(zip(headers, raw_row, strict=False))
        code = _clean(row.get("onec_item_code"))
        if code:
            rows[code] = row
    return rows


def _load_csv_by_code(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open(encoding="utf-8-sig", newline="") as stream:
        return {
            _clean(row.get("onec_item_code")): row
            for row in csv.DictReader(stream)
            if _clean(row.get("onec_item_code"))
        }


def _fetch_onec_order_lines(engine, order_number: str) -> list[dict[str, Any]]:
    query = text("""
        WITH target_doc AS (
          SELECT TOP 1 doc._IDRRef
          FROM dbo._Document133 AS doc WITH (NOLOCK)
          WHERE LTRIM(RTRIM(doc._Number)) = :order_number
          ORDER BY doc._Date_Time DESC
        )
        SELECT
          LTRIM(RTRIM(doc._Number)) AS order_number,
          vt._LineNo2516 AS line_no,
          item._Code AS onec_item_code,
          LTRIM(RTRIM(item._Fld836)) AS article_1c,
          LTRIM(RTRIM(item._Fld9945)) AS sku_1c,
          item._Description AS item_name_1c,
          barcode._Fld6984 AS barcode_1c,
          unit._Description AS unit,
          vt._Fld2520 AS qty
        FROM target_doc
        JOIN dbo._Document133 AS doc WITH (NOLOCK)
          ON doc._IDRRef = target_doc._IDRRef
        JOIN dbo._Document133_VT2515 AS vt WITH (NOLOCK)
          ON vt._Document133_IDRRef = doc._IDRRef
        LEFT JOIN dbo._Reference62 AS item WITH (NOLOCK)
          ON item._IDRRef = vt._Fld2523RRef
        LEFT JOIN dbo._Reference41 AS unit WITH (NOLOCK)
          ON unit._IDRRef = vt._Fld2517RRef
        OUTER APPLY (
          SELECT TOP 1
            LTRIM(RTRIM(barcode_row._Fld6984)) AS _Fld6984
          FROM dbo._InfoRg6983 AS barcode_row WITH (NOLOCK)
          WHERE barcode_row._Fld6985_RRRef = item._IDRRef
            AND LTRIM(RTRIM(barcode_row._Fld6984)) <> ''
          ORDER BY barcode_row._Fld6984 ASC
        ) AS barcode
        ORDER BY vt._LineNo2516 ASC
    """)
    with engine.connect() as connection:
        return [
            dict(row)
            for row in connection.execute(query, {"order_number": order_number}).mappings()
        ]


def _load_products_by_code(
    session: Session,
    *,
    codes: Iterable[str],
    articles: Iterable[str],
    skus: Iterable[str],
) -> dict[str, Product]:
    clean_codes = [value for value in codes if value]
    clean_articles = [value for value in articles if value]
    clean_skus = [value for value in skus if value]
    query = (
        select(Product)
        .options(selectinload(Product.compatibilities), selectinload(Product.sku_plans))
        .where(
            or_(
                Product.code_1c.in_(clean_codes),
                Product.article.in_(clean_articles),
                Product.fact_sku.in_(clean_skus),
                Product.planned_sku.in_(clean_skus),
            )
        )
    )
    products = session.execute(query).scalars().all()
    return {product.code_1c: product for product in products if product.code_1c}


def _build_master_rows(
    *,
    session: Session,
    order_number: str,
    onec_lines: list[dict[str, Any]],
    current_rows: dict[str, dict[str, Any]],
    old_sku_rows: dict[str, dict[str, str]],
    old_tech_rows: dict[str, dict[str, str]],
    old_normalization_rows: dict[str, dict[str, str]],
    products_by_code: dict[str, Product],
) -> list[dict[str, Any]]:
    master_rows: list[dict[str, Any]] = []
    for line in onec_lines:
        code = _clean(line.get("onec_item_code"))
        current = current_rows.get(code, {})
        old_sku = old_sku_rows.get(code, {})
        old_tech = old_tech_rows.get(code, {})
        old_norm = old_normalization_rows.get(code, {})
        product = products_by_code.get(code)
        generated_sku = ""
        generation_status = ""
        generation_reasons = ""
        planned_sku_db = ""
        fact_sku_db = ""
        compatibilities: list[str] = []
        if product is not None:
            result = generate_sku_for_product(session, product)
            generated_sku = _clean(result.planned_sku)
            generation_status = result.status
            generation_reasons = "; ".join(result.reasons)
            planned_sku_db = _clean(product.planned_sku)
            fact_sku_db = _clean(product.fact_sku)
            compatibilities = [
                _clean(item.value) for item in product.compatibilities if _clean(item.value)
            ]

        sku_1c = _clean(line.get("sku_1c"))
        article_1c = _clean(line.get("article_1c"))
        external_sku = _clean(old_sku.get("ved_register_sku"))
        trade_name = _clean(old_sku.get("trade_name") or old_tech.get("trade_name"))
        compatibility = _clean(
            current.get("compatibility_model")
            or old_sku.get("compatibility")
            or old_tech.get("compatibility")
            or "; ".join(compatibilities)
        )
        capacity = _clean(
            current.get("capacity_mAh")
            or old_sku.get("capacity_mah")
            or old_tech.get("capacity_mah")
        )
        voltage = _clean(current.get("voltage_V") or old_tech.get("voltage_v"))
        energy = _clean(current.get("energy_Wh") or old_tech.get("wh"))
        tnved = _clean(
            current.get("tnved_expected") or old_sku.get("hs_code") or old_tech.get("hs_code")
        )
        un_code = _clean(old_sku.get("un_code") or old_tech.get("un_code") or "UN3480")
        row = {
            "line_no": _clean(line.get("line_no")),
            "onec_order": order_number,
            "onec_item_code": code,
            "article_1c": article_1c,
            "sku_1c": sku_1c,
            "item_name_1c": _clean(line.get("item_name_1c")),
            "qty": _clean(line.get("qty") or current.get("qty")),
            "unit": _clean(line.get("unit") or current.get("unit") or "шт"),
            "barcode_1c": _clean(line.get("barcode_1c") or current.get("barcode_1c")),
            "barcode_source": _clean(current.get("barcode_source") or "1c_internal"),
            "generated_sku": generated_sku,
            "generation_status": generation_status,
            "generation_reasons": generation_reasons,
            "planned_sku_db": planned_sku_db,
            "fact_sku_db": fact_sku_db,
            "external_sku_candidate": external_sku,
            "old_onec_sku": _clean(old_sku.get("onec_sku")),
            "old_sku_relation": _clean(old_sku.get("sku_relation")),
            "trade_name_family": trade_name,
            "compatibility": compatibility,
            "capacity_mah": capacity,
            "voltage_v": voltage,
            "energy_wh": energy,
            "dim": _clean(old_tech.get("dim")),
            "bms": _clean(old_tech.get("bms")),
            "connector": _clean(old_tech.get("connector")),
            "tnved": tnved,
            "un_code": un_code,
            "current_eaeu_ds_number": _clean(current.get("current_ds_number")),
            "current_eaeu_ds_type": _clean(current.get("current_ds_type")),
            "gost_r_ds_number": _clean(current.get("gost_r_ds_number")),
            "gost_r_ds_covers_sku": _clean(current.get("gost_r_ds_covers_sku")),
            "gtin_ean13": "",
            "old_normalization_status": _clean(old_norm.get("status")),
            "old_normalization_decision": _clean(old_norm.get("decision")),
        }
        missing_fields = _missing_fields(row)
        row["missing_fields"] = "; ".join(missing_fields)
        row["review_status"] = _review_status(row, missing_fields)
        row["next_action"] = _next_action(row, missing_fields)
        master_rows.append(row)
    return master_rows


def _missing_fields(row: dict[str, Any]) -> list[str]:
    return [field for field in REQUIRED_FIELDS if not _clean(row.get(field))]


def _review_status(row: dict[str, Any], missing_fields: list[str]) -> str:
    if "sku_1c" in missing_fields:
        return "missing_sku_1c"
    if _has_sku_policy_conflict(row):
        return "sku_policy_conflict"
    if "gost_r_ds_number" in missing_fields or "gost_r_ds_covers_sku" in missing_fields:
        return "missing_gost_r_ds"
    if missing_fields:
        return "missing_fields"
    return "ready_for_broker_review"


def _has_sku_policy_conflict(row: dict[str, Any]) -> bool:
    sku_1c = _clean(row.get("sku_1c"))
    generated = _clean(row.get("generated_sku"))
    external = _clean(row.get("external_sku_candidate"))
    old_onec = _clean(row.get("old_onec_sku"))
    generation_status = _clean(row.get("generation_status"))
    return any(
        (
            bool(generated and sku_1c and generated != sku_1c),
            bool(external and sku_1c and external != sku_1c),
            bool(old_onec and sku_1c and old_onec != sku_1c),
            generation_status == "conflict",
        )
    )


def _next_action(row: dict[str, Any], missing_fields: list[str]) -> str:
    if "sku_1c" in missing_fields:
        return "Заполнить отдельное поле SKU в 1С или подтвердить, что SKU не нужен."
    if _has_sku_policy_conflict(row):
        return "Подтвердить канонический внешний SKU для ДС/инвойса; F5-BAT не применять автоматически."
    if "gost_r_ds_number" in missing_fields or "gost_r_ds_covers_sku" in missing_fields:
        return "Получить номер новой ДС ГОСТ Р и подтвердить покрытие этой строки."
    if missing_fields:
        return "Дозаполнить обязательные свойства перед передачей брокеру."
    return "Можно передавать брокеру на финальную проверку покрытия."


def _field_title(field: str) -> str:
    return COLUMN_TITLES.get(field, field)


def _translate_missing_fields(value: Any) -> str:
    fields = [_clean(item) for item in _clean(value).split(";") if _clean(item)]
    return "; ".join(_field_title(field) for field in fields)


def _display_cell(column: str, value: Any) -> Any:
    if column == "missing_fields":
        return _translate_missing_fields(value)
    if column == "review_status":
        return REVIEW_STATUS_TITLES.get(_clean(value), value)
    if column == "generation_status":
        return GENERATION_STATUS_TITLES.get(_clean(value), value)
    if column == "barcode_source":
        return BARCODE_SOURCE_TITLES.get(_clean(value), value)
    if column == "old_normalization_status":
        return OLD_NORMALIZATION_STATUS_TITLES.get(_clean(value), value)
    if column == "old_normalization_decision":
        return OLD_NORMALIZATION_DECISION_TITLES.get(_clean(value), value)
    return value


def _build_workbook(master_rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    master_sheet = workbook.active
    master_sheet.title = SHEET_TITLES["master"]
    _write_table(master_sheet, MASTER_COLUMNS, master_rows)

    missing_rows = [
        {
            "line_no": row["line_no"],
            "onec_item_code": row["onec_item_code"],
            "article_1c": row["article_1c"],
            "sku_1c": row["sku_1c"],
            "review_status": row["review_status"],
            "missing_fields": row["missing_fields"],
            "next_action": row["next_action"],
        }
        for row in master_rows
        if row["missing_fields"]
    ]
    _write_table(
        workbook.create_sheet(SHEET_TITLES["missing"]),
        (
            "line_no",
            "onec_item_code",
            "article_1c",
            "sku_1c",
            "review_status",
            "missing_fields",
            "next_action",
        ),
        missing_rows,
    )

    conflict_rows = [
        {
            "line_no": row["line_no"],
            "onec_item_code": row["onec_item_code"],
            "article_1c": row["article_1c"],
            "sku_1c": row["sku_1c"],
            "generated_sku": row["generated_sku"],
            "external_sku_candidate": row["external_sku_candidate"],
            "old_onec_sku": row["old_onec_sku"],
            "generation_status": row["generation_status"],
            "generation_reasons": row["generation_reasons"],
            "old_sku_relation": row["old_sku_relation"],
        }
        for row in master_rows
        if _has_sku_policy_conflict(row)
    ]
    _write_table(
        workbook.create_sheet(SHEET_TITLES["conflicts"]),
        (
            "line_no",
            "onec_item_code",
            "article_1c",
            "sku_1c",
            "generated_sku",
            "external_sku_candidate",
            "old_onec_sku",
            "generation_status",
            "generation_reasons",
            "old_sku_relation",
        ),
        conflict_rows,
    )

    broker_rows = [
        {
            "line_no": row["line_no"],
            "sku_1c": row["sku_1c"],
            "trade_name_family": row["trade_name_family"],
            "item_name_1c": row["item_name_1c"],
            "compatibility": row["compatibility"],
            "capacity_mah": row["capacity_mah"],
            "voltage_v": row["voltage_v"],
            "energy_wh": row["energy_wh"],
            "tnved": row["tnved"],
            "un_code": row["un_code"],
            "qty": row["qty"],
            "gost_r_ds_number": row["gost_r_ds_number"],
            "gost_r_ds_covers_sku": row["gost_r_ds_covers_sku"],
            "question_for_broker": "Подтвердите покрытие SKU новой ДС ГОСТ Р.",
        }
        for row in master_rows
    ]
    _write_table(
        workbook.create_sheet(SHEET_TITLES["broker"]),
        (
            "line_no",
            "sku_1c",
            "trade_name_family",
            "item_name_1c",
            "compatibility",
            "capacity_mah",
            "voltage_v",
            "energy_wh",
            "tnved",
            "un_code",
            "qty",
            "gost_r_ds_number",
            "gost_r_ds_covers_sku",
            "question_for_broker",
        ),
        broker_rows,
    )

    _write_table(
        workbook.create_sheet(SHEET_TITLES["onec"]),
        ("line_no", "onec_item_code", "property_name", "candidate_value", "reason"),
        _build_onec_candidate_rows(master_rows),
    )
    return workbook


def _build_onec_candidate_rows(master_rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in master_rows:
        if not row["sku_1c"] and row["generated_sku"]:
            rows.append(
                {
                    "line_no": row["line_no"],
                    "onec_item_code": row["onec_item_code"],
                    "property_name": "SKU",
                    "candidate_value": row["generated_sku"],
                    "reason": "SKU пустой в 1С; можно рассмотреть кандидат генератора.",
                }
            )
        for property_name, source_field in (("Напряжение", "voltage_v"), ("Wh", "energy_wh")):
            if row[source_field]:
                rows.append(
                    {
                        "line_no": row["line_no"],
                        "onec_item_code": row["onec_item_code"],
                        "property_name": property_name,
                        "candidate_value": row[source_field],
                        "reason": "Кандидат из текущей сверки ДС; применять только после ручного утверждения.",
                    }
                )
    return rows


def _write_table(worksheet, columns: Iterable[str], rows: list[dict[str, Any]]) -> None:
    column_list = list(columns)
    worksheet.append([_field_title(column) for column in column_list])
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        worksheet.append([_display_cell(column, row.get(column, "")) for column in column_list])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths: defaultdict[int, int] = defaultdict(lambda: 10)
    for row in worksheet.iter_rows():
        for cell in row:
            value_len = len(_clean(cell.value))
            widths[cell.column] = min(max(widths[cell.column], value_len + 2), 60)
    for index, width in widths.items():
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width


if __name__ == "__main__":
    started_at = datetime.now()
    main()
    elapsed = datetime.now() - started_at
    print({"elapsed_seconds": round(elapsed.total_seconds(), 3)})
