"""Export dynamic min/max shadow recommendations to a review-friendly XLSX."""

from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(size=11, italic=True, color="595959")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
READ_ONLY_FILL = PatternFill("solid", fgColor="E2F0D9")
ALERT_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
DECISION_OPTIONS = (
    "Не проверено",
    "Подтвердить",
    "Уменьшить",
    "Отклонить",
    "Нужны данные",
)
STATUS_LABELS = {
    "sales_start": "Пошли продажи",
    "sale": "Растим",
    "working": "Поддерживаем",
}


def _decimal(value: Any) -> Decimal:
    text = str(value or "").strip().replace(" ", "").replace(",", ".")
    return Decimal(text or "0")


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _load_unit_economics(path: Path | None) -> dict[tuple[str, str], dict[str, Decimal]]:
    if path is None or not path.exists():
        return {}
    return {
        (
            str(row.get("decision_date") or "").strip(),
            str(row.get("nomenclature_code") or "").strip(),
        ): {
            "unit_cost": _decimal(row.get("inventory_cost_per_unit_rub")),
            "unit_margin": _decimal(row.get("gross_margin_per_unit_rub")),
        }
        for row in _read_rows(path)
    }


def _date(value: Any) -> date | None:
    text = str(value or "").strip()
    return date.fromisoformat(text) if text else None


def _style_header(sheet, row: int) -> None:
    for cell in sheet[row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_body(sheet, *, min_row: int, max_row: int, max_col: int) -> None:
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def _add_table(sheet, *, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _add_decision_validation(sheet, column: str, max_row: int) -> None:
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(DECISION_OPTIONS) + '"',
        allow_blank=False,
    )
    validation.promptTitle = "Решение закупщика"
    validation.prompt = "Выберите итог проверки из списка"
    validation.error = "Нужно выбрать одно из предложенных решений"
    validation.errorTitle = "Недопустимое значение"
    sheet.add_data_validation(validation)
    validation.add(f"{column}2:{column}{max(max_row, 2)}")


def _prepare_rows(
    rows: Sequence[Mapping[str, Any]],
    *,
    unit_economics: Mapping[tuple[str, str], Mapping[str, Decimal]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    unit_economics = unit_economics or {}
    by_code: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        by_code[str(row.get("nomenclature_code") or "")].append(row)

    sku_rows: list[dict[str, Any]] = []
    sku_stats: dict[str, dict[str, Any]] = {}
    for code, group in by_code.items():
        ordered = sorted(group, key=lambda row: str(row.get("decision_date") or ""))
        latest = ordered[-1]
        economics = unit_economics.get((str(latest.get("decision_date") or ""), code), {})
        unit_cost = _decimal(economics.get("unit_cost"))
        unit_margin = _decimal(economics.get("unit_margin"))
        shortages = [_decimal(row.get("projected_shortage_qty")) for row in group]
        increments = [_decimal(row.get("dynamic_minmax_increment_qty")) for row in group]
        ratios = [
            (
                _decimal(row.get("recent_rate")) / _decimal(row.get("baseline_rate"))
                if _decimal(row.get("baseline_rate")) > 0
                else Decimal("999") if _decimal(row.get("recent_rate")) > 0 else Decimal("0")
            )
            for row in group
        ]
        stats = {
            "signal_count": len(group),
            "first_signal": _date(ordered[0].get("decision_date")),
            "last_signal": _date(latest.get("decision_date")),
            "total_shortage": sum(shortages, Decimal("0")),
            "max_shortage": max(shortages, default=Decimal("0")),
            "total_increment": sum(increments, Decimal("0")),
            "max_increment": max(increments, default=Decimal("0")),
            "max_rate_ratio": max(ratios, default=Decimal("0")),
        }
        sku_stats[code] = stats
        sku_rows.append(
            {
                "decision": DECISION_OPTIONS[0],
                "confirmed_qty": None,
                "comment": "",
                "reviewer": "",
                "review_date": None,
                "nomenclature_code": code,
                "name": latest.get("name"),
                "latest_status": STATUS_LABELS.get(str(latest.get("status")), latest.get("status")),
                **stats,
                "latest_free_stock": _decimal(latest.get("free_stock_qty")),
                "latest_ordinary_min": _decimal(latest.get("ordinary_min_stock_qty")),
                "latest_shortage": _decimal(latest.get("projected_shortage_qty")),
                "latest_increment": _decimal(latest.get("dynamic_minmax_increment_qty")),
                "latest_unit_cost": unit_cost,
                "latest_unit_margin": unit_margin,
                "latest_required_capital": (
                    _decimal(latest.get("dynamic_minmax_increment_qty")) * unit_cost
                ),
                "latest_margin_at_risk": (
                    _decimal(latest.get("projected_shortage_qty")) * unit_margin
                ),
                "lead_quantile": latest.get("lead_quantile"),
                "production_action": latest.get("production_action"),
            }
        )
    sku_rows.sort(
        key=lambda row: (
            -int(row["signal_count"]),
            -_decimal(row["total_shortage"]),
            str(row["nomenclature_code"]),
        )
    )

    detail_rows: list[dict[str, Any]] = []
    for row in rows:
        code = str(row.get("nomenclature_code") or "")
        stats = sku_stats[code]
        economics = unit_economics.get((str(row.get("decision_date") or ""), code), {})
        unit_cost = _decimal(economics.get("unit_cost"))
        unit_margin = _decimal(economics.get("unit_margin"))
        shortage = _decimal(row.get("projected_shortage_qty"))
        increment = _decimal(row.get("dynamic_minmax_increment_qty"))
        detail_rows.append(
            {
                "decision": DECISION_OPTIONS[0],
                "confirmed_qty": None,
                "comment": "",
                "reviewer": "",
                "review_date": None,
                "decision_date": _date(row.get("decision_date")),
                "nomenclature_code": code,
                "name": row.get("name"),
                "status": STATUS_LABELS.get(str(row.get("status")), row.get("status")),
                "signal_count": stats["signal_count"],
                "ordinary_min": _decimal(row.get("ordinary_min_stock_qty")),
                "free_stock": _decimal(row.get("free_stock_qty")),
                "recent_sales": _decimal(row.get("recent_sales_qty")),
                "baseline_sales": _decimal(row.get("baseline_sales_qty")),
                "recent_rate": _decimal(row.get("recent_rate")),
                "baseline_rate": _decimal(row.get("baseline_rate")),
                "rate_ratio": (
                    _decimal(row.get("recent_rate")) / _decimal(row.get("baseline_rate"))
                    if _decimal(row.get("baseline_rate")) > 0
                    else Decimal("999") if _decimal(row.get("recent_rate")) > 0 else Decimal("0")
                ),
                "lead_quantile": row.get("lead_quantile"),
                "projected_demand": _decimal(row.get("projected_demand_qty")),
                "inventory_position": _decimal(row.get("inventory_position_qty")),
                "projected_shortage": shortage,
                "increment": increment,
                "unit_cost": unit_cost,
                "unit_margin": unit_margin,
                "required_capital": increment * unit_cost,
                "margin_at_risk": shortage * unit_margin,
                "reason": row.get("reason"),
                "human_check": row.get("human_check"),
                "production_action": row.get("production_action"),
            }
        )
    detail_rows.sort(
        key=lambda row: (
            -int(row["signal_count"]),
            -_decimal(row["projected_shortage"]),
            row["decision_date"] or date.min,
            str(row["nomenclature_code"]),
        )
    )
    return sku_rows, detail_rows


def _write_summary_sheet(
    workbook: Workbook,
    rows: Sequence[Mapping[str, Any]],
    sku_rows: Sequence[Mapping[str, Any]],
) -> None:
    sheet = workbook.active
    sheet.title = "Сводка"
    sheet["A1"] = "Dynamic min/max — read-only shadow"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = (
        "Историческая очередь для проверки правила. Это не текущий заказ и не разрешение "
        "на запись в 1С, Bitrix24 или Telegram."
    )
    sheet["A2"].font = SUBTITLE_FONT
    sheet.merge_cells("A1:F1")
    sheet.merge_cells("A2:F2")
    sheet["A4"] = "Показатель"
    sheet["B4"] = "Значение"
    metrics = [
        ("Исторических сигналов", len(rows)),
        ("Уникальных SKU", len(sku_rows)),
        ("SKU с повторными сигналами", sum(int(row["signal_count"]) > 1 for row in sku_rows)),
        (
            "Максимум сигналов на SKU",
            max((int(row["signal_count"]) for row in sku_rows), default=0),
        ),
        (
            "Сумма рекомендованной добавки, шт.",
            sum((_decimal(row.get("dynamic_minmax_increment_qty")) for row in rows), Decimal("0")),
        ),
        (
            "Сумма прогнозируемого дефицита, шт.",
            sum((_decimal(row.get("projected_shortage_qty")) for row in rows), Decimal("0")),
        ),
        (
            "Капитал последних рекомендаций по SKU, ₽",
            sum((_decimal(row.get("latest_required_capital")) for row in sku_rows), Decimal("0")),
        ),
        ("Production action", "none_read_only"),
    ]
    for index, (label, value) in enumerate(metrics, start=5):
        sheet.cell(index, 1, label)
        sheet.cell(index, 2, float(value) if isinstance(value, Decimal) else value)
    _style_header(sheet, 4)
    _style_body(sheet, min_row=5, max_row=4 + len(metrics), max_col=2)

    monthly: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"signals": 0, "skus": set(), "shortage": Decimal("0"), "increment": Decimal("0")}
    )
    for row in rows:
        month = str(row.get("decision_date") or "")[:7]
        bucket = monthly[month]
        bucket["signals"] += 1
        bucket["skus"].add(str(row.get("nomenclature_code") or ""))
        bucket["shortage"] += _decimal(row.get("projected_shortage_qty"))
        bucket["increment"] += _decimal(row.get("dynamic_minmax_increment_qty"))
    start = 14
    headers = ["Месяц", "Сигналов", "SKU", "Дефицит, шт.", "Добавка, шт."]
    for column, header in enumerate(headers, start=1):
        sheet.cell(start, column, header)
    for month in sorted(monthly):
        values = monthly[month]
        sheet.append(
            [
                month,
                values["signals"],
                len(values["skus"]),
                float(values["shortage"]),
                float(values["increment"]),
            ]
        )
    _style_header(sheet, start)
    _style_body(sheet, min_row=start + 1, max_row=start + len(monthly), max_col=5)
    chart = BarChart()
    chart.title = "Исторические сигналы по месяцам"
    chart.y_axis.title = "Количество сигналов"
    chart.x_axis.title = "Месяц"
    chart.add_data(
        Reference(sheet, min_col=2, min_row=start, max_row=start + len(monthly)),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(sheet, min_col=1, min_row=start + 1, max_row=start + len(monthly))
    )
    chart.height = 7
    chart.width = 14
    sheet.add_chart(chart, "H4")
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 20
    for column in range(3, 7):
        sheet.column_dimensions[get_column_letter(column)].width = 16
    sheet.freeze_panes = "A4"


def _write_sku_sheet(workbook: Workbook, rows: Sequence[Mapping[str, Any]]) -> None:
    sheet = workbook.create_sheet("SKU-сводка")
    headers = [
        "Решение",
        "Подтверждено, шт.",
        "Комментарий",
        "Проверил",
        "Дата проверки",
        "Код SKU",
        "Наименование",
        "Последний статус",
        "Сигналов",
        "Первый сигнал",
        "Последний сигнал",
        "Суммарный дефицит, шт.",
        "Макс. дефицит, шт.",
        "Суммарная добавка, шт.",
        "Макс. добавка, шт.",
        "Макс. ускорение, x",
        "Последний свободный запас",
        "Последний обычный min",
        "Последний дефицит, шт.",
        "Последняя добавка, шт.",
        "Себестоимость на дату сигнала, ₽",
        "Маржа на дату сигнала, ₽",
        "Капитал последней добавки, ₽",
        "Маржа под риском, ₽",
        "Lead quantile",
        "Production action",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row["decision"],
                row["confirmed_qty"],
                row["comment"],
                row["reviewer"],
                row["review_date"],
                row["nomenclature_code"],
                row["name"],
                row["latest_status"],
                row["signal_count"],
                row["first_signal"],
                row["last_signal"],
                float(row["total_shortage"]),
                float(row["max_shortage"]),
                float(row["total_increment"]),
                float(row["max_increment"]),
                float(row["max_rate_ratio"]),
                float(row["latest_free_stock"]),
                float(row["latest_ordinary_min"]),
                float(row["latest_shortage"]),
                float(row["latest_increment"]),
                float(row["latest_unit_cost"]),
                float(row["latest_unit_margin"]),
                float(row["latest_required_capital"]),
                float(row["latest_margin_at_risk"]),
                row["lead_quantile"],
                row["production_action"],
            ]
        )
    _style_header(sheet, 1)
    _style_body(sheet, min_row=2, max_row=sheet.max_row, max_col=len(headers))
    sheet.freeze_panes = "F2"
    table_ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    _add_table(sheet, ref=table_ref, name="SkuSummary")
    sheet.auto_filter.ref = table_ref
    _add_decision_validation(sheet, "A", sheet.max_row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.fill = INPUT_FILL
    for column in (5, 10, 11):
        for cell in sheet.iter_cols(
            min_col=column, max_col=column, min_row=2, max_row=sheet.max_row
        ):
            for item in cell:
                item.number_format = "dd.mm.yyyy"
    sheet.conditional_formatting.add(
        f"I2:I{sheet.max_row}", DataBarRule(start_type="min", end_type="max", color="5B9BD5")
    )
    sheet.conditional_formatting.add(
        f"L2:L{sheet.max_row}",
        ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="F8696B"),
    )
    sheet.conditional_formatting.add(
        f"A2:A{sheet.max_row}", FormulaRule(formula=['$A2="Подтвердить"'], fill=READ_ONLY_FILL)
    )
    sheet.conditional_formatting.add(
        f"A2:A{sheet.max_row}", FormulaRule(formula=['$A2="Отклонить"'], fill=ALERT_FILL)
    )
    widths = {"A": 18, "B": 16, "C": 32, "D": 18, "E": 14, "F": 15, "G": 58, "H": 20}
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    for column in range(9, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18


def _write_detail_sheet(workbook: Workbook, rows: Sequence[Mapping[str, Any]]) -> None:
    sheet = workbook.create_sheet("История сигналов")
    headers = [
        "Решение",
        "Подтверждено, шт.",
        "Комментарий",
        "Проверил",
        "Дата проверки",
        "Дата сигнала",
        "Код SKU",
        "Наименование",
        "Статус",
        "Сигналов по SKU",
        "Обычный min",
        "Свободный запас",
        "Продажи recent",
        "Продажи baseline",
        "Скорость recent",
        "Скорость baseline",
        "Ускорение, x",
        "Lead quantile",
        "Прогноз спроса",
        "Inventory position",
        "Дефицит, шт.",
        "Добавка, шт.",
        "Себестоимость, ₽",
        "Маржа с единицы, ₽",
        "Требуемый капитал, ₽",
        "Маржа под риском, ₽",
        "Причина",
        "Что проверить",
        "Production action",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row["decision"],
                row["confirmed_qty"],
                row["comment"],
                row["reviewer"],
                row["review_date"],
                row["decision_date"],
                row["nomenclature_code"],
                row["name"],
                row["status"],
                row["signal_count"],
                float(row["ordinary_min"]),
                float(row["free_stock"]),
                float(row["recent_sales"]),
                float(row["baseline_sales"]),
                float(row["recent_rate"]),
                float(row["baseline_rate"]),
                float(row["rate_ratio"]),
                row["lead_quantile"],
                float(row["projected_demand"]),
                float(row["inventory_position"]),
                float(row["projected_shortage"]),
                float(row["increment"]),
                float(row["unit_cost"]),
                float(row["unit_margin"]),
                float(row["required_capital"]),
                float(row["margin_at_risk"]),
                row["reason"],
                row["human_check"],
                row["production_action"],
            ]
        )
    _style_header(sheet, 1)
    _style_body(sheet, min_row=2, max_row=sheet.max_row, max_col=len(headers))
    sheet.freeze_panes = "G2"
    table_ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    _add_table(sheet, ref=table_ref, name="SignalHistory")
    sheet.auto_filter.ref = table_ref
    _add_decision_validation(sheet, "A", sheet.max_row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.fill = INPUT_FILL
    for column in (5, 6):
        for cells in sheet.iter_cols(
            min_col=column, max_col=column, min_row=2, max_row=sheet.max_row
        ):
            for cell in cells:
                cell.number_format = "dd.mm.yyyy"
    sheet.conditional_formatting.add(
        f"U2:U{sheet.max_row}",
        ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="F8696B"),
    )
    sheet.conditional_formatting.add(
        f"Q2:Q{sheet.max_row}",
        ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="FFD966"),
    )
    widths = {"A": 18, "B": 16, "C": 30, "D": 18, "E": 14, "F": 14, "G": 15, "H": 58, "I": 20}
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    for column in range(10, 23):
        sheet.column_dimensions[get_column_letter(column)].width = 17
    sheet.column_dimensions["AA"].width = 48
    sheet.column_dimensions["AB"].width = 54
    sheet.column_dimensions["AC"].width = 20


def build_workbook(
    input_csv: Path,
    output_xlsx: Path,
    *,
    decision_inputs_csv: Path | None = None,
) -> dict[str, Any]:
    source_rows = _read_rows(input_csv)
    if not source_rows:
        raise ValueError("shadow recommendations CSV is empty")
    unit_economics = _load_unit_economics(decision_inputs_csv)
    sku_rows, detail_rows = _prepare_rows(source_rows, unit_economics=unit_economics)
    workbook = Workbook()
    workbook.properties.title = "Dynamic min/max read-only shadow"
    workbook.properties.subject = "Историческая очередь рекомендаций для закупщика"
    workbook.properties.creator = "pricing-service"
    workbook.calculation.fullCalcOnLoad = True
    _write_summary_sheet(workbook, source_rows, sku_rows)
    _write_sku_sheet(workbook, sku_rows)
    _write_detail_sheet(workbook, detail_rows)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    return {
        "schema": "display_auto_order_shadow_workbook.v1",
        "input": str(input_csv),
        "output": str(output_xlsx),
        "signal_rows": len(detail_rows),
        "unique_sku_count": len(sku_rows),
        "repeat_sku_count": sum(int(row["signal_count"]) > 1 for row in sku_rows),
        "unit_economics_match_count": sum(
            _decimal(row["unit_cost"]) > 0 or _decimal(row["unit_margin"]) > 0
            for row in detail_rows
        ),
        "production_action": "none_read_only",
    }


def validate_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    expected = ["Сводка", "SKU-сводка", "История сигналов"]
    if workbook.sheetnames != expected:
        raise ValueError(f"unexpected workbook sheets: {workbook.sheetnames}")
    sku_sheet = workbook["SKU-сводка"]
    detail_sheet = workbook["История сигналов"]
    if not sku_sheet.tables or not detail_sheet.tables:
        raise ValueError("review sheets must contain Excel tables with filters")
    if sku_sheet.freeze_panes != "F2" or detail_sheet.freeze_panes != "G2":
        raise ValueError("review sheet freeze panes are missing")
    return {
        "sheets": workbook.sheetnames,
        "sku_rows": max(0, sku_sheet.max_row - 1),
        "detail_rows": max(0, detail_sheet.max_row - 1),
        "validated": True,
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-csv", type=Path, required=True)
    parser.add_argument("--output-xlsx", type=Path, required=True)
    parser.add_argument("--decision-inputs-csv", type=Path)
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    decision_inputs_csv = args.decision_inputs_csv
    if decision_inputs_csv is None:
        candidate = args.input_csv.parent.parent / "preflight" / "decision-inputs.csv"
        decision_inputs_csv = candidate if candidate.exists() else None
    summary = build_workbook(
        args.input_csv,
        args.output_xlsx,
        decision_inputs_csv=decision_inputs_csv,
    )
    summary["validation"] = validate_workbook(args.output_xlsx)
    print(json.dumps(summary, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
