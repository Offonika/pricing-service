"""Экспорт workbook для ручного разбора матчинга дисплеев конкурентов."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.infrastructure.db import session_scope
from app.models import CompetitorItem, Product
from app.services.product_display_modification import display_frame_conflict
from tasks.match_competitor_items_embeddings import (
    _competitor_device_code_text,
    _competitor_display_backlight,
    _competitor_display_color,
    _competitor_display_construction,
    _competitor_display_has_frame,
    _competitor_display_has_touch,
    _competitor_display_mapped_1c_quality_raw,
    _competitor_display_matrix_tags,
    _competitor_display_quality,
    _competitor_display_quality_raw,
    _competitor_display_refresh_rate_hz,
    _competitor_display_type,
    _display_backlight_conflict,
    _display_color_conflict,
    _display_construction_conflict,
    _display_matrix_family_conflict,
    _display_matrix_tags_conflict,
    _display_model_code_blocks,
    _display_model_code_requires_review,
    _display_phone_model_conflict,
    _display_quality_conflict,
    _display_refresh_rate_conflict,
    _display_text_model_conflict,
    _display_touch_conflict,
    _display_word_is_feature,
    _extract_device_codes,
    _product_display_backlight,
    _product_display_color,
    _product_display_construction,
    _product_display_has_touch,
    _product_display_matrix_tags,
    _product_display_quality,
    _product_display_quality_raw,
    _product_display_refresh_rate_hz,
)

DEFAULT_INPUT = "reports/display_matching_extended_guardrails_full.csv"
DEFAULT_OUTPUT = "reports/display_matching_review_workbook.xlsx"

SUMMARY_FIELDS = ["metric", "value"]
REVIEW_FIELDS = [
    "bucket",
    "status",
    "review_reasons",
    "competitor_item_id",
    "competitor",
    "external_id",
    "competitor_name",
    "normalized_title",
    "product_id",
    "product_article",
    "product_name",
    "best_score",
    "gap",
    "second_product_id",
    "second_product_article",
    "second_product_name",
    "second_score",
    "competitor_frame",
    "product_frame",
    "product_modification_status",
    "competitor_color",
    "product_color",
    "competitor_quality_raw",
    "competitor_mapped_1c_quality_raw",
    "competitor_normalized_quality",
    "product_quality_raw",
    "product_normalized_quality",
    "competitor_quality",
    "product_quality",
    "competitor_touch",
    "product_touch",
    "competitor_backlight",
    "product_backlight",
    "competitor_matrix_tags",
    "product_matrix_tags",
    "competitor_construction",
    "product_construction",
    "competitor_refresh_rate_hz",
    "product_refresh_rate_hz",
    "competitor_device_codes",
    "product_device_codes",
    "guardrail_conflicts",
]

CORRECTION_FIELDS = [
    "product_id",
    "article",
    "name",
    "in_frame",
    "display_has_frame",
    "display_modification_status",
    "display_modification_source",
    "display_modification_confidence",
]


def _read_report(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _float_value(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _int_value(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_candidates(value: str | None) -> list[dict[str, Any]]:
    if not value:
        return []
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return []
    return parsed if isinstance(parsed, list) else []


def _format_set(values: set[str]) -> str:
    return ", ".join(sorted(values))


def _load_entities(
    session: Session,
    rows: list[dict[str, str]],
) -> tuple[dict[int, CompetitorItem], dict[int, Product]]:
    item_ids = {
        item_id
        for row in rows
        if (item_id := _int_value(row.get("competitor_item_id"))) is not None
    }
    product_ids = {
        product_id
        for row in rows
        if (product_id := _int_value(row.get("best_product_id"))) is not None
    }
    items = {
        item.id: item
        for item in session.execute(
            select(CompetitorItem).where(CompetitorItem.id.in_(item_ids))
        ).scalars()
    }
    products = {
        product.id: product
        for product in session.execute(select(Product).where(Product.id.in_(product_ids))).scalars()
    }
    return items, products


def _guardrail_conflicts(item: CompetitorItem, product: Product) -> list[str]:
    conflicts = []
    if display_frame_conflict(product, _competitor_display_has_frame(item)):
        conflicts.append("frame")
    if _display_color_conflict(product, _competitor_display_color(item)):
        conflicts.append("color")
    if _display_model_code_blocks(item, product):
        conflicts.append("model_code")
    if _display_phone_model_conflict(item, product):
        conflicts.append("phone_model")
    if _display_text_model_conflict(item, product):
        conflicts.append("text_model")
    if _display_quality_conflict(product, _competitor_display_quality(item)):
        conflicts.append("quality")
    if _display_touch_conflict(product, _competitor_display_has_touch(item)):
        conflicts.append("touch")
    if _display_backlight_conflict(product, _competitor_display_backlight(item)):
        conflicts.append("backlight")
    if _display_matrix_tags_conflict(product, _competitor_display_matrix_tags(item)):
        conflicts.append("matrix_tags")
    if _display_construction_conflict(product, _competitor_display_construction(item)):
        conflicts.append("construction")
    if _display_matrix_family_conflict(
        product,
        _competitor_display_type(item),
        _competitor_display_construction(item),
    ):
        conflicts.append("matrix_family")
    if _display_refresh_rate_conflict(product, _competitor_display_refresh_rate_hz(item)):
        conflicts.append("refresh_rate")
    if _display_word_is_feature(" ".join(filter(None, [item.name, item.normalized_title]))):
        conflicts.append("false_display")
    return conflicts


def _review_reasons(
    row: dict[str, str],
    item: CompetitorItem | None,
    product: Product | None,
    *,
    safe_score: float,
    safe_gap: float,
) -> list[str]:
    reasons = []
    status = row.get("status") or ""
    score = _float_value(row.get("best_score"))
    gap = _float_value(row.get("gap"))
    candidates = _parse_candidates(row.get("candidates"))

    if status == "ambiguous":
        reasons.append("ambiguous_candidates")
    if status == "needs_review":
        reasons.append("needs_review")
    if score is None or score < safe_score:
        reasons.append("low_score")
    if gap is None or gap < safe_gap:
        reasons.append("low_gap")
    if len(candidates) > 1:
        reasons.append("has_alternatives")
    if product and product.display_modification_status == "conflict":
        reasons.append("product_display_modification_conflict")
    if item and product:
        reasons.extend(f"guardrail:{name}" for name in _guardrail_conflicts(item, product))
        if _display_model_code_requires_review(item, product):
            reasons.append("code_variant_review")
        if _competitor_display_has_frame(item) is None or product.display_has_frame is None:
            reasons.append("frame_unknown_side")
        competitor_quality = _competitor_display_quality(item)
        product_quality = _product_display_quality(product)
        if bool(competitor_quality) != bool(product_quality):
            reasons.append("quality_unknown_one_side")
        elif not competitor_quality and not product_quality:
            reasons.append("quality_unknown_both")
        if not _competitor_display_color(item) or not _product_display_color(product):
            reasons.append("color_unknown_side")
    return reasons


def _bucket_for_row(
    row: dict[str, str],
    item: CompetitorItem | None,
    product: Product | None,
    *,
    safe_score: float,
    safe_gap: float,
) -> str:
    status = row.get("status")
    reasons = _review_reasons(row, item, product, safe_score=safe_score, safe_gap=safe_gap)
    if any(reason.startswith("guardrail:") for reason in reasons):
        return "blocked_by_guardrail"
    if status == "suggested" and not any(reason.startswith("guardrail:") for reason in reasons):
        if not reasons:
            return "safe_suggested"
        return "suggested_review"
    if status == "needs_review":
        return "needs_review"
    if status == "ambiguous":
        return "ambiguous"
    return "other"


def _review_row(
    row: dict[str, str],
    item: CompetitorItem | None,
    product: Product | None,
    *,
    safe_score: float,
    safe_gap: float,
) -> dict[str, object]:
    candidates = _parse_candidates(row.get("candidates"))
    second = candidates[1] if len(candidates) > 1 else {}
    reasons = _review_reasons(row, item, product, safe_score=safe_score, safe_gap=safe_gap)
    bucket = _bucket_for_row(row, item, product, safe_score=safe_score, safe_gap=safe_gap)

    competitor_codes = _extract_device_codes(_competitor_device_code_text(item)) if item else set()
    product_codes = _extract_device_codes(product.name) if product else set()

    return {
        "bucket": bucket,
        "status": row.get("status"),
        "review_reasons": ", ".join(reasons),
        "competitor_item_id": row.get("competitor_item_id"),
        "competitor": row.get("competitor"),
        "external_id": row.get("external_id"),
        "competitor_name": row.get("name"),
        "normalized_title": row.get("normalized_title"),
        "product_id": row.get("best_product_id"),
        "product_article": row.get("best_product_article"),
        "product_name": row.get("best_product_name"),
        "best_score": _float_value(row.get("best_score")),
        "gap": _float_value(row.get("gap")),
        "second_product_id": second.get("product_id"),
        "second_product_article": second.get("article"),
        "second_product_name": second.get("name"),
        "second_score": second.get("score"),
        "competitor_frame": _competitor_display_has_frame(item) if item else None,
        "product_frame": product.display_has_frame if product else None,
        "product_modification_status": product.display_modification_status if product else None,
        "competitor_color": _competitor_display_color(item) if item else None,
        "product_color": _product_display_color(product) if product else None,
        "competitor_quality_raw": _competitor_display_quality_raw(item) if item else None,
        "competitor_mapped_1c_quality_raw": (
            _competitor_display_mapped_1c_quality_raw(item) if item else None
        ),
        "competitor_normalized_quality": _competitor_display_quality(item) if item else None,
        "product_quality_raw": _product_display_quality_raw(product) if product else None,
        "product_normalized_quality": _product_display_quality(product) if product else None,
        "competitor_quality": _competitor_display_quality(item) if item else None,
        "product_quality": _product_display_quality(product) if product else None,
        "competitor_touch": _competitor_display_has_touch(item) if item else None,
        "product_touch": _product_display_has_touch(product) if product else None,
        "competitor_backlight": _competitor_display_backlight(item) if item else None,
        "product_backlight": _product_display_backlight(product) if product else None,
        "competitor_matrix_tags": (
            _format_set(_competitor_display_matrix_tags(item)) if item else ""
        ),
        "product_matrix_tags": (
            _format_set(_product_display_matrix_tags(product)) if product else ""
        ),
        "competitor_construction": _competitor_display_construction(item) if item else None,
        "product_construction": _product_display_construction(product) if product else None,
        "competitor_refresh_rate_hz": _competitor_display_refresh_rate_hz(item) if item else None,
        "product_refresh_rate_hz": _product_display_refresh_rate_hz(product) if product else None,
        "competitor_device_codes": _format_set(competitor_codes),
        "product_device_codes": _format_set(product_codes),
        "guardrail_conflicts": (
            ", ".join(_guardrail_conflicts(item, product)) if item and product else ""
        ),
    }


def _correction_rows(session: Session) -> list[dict[str, object]]:
    products = session.execute(
        select(Product)
        .where(Product.display_modification_status == "conflict")
        .order_by(Product.article)
    ).scalars()
    return [
        {
            "product_id": product.id,
            "article": product.article,
            "name": product.name,
            "in_frame": product.in_frame,
            "display_has_frame": product.display_has_frame,
            "display_modification_status": product.display_modification_status,
            "display_modification_source": product.display_modification_source,
            "display_modification_confidence": product.display_modification_confidence,
        }
        for product in products
    ]


def _write_sheet(
    workbook: Workbook,
    title: str,
    fields: list[str],
    rows: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(fields)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append([row.get(field) for field in fields])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for idx, field in enumerate(fields, start=1):
        width = min(max(len(field) + 2, 12), 60)
        if field in {"competitor_name", "product_name", "normalized_title", "review_reasons"}:
            width = 55
        sheet.column_dimensions[get_column_letter(idx)].width = width
    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_workbook(
    session: Session,
    *,
    input_report: Path,
    output: Path,
    safe_score: float,
    safe_gap: float,
) -> dict[str, int]:
    rows = [row for row in _read_report(input_report) if row.get("item_type") == "display"]
    items, products = _load_entities(session, rows)

    review_rows = [
        _review_row(
            row,
            items.get(_int_value(row.get("competitor_item_id")) or -1),
            products.get(_int_value(row.get("best_product_id")) or -1),
            safe_score=safe_score,
            safe_gap=safe_gap,
        )
        for row in rows
    ]
    by_bucket: dict[str, list[dict[str, object]]] = {}
    for row in review_rows:
        by_bucket.setdefault(str(row["bucket"]), []).append(row)

    status_counts = Counter(str(row.get("status")) for row in review_rows)
    bucket_counts = Counter(str(row.get("bucket")) for row in review_rows)
    reason_counts = Counter(
        reason.strip()
        for row in review_rows
        for reason in str(row.get("review_reasons") or "").split(",")
        if reason.strip()
    )

    summary_rows = [
        {"metric": "input_report", "value": str(input_report)},
        {"metric": "safe_score", "value": safe_score},
        {"metric": "safe_gap", "value": safe_gap},
        {"metric": "display_rows", "value": len(review_rows)},
        *[
            {"metric": f"status:{status}", "value": count}
            for status, count in status_counts.most_common()
        ],
        *[
            {"metric": f"bucket:{bucket}", "value": count}
            for bucket, count in bucket_counts.most_common()
        ],
        *[
            {"metric": f"reason:{reason}", "value": count}
            for reason, count in reason_counts.most_common(30)
        ],
    ]

    correction_rows = _correction_rows(session)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "summary", SUMMARY_FIELDS, summary_rows)
    _write_sheet(workbook, "safe_suggested", REVIEW_FIELDS, by_bucket.get("safe_suggested", []))
    _write_sheet(
        workbook,
        "blocked_by_guardrail",
        REVIEW_FIELDS,
        by_bucket.get("blocked_by_guardrail", []),
    )
    _write_sheet(workbook, "suggested_review", REVIEW_FIELDS, by_bucket.get("suggested_review", []))
    _write_sheet(workbook, "needs_review", REVIEW_FIELDS, by_bucket.get("needs_review", []))
    _write_sheet(workbook, "ambiguous", REVIEW_FIELDS, by_bucket.get("ambiguous", []))
    _write_sheet(workbook, "1c_corrections", CORRECTION_FIELDS, correction_rows)
    workbook.save(output)

    return {
        "display_rows": len(review_rows),
        "safe_suggested": len(by_bucket.get("safe_suggested", [])),
        "blocked_by_guardrail": len(by_bucket.get("blocked_by_guardrail", [])),
        "suggested_review": len(by_bucket.get("suggested_review", [])),
        "needs_review": len(by_bucket.get("needs_review", [])),
        "ambiguous": len(by_bucket.get("ambiguous", [])),
        "onec_corrections": len(correction_rows),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default=DEFAULT_INPUT, help="CSV из match dry-run")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Путь к XLSX workbook")
    parser.add_argument("--safe-score", type=float, default=0.8)
    parser.add_argument("--safe-gap", type=float, default=0.02)
    args = parser.parse_args()

    with session_scope(read_only=True) as session:
        stats = export_workbook(
            session,
            input_report=Path(args.input),
            output=Path(args.output),
            safe_score=args.safe_score,
            safe_gap=args.safe_gap,
        )
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
