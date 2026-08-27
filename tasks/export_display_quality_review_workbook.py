from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from app.infrastructure.db import session_scope

try:
    from report_display_quality_mismatch_candidates import build_report
except ImportError:  # pragma: no cover - fallback for alternate launch modes
    from tasks.report_display_quality_mismatch_candidates import build_report

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
FILL_UNIQUE = PatternFill(fill_type="solid", fgColor="E2F0D9")
FILL_MULTI = PatternFill(fill_type="solid", fgColor="FFF2CC")
FILL_POOL_ONLY = PatternFill(fill_type="solid", fgColor="FCE4D6")
FILL_NO_POOL = PatternFill(fill_type="solid", fgColor="F4CCCC")
REVIEW_STATUS_OPTIONS = (
    "не проверено",
    "подтвердить текущий",
    "заменить на рекомендацию",
    "заменить вручную",
    "пропустить",
)


def _apply_header_style(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_body_style(sheet) -> None:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 50)


def _recommended_candidate(row: dict[str, object]) -> dict[str, object] | None:
    candidates = row.get("matching_quality_candidates") or []
    if len(candidates) == 1:
        return candidates[0]
    return None


def _aggregate_review_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, str, str], dict[str, object]] = {}

    for row in rows:
        key = (
            str(row.get("current_quality_raw") or row.get("current_quality") or ""),
            str(row.get("competitor_quality_price") or row.get("competitor_quality") or ""),
            str(row.get("competitor") or ""),
        )
        bucket = grouped.setdefault(
            key,
            {
                "current_quality_raw": key[0],
                "competitor_quality_price": key[1],
                "competitor": key[2],
                "row_count": 0,
                "rows_with_candidates": 0,
                "rows_with_matching_quality_candidates": 0,
                "rows_with_unique_recommendation": 0,
                "recommended_quality_counter": Counter(),
                "normalized_our_qualities": Counter(),
                "normalized_competitor_qualities": Counter(),
                "mapped_1c_raw_counter": Counter(),
                "match_ids": [],
                "product_ids": [],
                "parsed_keys": Counter(),
                "sample_candidates": [],
            },
        )
        bucket["row_count"] += 1
        if int(row["candidate_count"]) > 0:
            bucket["rows_with_candidates"] += 1
        if int(row["matching_quality_candidate_count"]) > 0:
            bucket["rows_with_matching_quality_candidates"] += 1

        recommended = _recommended_candidate(row)
        if recommended is not None:
            bucket["rows_with_unique_recommendation"] += 1
            recommended_quality = str(recommended.get("quality") or "").strip()
            if recommended_quality:
                bucket["recommended_quality_counter"][recommended_quality] += 1

        bucket["match_ids"].append(str(row["product_match_id"]))
        bucket["product_ids"].append(str(row["current_product_id"]))

        parsed_key = str(row.get("parsed_key") or "").strip()
        if parsed_key:
            bucket["parsed_keys"][parsed_key] += 1
        normalized_our_quality = str(row.get("current_quality") or "").strip()
        if normalized_our_quality:
            bucket["normalized_our_qualities"][normalized_our_quality] += 1
        normalized_quality = str(row.get("competitor_quality") or "").strip()
        if normalized_quality:
            bucket["normalized_competitor_qualities"][normalized_quality] += 1
        mapped_1c_quality_raw = str(row.get("mapped_1c_quality_raw") or "").strip()
        if mapped_1c_quality_raw:
            bucket["mapped_1c_raw_counter"][mapped_1c_quality_raw] += 1

        for candidate in row.get("candidate_pool") or []:
            sample_candidate = {
                "quality": candidate.get("quality"),
                "quality_matches_competitor": bool(candidate.get("quality_matches_competitor")),
                "source": row["competitor"],
            }
            if sample_candidate not in bucket["sample_candidates"]:
                bucket["sample_candidates"].append(sample_candidate)

    aggregated_rows: list[dict[str, object]] = []
    for bucket in grouped.values():
        recommended_quality_counter = bucket["recommended_quality_counter"]
        recommended_quality = ""
        recommendation = "кандидаты не найдены"
        has_unique_recommendation = False

        if recommended_quality_counter:
            recommendation_values = list(recommended_quality_counter.keys())
            if len(recommendation_values) == 1:
                recommended_quality = recommendation_values[0]
                recommendation = "заменить на рекомендованное качество"
                has_unique_recommendation = True
            else:
                recommended_quality = " / ".join(sorted(recommendation_values))
                recommendation = "есть разные рекомендации, нужна проверка"
        elif int(bucket["rows_with_candidates"]) > 0:
            recommendation = "кандидаты есть, но автоматической рекомендации нет"

        aggregated_rows.append(
            {
                "current_quality_raw": bucket["current_quality_raw"],
                "competitor_quality_price": bucket["competitor_quality_price"],
                "competitor": bucket["competitor"],
                "row_count": bucket["row_count"],
                "rows_with_candidates": bucket["rows_with_candidates"],
                "rows_with_matching_quality_candidates": bucket[
                    "rows_with_matching_quality_candidates"
                ],
                "rows_with_unique_recommendation": bucket["rows_with_unique_recommendation"],
                "recommended_quality": recommended_quality,
                "recommendation": recommendation,
                "has_unique_recommendation": has_unique_recommendation,
                "match_ids": ", ".join(bucket["match_ids"][:20]),
                "product_ids": ", ".join(bucket["product_ids"][:20]),
                "parsed_key": ", ".join(key for key, _ in bucket["parsed_keys"].most_common(5)),
                "normalized_our_quality": " / ".join(
                    quality for quality, _ in bucket["normalized_our_qualities"].most_common()
                ),
                "mapped_1c_quality_raw": " / ".join(
                    quality for quality, _ in bucket["mapped_1c_raw_counter"].most_common()
                ),
                "normalized_competitor_quality": " / ".join(
                    quality
                    for quality, _ in bucket["normalized_competitor_qualities"].most_common()
                ),
                "sample_candidates": bucket["sample_candidates"],
            }
        )

    return aggregated_rows


def _sort_key(row: dict[str, object]) -> tuple[str, str, str]:
    return (
        str(row.get("current_quality_raw") or ""),
        str(row.get("competitor_quality_price") or ""),
        str(row.get("competitor") or ""),
    )


def _quality_options(review_rows: list[dict[str, object]]) -> list[str]:
    values: set[str] = set()
    for row in review_rows:
        for key in ("normalized_our_quality", "recommended_quality"):
            value = str(row.get(key) or "").strip()
            if value:
                values.update(part.strip() for part in value.split("/") if part.strip())
    return sorted(values)


def _write_instruction_sheet(
    workbook: Workbook, stats: dict[str, int], review_rows: list[dict[str, object]]
) -> None:
    sheet = workbook.active
    sheet.title = "Инструкция"
    sheet["A1"] = "Проверка сопоставления качества дисплеев"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A3"] = "Сформировано"
    sheet["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet["A5"] = "Что смотреть"
    sheet["A5"].font = Font(bold=True)
    instructions = [
        "1. Рабочий лист: 'Проверка качества'. Одна строка = одна уникальная пара качеств внутри источника.",
        "2. Повторы были потому, что раньше выгрузка шла по каждому товару. Теперь одинаковые пары качеств схлопнуты.",
        "3. Основная проверка: 'Качество в 1С' -> 'Качество в прайсе конкурента' -> 'Источник'.",
        "4. 'Качество в прайсе конкурента' теперь берется в максимально сыром виде из названия/прайса конкурента.",
        "5. На видимом листе оставлены только исходные наименования качеств и источник.",
        "6. Технические поля и нормализация скрыты справа и не мешают просмотру.",
        "6. Если нужен свой вариант, используйте 'Ручное качество'.",
    ]
    for index, line in enumerate(instructions, start=6):
        sheet[f"A{index}"] = line
    sheet["A14"] = "Сводка"
    sheet["A14"].font = Font(bold=True)
    summary_rows = [
        ("Всего сопоставлений дисплеев", stats["display_matches"]),
        ("Расхождения по качеству", stats["quality_mismatches"]),
        ("Есть пул кандидатов", stats["with_candidate_pool"]),
        ("Есть кандидат с нужным качеством", stats["with_matching_quality_candidate"]),
        ("Есть единственная рекомендация", stats["with_unique_matching_quality_candidate"]),
        ("Строк в Excel после агрегации", len(review_rows)),
    ]
    for offset, (label, value) in enumerate(summary_rows, start=15):
        sheet[f"A{offset}"] = label
        sheet[f"B{offset}"] = value
    sheet.column_dimensions["A"].width = 58
    sheet.column_dimensions["B"].width = 24


def _write_reference_sheet(workbook: Workbook, quality_options: list[str]) -> None:
    sheet = workbook.create_sheet("Справочник")
    sheet.append(["Справочник качеств"])
    for value in quality_options:
        sheet.append([value])
    _apply_header_style(sheet)
    _apply_body_style(sheet)
    sheet.sheet_state = "hidden"


def _write_review_sheet(workbook: Workbook, review_rows: list[dict[str, object]]) -> None:
    sheet = workbook.create_sheet("Проверка качества")
    headers = [
        "Статус",
        "Комментарий",
        "Качество в 1С",
        "Качество в прайсе конкурента",
        "Источник",
        "Рек. качество в 1С (raw)",
        "Выбранное наше качество",
        "Рекомендация",
        "Рек. качество",
        "Ручное качество",
        "Кол-во строк",
        "Строк с кандидатами",
        "Строк с рекомендацией",
        "Норм. наше качество",
        "Норм. качество конкурента",
        "ID товаров",
        "ID match",
    ]
    sheet.append(headers)

    for row_index, row in enumerate(review_rows, start=2):
        sheet.append(
            [
                REVIEW_STATUS_OPTIONS[0],
                "",
                row["current_quality_raw"],
                row["competitor_quality_price"],
                row["competitor"],
                row["mapped_1c_quality_raw"],
                row["normalized_our_quality"],
                row["recommendation"],
                row["recommended_quality"],
                "",
                row["row_count"],
                row["rows_with_candidates"],
                row["rows_with_unique_recommendation"],
                row["normalized_our_quality"],
                row["normalized_competitor_quality"],
                row["product_ids"],
                row["match_ids"],
            ]
        )

        fill = FILL_NO_POOL
        if row["has_unique_recommendation"]:
            fill = FILL_UNIQUE
        elif int(row["rows_with_matching_quality_candidates"]) > 0:
            fill = FILL_MULTI
        elif int(row["rows_with_candidates"]) > 0:
            fill = FILL_POOL_ONLY
        for cell in sheet[row_index]:
            cell.fill = fill

    _apply_header_style(sheet)
    _apply_body_style(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    status_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(REVIEW_STATUS_OPTIONS) + '"',
        allow_blank=False,
    )
    status_validation.prompt = "Выберите итог проверки"
    status_validation.error = "Нужно выбрать значение из списка"
    sheet.add_data_validation(status_validation)
    status_validation.add(f"A2:A{max(sheet.max_row, 2)}")

    _autosize_columns(sheet)
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 14
    for col in ("F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"):
        sheet.column_dimensions[col].hidden = True


def _write_candidates_sheet(workbook: Workbook, review_rows: list[dict[str, object]]) -> None:
    sheet = workbook.create_sheet("Кандидаты")
    headers = [
        "Качество в 1С",
        "Качество в прайсе конкурента",
        "Качество кандидата",
        "Совпадает по качеству",
        "Источник",
    ]
    sheet.append(headers)

    for row in review_rows:
        for candidate in row.get("sample_candidates") or []:
            sheet.append(
                [
                    row["current_quality_raw"],
                    row["competitor_quality_price"],
                    candidate["quality"],
                    "да" if candidate["quality_matches_competitor"] else "нет",
                    candidate["source"],
                ]
            )
            if candidate["quality_matches_competitor"]:
                for cell in sheet[sheet.max_row]:
                    cell.fill = FILL_UNIQUE

    _apply_header_style(sheet)
    _apply_body_style(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _autosize_columns(sheet)
    sheet.sheet_state = "hidden"


def _write_raw_sheet(workbook: Workbook, review_rows: list[dict[str, object]]) -> None:
    sheet = workbook.create_sheet("RawJSON")
    sheet.append(["match_ids", "product_ids", "sample_candidates_json"])
    for row in review_rows:
        sheet.append(
            [
                row["match_ids"],
                row["product_ids"],
                json.dumps(row.get("sample_candidates") or [], ensure_ascii=False),
            ]
        )
    _apply_header_style(sheet)
    _apply_body_style(sheet)
    sheet.freeze_panes = "A2"
    _autosize_columns(sheet)
    sheet.sheet_state = "hidden"


def build_workbook(output_path: Path) -> dict[str, object]:
    with session_scope(read_only=True) as session:
        stats, rows = build_report(session)

    review_rows = sorted(_aggregate_review_rows(rows), key=_sort_key)
    workbook = Workbook()
    _write_instruction_sheet(workbook, stats, review_rows)
    quality_options = _quality_options(review_rows)
    _write_reference_sheet(workbook, quality_options)
    _write_review_sheet(workbook, review_rows)
    _write_candidates_sheet(workbook, review_rows)
    _write_raw_sheet(workbook, review_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    summary = dict(stats)
    summary["written_rows"] = len(review_rows)
    summary["quality_options"] = len(quality_options)
    summary["output"] = str(output_path)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export an Excel workbook for manual review of display quality matching."
    )
    parser.add_argument(
        "--output",
        default="reports/display_quality_mismatch_review.xlsx",
        help="XLSX output path",
    )
    args = parser.parse_args()

    summary = build_workbook(Path(args.output))
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
