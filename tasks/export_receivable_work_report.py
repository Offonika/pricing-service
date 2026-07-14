from __future__ import annotations

import argparse
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.infrastructure.db.engines import build_engine
from app.models import ReceivableWorkItem
from app.services.receivable_workflow import STATUS_CLOSED

DEFAULT_OUTPUT_DIR = Path("reports/receivables/buyers")
SUMMARY_SHEET_TITLE = "Контроль"
DETAILS_SHEET_TITLE = "Детализация"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
BORDER_SIDE = Side(style="thin", color="D9E2F3")
BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)


def _parse_date(value: str) -> date:
    return date.fromisoformat(value)


def _money(value: Decimal | None) -> float:
    return float(value or Decimal("0"))


def _date(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "date"):
        return value.date().isoformat()
    return str(value)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _autosize(ws) -> None:
    for column in ws.columns:
        width = 10
        letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 60))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[letter].width = width


def load_receivable_work_report_items(session: Session) -> list[ReceivableWorkItem]:
    return (
        session.execute(
            select(ReceivableWorkItem)
            .where(ReceivableWorkItem.status != STATUS_CLOSED)
            .order_by(
                ReceivableWorkItem.department_name,
                ReceivableWorkItem.overdue_days.desc().nullslast(),
                ReceivableWorkItem.current_balance.desc(),
                ReceivableWorkItem.counterparty_name,
            )
        )
        .scalars()
        .all()
    )


def export_receivable_work_report(
    items: list[ReceivableWorkItem],
    *,
    output_path: Path,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = SUMMARY_SHEET_TITLE
    summary_ws.append(["Клиент", "Сумма просрочки", "Дней просрочки", "Ответственный"])
    for item in items:
        summary_ws.append(
            [
                item.counterparty_name or item.counterparty_ref,
                _money(item.current_balance),
                item.overdue_days,
                item.current_manager_name or item.origin_manager_name or item.assigned_source or "",
            ]
        )
    _style_header(summary_ws)
    _autosize(summary_ws)
    summary_ws.freeze_panes = "A2"

    details_ws = workbook.create_sheet(DETAILS_SHEET_TITLE)
    details_ws.append(
        [
            "Подразделение",
            "Клиент",
            "Сумма",
            "Статус",
            "Срок оплаты",
            "Дней просрочки",
            "Обещанная дата оплаты",
            "Следующее действие",
            "Результат последнего контакта",
            "Ответственный",
            "SMS",
            "Дата SMS",
            "Нет обновления с",
            "Bitrix item id",
        ]
    )
    for item in items:
        details_ws.append(
            [
                item.department_name or item.department_ref or "",
                item.counterparty_name or item.counterparty_ref,
                _money(item.current_balance),
                item.status,
                _date(item.due_date),
                item.overdue_days,
                _date(item.promised_payment_date),
                _date(item.next_action_date),
                item.last_contact_comment or "",
                item.current_manager_name or item.origin_manager_name or item.assigned_source or "",
                item.last_sms_status or "",
                _date(item.last_sms_at),
                _date(item.last_manager_update_at),
                item.bitrix_item_id or "",
            ]
        )
    _style_header(details_ws)
    _autosize(details_ws)
    details_ws.freeze_panes = "A2"

    workbook.save(output_path)
    return output_path


def build_output_path(*, business_date: date, output_dir: Path) -> Path:
    return output_dir / business_date.isoformat() / f"Дебиторка покупателей {business_date}.xlsx"


def main() -> None:
    parser = argparse.ArgumentParser(description="Export buyer receivables work report")
    parser.add_argument("--date", type=_parse_date, default=date.today())
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    engine = build_engine(get_settings().database_url)
    with Session(engine) as session:
        items = load_receivable_work_report_items(session)
        path = export_receivable_work_report(
            items,
            output_path=build_output_path(business_date=args.date, output_dir=args.output_dir),
        )
    print(path)


if __name__ == "__main__":
    main()
