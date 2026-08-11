"""Create an access-controlled Excel export of encrypted SMS journal data."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.infrastructure.db.engines import build_engine
from app.services.sms_journal import SmsJournalCipher
from app.services.sms_journal_export import SmsJournalExportRow, load_sms_journal_export_rows

MOSCOW = ZoneInfo("Europe/Moscow")
MAX_PERIOD_DAYS = 31
MAX_ROWS = 50_000
SUMMARY_SHEET_TITLE = "Сводка"
SHEET_TITLE = "SMS"
METADATA_SHEET_TITLE = "Параметры"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
BORDER_SIDE = Side(style="thin", color="D9E2F3")
BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

HEADERS = (
    "Event ID",
    "Дата и время (МСК)",
    "Источник",
    "Тип объекта",
    "Идентификатор объекта",
    "Событие",
    "Исполнитель",
    "Телефон (маска)",
    "Текст SMS",
    "Fingerprint текста",
    "Символов",
    "Кодировка",
    "Расчётных сегментов",
    "Провайдер",
    "Provider message ID",
    "Статус отправки",
    "Статус доставки",
    "Код ошибки",
    "Номер попытки",
    "Отправлено (МСК)",
    "Доставлено (МСК)",
    "Оплаченных сегментов",
    "Цена сегмента",
    "Стоимость",
    "Период сверки",
)

SUMMARY_HEADERS = (
    "Источник",
    "Событие",
    "Fingerprint текста",
    "Текст SMS",
    "Кодировка",
    "Сегментов на SMS",
    "Попыток",
    "Расчётных сегментов всего",
    "Оплаченных сегментов всего",
    "Стоимость всего",
)


def _parse_date(value: str) -> date:
    return date.fromisoformat(value)


def _authorized_actors(value: str | None) -> set[str]:
    return {item.strip() for item in (value or "").split(",") if item.strip()}


def validate_export_request(
    *,
    date_from: date,
    date_to: date,
    actor: str,
    allowed_actors: set[str],
    confirmed: bool,
) -> None:
    if date_to < date_from:
        raise ValueError("date-to must not be earlier than date-from")
    if (date_to - date_from).days + 1 > MAX_PERIOD_DAYS:
        raise ValueError(f"export period must not exceed {MAX_PERIOD_DAYS} days")
    if not confirmed:
        raise PermissionError("sensitive export requires explicit confirmation")
    if not allowed_actors or actor not in allowed_actors:
        raise PermissionError("actor is not allowed to export sensitive SMS data")


def _utc_bounds(date_from: date, date_to: date) -> tuple[datetime, datetime]:
    start = datetime.combine(date_from, time.min, tzinfo=MOSCOW)
    end = datetime.combine(date_to + timedelta(days=1), time.min, tzinfo=MOSCOW)
    return (
        start.astimezone(UTC).replace(tzinfo=None),
        end.astimezone(UTC).replace(tzinfo=None),
    )


def _moscow_text(value: datetime | None) -> str:
    if value is None:
        return ""
    aware = value.replace(tzinfo=UTC) if value.tzinfo is None else value.astimezone(UTC)
    return aware.astimezone(MOSCOW).strftime("%Y-%m-%d %H:%M:%S")


def _decimal(value: Decimal | None) -> float | None:
    return None if value is None else float(value)


def _row_values(row: SmsJournalExportRow) -> list[object]:
    return [
        row.event_id,
        _moscow_text(row.created_at),
        row.source_system,
        row.source_entity_type,
        row.source_entity_id,
        row.event_type,
        row.actor_id or "",
        row.recipient_phone_masked,
        row.message_text,
        row.message_fingerprint,
        row.character_count,
        row.encoding,
        row.estimated_segments,
        row.provider,
        row.provider_message_id or "",
        row.send_status,
        row.delivery_status,
        row.provider_error_code or "",
        row.attempt_number,
        _moscow_text(row.sent_at),
        _moscow_text(row.delivered_at),
        row.billed_segments,
        _decimal(row.unit_price),
        _decimal(row.total_cost),
        row.reconciliation_period or "",
    ]


def _summary_values(rows: list[SmsJournalExportRow]) -> list[list[object]]:
    grouped: dict[tuple[str, str, str, str, str, int], dict[str, object]] = {}
    for row in rows:
        key = (
            row.source_system,
            row.event_type,
            row.message_fingerprint,
            row.message_text,
            row.encoding,
            row.estimated_segments,
        )
        item = grouped.setdefault(
            key,
            {
                "attempts": 0,
                "estimated_segments": 0,
                "billed_segments": 0,
                "total_cost": Decimal("0"),
            },
        )
        item["attempts"] = int(item["attempts"]) + 1
        item["estimated_segments"] = int(item["estimated_segments"]) + row.estimated_segments
        item["billed_segments"] = int(item["billed_segments"]) + (row.billed_segments or 0)
        item["total_cost"] = Decimal(item["total_cost"]) + (row.total_cost or Decimal("0"))

    result = []
    for key, item in grouped.items():
        source_system, event_type, fingerprint, text, encoding, segments = key
        result.append(
            [
                source_system,
                event_type,
                fingerprint,
                text,
                encoding,
                segments,
                item["attempts"],
                item["estimated_segments"],
                item["billed_segments"],
                float(Decimal(item["total_cost"])),
            ]
        )
    return sorted(result, key=lambda item: (-int(item[7]), str(item[3])))


def _force_text_cells(worksheet) -> None:
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.data_type = "s"


def _style_table(worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in worksheet.columns:
        width = 10
        letter = get_column_letter(column[0].column)
        for cell in column:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            width = max(width, min(len(str(cell.value or "")) + 2, 80))
        worksheet.column_dimensions[letter].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def export_sms_journal_xlsx(
    rows: list[SmsJournalExportRow],
    *,
    output_path: Path,
    actor: str,
    date_from: date,
    date_to: date,
    source_system: str | None,
    exported_at: datetime | None = None,
) -> tuple[Path, Path]:
    if output_path.exists():
        raise FileExistsError(f"refusing to overwrite existing export: {output_path}")
    audit_path = output_path.with_suffix(output_path.suffix + ".audit.json")
    if audit_path.exists():
        raise FileExistsError(f"refusing to overwrite existing audit: {audit_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = SUMMARY_SHEET_TITLE
    summary.append(list(SUMMARY_HEADERS))
    for values in _summary_values(rows):
        summary.append(values)
    _force_text_cells(summary)
    _style_table(summary)

    worksheet = workbook.create_sheet(SHEET_TITLE)
    worksheet.append(list(HEADERS))
    for row in rows:
        worksheet.append(_row_values(row))
    _force_text_cells(worksheet)
    _style_table(worksheet)

    exported_at = exported_at or datetime.now(UTC)
    metadata = workbook.create_sheet(METADATA_SHEET_TITLE)
    metadata_rows = [
        ("Назначение", "Закрытая аналитическая выгрузка SMS"),
        ("Исполнитель", actor),
        ("Период с", date_from.isoformat()),
        ("Период по", date_to.isoformat()),
        ("Источник", source_system or "все"),
        ("Строк", len(rows)),
        ("Сформировано UTC", exported_at.astimezone(UTC).isoformat()),
        ("Ограничение", "Содержит полный текст SMS; телефон только в маске"),
    ]
    for item in metadata_rows:
        metadata.append(item)
    _force_text_cells(metadata)
    metadata.column_dimensions["A"].width = 24
    metadata.column_dimensions["B"].width = 70

    previous_umask = os.umask(0o077)
    try:
        workbook.save(output_path)
        os.chmod(output_path, 0o600)
        digest = hashlib.sha256(output_path.read_bytes()).hexdigest()
        audit = {
            "export_kind": "sms_journal_sensitive_xlsx",
            "actor": actor,
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "source_system": source_system,
            "row_count": len(rows),
            "exported_at": exported_at.astimezone(UTC).isoformat(),
            "xlsx_sha256": digest,
        }
        audit_path.write_text(
            json.dumps(audit, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
            encoding="utf-8",
        )
        os.chmod(audit_path, 0o600)
    finally:
        os.umask(previous_umask)
    return output_path, audit_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Export protected SMS journal statistics to XLSX")
    parser.add_argument("--date-from", type=_parse_date, required=True)
    parser.add_argument("--date-to", type=_parse_date, required=True)
    parser.add_argument("--actor", required=True)
    parser.add_argument("--source-system")
    parser.add_argument("--limit", type=int, default=MAX_ROWS)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--confirm-sensitive-export", action="store_true")
    args = parser.parse_args()

    settings = get_settings()
    validate_export_request(
        date_from=args.date_from,
        date_to=args.date_to,
        actor=args.actor,
        allowed_actors=_authorized_actors(settings.sms_journal_export_allowed_actors),
        confirmed=args.confirm_sensitive_export,
    )
    if not settings.sms_journal_encryption_key or not settings.sms_journal_phone_hash_key:
        raise RuntimeError("SMS journal encryption is not configured")
    created_from, created_to = _utc_bounds(args.date_from, args.date_to)
    cipher = SmsJournalCipher(
        settings.sms_journal_encryption_key,
        settings.sms_journal_phone_hash_key,
    )
    engine = build_engine(settings.database_url)
    with Session(engine) as session:
        rows = load_sms_journal_export_rows(
            session,
            cipher,
            created_from=created_from,
            created_to=created_to,
            source_system=args.source_system,
            limit=args.limit,
        )
    output_path, audit_path = export_sms_journal_xlsx(
        rows,
        output_path=args.output,
        actor=args.actor,
        date_from=args.date_from,
        date_to=args.date_to,
        source_system=args.source_system,
    )
    print(f"XLSX={output_path};AUDIT={audit_path};ROWS={len(rows)}")


if __name__ == "__main__":
    main()
