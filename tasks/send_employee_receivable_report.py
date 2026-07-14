from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import DataBarRule, FormulaRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import distinct, func, select, text
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.infrastructure.db.engines import build_engine
from app.models import ReceivableCase
from app.services.return_scheme import send_return_scheme_telegram_report

DEFAULT_OUTPUT_DIR = Path("reports/receivables/employee")
SEGMENT_EMPLOYEE = "employee"
CONTROL_SHEET_TITLE = "Контроль долгов"
SUMMARY_SHEET_TITLE = "Сводка"
CURRENT_SHEET_TITLE = "Текущие долги"
CHANGES_SHEET_TITLE = "Изменения по долгам"
RELATED_DOCS_SHEET_TITLE = "Реализации и списания"
CHART_DATA_SHEET_TITLE = "_chart_data"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
STRIPE_FILL = PatternFill(fill_type="solid", fgColor="F7FBFF")
BORDER_SIDE = Side(style="thin", color="D9E2F3")
BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
STATUS_FILLS = {
    "Новый": PatternFill(fill_type="solid", fgColor="D9EAF7"),
    "Рост": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "Снижение": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "Закрыт": PatternFill(fill_type="solid", fgColor="D9D9D9"),
}


@dataclass(slots=True)
class EmployeeReceivableItem:
    snapshot_date: date
    counterparty_ref: str
    counterparty_name: str | None
    current_balance: Decimal
    aged_bucket: str
    activity_segment: str
    current_manager_ref: str | None
    current_manager_name: str | None
    origin_document_ref: str | None
    origin_document_number: str | None
    counterparty_code: str | None = None


@dataclass(slots=True)
class EmployeeReceivableChange:
    status: str
    counterparty_ref: str
    counterparty_name: str | None
    current_balance: Decimal
    previous_balance: Decimal
    delta_balance: Decimal
    aged_bucket: str
    activity_segment: str
    current_manager_ref: str | None
    current_manager_name: str | None
    origin_document_ref: str | None
    origin_document_number: str | None
    counterparty_code: str | None = None


@dataclass(slots=True)
class EmployeeReceivableRelatedDocument:
    counterparty_name: str | None
    document_kind: str
    document_date: datetime | None
    document_number: str | None
    document_ref: str | None
    responsible_name: str | None
    department_name: str | None
    organization_name: str | None
    item_name: str | None
    quantity: Decimal
    amount_delta: Decimal


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    return date.fromisoformat(value)


def _load_env(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    env: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def _build_onec_engine(settings) -> Engine | None:
    if not settings.onec_database_url:
        return None
    return build_engine(
        settings.onec_database_url,
        connect_args={
            "timeout": float(settings.onec_query_timeout_seconds),
            "login_timeout": float(settings.onec_login_timeout_seconds),
        },
    )


def _format_money(value: Decimal) -> str:
    return f"{value:,.2f}".replace(",", " ")


def _format_display_date(value: date | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d.%m.%Y")


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d.%m.%Y %H:%M")


def _previous_week_window(anchor_date: date) -> tuple[datetime, datetime]:
    current_week_start = anchor_date - timedelta(days=anchor_date.weekday())
    previous_week_start = current_week_start - timedelta(days=7)
    return (
        datetime.combine(previous_week_start, time.min),
        datetime.combine(current_week_start, time.min),
    )


def build_output_path(*, snapshot_date: date, output_dir: Path) -> Path:
    dated_dir = output_dir / snapshot_date.isoformat()
    return dated_dir / f"Долги сотрудников {_format_display_date(snapshot_date)}.xlsx"


def _resolve_default_telegram_target(env: dict[str, str]) -> tuple[str | None, str | None]:
    token = env.get("WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_TOKEN") or env.get(
        "SMARTPHONE_RELEASES_ALERT_TELEGRAM_TOKEN"
    )
    chat_id = env.get("WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_CHAT_ID") or env.get(
        "SMARTPHONE_RELEASES_ALERT_TELEGRAM_CHAT_ID"
    )
    return token, chat_id


def _build_in_clause(values: Sequence[str], *, prefix: str) -> tuple[str, dict[str, str]]:
    params = {f"{prefix}_{index}": value for index, value in enumerate(values)}
    placeholders = ", ".join(f":{name}" for name in params)
    return placeholders, params


def _ref_expr(engine: Engine) -> str:
    if engine.dialect.name == "mssql":
        return "master.dbo.fn_varbintohexstr(c._IDRRef)"
    return "c._IDRRef"


def _with_nolock(engine: Engine) -> str:
    return "WITH (NOLOCK)" if engine.dialect.name == "mssql" else ""


def _table_name(engine: Engine, table_name: str) -> str:
    return f"dbo.{table_name}" if engine.dialect.name == "mssql" else table_name


def _build_ref_filter_clause(
    engine: Engine,
    refs: Sequence[str],
    *,
    column_name: str,
    prefix: str,
) -> tuple[str, dict[str, str]]:
    if engine.dialect.name == "mssql":
        hex_refs = [value.upper() for value in refs if re.fullmatch(r"0x[0-9A-Fa-f]+", value)]
        if not hex_refs:
            return "1 = 0", {}
        return f"{column_name} IN ({', '.join(hex_refs)})", {}

    params = {f"{prefix}_{index}": value for index, value in enumerate(refs)}
    placeholders = ", ".join(f":{name}" for name in params)
    return f"{column_name} IN ({placeholders})", params


def fetch_counterparty_code_mapping_by_ref(
    onec_engine: Engine,
    *,
    counterparty_refs: Sequence[str],
) -> dict[str, str]:
    refs = sorted({value for value in counterparty_refs if value})
    if not refs:
        return {}

    where_clause, params = _build_ref_filter_clause(
        onec_engine,
        refs,
        column_name="c._IDRRef",
        prefix="counterparty_ref",
    )
    stmt = text(f"""
        SELECT
            {_ref_expr(onec_engine)} AS counterparty_ref,
            RTRIM(c._Code) AS counterparty_code
        FROM dbo._Reference54 AS c {_with_nolock(onec_engine)}
        WHERE {where_clause}
    """)
    with onec_engine.connect() as connection:
        rows = connection.execute(stmt, params).mappings().all()
    return {
        str(row["counterparty_ref"]).strip().upper(): str(row["counterparty_code"]).strip()
        for row in rows
        if row.get("counterparty_ref") and row.get("counterparty_code")
    }


def resolve_employee_snapshot_dates(
    session: Session,
    *,
    requested_date: date | None,
    latest_not_after: date | None = None,
) -> tuple[date, date | None]:
    available_dates = (
        session.execute(
            select(distinct(ReceivableCase.snapshot_date))
            .where(ReceivableCase.segment == SEGMENT_EMPLOYEE)
            .order_by(ReceivableCase.snapshot_date.desc())
        )
        .scalars()
        .all()
    )
    if not available_dates:
        raise RuntimeError("No employee receivable snapshots found in receivable_case")

    if requested_date is not None and latest_not_after is not None:
        raise ValueError("requested_date and latest_not_after are mutually exclusive")

    if requested_date is not None:
        snapshot_date = requested_date
        if snapshot_date not in available_dates:
            raise RuntimeError(
                f"Employee receivable snapshot {snapshot_date.isoformat()} not found"
            )
    elif latest_not_after is not None:
        snapshot_date = next((item for item in available_dates if item <= latest_not_after), None)
        if snapshot_date is None:
            raise RuntimeError(
                "No employee receivable snapshot found on or before "
                f"{latest_not_after.isoformat()}"
            )
    else:
        snapshot_date = available_dates[0]

    previous_date = next((item for item in available_dates if item < snapshot_date), None)
    return snapshot_date, previous_date


def load_employee_items(session: Session, *, snapshot_date: date) -> list[EmployeeReceivableItem]:
    rows = (
        session.execute(
            select(ReceivableCase)
            .where(
                ReceivableCase.snapshot_date == snapshot_date,
                ReceivableCase.segment == SEGMENT_EMPLOYEE,
            )
            .order_by(ReceivableCase.current_balance.desc(), ReceivableCase.counterparty_ref)
        )
        .scalars()
        .all()
    )
    return [
        EmployeeReceivableItem(
            snapshot_date=row.snapshot_date,
            counterparty_ref=row.counterparty_ref,
            counterparty_name=row.counterparty_name,
            current_balance=Decimal(row.current_balance),
            aged_bucket=row.aged_bucket,
            activity_segment=row.activity_segment,
            current_manager_ref=row.current_manager_ref,
            current_manager_name=row.current_manager_name,
            origin_document_ref=row.origin_document_ref,
            origin_document_number=row.origin_document_number,
        )
        for row in rows
    ]


def enrich_employee_items_with_counterparty_codes(
    items: list[EmployeeReceivableItem],
    *,
    onec_engine: Engine | None,
) -> None:
    if onec_engine is None or not items:
        return
    mapping = fetch_counterparty_code_mapping_by_ref(
        onec_engine,
        counterparty_refs=[item.counterparty_ref for item in items],
    )
    for item in items:
        item.counterparty_code = mapping.get((item.counterparty_ref or "").upper())


def load_employee_snapshot_history(
    session: Session,
    *,
    snapshot_date: date,
    limit: int = 7,
) -> list[tuple[date, Decimal]]:
    snapshot_dates = (
        session.execute(
            select(distinct(ReceivableCase.snapshot_date))
            .where(
                ReceivableCase.segment == SEGMENT_EMPLOYEE,
                ReceivableCase.snapshot_date <= snapshot_date,
            )
            .order_by(ReceivableCase.snapshot_date.desc())
        )
        .scalars()
        .all()
    )
    target_dates = list(reversed(snapshot_dates[:limit]))
    if not target_dates:
        return []

    totals = {
        item_date: Decimal(total or 0)
        for item_date, total in session.execute(
            select(ReceivableCase.snapshot_date, func.sum(ReceivableCase.current_balance))
            .where(
                ReceivableCase.segment == SEGMENT_EMPLOYEE,
                ReceivableCase.snapshot_date.in_(target_dates),
            )
            .group_by(ReceivableCase.snapshot_date)
            .order_by(ReceivableCase.snapshot_date.asc())
        )
    }
    return [(item_date, totals.get(item_date, Decimal("0.00"))) for item_date in target_dates]


def load_employee_related_documents(
    *,
    onec_engine: Engine | None,
    snapshot_date: date,
) -> list[EmployeeReceivableRelatedDocument]:
    if onec_engine is None:
        return []
    previous_week_start, previous_week_end = _previous_week_window(snapshot_date)
    sale_table = _table_name(onec_engine, "_Document203")
    sale_line_table = _table_name(onec_engine, "_Document203_VT4966")
    organization_table = _table_name(onec_engine, "_Reference66")
    department_table = _table_name(onec_engine, "_Reference80")
    counterparty_table = _table_name(onec_engine, "_Reference54")
    item_table = _table_name(onec_engine, "_Reference62")
    responsible_table = _table_name(onec_engine, "_Reference69")
    document_ref_expr = (
        "master.dbo.fn_varbintohexstr(sale._IDRRef)"
        if onec_engine.dialect.name == "mssql"
        else "sale._IDRRef"
    )
    stmt = text(f"""
        SELECT
            counterparty._Description AS counterparty_name,
            sale._Date_Time AS document_date,
            RTRIM(sale._Number) AS document_number,
            {document_ref_expr} AS document_ref,
            responsible._Description AS responsible_name,
            department._Description AS department_name,
            organization._Description AS organization_name,
            item._Description AS item_name,
            sale_line._Fld4971 AS quantity,
            sale_line._Fld4982 AS amount_delta
        FROM {sale_table} AS sale {_with_nolock(onec_engine)}
        JOIN {sale_line_table} AS sale_line {_with_nolock(onec_engine)}
            ON sale_line._Document203_IDRRef = sale._IDRRef
        LEFT JOIN {organization_table} AS organization {_with_nolock(onec_engine)}
            ON organization._IDRRef = sale._Fld4932RRef
        LEFT JOIN {department_table} AS department {_with_nolock(onec_engine)}
            ON department._IDRRef = sale._Fld4940RRef
        LEFT JOIN {counterparty_table} AS counterparty {_with_nolock(onec_engine)}
            ON counterparty._IDRRef = sale._Fld4942RRef
        LEFT JOIN {responsible_table} AS responsible {_with_nolock(onec_engine)}
            ON responsible._IDRRef = sale._Fld4950RRef
        LEFT JOIN {item_table} AS item {_with_nolock(onec_engine)}
            ON item._IDRRef = sale_line._Fld4974RRef
        WHERE sale._Date_Time >= :previous_week_start
          AND sale._Date_Time < :previous_week_end
          AND sale._Posted = :posted
          AND sale._Marked = :marked
          AND counterparty._Description = :counterparty_name
        ORDER BY sale._Date_Time DESC, sale._Number DESC, sale_line._LineNo4967 ASC
        """)
    with onec_engine.connect() as connection:
        rows = (
            connection.execute(
                stmt,
                {
                    "previous_week_start": previous_week_start,
                    "previous_week_end": previous_week_end,
                    "posted": b"\x01" if onec_engine.dialect.name == "mssql" else 1,
                    "marked": b"\x00" if onec_engine.dialect.name == "mssql" else 0,
                    "counterparty_name": "Списание товара",
                },
            )
            .mappings()
            .all()
        )

    return [
        EmployeeReceivableRelatedDocument(
            counterparty_name=row.get("counterparty_name"),
            document_kind="Реализация товаров и услуг",
            document_date=row.get("document_date"),
            document_number=row.get("document_number"),
            document_ref=row.get("document_ref"),
            responsible_name=row.get("responsible_name"),
            department_name=row.get("department_name"),
            organization_name=row.get("organization_name"),
            item_name=row.get("item_name"),
            quantity=Decimal(row.get("quantity") or 0),
            amount_delta=Decimal(row.get("amount_delta") or 0),
        )
        for row in rows
    ]


def build_employee_receivable_changes(
    current_items: list[EmployeeReceivableItem],
    previous_items: list[EmployeeReceivableItem],
) -> list[EmployeeReceivableChange]:
    previous_by_ref = {item.counterparty_ref: item for item in previous_items}
    changes: list[EmployeeReceivableChange] = []

    for item in current_items:
        previous = previous_by_ref.pop(item.counterparty_ref, None)
        previous_balance = previous.current_balance if previous is not None else Decimal("0.00")
        delta_balance = item.current_balance - previous_balance
        if previous is None:
            status = "new"
        elif delta_balance > 0:
            status = "increased"
        elif delta_balance < 0:
            status = "decreased"
        else:
            status = "unchanged"
        changes.append(
            EmployeeReceivableChange(
                status=status,
                counterparty_ref=item.counterparty_ref,
                counterparty_name=item.counterparty_name,
                current_balance=item.current_balance,
                previous_balance=previous_balance,
                delta_balance=delta_balance,
                aged_bucket=item.aged_bucket,
                activity_segment=item.activity_segment,
                current_manager_ref=item.current_manager_ref,
                current_manager_name=item.current_manager_name,
                origin_document_ref=item.origin_document_ref,
                origin_document_number=item.origin_document_number,
                counterparty_code=(
                    item.counterparty_code or previous.counterparty_code
                    if previous is not None
                    else item.counterparty_code
                ),
            )
        )

    for previous in previous_by_ref.values():
        changes.append(
            EmployeeReceivableChange(
                status="closed",
                counterparty_ref=previous.counterparty_ref,
                counterparty_name=previous.counterparty_name,
                current_balance=Decimal("0.00"),
                previous_balance=previous.current_balance,
                delta_balance=-previous.current_balance,
                aged_bucket=previous.aged_bucket,
                activity_segment=previous.activity_segment,
                current_manager_ref=previous.current_manager_ref,
                current_manager_name=previous.current_manager_name,
                origin_document_ref=previous.origin_document_ref,
                origin_document_number=previous.origin_document_number,
                counterparty_code=previous.counterparty_code,
            )
        )

    return sorted(
        changes,
        key=lambda item: (
            abs(item.delta_balance),
            item.current_balance,
            item.counterparty_name or item.counterparty_ref,
        ),
        reverse=True,
    )


def _status_label(status: str) -> str:
    mapping = {
        "new": "Новый",
        "increased": "Рост",
        "decreased": "Снижение",
        "unchanged": "Без изменений",
        "closed": "Закрыт",
    }
    return mapping.get(status, status)


def _make_sheet(
    workbook: Workbook,
    *,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    amount_columns: set[int],
) -> Any:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    max_row = sheet.max_row
    max_column = sheet.max_column

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(2, max_row + 1):
        striped = row_index % 2 == 0
        for column_index in range(1, max_column + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if striped:
                cell.fill = STRIPE_FILL
            if column_index in amount_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="top")

    if max_row > 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"

    sheet.sheet_view.zoomScale = 90
    sheet.row_dimensions[1].height = 24

    for column_index in range(1, max_column + 1):
        width = 10
        for row_index in range(1, max_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            rendered = f"{value:,.2f}" if isinstance(value, float) else str(value)
            width = max(width, len(rendered) + 2)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(width, 42)

    return sheet


def _add_table_conditional_formatting(
    sheet: Any, *, status_column: str, amount_columns: list[str]
) -> None:
    if sheet.max_row < 2:
        return

    for column_letter in amount_columns:
        sheet.conditional_formatting.add(
            f"{column_letter}2:{column_letter}{sheet.max_row}",
            DataBarRule(
                start_type="min",
                start_value=0,
                end_type="max",
                end_value=0,
                color="5B9BD5",
                showValue=True,
            ),
        )

    if "E" in amount_columns or sheet.max_column >= 5:
        sheet.conditional_formatting.add(
            f"E2:E{sheet.max_row}",
            IconSetRule("3Arrows", "num", [-1, 0, 1], showValue=True),
        )

    for status_label, fill in STATUS_FILLS.items():
        sheet.conditional_formatting.add(
            f"A2:{get_column_letter(sheet.max_column)}{sheet.max_row}",
            FormulaRule(
                formula=[f'${status_column}2="{status_label}"'],
                stopIfTrue=False,
                fill=fill,
            ),
        )


def _write_chart_block(
    sheet: Any,
    *,
    start_row: int,
    start_col: int,
    headers: list[str],
    rows: list[list[Any]],
) -> tuple[int, int]:
    for col_offset, value in enumerate(headers, start=start_col):
        cell = sheet.cell(row=start_row, column=col_offset, value=value)
        cell.font = Font(bold=True)
    for row_offset, row in enumerate(rows, start=1):
        for col_offset, value in enumerate(row, start=start_col):
            sheet.cell(row=start_row + row_offset, column=col_offset, value=value)
    return start_row + 1, start_row + max(len(rows), 1)


def _add_chart(
    dashboard: Any,
    chart_data_sheet: Any,
    *,
    title: str,
    anchor: str,
    category_col: int,
    value_col: int,
    start_row: int,
    end_row: int,
    chart_type: str = "bar",
) -> None:
    if end_row < start_row:
        return

    if chart_type == "line":
        chart = LineChart()
        chart.style = 10
        chart.height = 6
        chart.width = 11
    else:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.height = 6
        chart.width = 11
    chart.title = title
    chart.y_axis.title = ""
    chart.legend = None
    data = Reference(chart_data_sheet, min_col=value_col, min_row=start_row - 1, max_row=end_row)
    categories = Reference(
        chart_data_sheet,
        min_col=category_col,
        min_row=start_row,
        max_row=end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    dashboard.add_chart(chart, anchor)


def _set_card(sheet: Any, *, cell: str, title: str, value: str, fill: str) -> None:
    sheet[cell] = title
    sheet[cell].font = Font(size=10, bold=True, color="FFFFFF")
    sheet[cell].fill = PatternFill(fill_type="solid", fgColor=fill)
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet[cell].border = BORDER
    value_cell = sheet.cell(row=sheet[cell].row + 1, column=sheet[cell].column, value=value)
    value_cell.font = Font(size=14, bold=True)
    value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    value_cell.border = BORDER


def _build_dashboard(
    workbook: Workbook,
    *,
    snapshot_date: date,
    previous_date: date | None,
    current_items: list[EmployeeReceivableItem],
    changes: list[EmployeeReceivableChange],
    snapshot_history: list[tuple[date, Decimal]],
    related_documents: list[EmployeeReceivableRelatedDocument] | None = None,
) -> None:
    dashboard = workbook.create_sheet(title=CONTROL_SHEET_TITLE[:31], index=0)
    chart_data_sheet = workbook.create_sheet(title=CHART_DATA_SHEET_TITLE)
    chart_data_sheet.sheet_state = "hidden"

    current_total = sum((item.current_balance for item in current_items), start=Decimal("0.00"))
    previous_total = sum((item.previous_balance for item in changes), start=Decimal("0.00"))
    counts = defaultdict(int)
    for item in changes:
        counts[_status_label(item.status)] += 1

    dashboard["A1"] = "Контроль долгов сотрудников"
    dashboard["A1"].font = Font(size=16, bold=True)
    dashboard["A2"] = (
        f"Срез: {_format_display_date(snapshot_date)}"
        f" | Сравнение: {_format_display_date(previous_date) if previous_date else 'нет'}"
    )

    _set_card(
        dashboard,
        cell="A4",
        title="Текущий долг",
        value=f"{_format_money(current_total)} ₽",
        fill="1F4E78",
    )
    _set_card(
        dashboard,
        cell="C4",
        title="Кейсов",
        value=str(len(current_items)),
        fill="2E75B6",
    )
    _set_card(
        dashboard,
        cell="E4",
        title="Рост к прошлому срезу",
        value=f"{_format_money(current_total - previous_total)} ₽",
        fill="C55A11",
    )
    _set_card(
        dashboard,
        cell="G4",
        title="Новые / закрытые",
        value=f"{counts['Новый']} / {counts['Закрыт']}",
        fill="548235",
    )

    links = [
        (SUMMARY_SHEET_TITLE, "Сводка"),
        (CURRENT_SHEET_TITLE, "Текущие долги"),
        (CHANGES_SHEET_TITLE, "Изменения по долгам"),
        (RELATED_DOCS_SHEET_TITLE, "Реализации и списания"),
    ]
    for index, (sheet_name, label) in enumerate(links, start=1):
        cell = dashboard.cell(row=8, column=index, value=label)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.style = "Hyperlink"

    top_current = sorted(current_items, key=lambda item: item.current_balance, reverse=True)[:5]
    top_growth = [
        item for item in changes if item.status in {"new", "increased"} and item.delta_balance > 0
    ][:5]
    by_manager: dict[str, Decimal] = defaultdict(lambda: Decimal("0.00"))
    for item in current_items:
        manager_label = item.current_manager_name or item.current_manager_ref or "Не назначен"
        by_manager[manager_label] += item.current_balance
    top_managers = sorted(by_manager.items(), key=lambda item: item[1], reverse=True)[:5]

    aged_totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0.00"))
    for item in current_items:
        aged_totals[item.aged_bucket or "unknown"] += item.current_balance
    aged_rows = sorted(aged_totals.items(), key=lambda item: item[1], reverse=True)

    dashboard["A10"] = "Крупнейшие долги"
    dashboard["E10"] = "Наибольший рост"
    dashboard["I10"] = "По ответственным"
    dashboard["M10"] = "По возрасту"
    for label_cell in ("A10", "E10", "I10", "M10"):
        dashboard[label_cell].font = Font(bold=True)

    for row_index, item in enumerate(top_current, start=11):
        dashboard.cell(
            row=row_index, column=1, value=item.counterparty_name or item.counterparty_ref
        )
        dashboard.cell(row=row_index, column=2, value=float(item.current_balance)).number_format = (
            "#,##0.00"
        )
        dashboard.cell(row=row_index, column=3, value=item.counterparty_code or "")

    for row_index, item in enumerate(top_growth, start=11):
        dashboard.cell(
            row=row_index, column=5, value=item.counterparty_name or item.counterparty_ref
        )
        dashboard.cell(row=row_index, column=6, value=float(item.delta_balance)).number_format = (
            "#,##0.00"
        )
        dashboard.cell(row=row_index, column=7, value=item.counterparty_code or "")

    for row_index, (manager_name, total_balance) in enumerate(top_managers, start=11):
        dashboard.cell(row=row_index, column=9, value=manager_name)
        dashboard.cell(row=row_index, column=10, value=float(total_balance)).number_format = (
            "#,##0.00"
        )

    for row_index, (aged_bucket, total_balance) in enumerate(aged_rows[:5], start=11):
        dashboard.cell(row=row_index, column=13, value=aged_bucket)
        dashboard.cell(row=row_index, column=14, value=float(total_balance)).number_format = (
            "#,##0.00"
        )

    history_start, history_end = _write_chart_block(
        chart_data_sheet,
        start_row=1,
        start_col=1,
        headers=["Дата", "Общий долг"],
        rows=[
            [_format_display_date(history_date), float(total)]
            for history_date, total in snapshot_history
        ]
        or [["Нет данных", 0.0]],
    )
    top_current_start, top_current_end = _write_chart_block(
        chart_data_sheet,
        start_row=12,
        start_col=1,
        headers=["Контрагент", "Текущий долг"],
        rows=[
            [item.counterparty_name or item.counterparty_ref, float(item.current_balance)]
            for item in top_current
        ]
        or [["Нет данных", 0.0]],
    )
    growth_start, growth_end = _write_chart_block(
        chart_data_sheet,
        start_row=23,
        start_col=1,
        headers=["Контрагент", "Рост долга"],
        rows=[
            [item.counterparty_name or item.counterparty_ref, float(item.delta_balance)]
            for item in top_growth
        ]
        or [["Нет данных", 0.0]],
    )
    status_start, status_end = _write_chart_block(
        chart_data_sheet,
        start_row=34,
        start_col=1,
        headers=["Статус", "Количество"],
        rows=[[status, count] for status, count in counts.items()] or [["Нет данных", 0]],
    )
    aged_start, aged_end = _write_chart_block(
        chart_data_sheet,
        start_row=45,
        start_col=1,
        headers=["Возраст", "Сумма"],
        rows=[[bucket, float(total)] for bucket, total in aged_rows[:5]] or [["Нет данных", 0.0]],
    )

    _add_chart(
        dashboard,
        chart_data_sheet,
        title="Динамика долга",
        anchor="A18",
        category_col=1,
        value_col=2,
        start_row=history_start,
        end_row=history_end,
        chart_type="line",
    )
    _add_chart(
        dashboard,
        chart_data_sheet,
        title="Крупнейшие долги",
        anchor="H18",
        category_col=1,
        value_col=2,
        start_row=top_current_start,
        end_row=top_current_end,
    )
    _add_chart(
        dashboard,
        chart_data_sheet,
        title="Рост по кейсам",
        anchor="A33",
        category_col=1,
        value_col=2,
        start_row=growth_start,
        end_row=growth_end,
    )
    _add_chart(
        dashboard,
        chart_data_sheet,
        title="Статусы изменений",
        anchor="H33",
        category_col=1,
        value_col=2,
        start_row=status_start,
        end_row=status_end,
    )
    _add_chart(
        dashboard,
        chart_data_sheet,
        title="Возраст долга",
        anchor="O18",
        category_col=1,
        value_col=2,
        start_row=aged_start,
        end_row=aged_end,
    )

    for column_index in range(1, 16):
        dashboard.column_dimensions[get_column_letter(column_index)].width = 18
    for cell in ("A1", "A2", "A10", "E10", "I10", "M10"):
        dashboard[cell].alignment = Alignment(vertical="center")


def export_employee_receivable_report(
    *,
    snapshot_date: date,
    previous_date: date | None,
    current_items: list[EmployeeReceivableItem],
    changes: list[EmployeeReceivableChange],
    output_path: Path,
    snapshot_history: list[tuple[date, Decimal]] | None = None,
    related_documents: list[EmployeeReceivableRelatedDocument] | None = None,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _build_dashboard(
        workbook,
        snapshot_date=snapshot_date,
        previous_date=previous_date,
        current_items=current_items,
        changes=changes,
        snapshot_history=snapshot_history or [],
        related_documents=related_documents or [],
    )

    summary_headers = ["Показатель", "Значение"]
    current_total = sum((item.current_balance for item in current_items), start=Decimal("0.00"))
    previous_total = sum((item.previous_balance for item in changes), start=Decimal("0.00"))
    summary_rows = [
        ["Дата среза", _format_display_date(snapshot_date)],
        [
            "Сравнение с",
            _format_display_date(previous_date) if previous_date else "нет предыдущего среза",
        ],
        ["Текущих сотрудников с долгом", len(current_items)],
        ["Текущий итог, ₽", float(current_total)],
        ["Предыдущий итог, ₽", float(previous_total)],
        ["Новые", sum(1 for item in changes if item.status == "new")],
        ["Рост", sum(1 for item in changes if item.status == "increased")],
        ["Снижение", sum(1 for item in changes if item.status == "decreased")],
        ["Без изменений", sum(1 for item in changes if item.status == "unchanged")],
        ["Закрытые", sum(1 for item in changes if item.status == "closed")],
    ]
    _make_sheet(
        workbook,
        title=SUMMARY_SHEET_TITLE,
        headers=summary_headers,
        rows=summary_rows,
        amount_columns={2},
    )

    current_headers = [
        "№",
        "Контрагент",
        "Текущий долг, ₽",
        "Было ранее, ₽",
        "Изменение, ₽",
        "Статус",
        "Возраст долга",
        "Активность",
        "Текущий ответственный",
        "Код 1С контрагента",
    ]
    change_by_ref = {item.counterparty_ref: item for item in changes}
    current_rows: list[list[Any]] = []
    for index, item in enumerate(
        sorted(current_items, key=lambda row: row.current_balance, reverse=True),
        start=1,
    ):
        change = change_by_ref[item.counterparty_ref]
        current_rows.append(
            [
                index,
                item.counterparty_name or item.counterparty_ref,
                float(item.current_balance),
                float(change.previous_balance),
                float(change.delta_balance),
                _status_label(change.status),
                item.aged_bucket,
                item.activity_segment,
                item.current_manager_name or item.current_manager_ref or "",
                item.counterparty_code or "",
            ]
        )
    current_sheet = _make_sheet(
        workbook,
        title=CURRENT_SHEET_TITLE,
        headers=current_headers,
        rows=current_rows,
        amount_columns={3, 4, 5},
    )

    change_headers = [
        "Статус",
        "Контрагент",
        "Текущий долг, ₽",
        "Было ранее, ₽",
        "Изменение, ₽",
        "Возраст долга",
        "Активность",
        "Текущий ответственный",
        "Код 1С контрагента",
    ]
    change_rows = [
        [
            _status_label(item.status),
            item.counterparty_name or item.counterparty_ref,
            float(item.current_balance),
            float(item.previous_balance),
            float(item.delta_balance),
            item.aged_bucket,
            item.activity_segment,
            item.current_manager_name or item.current_manager_ref or "",
            item.counterparty_code or "",
        ]
        for item in changes
        if item.status != "unchanged"
    ]
    changes_sheet = _make_sheet(
        workbook,
        title=CHANGES_SHEET_TITLE,
        headers=change_headers,
        rows=change_rows,
        amount_columns={3, 4, 5},
    )

    related_document_rows = [
        [
            index,
            item.document_kind,
            item.document_number or "",
            _format_datetime(item.document_date),
            item.counterparty_name or "",
            item.responsible_name or "",
            item.department_name or "",
            item.organization_name or "",
            item.item_name or "",
            float(item.quantity),
            float(item.amount_delta),
        ]
        for index, item in enumerate(related_documents or [], start=1)
    ]
    related_documents_sheet = _make_sheet(
        workbook,
        title=RELATED_DOCS_SHEET_TITLE,
        headers=[
            "№",
            "Документ",
            "Код документа",
            "Дата документа",
            "Контрагент",
            "Ответственный",
            "Подразделение",
            "Организация",
            "Номенклатура",
            "Количество",
            "Сумма, ₽",
        ],
        rows=related_document_rows,
        amount_columns={10},
    )

    _add_table_conditional_formatting(current_sheet, status_column="F", amount_columns=["C", "E"])
    _add_table_conditional_formatting(changes_sheet, status_column="A", amount_columns=["C", "E"])
    if related_documents_sheet.max_row >= 2:
        related_documents_sheet.conditional_formatting.add(
            f"J2:J{related_documents_sheet.max_row}",
            DataBarRule(
                start_type="min",
                start_value=0,
                end_type="max",
                end_value=0,
                color="70AD47",
                showValue=True,
            ),
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def build_telegram_message(
    *,
    snapshot_date: date,
    previous_date: date | None,
    current_items: list[EmployeeReceivableItem],
    changes: list[EmployeeReceivableChange],
) -> str:
    current_total = sum((item.current_balance for item in current_items), start=Decimal("0.00"))
    changed_count = sum(1 for item in changes if item.status != "unchanged")
    previous_label = _format_display_date(previous_date) if previous_date else "нет"
    return "\n".join(
        [
            "Отчет по долгам сотрудников",
            f"Дата среза: {_format_display_date(snapshot_date)}",
            f"Сравнение с: {previous_label}",
            f"Текущих кейсов: {len(current_items)}",
            f"Итог: {_format_money(current_total)} ₽",
            f"Изменений: {changed_count}",
        ]
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build employee receivable Excel report and optionally send it to Telegram"
    )
    parser.add_argument("--date", dest="snapshot_date", help="Snapshot date in YYYY-MM-DD")
    parser.add_argument(
        "--output",
        help="Output path for generated report. Defaults to reports/receivables/employee/<date>/...",
    )
    parser.add_argument(
        "--send-telegram",
        action="store_true",
        help="Send generated report to Telegram using explicit args or configured alert chat",
    )
    parser.add_argument("--telegram-token", help="Telegram bot token override")
    parser.add_argument("--telegram-chat-id", help="Telegram chat id override")
    parser.add_argument(
        "--env-file",
        default=".env",
        help="Path to .env for Telegram defaults (default: .env)",
    )
    args = parser.parse_args()

    settings = get_settings()
    engine = build_engine(settings.database_url)
    onec_engine = _build_onec_engine(settings)

    with Session(engine) as session:
        snapshot_date, previous_date = resolve_employee_snapshot_dates(
            session,
            requested_date=_parse_date(args.snapshot_date),
        )
        current_items = load_employee_items(session, snapshot_date=snapshot_date)
        previous_items = (
            load_employee_items(session, snapshot_date=previous_date) if previous_date else []
        )
        related_documents = load_employee_related_documents(
            onec_engine=onec_engine,
            snapshot_date=snapshot_date,
        )
        snapshot_history = load_employee_snapshot_history(
            session,
            snapshot_date=snapshot_date,
            limit=7,
        )

    enrich_employee_items_with_counterparty_codes(current_items, onec_engine=onec_engine)
    enrich_employee_items_with_counterparty_codes(previous_items, onec_engine=onec_engine)
    changes = build_employee_receivable_changes(current_items, previous_items)
    output_path = (
        Path(args.output)
        if args.output
        else build_output_path(snapshot_date=snapshot_date, output_dir=DEFAULT_OUTPUT_DIR)
    )
    export_employee_receivable_report(
        snapshot_date=snapshot_date,
        previous_date=previous_date,
        current_items=current_items,
        changes=changes,
        output_path=output_path,
        snapshot_history=snapshot_history,
        related_documents=related_documents,
    )

    print(f"report_path={output_path}")
    print(f"snapshot_date={snapshot_date.isoformat()}")
    print(f"previous_date={previous_date.isoformat() if previous_date else ''}")
    print(f"current_count={len(current_items)}")
    print(f"changed_count={sum(1 for item in changes if item.status != 'unchanged')}")

    if not args.send_telegram:
        return

    env = _load_env(Path(args.env_file))
    token = args.telegram_token or _resolve_default_telegram_target(env)[0]
    chat_id = args.telegram_chat_id or _resolve_default_telegram_target(env)[1]
    if not token or not chat_id:
        raise RuntimeError("Telegram token/chat_id are not configured")

    message = build_telegram_message(
        snapshot_date=snapshot_date,
        previous_date=previous_date,
        current_items=current_items,
        changes=changes,
    )
    send_return_scheme_telegram_report(
        token=token,
        chat_id=chat_id,
        message=message,
        report_path=output_path,
    )
    print("telegram_sent=true")


if __name__ == "__main__":
    main()
