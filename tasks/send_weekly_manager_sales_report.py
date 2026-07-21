from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any, Callable, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import DataBarRule, FormulaRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select, text
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.infrastructure.db.engines import build_engine
from app.models import OneCSalesDailyKpi
from app.services.return_scheme import send_return_scheme_telegram_report

try:
    from tasks.send_employee_receivable_report import (
        build_employee_receivable_changes,
        enrich_employee_items_with_counterparty_codes,
        export_employee_receivable_report,
        load_employee_items,
        load_employee_related_documents,
        load_employee_snapshot_history,
        resolve_employee_snapshot_dates,
    )
    from tasks.send_employee_receivable_report import (
        build_telegram_message as build_employee_telegram_message,
    )
except ModuleNotFoundError:  # pragma: no cover - script execution fallback
    from send_employee_receivable_report import (
        build_employee_receivable_changes,
        enrich_employee_items_with_counterparty_codes,
        export_employee_receivable_report,
        load_employee_items,
        load_employee_related_documents,
        load_employee_snapshot_history,
        resolve_employee_snapshot_dates,
    )
    from send_employee_receivable_report import (
        build_telegram_message as build_employee_telegram_message,
    )

DEFAULT_OUTPUT_DIR = Path("reports/sales/weekly")
ZERO_MONEY = Decimal("0.00")
ZERO_QTY = Decimal("0.000")
MONEY_QUANT = Decimal("0.01")
QTY_QUANT = Decimal("0.001")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
STRIPE_FILL = PatternFill(fill_type="solid", fgColor="F7FBFF")
SUMMARY_SHEET_TITLE = "Сводка"
DASHBOARD_SHEET_TITLE = "Дашборд"
MANAGERS_SHEET_TITLE = "Личные продажи"
ATTENTION_SHEET_TITLE = "Зона внимания"
STORE_DETAIL_SHEET_TITLE = "Продажи по магазинам"
CASH_ORDERS_SHEET_TITLE = "РКО излишек-недостача"
CHART_DATA_SHEET_TITLE = "_chart_data"
SHORTAGE_CASH_ORDERS_SQL = text("""
SELECT
    pko._Number AS doc_number,
    pko._Date_Time AS doc_date,
    cp._Description AS counterparty_name,
    RTRIM(cp._Code) AS counterparty_code,
    CAST(pko._Fld4688 AS decimal(18, 2)) AS amount,
    employee._Description AS employee_name,
    cashbox._Description AS cashbox_name,
    store._Description AS store_name,
    base_sale._Number AS base_sale_number
FROM dbo._Document196 AS pko
JOIN dbo._Reference54 AS cp
    ON cp._IDRRef = pko._Fld4684_RRRef
LEFT JOIN dbo._Reference69 AS employee
    ON employee._IDRRef = pko._Fld4695RRef
LEFT JOIN dbo._Reference45 AS cashbox
    ON cashbox._IDRRef = pko._Fld4681RRef
LEFT JOIN dbo._Reference68 AS store
    ON store._IDRRef = pko._Fld4682RRef
LEFT JOIN dbo._Document203 AS base_sale
    ON pko._Fld4697_RTRef = 0x000000CB
   AND base_sale._IDRRef = pko._Fld4697_RRRef
WHERE pko._Marked = 0x00
  AND pko._Posted = 0x01
  AND pko._Date_Time >= :date_from
  AND pko._Date_Time < :date_to
  AND (
      cp._Description LIKE N'%лишек%'
      OR cp._Description LIKE N'%недостач%'
  )
ORDER BY pko._Date_Time DESC, pko._Number DESC
""")


@dataclass(slots=True)
class WeeklySalesWindow:
    week_start: date
    week_end: date
    compare_week_start: date
    compare_week_end: date


@dataclass(slots=True)
class SalesKpiRecord:
    sales_date: date
    manager_ref: str | None
    manager_name: str | None
    manager_code: str | None
    store_ref: str | None
    store_name: str | None
    store_code: str | None
    revenue: Decimal
    sales_count: Decimal


@dataclass(slots=True)
class WeeklyManagerSalesItem:
    manager_ref: str | None
    manager_name: str | None
    manager_code: str | None
    store_names: tuple[str, ...]
    current_revenue: Decimal
    current_sales_count: Decimal
    current_avg_ticket: Decimal
    previous_revenue: Decimal
    previous_sales_count: Decimal
    previous_avg_ticket: Decimal
    revenue_delta: Decimal
    sales_count_delta: Decimal
    avg_ticket_delta: Decimal
    signal: str = "Норма"
    revenue_rank: int = 0


@dataclass(slots=True)
class WeeklyManagerStoreSalesItem:
    manager_ref: str | None
    manager_name: str | None
    manager_code: str | None
    store_ref: str | None
    store_name: str | None
    store_code: str | None
    current_revenue: Decimal
    current_sales_count: Decimal
    current_avg_ticket: Decimal
    previous_revenue: Decimal
    previous_sales_count: Decimal
    previous_avg_ticket: Decimal
    revenue_delta: Decimal
    sales_count_delta: Decimal
    avg_ticket_delta: Decimal
    signal: str = "Норма"


@dataclass(slots=True)
class ShortageCashOrderItem:
    document_number: str | None
    document_date: datetime | None
    counterparty_name: str | None
    counterparty_code: str | None
    amount: Decimal
    employee_name: str | None
    cashbox_name: str | None
    store_name: str | None
    base_sale_number: str | None


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    return date.fromisoformat(value)


def _load_env(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    env: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def _resolve_default_telegram_target(env: dict[str, str]) -> tuple[str | None, str | None]:
    token = (
        env.get("WEEKLY_MANAGER_SALES_REPORT_TELEGRAM_TOKEN")
        or env.get("WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_TOKEN")
        or env.get("SMARTPHONE_RELEASES_ALERT_TELEGRAM_TOKEN")
    )
    chat_id = (
        env.get("WEEKLY_MANAGER_SALES_REPORT_TELEGRAM_CHAT_ID")
        or env.get("WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_CHAT_ID")
        or env.get("SMARTPHONE_RELEASES_ALERT_TELEGRAM_CHAT_ID")
    )
    return token, chat_id


def _parse_telegram_chat_ids(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [chunk.strip() for chunk in raw.split(",") if chunk.strip()]


def _quantize_money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _quantize_qty(value: Decimal) -> Decimal:
    return value.quantize(QTY_QUANT, rounding=ROUND_HALF_UP)


def _avg_ticket(revenue: Decimal, sales_count: Decimal) -> Decimal:
    if sales_count == 0:
        return ZERO_MONEY
    return _quantize_money(revenue / sales_count)


def _format_money(value: Decimal) -> str:
    return f"{_quantize_money(value):,.2f}".replace(",", " ")


def _format_qty(value: Decimal) -> str:
    rendered = f"{_quantize_qty(value):,.3f}".replace(",", " ")
    if "." in rendered:
        rendered = rendered.rstrip("0").rstrip(".")
    return rendered


def _coerce_decimal(value: Any, *, default: Decimal = ZERO_MONEY) -> Decimal:
    if value is None:
        return default
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d.%m.%Y %H:%M")


def _format_display_date(value: date | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d.%m.%Y")


def _week_start(value: date) -> date:
    return value - timedelta(days=value.weekday())


def _last_completed_week_end(value: date) -> date:
    return value - timedelta(days=(value.weekday() + 1) % 7)


def _build_output_path(*, window: WeeklySalesWindow, output_dir: Path) -> Path:
    dated_dir = output_dir / window.week_end.isoformat()
    filename = (
        f"Личные продажи менеджеров "
        f"{_format_display_date(window.week_start)}-{_format_display_date(window.week_end)}.xlsx"
    )
    return dated_dir / filename


def _resolve_report_window(
    session: Session,
    *,
    requested_date: date | None,
) -> WeeklySalesWindow:
    latest_sales_date = session.execute(select(func.max(OneCSalesDailyKpi.sales_date))).scalar_one()
    if latest_sales_date is None:
        raise RuntimeError("No sales KPI rows found in onec_sales_daily_kpi")

    report_anchor = requested_date or _last_completed_week_end(latest_sales_date)
    current_week_start = _week_start(report_anchor)
    current_week_end = current_week_start + timedelta(days=6)
    compare_week_end = current_week_start - timedelta(days=1)
    compare_week_start = compare_week_end - timedelta(days=6)
    return WeeklySalesWindow(
        week_start=current_week_start,
        week_end=current_week_end,
        compare_week_start=compare_week_start,
        compare_week_end=compare_week_end,
    )


def _build_onec_engine(settings) -> Engine | None:
    if not settings.onec_database_url:
        return None
    return build_engine(
        settings.onec_database_url,
        connect_args={
            "timeout": float(settings.onec_query_timeout_seconds),
            "login_timeout": float(settings.onec_login_timeout_seconds),
        },
    )


def _build_in_clause(values: Sequence[str], *, prefix: str) -> tuple[str, dict[str, str]]:
    params = {f"{prefix}_{index}": value for index, value in enumerate(values)}
    placeholders = ", ".join(f":{name}" for name in params)
    return placeholders, params


def _ref_expr(engine: Engine, alias: str) -> str:
    if engine.dialect.name == "mssql":
        return f"master.dbo.fn_varbintohexstr({alias}._IDRRef)"
    return f"{alias}._IDRRef"


def _with_nolock(engine: Engine) -> str:
    return "WITH (NOLOCK)" if engine.dialect.name == "mssql" else ""


def _build_ref_filter_clause(
    engine: Engine,
    refs: Sequence[str],
    *,
    column_name: str,
    prefix: str,
) -> tuple[str, dict[str, str]]:
    if engine.dialect.name == "mssql":
        hex_refs = [value.upper() for value in refs if re.fullmatch(r"0x[0-9A-Fa-f]+", value)]
        if not hex_refs:
            return "1 = 0", {}
        return f"{column_name} IN ({', '.join(hex_refs)})", {}

    params = {f"{prefix}_{index}": value for index, value in enumerate(refs)}
    placeholders = ", ".join(f":{name}" for name in params)
    return f"{column_name} IN ({placeholders})", params


def fetch_reference_code_mapping_by_ref(
    onec_engine: Engine,
    *,
    refs: Sequence[str],
    table_name: str,
    alias: str = "ref_item",
) -> dict[str, str]:
    values = sorted({value for value in refs if value})
    if not values:
        return {}

    where_clause, params = _build_ref_filter_clause(
        onec_engine,
        values,
        column_name=f"{alias}._IDRRef",
        prefix="object_ref",
    )
    ref_expr = _ref_expr(onec_engine, alias)
    stmt = text(f"""
        SELECT
            {ref_expr} AS object_ref,
            RTRIM({alias}._Code) AS object_code
        FROM dbo.{table_name} AS {alias} {_with_nolock(onec_engine)}
        WHERE {where_clause}
    """)
    with onec_engine.connect() as connection:
        rows = connection.execute(stmt, params).mappings().all()
    return {
        str(row["object_ref"]).strip().upper(): str(row["object_code"]).strip()
        for row in rows
        if row.get("object_ref") and row.get("object_code")
    }


def enrich_sales_records_with_codes(
    records: list[SalesKpiRecord],
    *,
    onec_engine: Engine | None,
) -> None:
    if onec_engine is None or not records:
        return

    manager_codes = fetch_reference_code_mapping_by_ref(
        onec_engine,
        refs=[record.manager_ref for record in records],
        table_name="_Reference69",
        alias="manager_item",
    )
    store_codes = fetch_reference_code_mapping_by_ref(
        onec_engine,
        refs=[record.store_ref for record in records],
        table_name="_Reference80",
        alias="store_item",
    )
    for record in records:
        record.manager_code = manager_codes.get((record.manager_ref or "").upper())
        record.store_code = store_codes.get((record.store_ref or "").upper())


def load_weekly_sales_history(
    session: Session,
    *,
    week_end: date,
    limit: int = 4,
) -> list[tuple[date, Decimal, Decimal]]:
    week_start = week_end - timedelta(days=6)
    first_week_start = week_start - timedelta(days=7 * (limit - 1))
    rows = session.execute(
        select(
            OneCSalesDailyKpi.sales_date, OneCSalesDailyKpi.revenue, OneCSalesDailyKpi.sales_count
        )
        .where(
            OneCSalesDailyKpi.sales_date >= first_week_start,
            OneCSalesDailyKpi.sales_date <= week_end,
        )
        .order_by(OneCSalesDailyKpi.sales_date.asc())
    ).all()
    by_week: dict[date, dict[str, Decimal]] = {
        first_week_start
        + timedelta(days=7 * index): {
            "revenue": ZERO_MONEY,
            "sales_count": ZERO_QTY,
        }
        for index in range(limit)
    }
    for sales_date, revenue, sales_count in rows:
        bucket = sales_date - timedelta(days=sales_date.weekday())
        if bucket not in by_week:
            continue
        by_week[bucket]["revenue"] += Decimal(revenue)
        by_week[bucket]["sales_count"] += Decimal(sales_count)
    return [
        (bucket, _quantize_money(values["revenue"]), _quantize_qty(values["sales_count"]))
        for bucket, values in sorted(by_week.items())
    ]


def _load_sales_records(
    session: Session,
    *,
    date_from: date,
    date_to: date,
) -> list[SalesKpiRecord]:
    rows = (
        session.execute(
            select(OneCSalesDailyKpi)
            .where(
                OneCSalesDailyKpi.sales_date >= date_from,
                OneCSalesDailyKpi.sales_date <= date_to,
            )
            .order_by(
                OneCSalesDailyKpi.sales_date.asc(),
                OneCSalesDailyKpi.manager_name.asc(),
                OneCSalesDailyKpi.store_name.asc(),
            )
        )
        .scalars()
        .all()
    )
    return [
        SalesKpiRecord(
            sales_date=row.sales_date,
            manager_ref=row.manager_ref,
            manager_name=row.manager_name,
            manager_code=None,
            store_ref=row.store_ref,
            store_name=row.store_name,
            store_code=None,
            revenue=Decimal(row.revenue),
            sales_count=Decimal(row.sales_count),
        )
        for row in rows
    ]


def fetch_onec_shortage_cash_orders(
    onec_engine: Engine,
    *,
    date_from: date,
    date_to: date,
) -> list[ShortageCashOrderItem]:
    window_start = datetime.combine(date_from, time.min)
    window_end = datetime.combine(date_to + timedelta(days=1), time.min)
    with onec_engine.connect() as connection:
        rows = (
            connection.execute(
                SHORTAGE_CASH_ORDERS_SQL,
                {
                    "date_from": window_start,
                    "date_to": window_end,
                },
            )
            .mappings()
            .all()
        )

    items: list[ShortageCashOrderItem] = []
    for row in rows:
        document_date = row["doc_date"]
        if isinstance(document_date, date) and not isinstance(document_date, datetime):
            document_date = datetime.combine(document_date, time.min)
        items.append(
            ShortageCashOrderItem(
                document_number=row["doc_number"],
                document_date=document_date,
                counterparty_name=row["counterparty_name"],
                counterparty_code=row["counterparty_code"],
                amount=_quantize_money(_coerce_decimal(row["amount"])),
                employee_name=row["employee_name"],
                cashbox_name=row["cashbox_name"],
                store_name=row["store_name"],
                base_sale_number=row["base_sale_number"],
            )
        )
    return items


def _manager_sort_name(manager_name: str | None, manager_ref: str | None) -> str:
    return (manager_name or manager_ref or "Не назначен").strip()


def _manager_key(record: SalesKpiRecord) -> tuple[str | None, str | None]:
    return record.manager_ref, record.manager_name


def _manager_store_key(
    record: SalesKpiRecord,
) -> tuple[str | None, str | None, str | None, str | None]:
    return record.manager_ref, record.manager_name, record.store_ref, record.store_name


def _aggregate_records(
    records: list[SalesKpiRecord],
    *,
    key_fn: Callable[[SalesKpiRecord], tuple[Any, ...]],
) -> dict[tuple[Any, ...], dict[str, Any]]:
    aggregated: dict[tuple[Any, ...], dict[str, Any]] = {}
    for record in records:
        key = key_fn(record)
        item = aggregated.setdefault(
            key,
            {
                "manager_ref": record.manager_ref,
                "manager_name": record.manager_name,
                "manager_code": record.manager_code,
                "store_ref": record.store_ref,
                "store_name": record.store_name,
                "store_code": record.store_code,
                "store_names": set(),
                "revenue": ZERO_MONEY,
                "sales_count": ZERO_QTY,
            },
        )
        if record.store_name:
            item["store_names"].add(record.store_name)
        elif record.store_ref:
            item["store_names"].add(record.store_ref)
        item["revenue"] = item["revenue"] + record.revenue
        item["sales_count"] = item["sales_count"] + record.sales_count
    return aggregated


def _classify_manager_items(items: list[WeeklyManagerSalesItem]) -> None:
    if not items:
        return
    ranked_desc = sorted(
        items,
        key=lambda item: (
            item.current_revenue,
            item.current_sales_count,
            _manager_sort_name(item.manager_name, item.manager_ref),
        ),
        reverse=True,
    )
    for index, item in enumerate(ranked_desc, start=1):
        item.revenue_rank = index

    positive_volume_items = [
        item for item in reversed(ranked_desc) if item.current_revenue > ZERO_MONEY
    ]
    low_volume_count = (
        max(1, len(positive_volume_items) // 4) if len(positive_volume_items) > 1 else 0
    )
    low_volume_keys = {
        (item.manager_ref, item.manager_name) for item in positive_volume_items[:low_volume_count]
    }
    for item in items:
        item_key = (item.manager_ref, item.manager_name)
        if item.current_revenue <= ZERO_MONEY or item.current_sales_count <= ZERO_QTY:
            item.signal = "Нет продаж"
        elif item.revenue_delta < ZERO_MONEY and item.sales_count_delta < ZERO_QTY:
            item.signal = "Просадка"
        elif item_key in low_volume_keys:
            item.signal = "Низкий объем"
        elif item.revenue_delta < ZERO_MONEY:
            item.signal = "Снижение выручки"
        else:
            item.signal = "Норма"


def _classify_store_items(items: list[WeeklyManagerStoreSalesItem]) -> None:
    for item in items:
        if item.current_revenue <= ZERO_MONEY or item.current_sales_count <= ZERO_QTY:
            item.signal = "Нет продаж"
        elif item.revenue_delta < ZERO_MONEY and item.sales_count_delta < ZERO_QTY:
            item.signal = "Просадка"
        elif item.revenue_delta < ZERO_MONEY:
            item.signal = "Снижение выручки"
        else:
            item.signal = "Норма"


def build_weekly_manager_sales_items(
    current_records: list[SalesKpiRecord],
    previous_records: list[SalesKpiRecord],
) -> list[WeeklyManagerSalesItem]:
    current_map = _aggregate_records(current_records, key_fn=_manager_key)
    previous_map = _aggregate_records(previous_records, key_fn=_manager_key)
    keys = set(current_map) | set(previous_map)
    items: list[WeeklyManagerSalesItem] = []

    for key in keys:
        current = current_map.get(key)
        previous = previous_map.get(key)
        current_revenue = _quantize_money(current["revenue"]) if current else ZERO_MONEY
        current_sales_count = _quantize_qty(current["sales_count"]) if current else ZERO_QTY
        previous_revenue = _quantize_money(previous["revenue"]) if previous else ZERO_MONEY
        previous_sales_count = _quantize_qty(previous["sales_count"]) if previous else ZERO_QTY
        store_names = tuple(
            sorted(
                (current.get("store_names") if current else set())
                | (previous.get("store_names") if previous else set())
            )
        )
        items.append(
            WeeklyManagerSalesItem(
                manager_ref=(current or previous).get("manager_ref"),
                manager_name=(current or previous).get("manager_name"),
                manager_code=(current or previous).get("manager_code"),
                store_names=store_names,
                current_revenue=current_revenue,
                current_sales_count=current_sales_count,
                current_avg_ticket=_avg_ticket(current_revenue, current_sales_count),
                previous_revenue=previous_revenue,
                previous_sales_count=previous_sales_count,
                previous_avg_ticket=_avg_ticket(previous_revenue, previous_sales_count),
                revenue_delta=_quantize_money(current_revenue - previous_revenue),
                sales_count_delta=_quantize_qty(current_sales_count - previous_sales_count),
                avg_ticket_delta=_quantize_money(
                    _avg_ticket(current_revenue, current_sales_count)
                    - _avg_ticket(previous_revenue, previous_sales_count)
                ),
            )
        )

    _classify_manager_items(items)
    return sorted(
        items,
        key=lambda item: (
            item.current_revenue,
            item.current_sales_count,
            _manager_sort_name(item.manager_name, item.manager_ref),
        ),
        reverse=True,
    )


def build_weekly_manager_store_sales_items(
    current_records: list[SalesKpiRecord],
    previous_records: list[SalesKpiRecord],
) -> list[WeeklyManagerStoreSalesItem]:
    current_map = _aggregate_records(current_records, key_fn=_manager_store_key)
    previous_map = _aggregate_records(previous_records, key_fn=_manager_store_key)
    keys = set(current_map) | set(previous_map)
    items: list[WeeklyManagerStoreSalesItem] = []

    for key in keys:
        current = current_map.get(key)
        previous = previous_map.get(key)
        current_revenue = _quantize_money(current["revenue"]) if current else ZERO_MONEY
        current_sales_count = _quantize_qty(current["sales_count"]) if current else ZERO_QTY
        previous_revenue = _quantize_money(previous["revenue"]) if previous else ZERO_MONEY
        previous_sales_count = _quantize_qty(previous["sales_count"]) if previous else ZERO_QTY
        items.append(
            WeeklyManagerStoreSalesItem(
                manager_ref=(current or previous).get("manager_ref"),
                manager_name=(current or previous).get("manager_name"),
                manager_code=(current or previous).get("manager_code"),
                store_ref=(current or previous).get("store_ref"),
                store_name=(current or previous).get("store_name"),
                store_code=(current or previous).get("store_code"),
                current_revenue=current_revenue,
                current_sales_count=current_sales_count,
                current_avg_ticket=_avg_ticket(current_revenue, current_sales_count),
                previous_revenue=previous_revenue,
                previous_sales_count=previous_sales_count,
                previous_avg_ticket=_avg_ticket(previous_revenue, previous_sales_count),
                revenue_delta=_quantize_money(current_revenue - previous_revenue),
                sales_count_delta=_quantize_qty(current_sales_count - previous_sales_count),
                avg_ticket_delta=_quantize_money(
                    _avg_ticket(current_revenue, current_sales_count)
                    - _avg_ticket(previous_revenue, previous_sales_count)
                ),
            )
        )

    _classify_store_items(items)
    return sorted(
        items,
        key=lambda item: (
            item.current_revenue,
            item.current_sales_count,
            _manager_sort_name(item.manager_name, item.manager_ref),
            item.store_name or item.store_ref or "",
        ),
        reverse=True,
    )


def _signal_priority(value: str) -> int:
    priorities = {
        "Нет продаж": 0,
        "Просадка": 1,
        "Низкий объем": 2,
        "Снижение выручки": 3,
        "Норма": 4,
    }
    return priorities.get(value, 99)


def build_attention_manager_sales_items(
    items: list[WeeklyManagerSalesItem],
) -> list[WeeklyManagerSalesItem]:
    attention_items = [item for item in items if item.signal != "Норма"]
    return sorted(
        attention_items,
        key=lambda item: (
            _signal_priority(item.signal),
            item.current_revenue,
            item.revenue_delta,
            _manager_sort_name(item.manager_name, item.manager_ref),
        ),
    )


def _make_sheet(
    workbook: Workbook,
    *,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    money_columns: set[int] | None = None,
    qty_columns: set[int] | None = None,
) -> Any:
    money_columns = money_columns or set()
    qty_columns = qty_columns or set()

    sheet = workbook.create_sheet(title=title[:31])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    max_row = sheet.max_row
    max_column = sheet.max_column
    negative_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    positive_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    warning_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    critical_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    header_font = Font(bold=True, color="FFFFFF")
    border_side = Side(style="thin", color="D9E2F3")
    border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = HEADER_FILL
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(2, max_row + 1):
        striped = row_index % 2 == 0
        for column_index in range(1, max_column + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            header_value = str(sheet.cell(row=1, column=column_index).value or "")
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if striped:
                cell.fill = STRIPE_FILL
            if column_index in money_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif column_index in qty_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.000"
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif header_value in {"№", "Место по выручке"}:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

            if header_value == "Сигнал":
                if cell.value == "Нет продаж":
                    cell.fill = critical_fill
                elif cell.value in {"Просадка", "Низкий объем", "Снижение выручки"}:
                    cell.fill = warning_fill

            if header_value.startswith("Дельта ") and isinstance(cell.value, (int, float)):
                if cell.value < 0:
                    cell.fill = negative_fill
                elif cell.value > 0:
                    cell.fill = positive_fill

    if max_row >= 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"

    sheet.sheet_view.zoomScale = 90
    sheet.row_dimensions[1].height = 24

    for column_index in range(1, max_column + 1):
        width = 10
        for row_index in range(1, max_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            if isinstance(value, float):
                rendered = f"{value:,.2f}"
            else:
                rendered = str(value)
            width = max(width, len(rendered) + 2)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(width, 42)

    return sheet


def _add_weekly_sheet_conditional_formatting(
    sheet: Any,
    *,
    signal_column: str,
    revenue_column: str,
    delta_columns: list[str],
) -> None:
    if sheet.max_row < 2:
        return

    sheet.conditional_formatting.add(
        f"{revenue_column}2:{revenue_column}{sheet.max_row}",
        DataBarRule(
            start_type="min",
            start_value=0,
            end_type="max",
            end_value=0,
            color="5B9BD5",
            showValue=True,
        ),
    )
    for column_letter in delta_columns:
        sheet.conditional_formatting.add(
            f"{column_letter}2:{column_letter}{sheet.max_row}",
            IconSetRule("3Arrows", "num", [-1, 0, 1], showValue=True),
        )

    signal_fills = {
        "Нет продаж": PatternFill(fill_type="solid", fgColor="F4CCCC"),
        "Просадка": PatternFill(fill_type="solid", fgColor="FCE4D6"),
        "Низкий объем": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        "Снижение выручки": PatternFill(fill_type="solid", fgColor="FFE699"),
    }
    for signal_value, fill in signal_fills.items():
        sheet.conditional_formatting.add(
            f"A2:{get_column_letter(sheet.max_column)}{sheet.max_row}",
            FormulaRule(
                formula=[f'${signal_column}2="{signal_value}"'],
                stopIfTrue=False,
                fill=fill,
            ),
        )


def _write_chart_block(
    sheet: Any,
    *,
    start_row: int,
    start_col: int,
    headers: list[str],
    rows: list[list[Any]],
) -> tuple[int, int]:
    for col_offset, value in enumerate(headers, start=start_col):
        cell = sheet.cell(row=start_row, column=col_offset, value=value)
        cell.font = Font(bold=True)
    for row_offset, row in enumerate(rows, start=1):
        for col_offset, value in enumerate(row, start=start_col):
            sheet.cell(row=start_row + row_offset, column=col_offset, value=value)
    return start_row + 1, start_row + max(len(rows), 1)


def _add_dashboard_chart(
    dashboard: Any,
    chart_data_sheet: Any,
    *,
    title: str,
    anchor: str,
    category_col: int,
    value_col: int,
    start_row: int,
    end_row: int,
    chart_type: str = "bar",
) -> None:
    if end_row < start_row:
        return
    if chart_type == "line":
        chart = LineChart()
        chart.style = 10
        chart.height = 6
        chart.width = 11
    else:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.height = 6
        chart.width = 11
    chart.title = title
    chart.legend = None
    data = Reference(chart_data_sheet, min_col=value_col, min_row=start_row - 1, max_row=end_row)
    categories = Reference(
        chart_data_sheet, min_col=category_col, min_row=start_row, max_row=end_row
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    dashboard.add_chart(chart, anchor)


def _set_dashboard_card(sheet: Any, *, cell: str, title: str, value: str, fill: str) -> None:
    sheet[cell] = title
    sheet[cell].font = Font(size=10, bold=True, color="FFFFFF")
    sheet[cell].fill = PatternFill(fill_type="solid", fgColor=fill)
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet[cell].border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    value_cell = sheet.cell(row=sheet[cell].row + 1, column=sheet[cell].column, value=value)
    value_cell.font = Font(size=14, bold=True)
    value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    value_cell.border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )


def _build_weekly_dashboard(
    workbook: Workbook,
    *,
    window: WeeklySalesWindow,
    manager_items: list[WeeklyManagerSalesItem],
    attention_items: list[WeeklyManagerSalesItem],
    manager_store_items: list[WeeklyManagerStoreSalesItem],
    cash_order_items: list[ShortageCashOrderItem],
    weekly_history: list[tuple[date, Decimal, Decimal]],
) -> None:
    dashboard = workbook.create_sheet(title=DASHBOARD_SHEET_TITLE[:31], index=0)
    chart_data_sheet = workbook.create_sheet(title=CHART_DATA_SHEET_TITLE)
    chart_data_sheet.sheet_state = "hidden"

    total_current_revenue = sum((item.current_revenue for item in manager_items), start=ZERO_MONEY)
    total_previous_revenue = sum(
        (item.previous_revenue for item in manager_items), start=ZERO_MONEY
    )
    total_current_sales = sum((item.current_sales_count for item in manager_items), start=ZERO_QTY)
    total_cash_orders = sum((item.amount for item in cash_order_items), start=ZERO_MONEY)

    dashboard["A1"] = "Управленческий weekly по личным продажам"
    dashboard["A1"].font = Font(size=16, bold=True)
    dashboard["A2"] = (
        f"Период: {_format_display_date(window.week_start)} - {_format_display_date(window.week_end)}"
        f" | Сравнение: {_format_display_date(window.compare_week_start)} - "
        f"{_format_display_date(window.compare_week_end)}"
    )

    _set_dashboard_card(
        dashboard,
        cell="A4",
        title="Выручка недели",
        value=f"{_format_money(total_current_revenue)} ₽",
        fill="1F4E78",
    )
    _set_dashboard_card(
        dashboard,
        cell="C4",
        title="Дельта к прошлой неделе",
        value=f"{_format_money(total_current_revenue - total_previous_revenue)} ₽",
        fill="C55A11",
    )
    _set_dashboard_card(
        dashboard,
        cell="E4",
        title="Продажи, шт",
        value=_format_qty(total_current_sales),
        fill="2E75B6",
    )
    _set_dashboard_card(
        dashboard,
        cell="G4",
        title="Менеджеры в зоне внимания",
        value=str(len(attention_items)),
        fill="BF9000",
    )
    _set_dashboard_card(
        dashboard,
        cell="I4",
        title="РКО излишек/недостача",
        value=f"{len(cash_order_items)} / {_format_money(total_cash_orders)} ₽",
        fill="548235",
    )

    for index, (sheet_name, label) in enumerate(
        [
            (SUMMARY_SHEET_TITLE, "Сводка"),
            (MANAGERS_SHEET_TITLE, "Личные продажи"),
            (ATTENTION_SHEET_TITLE, "Зона внимания"),
            (STORE_DETAIL_SHEET_TITLE, "Продажи по магазинам"),
            (CASH_ORDERS_SHEET_TITLE, "РКО"),
        ],
        start=1,
    ):
        cell = dashboard.cell(row=8, column=index, value=label)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.style = "Hyperlink"

    store_totals: dict[tuple[str, str], dict[str, Decimal]] = defaultdict(
        lambda: {"revenue_delta": ZERO_MONEY, "current_revenue": ZERO_MONEY}
    )
    for item in manager_store_items:
        store_key = (item.store_name or item.store_ref or "Не указан", item.store_code or "")
        store_totals[store_key]["revenue_delta"] += item.revenue_delta
        store_totals[store_key]["current_revenue"] += item.current_revenue
    problem_stores = sorted(
        [
            (name, code, values["revenue_delta"], values["current_revenue"])
            for (name, code), values in store_totals.items()
        ],
        key=lambda item: (item[2], item[3]),
    )[:5]

    rko_by_employee: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"count": Decimal("0"), "amount": ZERO_MONEY}
    )
    rko_by_store: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"count": Decimal("0"), "amount": ZERO_MONEY}
    )
    for item in cash_order_items:
        employee_label = item.employee_name or "Не указан"
        store_label = item.store_name or "Не указан"
        rko_by_employee[employee_label]["count"] += Decimal("1")
        rko_by_employee[employee_label]["amount"] += item.amount
        rko_by_store[store_label]["count"] += Decimal("1")
        rko_by_store[store_label]["amount"] += item.amount
    top_rko_employees = sorted(
        rko_by_employee.items(),
        key=lambda row: (row[1]["count"], row[1]["amount"]),
        reverse=True,
    )[:5]
    top_rko_stores = sorted(
        rko_by_store.items(),
        key=lambda row: (row[1]["count"], row[1]["amount"]),
        reverse=True,
    )[:5]

    dashboard["A10"] = "Проблемные менеджеры"
    dashboard["F10"] = "Проблемные магазины"
    dashboard["K10"] = "РКО-контроль"
    for cell_name in ("A10", "F10", "K10"):
        dashboard[cell_name].font = Font(bold=True)

    for row_index, item in enumerate(attention_items[:5], start=11):
        dashboard.cell(
            row=row_index, column=1, value=_manager_sort_name(item.manager_name, item.manager_ref)
        )
        dashboard.cell(row=row_index, column=2, value=item.signal)
        dashboard.cell(row=row_index, column=3, value=float(item.revenue_delta)).number_format = (
            "#,##0.00"
        )
        dashboard.cell(row=row_index, column=4, value=item.manager_code or "")

    for row_index, (store_name, store_code, delta_value, current_value) in enumerate(
        problem_stores, start=11
    ):
        dashboard.cell(row=row_index, column=6, value=store_name)
        dashboard.cell(row=row_index, column=7, value=float(delta_value)).number_format = "#,##0.00"
        dashboard.cell(row=row_index, column=8, value=float(current_value)).number_format = (
            "#,##0.00"
        )
        dashboard.cell(row=row_index, column=9, value=store_code)

    for row_index, (employee_name, metrics) in enumerate(top_rko_employees, start=11):
        dashboard.cell(row=row_index, column=11, value=employee_name)
        dashboard.cell(row=row_index, column=12, value=int(metrics["count"]))
        dashboard.cell(row=row_index, column=13, value=float(metrics["amount"])).number_format = (
            "#,##0.00"
        )

    history_start, history_end = _write_chart_block(
        chart_data_sheet,
        start_row=1,
        start_col=1,
        headers=["Неделя", "Выручка"],
        rows=[
            [
                f"{_format_display_date(week_start)}-{_format_display_date(week_start + timedelta(days=6))}",
                float(revenue),
            ]
            for week_start, revenue, _sales in weekly_history
        ]
        or [["Нет данных", 0.0]],
    )
    problem_manager_start, problem_manager_end = _write_chart_block(
        chart_data_sheet,
        start_row=12,
        start_col=1,
        headers=["Менеджер", "Дельта выручки"],
        rows=[
            [_manager_sort_name(item.manager_name, item.manager_ref), float(item.revenue_delta)]
            for item in attention_items[:5]
        ]
        or [["Нет данных", 0.0]],
    )
    problem_store_start, problem_store_end = _write_chart_block(
        chart_data_sheet,
        start_row=23,
        start_col=1,
        headers=["Магазин", "Дельта выручки"],
        rows=[
            [store_name, float(delta_value)]
            for store_name, _code, delta_value, _current in problem_stores
        ]
        or [["Нет данных", 0.0]],
    )
    rko_employee_start, rko_employee_end = _write_chart_block(
        chart_data_sheet,
        start_row=34,
        start_col=1,
        headers=["Кто оформил", "Сумма РКО"],
        rows=[
            [employee_name, float(metrics["amount"])]
            for employee_name, metrics in top_rko_employees
        ]
        or [["Нет данных", 0.0]],
    )
    rko_store_start, rko_store_end = _write_chart_block(
        chart_data_sheet,
        start_row=45,
        start_col=1,
        headers=["Магазин", "Количество РКО"],
        rows=[[store_name, int(metrics["count"])] for store_name, metrics in top_rko_stores]
        or [["Нет данных", 0]],
    )

    _add_dashboard_chart(
        dashboard,
        chart_data_sheet,
        title="Тренд по выручке",
        anchor="A18",
        category_col=1,
        value_col=2,
        start_row=history_start,
        end_row=history_end,
        chart_type="line",
    )
    _add_dashboard_chart(
        dashboard,
        chart_data_sheet,
        title="Просадка по менеджерам",
        anchor="H18",
        category_col=1,
        value_col=2,
        start_row=problem_manager_start,
        end_row=problem_manager_end,
    )
    _add_dashboard_chart(
        dashboard,
        chart_data_sheet,
        title="Просадка по магазинам",
        anchor="O18",
        category_col=1,
        value_col=2,
        start_row=problem_store_start,
        end_row=problem_store_end,
    )
    _add_dashboard_chart(
        dashboard,
        chart_data_sheet,
        title="РКО по сотрудникам",
        anchor="A33",
        category_col=1,
        value_col=2,
        start_row=rko_employee_start,
        end_row=rko_employee_end,
    )
    _add_dashboard_chart(
        dashboard,
        chart_data_sheet,
        title="РКО по магазинам",
        anchor="H33",
        category_col=1,
        value_col=2,
        start_row=rko_store_start,
        end_row=rko_store_end,
    )

    for column_index in range(1, 15):
        dashboard.column_dimensions[get_column_letter(column_index)].width = 18


def export_weekly_manager_sales_report(
    *,
    window: WeeklySalesWindow,
    manager_items: list[WeeklyManagerSalesItem],
    attention_items: list[WeeklyManagerSalesItem],
    manager_store_items: list[WeeklyManagerStoreSalesItem],
    cash_order_items: list[ShortageCashOrderItem],
    output_path: Path,
    weekly_history: list[tuple[date, Decimal, Decimal]] | None = None,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _build_weekly_dashboard(
        workbook,
        window=window,
        manager_items=manager_items,
        attention_items=attention_items,
        manager_store_items=manager_store_items,
        cash_order_items=cash_order_items,
        weekly_history=weekly_history or [],
    )

    total_current_revenue = sum((item.current_revenue for item in manager_items), start=ZERO_MONEY)
    total_current_sales = sum((item.current_sales_count for item in manager_items), start=ZERO_QTY)
    total_previous_revenue = sum(
        (item.previous_revenue for item in manager_items), start=ZERO_MONEY
    )
    total_previous_sales = sum(
        (item.previous_sales_count for item in manager_items), start=ZERO_QTY
    )
    total_cash_order_amount = sum((item.amount for item in cash_order_items), start=ZERO_MONEY)

    leader = manager_items[0] if manager_items else None
    outsider = manager_items[-1] if manager_items else None
    deepest_drop = min(
        manager_items,
        key=lambda item: (item.revenue_delta, item.current_revenue),
        default=None,
    )

    summary_rows: list[list[Any]] = [
        [
            "Период отчета",
            f"{_format_display_date(window.week_start)} - {_format_display_date(window.week_end)}",
        ],
        [
            "Период сравнения",
            (
                f"{_format_display_date(window.compare_week_start)} - "
                f"{_format_display_date(window.compare_week_end)}"
            ),
        ],
        ["Менеджеров в отчете", len(manager_items)],
        ["Менеджеров в зоне внимания", len(attention_items)],
        ["РКО Излишек/недостача, шт", len(cash_order_items)],
        ["РКО Излишек/недостача, сумма, ₽", float(total_cash_order_amount)],
        ["Выручка недели, ₽", float(total_current_revenue)],
        ["Выручка прошлой недели, ₽", float(total_previous_revenue)],
        ["Дельта выручки, ₽", float(total_current_revenue - total_previous_revenue)],
        ["Продажи недели, шт", float(total_current_sales)],
        ["Продажи прошлой недели, шт", float(total_previous_sales)],
        ["Дельта продаж, шт", float(total_current_sales - total_previous_sales)],
        ["Средний чек недели, ₽", float(_avg_ticket(total_current_revenue, total_current_sales))],
        [
            "Средний чек прошлой недели, ₽",
            float(_avg_ticket(total_previous_revenue, total_previous_sales)),
        ],
        [
            "Дельта среднего чека, ₽",
            float(
                _avg_ticket(total_current_revenue, total_current_sales)
                - _avg_ticket(total_previous_revenue, total_previous_sales)
            ),
        ],
        [
            "Лидер по выручке",
            (
                f"{_manager_sort_name(leader.manager_name, leader.manager_ref)} "
                f"({_format_money(leader.current_revenue)} ₽)"
                if leader
                else ""
            ),
        ],
        [
            "Минимальная выручка",
            (
                f"{_manager_sort_name(outsider.manager_name, outsider.manager_ref)} "
                f"({_format_money(outsider.current_revenue)} ₽)"
                if outsider
                else ""
            ),
        ],
        [
            "Самая сильная просадка",
            (
                f"{_manager_sort_name(deepest_drop.manager_name, deepest_drop.manager_ref)} "
                f"({_format_money(deepest_drop.revenue_delta)} ₽)"
                if deepest_drop
                else ""
            ),
        ],
    ]
    _make_sheet(
        workbook,
        title=SUMMARY_SHEET_TITLE,
        headers=["Показатель", "Значение"],
        rows=summary_rows,
        money_columns={2},
        qty_columns=set(),
    )
    summary_sheet = workbook[SUMMARY_SHEET_TITLE]
    summary_money_labels = {
        "Выручка недели, ₽",
        "Выручка прошлой недели, ₽",
        "Дельта выручки, ₽",
        "РКО Излишек/недостача, сумма, ₽",
        "Средний чек недели, ₽",
        "Средний чек прошлой недели, ₽",
        "Дельта среднего чека, ₽",
    }
    summary_qty_labels = {
        "Продажи недели, шт",
        "Продажи прошлой недели, шт",
        "Дельта продаж, шт",
    }
    summary_integer_labels = {
        "Менеджеров в отчете",
        "Менеджеров в зоне внимания",
        "РКО Излишек/недостача, шт",
    }
    for row_index in range(2, summary_sheet.max_row + 1):
        label = summary_sheet.cell(row=row_index, column=1).value
        value_cell = summary_sheet.cell(row=row_index, column=2)
        if not isinstance(value_cell.value, (int, float)):
            continue
        if label in summary_money_labels:
            value_cell.number_format = "#,##0.00"
            value_cell.alignment = Alignment(horizontal="right", vertical="top")
        elif label in summary_qty_labels:
            value_cell.number_format = "#,##0.000"
            value_cell.alignment = Alignment(horizontal="right", vertical="top")
        elif label in summary_integer_labels:
            value_cell.number_format = "#,##0"
            value_cell.alignment = Alignment(horizontal="right", vertical="top")

    manager_rows = [
        [
            index,
            item.signal,
            _manager_sort_name(item.manager_name, item.manager_ref),
            ", ".join(item.store_names),
            float(item.current_revenue),
            float(item.current_sales_count),
            float(item.current_avg_ticket),
            float(item.previous_revenue),
            float(item.previous_sales_count),
            float(item.previous_avg_ticket),
            float(item.revenue_delta),
            float(item.sales_count_delta),
            float(item.avg_ticket_delta),
            item.revenue_rank,
            item.manager_code or "",
        ]
        for index, item in enumerate(manager_items, start=1)
    ]
    manager_sheet = _make_sheet(
        workbook,
        title=MANAGERS_SHEET_TITLE,
        headers=[
            "№",
            "Сигнал",
            "Менеджер",
            "Магазины",
            "Выручка недели, ₽",
            "Продажи недели, шт",
            "Ср. чек недели, ₽",
            "Выручка прошлая неделя, ₽",
            "Продажи прошлая неделя, шт",
            "Ср. чек прошлая неделя, ₽",
            "Дельта выручки, ₽",
            "Дельта продаж, шт",
            "Дельта ср. чека, ₽",
            "Место по выручке",
            "Код 1С менеджера",
        ],
        rows=manager_rows,
        money_columns={5, 7, 8, 10, 11, 13},
        qty_columns={6, 9, 12},
    )

    attention_rows = [
        [
            index,
            item.signal,
            _manager_sort_name(item.manager_name, item.manager_ref),
            ", ".join(item.store_names),
            float(item.current_revenue),
            float(item.current_sales_count),
            float(item.current_avg_ticket),
            float(item.revenue_delta),
            float(item.sales_count_delta),
            float(item.avg_ticket_delta),
            item.manager_code or "",
        ]
        for index, item in enumerate(attention_items, start=1)
    ]
    attention_sheet = _make_sheet(
        workbook,
        title=ATTENTION_SHEET_TITLE,
        headers=[
            "№",
            "Сигнал",
            "Менеджер",
            "Магазины",
            "Выручка недели, ₽",
            "Продажи недели, шт",
            "Ср. чек недели, ₽",
            "Дельта выручки, ₽",
            "Дельта продаж, шт",
            "Дельта ср. чека, ₽",
            "Код 1С менеджера",
        ],
        rows=attention_rows,
        money_columns={5, 7, 8, 10},
        qty_columns={6, 9},
    )

    detail_rows = [
        [
            index,
            item.signal,
            _manager_sort_name(item.manager_name, item.manager_ref),
            item.store_name or item.store_ref or "",
            float(item.current_revenue),
            float(item.current_sales_count),
            float(item.current_avg_ticket),
            float(item.previous_revenue),
            float(item.previous_sales_count),
            float(item.previous_avg_ticket),
            float(item.revenue_delta),
            float(item.sales_count_delta),
            float(item.avg_ticket_delta),
            item.manager_code or "",
            item.store_code or "",
        ]
        for index, item in enumerate(manager_store_items, start=1)
    ]
    store_sheet = _make_sheet(
        workbook,
        title=STORE_DETAIL_SHEET_TITLE,
        headers=[
            "№",
            "Сигнал",
            "Менеджер",
            "Магазин",
            "Выручка недели, ₽",
            "Продажи недели, шт",
            "Ср. чек недели, ₽",
            "Выручка прошлая неделя, ₽",
            "Продажи прошлая неделя, шт",
            "Ср. чек прошлая неделя, ₽",
            "Дельта выручки, ₽",
            "Дельта продаж, шт",
            "Дельта ср. чека, ₽",
            "Код 1С менеджера",
            "Код 1С магазина",
        ],
        rows=detail_rows,
        money_columns={5, 7, 8, 10, 11, 13},
        qty_columns={6, 9, 12},
    )

    cash_order_rows = [
        [
            index,
            _format_datetime(item.document_date),
            item.document_number or "",
            item.counterparty_name or "",
            item.counterparty_code or "",
            float(item.amount),
            item.employee_name or "",
            item.cashbox_name or "",
            item.store_name or "",
            item.base_sale_number or "",
        ]
        for index, item in enumerate(cash_order_items, start=1)
    ]
    cash_orders_sheet = _make_sheet(
        workbook,
        title=CASH_ORDERS_SHEET_TITLE,
        headers=[
            "№",
            "Дата РКО",
            "Номер РКО",
            "Контрагент",
            "Код 1С контрагента",
            "Сумма, ₽",
            "Кто оформил",
            "Касса",
            "Магазин",
            "Документ-основание",
        ],
        rows=cash_order_rows,
        money_columns={6},
        qty_columns=set(),
    )

    _add_weekly_sheet_conditional_formatting(
        manager_sheet,
        signal_column="B",
        revenue_column="E",
        delta_columns=["K", "L", "M"],
    )
    _add_weekly_sheet_conditional_formatting(
        attention_sheet,
        signal_column="B",
        revenue_column="E",
        delta_columns=["H", "I", "J"],
    )
    _add_weekly_sheet_conditional_formatting(
        store_sheet,
        signal_column="B",
        revenue_column="E",
        delta_columns=["K", "L", "M"],
    )
    if cash_orders_sheet.max_row >= 2:
        cash_orders_sheet.conditional_formatting.add(
            f"F2:F{cash_orders_sheet.max_row}",
            DataBarRule(
                start_type="min",
                start_value=0,
                end_type="max",
                end_value=0,
                color="70AD47",
                showValue=True,
            ),
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def build_telegram_message(
    *,
    window: WeeklySalesWindow,
    manager_items: list[WeeklyManagerSalesItem],
    attention_items: list[WeeklyManagerSalesItem],
    cash_order_items: list[ShortageCashOrderItem] | None = None,
) -> str:
    total_current_revenue = sum((item.current_revenue for item in manager_items), start=ZERO_MONEY)
    total_current_sales = sum((item.current_sales_count for item in manager_items), start=ZERO_QTY)
    resolved_cash_order_items = cash_order_items or []
    return "\n".join(
        [
            "Личные продажи менеджеров",
            (
                f"Период: {_format_display_date(window.week_start)} - "
                f"{_format_display_date(window.week_end)}"
            ),
            (
                "Сравнение: "
                f"{_format_display_date(window.compare_week_start)} - "
                f"{_format_display_date(window.compare_week_end)}"
            ),
            f"Менеджеров: {len(manager_items)}",
            f"Выручка: {_format_money(total_current_revenue)} ₽",
            f"Продано: {_format_qty(total_current_sales)} шт.",
            f"В зоне внимания: {len(attention_items)}",
            f"РКО Излишек/недостача: {len(resolved_cash_order_items)}",
        ]
    )


def _build_employee_attachment_path(*, window: WeeklySalesWindow, snapshot_date: date) -> Path:
    return (
        DEFAULT_OUTPUT_DIR
        / window.week_end.isoformat()
        / f"Долги сотрудников {_format_display_date(snapshot_date)}.xlsx"
    )


def send_weekly_reports_to_telegram(
    *,
    token: str,
    chat_ids: Sequence[str] | str,
    weekly_message: str,
    weekly_report_path: Path,
    employee_message: str | None = None,
    employee_report_path: Path | None = None,
    sender: Callable[..., None] = send_return_scheme_telegram_report,
) -> int:
    resolved_chat_ids = (
        _parse_telegram_chat_ids(chat_ids)
        if isinstance(chat_ids, str)
        else [str(item).strip() for item in chat_ids if str(item).strip()]
    )
    sent_count = 0
    for chat_id in resolved_chat_ids:
        sender(
            token=token,
            chat_id=chat_id,
            message=weekly_message,
            report_path=weekly_report_path,
        )
        sent_count += 1
        if employee_message and employee_report_path is not None:
            sender(
                token=token,
                chat_id=chat_id,
                message=employee_message,
                report_path=employee_report_path,
            )
            sent_count += 1
    return sent_count


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build weekly Excel report with personal sales of managers"
    )
    parser.add_argument(
        "--date",
        dest="report_date",
        help="Any date inside the target week in YYYY-MM-DD. By default use the latest completed week.",
    )
    parser.add_argument(
        "--output",
        help="Output path for generated report. Defaults to reports/sales/weekly/<week_end>/...",
    )
    parser.add_argument(
        "--send-telegram",
        action="store_true",
        help="Send generated report to Telegram using explicit args or configured env vars",
    )
    parser.add_argument("--telegram-token", help="Telegram bot token override")
    parser.add_argument("--telegram-chat-id", help="Telegram chat id override")
    parser.add_argument(
        "--env-file",
        default=".env",
        help="Path to .env for Telegram defaults (default: .env)",
    )
    args = parser.parse_args()

    settings = get_settings()
    engine = build_engine(settings.database_url)
    onec_engine = _build_onec_engine(settings)

    with Session(engine) as session:
        window = _resolve_report_window(session, requested_date=_parse_date(args.report_date))
        current_records = _load_sales_records(
            session,
            date_from=window.week_start,
            date_to=window.week_end,
        )
        previous_records = _load_sales_records(
            session,
            date_from=window.compare_week_start,
            date_to=window.compare_week_end,
        )
        employee_snapshot_date, employee_previous_date = resolve_employee_snapshot_dates(
            session,
            requested_date=None,
            latest_not_after=window.week_end,
        )
        employee_current_items = load_employee_items(
            session,
            snapshot_date=employee_snapshot_date,
        )
        employee_previous_items = (
            load_employee_items(session, snapshot_date=employee_previous_date)
            if employee_previous_date
            else []
        )
        employee_related_documents = load_employee_related_documents(
            onec_engine=onec_engine,
            snapshot_date=employee_snapshot_date,
        )
        weekly_history = load_weekly_sales_history(
            session,
            week_end=window.week_end,
            limit=4,
        )
        employee_snapshot_history = load_employee_snapshot_history(
            session,
            snapshot_date=employee_snapshot_date,
            limit=7,
        )
    cash_order_items = (
        fetch_onec_shortage_cash_orders(
            onec_engine,
            date_from=window.week_start,
            date_to=window.week_end,
        )
        if onec_engine is not None
        else []
    )
    enrich_sales_records_with_codes(current_records, onec_engine=onec_engine)
    enrich_sales_records_with_codes(previous_records, onec_engine=onec_engine)
    enrich_employee_items_with_counterparty_codes(employee_current_items, onec_engine=onec_engine)
    enrich_employee_items_with_counterparty_codes(employee_previous_items, onec_engine=onec_engine)

    manager_items = build_weekly_manager_sales_items(current_records, previous_records)
    manager_store_items = build_weekly_manager_store_sales_items(current_records, previous_records)
    attention_items = build_attention_manager_sales_items(manager_items)
    employee_changes = build_employee_receivable_changes(
        employee_current_items,
        employee_previous_items,
    )

    output_path = (
        Path(args.output)
        if args.output
        else _build_output_path(window=window, output_dir=DEFAULT_OUTPUT_DIR)
    )
    export_weekly_manager_sales_report(
        window=window,
        manager_items=manager_items,
        attention_items=attention_items,
        manager_store_items=manager_store_items,
        cash_order_items=cash_order_items,
        output_path=output_path,
        weekly_history=weekly_history,
    )
    employee_report_path = _build_employee_attachment_path(
        window=window,
        snapshot_date=employee_snapshot_date,
    )
    export_employee_receivable_report(
        snapshot_date=employee_snapshot_date,
        previous_date=employee_previous_date,
        current_items=employee_current_items,
        changes=employee_changes,
        output_path=employee_report_path,
        snapshot_history=employee_snapshot_history,
        related_documents=employee_related_documents,
    )

    print(f"report_path={output_path}")
    print(f"employee_report_path={employee_report_path}")
    print(f"week_start={window.week_start.isoformat()}")
    print(f"week_end={window.week_end.isoformat()}")
    print(f"compare_week_start={window.compare_week_start.isoformat()}")
    print(f"compare_week_end={window.compare_week_end.isoformat()}")
    print(f"manager_count={len(manager_items)}")
    print(f"attention_count={len(attention_items)}")
    print(f"cash_order_count={len(cash_order_items)}")
    print(f"employee_snapshot_date={employee_snapshot_date.isoformat()}")
    print(
        f"employee_previous_date={employee_previous_date.isoformat() if employee_previous_date else ''}"
    )
    print(f"employee_case_count={len(employee_current_items)}")

    if not args.send_telegram:
        return

    env = _load_env(Path(args.env_file))
    token = args.telegram_token or _resolve_default_telegram_target(env)[0]
    chat_ids = (
        _parse_telegram_chat_ids(args.telegram_chat_id)
        if args.telegram_chat_id
        else _parse_telegram_chat_ids(_resolve_default_telegram_target(env)[1])
    )
    if not token or not chat_ids:
        raise RuntimeError("Telegram token/chat_id are not configured")

    message = build_telegram_message(
        window=window,
        manager_items=manager_items,
        attention_items=attention_items,
        cash_order_items=cash_order_items,
    )
    employee_message = build_employee_telegram_message(
        snapshot_date=employee_snapshot_date,
        previous_date=employee_previous_date,
        current_items=employee_current_items,
        changes=employee_changes,
    )
    sent_count = send_weekly_reports_to_telegram(
        token=token,
        chat_ids=chat_ids,
        weekly_message=message,
        weekly_report_path=output_path,
        employee_message=employee_message,
        employee_report_path=employee_report_path,
    )
    print("telegram_sent=true")
    print(f"telegram_documents_sent={sent_count}")
    print(f"telegram_chat_count={len(chat_ids)}")


if __name__ == "__main__":
    main()
