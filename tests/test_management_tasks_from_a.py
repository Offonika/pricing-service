from __future__ import annotations

import json
import urllib.error
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from infra.cron import management_tasks_from_a as management_tasks_module
from infra.cron.management_tasks_from_a import (
    FallbackableTaskSyncError,
    _build_description,
    _export_receivable_batch_xlsx,
    _fingerprint_payload,
    _load_team_role_map,
    _task_fields,
    render_summary,
    sync_management_tasks,
)


def _task_payload(
    *,
    dedupe_key: str,
    rule_code: str = "custom_rule",
    owner_code: str = "finance",
    watcher_codes: list[str] | None = None,
    title: str = "Task title",
    summary: str = "Task summary",
) -> dict[str, object]:
    return {
        "rule_code": rule_code,
        "source_type": "receivable_case",
        "entity_ref": "cp-1",
        "entity_name": "Контрагент 1",
        "severity": "high",
        "owner_code": owner_code,
        "watcher_codes": watcher_codes or [],
        "title": title,
        "summary": summary,
        "reaction_deadline_at": "2026-03-20T12:00:00",
        "due_at": "2026-03-21T18:00:00",
        "dedupe_key": dedupe_key,
        "tags": ["management", "receivables"],
        "metrics": {"current_balance": "12000"},
        "references": [{"kind": "origin_document", "document_ref": "doc-1"}],
    }


def test_sync_management_tasks_creates_and_updates_state(tmp_path: Path) -> None:
    state_path = tmp_path / "state.json"
    created: list[tuple[int, list[int], str]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        assert path == "/api/management/task-payloads"
        assert params == {"date": "2026-03-20"}
        return {
            "payload": [_task_payload(dedupe_key="finance|2026-03-20|cp-1", watcher_codes=["hr"])]
        }

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        assert webhook_url == "https://bitrix.example/rest/1/token"
        created.append((assignee_id, observer_ids, str(payload["title"])))
        return 501

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        owner_overrides={},
        watcher_overrides={"hr": 777},
        team_roles={"cfo": 10105, "coo": 21},
        default_responsible_id=10105,
        default_observer_ids=[],
        default_created_by_id=None,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        dry_run=False,
        create_task=create_task,
    )

    assert summary["created"] == 1
    assert summary["updated"] == 0
    assert summary["noop"] == 0
    assert created == [(10105, [777], "Task title")]

    state = json.loads(state_path.read_text(encoding="utf-8"))
    assert state["tasks"]["finance|2026-03-20|cp-1"]["task_id"] == 501


def test_sync_management_tasks_noops_on_same_fingerprint(tmp_path: Path) -> None:
    state_path = tmp_path / "state.json"
    payload = _task_payload(dedupe_key="finance|2026-03-20|cp-1", watcher_codes=["hr"])
    state_path.write_text(
        json.dumps(
            {
                "tasks": {
                    "finance|2026-03-20|cp-1": {
                        "task_id": 501,
                        "fingerprint": _fingerprint_payload(payload),
                    }
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [payload]}

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        owner_overrides={},
        watcher_overrides={"hr": 777},
        team_roles={"cfo": 10105},
        default_responsible_id=10105,
        default_observer_ids=[],
        default_created_by_id=None,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        dry_run=False,
    )

    assert summary["created"] == 0
    assert summary["updated"] == 0
    assert summary["noop"] == 1


def test_sync_management_tasks_updates_existing_task(tmp_path: Path) -> None:
    state_path = tmp_path / "state.json"
    state_path.write_text(
        json.dumps(
            {
                "tasks": {
                    "finance|2026-03-20|cp-1": {
                        "task_id": 501,
                        "fingerprint": "old",
                    }
                }
            }
        ),
        encoding="utf-8",
    )
    updated: list[tuple[int, int, str]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {
            "payload": [
                _task_payload(
                    dedupe_key="finance|2026-03-20|cp-1",
                    title="Updated title",
                    summary="Updated summary",
                )
            ]
        }

    def update_task(
        *,
        webhook_url: str,
        task_id: int,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> None:
        updated.append((task_id, assignee_id, str(payload["title"])))

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        owner_overrides={},
        watcher_overrides={},
        team_roles={"cfo": 10105},
        default_responsible_id=10105,
        default_observer_ids=[],
        default_created_by_id=None,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        dry_run=False,
        update_task=update_task,
    )

    assert summary["created"] == 0
    assert summary["updated"] == 1
    assert summary["noop"] == 0
    assert updated == [(501, 10105, "Updated title")]


def test_render_summary_reports_rule_counters(tmp_path: Path) -> None:
    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {
            "payload": [
                _task_payload(dedupe_key="finance|2026-03-20|cp-1"),
                _task_payload(
                    dedupe_key="retail|2026-03-20|store-1",
                    owner_code="retail_supervisor",
                    title="Staffing",
                ),
            ]
        }

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 20),
        state_path=tmp_path / "state.json",
        owner_overrides={"retail_supervisor": 21},
        watcher_overrides={},
        team_roles={"cfo": 10105, "coo": 21},
        default_responsible_id=10105,
        default_observer_ids=[1],
        default_created_by_id=None,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        dry_run=True,
    )
    rendered = render_summary(summary)

    assert summary["created"] == 2
    assert "Payload'ов: 2" in rendered
    assert "custom_rule=2" in rendered


def test_sync_management_tasks_batches_new_daily_receivables(tmp_path: Path) -> None:
    state_path = tmp_path / "state.json"
    created: list[tuple[int, list[int], dict[str, object]]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        assert path == "/api/management/task-payloads"
        return {
            "payload": [
                {
                    **_task_payload(
                        dedupe_key="receivable_new_daily|2026-03-20|cp-1",
                        owner_code="finance",
                        title="Дебиторка: Появилась новая дебиторка. — Контрагент 1",
                        summary="Появилась новая дебиторка.",
                    ),
                    "rule_code": "receivable_new_daily",
                    "entity_ref": "cp-1",
                    "entity_name": "Контрагент 1",
                    "severity": "warning",
                    "metrics": {"current_balance": "12000"},
                },
                {
                    **_task_payload(
                        dedupe_key="receivable_new_daily|2026-03-20|cp-2",
                        owner_code="finance",
                        title="Дебиторка: Появилась новая дебиторка. — Контрагент 2",
                        summary="Появилась новая дебиторка.",
                    ),
                    "rule_code": "receivable_new_daily",
                    "entity_ref": "cp-2",
                    "entity_name": "Контрагент 2",
                    "severity": "critical",
                    "metrics": {"current_balance": "25000"},
                },
            ]
        }

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append((assignee_id, observer_ids, payload | {"created_by_id": created_by_id}))
        return 700

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        owner_overrides={},
        watcher_overrides={},
        team_roles={"cfo": 10105},
        default_responsible_id=10105,
        default_observer_ids=[],
        default_created_by_id=7869,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        dry_run=False,
        create_task=create_task,
    )

    assert summary["created"] == 1
    assert summary["payload_count"] == 1
    assert len(created) == 1
    _, _, payload = created[0]
    assert payload["rule_code"] == "receivable_new_daily_batch"
    assert payload["dedupe_key"] == "receivable_new_daily_batch|2026-03-20"
    assert payload["title"] == "Дебиторка: новая дебиторка за 2026-03-20"
    assert payload["severity"] == "critical"
    assert payload["created_by_id"] == 7869
    assert payload["metrics"] == {
        "counterparty_count": 2,
        "current_balance_total": "37000.00",
    }
    assert "Новых контрагентов с дебиторкой: 2 на 37 000" in str(payload["summary"])

    state = json.loads(state_path.read_text(encoding="utf-8"))
    assert state["tasks"]["receivable_new_daily_batch|2026-03-20"]["task_id"] == 700


def test_batch_description_is_human_readable() -> None:
    payload = {
        "rule_code": "receivable_new_daily_batch",
        "dedupe_key": "receivable_new_daily_batch|2026-03-20",
        "entity_ref": "receivables:new_daily:2026-03-20",
        "due_at": "2026-03-21T18:00:00",
        "reaction_deadline_at": "2026-03-20T12:00:00",
        "metrics": {
            "counterparty_count": 2,
            "current_balance_total": "37000.00",
        },
        "references": [
            {
                "counterparty_name": "Контрагент 2",
                "current_balance": "25000",
            },
            {
                "counterparty_name": "Контрагент 1",
                "current_balance": "12000",
            },
        ],
    }

    description = _build_description(payload)

    assert "Что произошло" in description
    assert "Как отобрано" in description
    assert 'сегмента "новая дебиторка за день"' in description
    assert "За 20.03.2026 выявлено 2 новых контрагентов с дебиторкой на 37 000 ₽." in description
    assert "Топ-10 по сумме" in description
    assert "1. Контрагент 2 — 25 000 ₽" in description
    assert "kind=receivable_case_batch_item" not in description


def test_finance_daily_batch_description_is_human_readable() -> None:
    payload = {
        "rule_code": "receivable_finance_daily_batch",
        "dedupe_key": "receivable_finance_daily_batch|2026-03-21",
        "due_at": "2026-03-23T18:00:00",
        "reaction_deadline_at": "2026-03-21T12:00:00",
        "metrics": {
            "counterparty_count": 4,
            "current_balance_total": "53000.00",
            "sections": {
                "receivable_new_daily_batch": {
                    "counterparty_count": 1,
                    "current_balance_total": "12000.00",
                },
                "receivable_employee_batch": {
                    "counterparty_count": 1,
                    "current_balance_total": "5000.00",
                },
                "receivable_fired_manager_batch": {
                    "counterparty_count": 1,
                    "current_balance_total": "7000.00",
                },
                "receivable_adjustment_candidate_batch": {
                    "counterparty_count": 1,
                    "current_balance_total": "29000.00",
                },
            },
        },
    }

    description = _build_description(payload)

    assert "единый дневной финансовый пакет" in description
    assert "Новая дебиторка: 1 на 12 000 ₽." in description
    assert "Долги сотрудников: 1 на 5 000 ₽." in description
    assert "Уволенные менеджеры: 1 на 7 000 ₽." in description
    assert "Корректировка: 1 на 29 000 ₽." in description


def test_export_receivable_batch_xlsx(tmp_path: Path) -> None:
    payload = {
        "references": [
            {
                "counterparty_name": "Контрагент 2",
                "counterparty_ref": "cp-2",
                "current_balance": "25000",
                "original_task_key": "receivable_new_daily|2026-03-20|cp-2",
            },
            {
                "counterparty_name": "Контрагент 1",
                "counterparty_ref": "cp-1",
                "current_balance": "12000",
                "original_task_key": "receivable_new_daily|2026-03-20|cp-1",
            },
        ]
    }
    output_path = tmp_path / "receivables.xlsx"

    _export_receivable_batch_xlsx(payload, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("№", "Контрагент", "Сумма долга, ₽", "Контрагент ref", "Исходный dedupe key")
    assert rows[1] == (1, "Контрагент 2", 25000, "cp-2", "receivable_new_daily|2026-03-20|cp-2")
    assert rows[2] == (2, "Контрагент 1", 12000, "cp-1", "receivable_new_daily|2026-03-20|cp-1")
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:E3"
    assert sheet["A1"].font.bold is True
    assert sheet.column_dimensions["B"].width >= 16


def test_export_receivable_finance_daily_batch_xlsx(tmp_path: Path) -> None:
    payload = {
        "rule_code": "receivable_finance_daily_batch",
        "references": [
            {
                "batch_rule_code": "receivable_employee_batch",
                "counterparty_name": "Сотрудник 1",
                "counterparty_ref": "cp-emp",
                "current_balance": "5000",
                "aged_bucket": "8-30",
                "activity_segment": "inactive",
                "current_manager_name": "Менеджер 1",
                "origin_document_number": "S-1",
                "original_task_key": "receivable_employee|2026-03-21|cp-emp",
            },
            {
                "batch_rule_code": "receivable_adjustment_candidate_batch",
                "counterparty_name": "Контрагент adj",
                "counterparty_ref": "cp-adj",
                "current_balance": "9000",
                "aged_bucket": "90+",
                "activity_segment": "inactive",
                "current_manager_name": "Менеджер 3",
                "origin_document_number": "S-3",
                "original_task_key": "receivable_adjustment_candidate|2026-03-21|cp-adj",
            },
        ],
    }
    output_path = tmp_path / "receivables-finance-daily.xlsx"

    _export_receivable_batch_xlsx(payload, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Категория",
        "№",
        "Контрагент",
        "Сумма долга, ₽",
        "Возраст долга",
        "Активность",
        "Текущий ответственный",
        "Исходный документ",
        "Контрагент ref",
        "Исходный dedupe key",
    )
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:J3"
    assert sheet["D2"].number_format == "#,##0.00"
    assert rows[1][0] == "Долги сотрудников"
    assert rows[2][0] == "Корректировка"


def test_sync_management_tasks_batches_overdue_receivables(tmp_path: Path) -> None:
    state_path = tmp_path / "state.json"
    created: list[tuple[int, list[int], dict[str, object]]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        assert path == "/api/management/task-payloads"
        return {
            "payload": [
                {
                    **_task_payload(
                        dedupe_key="receivable_overdue|2026-03-21|cp-1",
                        owner_code="sales_manager",
                        title="Дебиторка: Просрочено — Контрагент 1",
                        summary="Долг просрочен.",
                    ),
                    "rule_code": "receivable_overdue",
                    "entity_ref": "cp-1",
                    "entity_name": "Контрагент 1",
                    "severity": "high",
                    "metrics": {
                        "current_balance": "12000",
                        "aged_bucket": "31-60",
                        "due_date": "2026-03-15T00:00:00",
                        "overdue_days": 6,
                        "payment_term_source": "planned_payment_date",
                        "shipment_ban": False,
                    },
                },
                {
                    **_task_payload(
                        dedupe_key="receivable_overdue|2026-03-21|cp-2",
                        owner_code="sales_manager",
                        title="Дебиторка: Просрочено — Контрагент 2",
                        summary="Долг просрочен.",
                    ),
                    "rule_code": "receivable_overdue",
                    "entity_ref": "cp-2",
                    "entity_name": "Контрагент 2",
                    "severity": "critical",
                    "metrics": {
                        "current_balance": "25000",
                        "aged_bucket": "61-90",
                        "due_date": "2026-03-10T00:00:00",
                        "overdue_days": 11,
                        "payment_term_source": "credit_depth_days",
                        "shipment_ban": True,
                    },
                },
            ]
        }

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append((assignee_id, observer_ids, payload | {"created_by_id": created_by_id}))
        return 701

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 21),
        state_path=state_path,
        owner_overrides={},
        watcher_overrides={"finance": 10105},
        team_roles={"cco": 1, "cfo": 10105},
        default_responsible_id=10105,
        default_observer_ids=[],
        default_created_by_id=7869,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        overdue_batch_weekday=6,
        dry_run=False,
        create_task=create_task,
    )

    assert summary["created"] == 1
    assert summary["payload_count"] == 1
    assert len(created) == 1
    assignee_id, observer_ids, payload = created[0]
    assert assignee_id == 1
    assert observer_ids == [10105]
    assert payload["rule_code"] == "receivable_overdue_batch"
    assert payload["dedupe_key"] == "receivable_overdue_batch|2026-03-21"
    assert payload["title"] == "Дебиторка: просроченная дебиторка на 2026-03-21"
    assert payload["severity"] == "critical"
    assert payload["created_by_id"] == 7869
    assert payload["owner_code"] == "cco"
    assert payload["metrics"] == {
        "counterparty_count": 2,
        "current_balance_total": "37000.00",
        "aged_bucket_counts": {"31-60": 1, "61-90": 1},
        "aged_bucket_totals": {"31-60": "12000.00", "61-90": "25000.00"},
    }
    assert "31-60: 1 на 12 000 ₽" in str(payload["summary"])
    assert "61-90: 1 на 25 000 ₽" in str(payload["summary"])

    state = json.loads(state_path.read_text(encoding="utf-8"))
    assert state["tasks"]["receivable_overdue_batch|2026-03-21"]["task_id"] == 701


def test_sync_management_tasks_skips_overdue_outside_weekly_day(tmp_path: Path) -> None:
    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {
            "payload": [
                {
                    **_task_payload(
                        dedupe_key="receivable_overdue|2026-03-21|cp-1",
                        owner_code="sales_manager",
                        title="Дебиторка: Просрочено — Контрагент 1",
                        summary="Долг просрочен.",
                    ),
                    "rule_code": "receivable_overdue",
                    "entity_ref": "cp-1",
                    "entity_name": "Контрагент 1",
                    "severity": "high",
                    "metrics": {
                        "current_balance": "12000",
                        "aged_bucket": "31-60",
                        "due_date": "2026-03-15T00:00:00",
                        "overdue_days": 6,
                        "payment_term_source": "planned_payment_date",
                        "shipment_ban": False,
                    },
                }
            ]
        }

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 21),
        state_path=tmp_path / "state.json",
        owner_overrides={},
        watcher_overrides={"finance": 10105},
        team_roles={"cco": 1, "cfo": 10105},
        default_responsible_id=10105,
        default_observer_ids=[],
        default_created_by_id=7869,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        overdue_batch_weekday=1,
        dry_run=True,
    )

    assert summary["payload_count"] == 0
    assert summary["created"] == 0
    assert summary["updated"] == 0
    assert summary["noop"] == 0


def test_sync_management_tasks_batches_employee_fired_and_adjustment_receivables(
    tmp_path: Path,
) -> None:
    created: list[dict[str, object]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {
            "payload": [
                {
                    **_task_payload(
                        dedupe_key="receivable_employee|2026-03-21|cp-emp",
                        owner_code="finance",
                        title="Сотрудник",
                        summary="Сотрудник",
                    ),
                    "rule_code": "receivable_employee",
                    "entity_ref": "cp-emp",
                    "entity_name": "Сотрудник 1",
                    "watcher_codes": ["hr"],
                    "metrics": {
                        "current_balance": "5000",
                        "aged_bucket": "8-30",
                        "activity_segment": "inactive",
                    },
                    "references": [
                        {"kind": "current_manager", "current_manager_name": "Менеджер 1"},
                        {"kind": "origin_document", "document_number": "S-1"},
                    ],
                },
                {
                    **_task_payload(
                        dedupe_key="receivable_fired_manager|2026-03-21|cp-fired",
                        owner_code="finance_pool",
                        title="Уволенный",
                        summary="Уволенный",
                    ),
                    "rule_code": "receivable_fired_manager",
                    "entity_ref": "cp-fired",
                    "entity_name": "Контрагент fired",
                    "metrics": {
                        "current_balance": "7000",
                        "aged_bucket": "31-60",
                        "activity_segment": "inactive",
                    },
                    "references": [
                        {"kind": "current_manager", "current_manager_name": "Менеджер 2"},
                        {"kind": "origin_document", "document_number": "S-2"},
                    ],
                },
                {
                    **_task_payload(
                        dedupe_key="receivable_adjustment_candidate|2026-03-21|cp-adj",
                        owner_code="finance",
                        title="Корректировка",
                        summary="Корректировка",
                    ),
                    "rule_code": "receivable_adjustment_candidate",
                    "entity_ref": "cp-adj",
                    "entity_name": "Контрагент adj",
                    "metrics": {
                        "current_balance": "9000",
                        "aged_bucket": "90+",
                        "activity_segment": "inactive",
                    },
                    "references": [
                        {"kind": "current_manager", "current_manager_name": "Менеджер 3"},
                        {"kind": "origin_document", "document_number": "S-3"},
                    ],
                },
                {
                    **_task_payload(
                        dedupe_key="receivable_adjustment_candidate_large|cp-adj-big",
                        owner_code="retail_network_head",
                        title="Крупная корректировка",
                        summary="Крупная корректировка",
                    ),
                    "rule_code": "receivable_adjustment_candidate_large",
                    "entity_ref": "cp-adj-big",
                    "entity_name": "Контрагент adj big",
                    "watcher_codes": ["ceo"],
                    "created_by_code": "cfo",
                    "suppress_default_observers": True,
                    "allow_assignee_change_deadline": True,
                    "due_at": "2026-04-04T18:00:00",
                    "metrics": {
                        "current_balance": "15000",
                        "aged_bucket": "90+",
                        "activity_segment": "inactive",
                    },
                    "references": [
                        {"kind": "current_manager", "current_manager_name": "Менеджер 4"},
                        {"kind": "origin_document", "document_number": "S-4"},
                    ],
                },
            ]
        }

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append(
            {
                "payload": payload,
                "assignee_id": assignee_id,
                "observer_ids": observer_ids,
                "created_by_id": created_by_id,
            }
        )
        return 800 + len(created)

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 21),
        state_path=tmp_path / "state.json",
        owner_overrides={},
        watcher_overrides={"hr": 777},
        team_roles={"cfo": 10105, "ceo": 7869, "retail_network_head": 6759},
        default_responsible_id=10105,
        default_observer_ids=[],
        default_created_by_id=7869,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        overdue_batch_weekday=6,
        dry_run=False,
        create_task=create_task,
    )

    assert summary["payload_count"] == 1
    assert summary["created"] == 1
    assert len(created) == 1

    by_rule = {item["payload"]["rule_code"]: item for item in created}
    assert set(by_rule) == {"receivable_adjustment_candidate_large"}

    large_adjustment_item = by_rule["receivable_adjustment_candidate_large"]
    assert large_adjustment_item["assignee_id"] == 6759
    assert large_adjustment_item["observer_ids"] == [7869]
    assert large_adjustment_item["created_by_id"] == 10105
    assert large_adjustment_item["payload"]["allow_assignee_change_deadline"] is True
    assert large_adjustment_item["payload"]["due_at"] == "2026-04-04T18:00:00"
    assert (
        large_adjustment_item["payload"]["dedupe_key"]
        == "receivable_adjustment_candidate_large|cp-adj-big"
    )


def test_sync_management_tasks_updates_large_adjustment_by_stable_case_key(
    tmp_path: Path,
) -> None:
    state_path = tmp_path / "state.json"
    state_path.write_text(
        json.dumps(
            {
                "tasks": {
                    "receivable_adjustment_candidate_large|cp-adj-big": {
                        "task_id": 912,
                        "fingerprint": "old",
                    }
                }
            }
        ),
        encoding="utf-8",
    )
    updated: list[dict[str, object]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {
            "payload": [
                {
                    **_task_payload(
                        dedupe_key="receivable_adjustment_candidate_large|cp-adj-big",
                        owner_code="retail_network_head",
                        title="Крупная корректировка",
                        summary="Крупная корректировка",
                    ),
                    "rule_code": "receivable_adjustment_candidate_large",
                    "entity_ref": "cp-adj-big",
                    "entity_name": "Контрагент adj big",
                    "watcher_codes": ["ceo"],
                    "created_by_code": "cfo",
                    "suppress_default_observers": True,
                    "allow_assignee_change_deadline": True,
                    "metrics": {
                        "current_balance": "17000",
                        "aged_bucket": "90+",
                        "activity_segment": "inactive",
                    },
                }
            ]
        }

    def update_task(
        *,
        webhook_url: str,
        task_id: int,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> None:
        updated.append(
            {
                "task_id": task_id,
                "assignee_id": assignee_id,
                "observer_ids": observer_ids,
                "created_by_id": created_by_id,
                "dedupe_key": payload.get("dedupe_key"),
            }
        )

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 22),
        state_path=state_path,
        owner_overrides={},
        watcher_overrides={},
        team_roles={"cfo": 10105, "ceo": 7869, "retail_network_head": 6759},
        default_responsible_id=10105,
        default_observer_ids=[],
        default_created_by_id=None,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        dry_run=False,
        update_task=update_task,
    )

    assert summary["created"] == 0
    assert summary["updated"] == 1
    assert updated == [
        {
            "task_id": 912,
            "assignee_id": 6759,
            "observer_ids": [7869],
            "created_by_id": 10105,
            "dedupe_key": "receivable_adjustment_candidate_large|cp-adj-big",
        }
    ]


def test_task_fields_enable_assignee_deadline_change() -> None:
    params = _task_fields(
        {
            **_task_payload(dedupe_key="receivable_adjustment_candidate_large|cp-1"),
            "allow_assignee_change_deadline": True,
        },
        assignee_id=130751,
        observer_ids=[4241, 115204],
        created_by_id=130746,
    )

    assert ("fields[ALLOW_CHANGE_DEADLINE]", "Y") in params


def test_overdue_batch_description_is_human_readable() -> None:
    payload = {
        "rule_code": "receivable_overdue_batch",
        "dedupe_key": "receivable_overdue_batch|2026-03-21",
        "entity_ref": "receivables:overdue:2026-03-21",
        "due_at": "2026-03-23T18:00:00",
        "reaction_deadline_at": "2026-03-21T12:00:00",
        "metrics": {
            "counterparty_count": 2,
            "current_balance_total": "37000.00",
            "aged_bucket_counts": {"31-60": 1, "61-90": 1},
            "aged_bucket_totals": {"31-60": "12000.00", "61-90": "25000.00"},
        },
        "references": [
            {
                "counterparty_name": "Контрагент 2",
                "current_balance": "25000",
                "due_date": "2026-03-10",
                "overdue_days": 11,
            },
            {
                "counterparty_name": "Контрагент 1",
                "current_balance": "12000",
                "due_date": "2026-03-15",
                "overdue_days": 6,
            },
        ],
    }

    description = _build_description(payload)

    assert "Что произошло" in description
    assert "Как отобрано" in description
    assert "истёк согласованный срок оплаты или глубина кредита" in description
    assert "Разбивка по возрасту долга: 31-60: 1 на 12 000 ₽; 61-90: 1 на 25 000 ₽." in description
    assert "Топ-20 по сумме" in description
    assert "1. Контрагент 2 — 25 000 ₽; срок 10.03.2026; просрочка 11 дн." in description


def test_export_receivable_overdue_batch_xlsx(tmp_path: Path) -> None:
    payload = {
        "rule_code": "receivable_overdue_batch",
        "references": [
            {
                "counterparty_name": "Контрагент 2",
                "counterparty_ref": "cp-2",
                "current_balance": "25000",
                "due_date": "2026-03-10T00:00:00",
                "overdue_days": 11,
                "payment_term_source": "credit_depth_days",
                "aged_bucket": "61-90",
                "shipment_ban": True,
                "original_task_key": "receivable_overdue|2026-03-21|cp-2",
            },
            {
                "counterparty_name": "Контрагент 1",
                "counterparty_ref": "cp-1",
                "current_balance": "12000",
                "due_date": "2026-03-15T00:00:00",
                "overdue_days": 6,
                "payment_term_source": "planned_payment_date",
                "aged_bucket": "31-60",
                "shipment_ban": False,
                "original_task_key": "receivable_overdue|2026-03-21|cp-1",
            },
        ],
    }
    output_path = tmp_path / "receivables-overdue.xlsx"

    _export_receivable_batch_xlsx(payload, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "№",
        "Контрагент",
        "Сумма долга, ₽",
        "Срок оплаты",
        "Дней просрочки",
        "Источник срока",
        "Возраст долга",
        "Запрет отгрузки",
        "Контрагент ref",
        "Исходный dedupe key",
    )
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:J3"
    assert sheet["C2"].number_format == "#,##0.00"
    assert rows[1] == (
        1,
        "Контрагент 2",
        25000,
        "2026-03-10T00:00:00",
        11,
        "Глубина кредита",
        "61-90",
        "Да",
        "cp-2",
        "receivable_overdue|2026-03-21|cp-2",
    )
    assert rows[2] == (
        2,
        "Контрагент 1",
        12000,
        "2026-03-15T00:00:00",
        6,
        "Согласованная дата оплаты",
        "31-60",
        "Нет",
        "cp-1",
        "receivable_overdue|2026-03-21|cp-1",
    )


def test_load_team_role_map_supports_dual_bitrix_fields(tmp_path: Path) -> None:
    team_path = tmp_path / "team.yaml"
    team_path.write_text(
        """
employees:
  - role_code: cfo
    bitrix24_id: 10105
    bitrix_cloud_user_id: 1001
    bitrix_box_user_id: 2001
    bitrix_box_assistant_enabled: true
    allowed_delivery_channels:
      - bitrix_task
      - telegram_digest
""".strip(),
        encoding="utf-8",
    )

    role_map = _load_team_role_map(str(team_path))

    assert role_map["cfo"]["legacy"] == 10105
    assert role_map["cfo"]["cloud"] == 1001
    assert role_map["cfo"]["box"] == 2001
    assert role_map["cfo"]["bitrix_box_assistant_enabled"] is True
    assert role_map["cfo"]["allowed_delivery_channels"] == ["bitrix_task", "telegram_digest"]


def test_load_team_role_map_does_not_backfill_box_from_legacy_id(tmp_path: Path) -> None:
    team_path = tmp_path / "team.yaml"
    team_path.write_text(
        """
employees:
  - role_code: cfo
    bitrix24_id: 10105
    bitrix_cloud_user_id: 1001
""".strip(),
        encoding="utf-8",
    )

    role_map = _load_team_role_map(str(team_path))

    assert role_map["cfo"]["legacy"] == 10105
    assert role_map["cfo"]["cloud"] == 1001
    assert role_map["cfo"]["box"] is None


def test_build_env_delivery_targets_box_does_not_inherit_shared_cloud_ids() -> None:
    migration_state, delivery_targets = management_tasks_module._build_env_delivery_targets(
        {
            "MANAGEMENT_B24_TASK_MIGRATION_STATE": "box_shadow",
            "MANAGEMENT_B24_WEBHOOK_URL": "https://bitrix-cloud.example/rest/1/token",
            "MANAGEMENT_B24_BOX_WEBHOOK_URL": "https://bitrix-box.example/rest/115204/token",
            "MANAGEMENT_B24_OWNER_OVERRIDES": "finance:10105",
            "MANAGEMENT_B24_WATCHER_OVERRIDES": "hr:777",
            "MANAGEMENT_B24_DEFAULT_RESPONSIBLE_ID": "10105",
            "MANAGEMENT_B24_DEFAULT_OBSERVER_IDS": "1,2",
            "MANAGEMENT_B24_CREATED_BY_ID": "55",
        },
        team_roles={
            "cfo": {
                "legacy": 10105,
                "cloud": 10105,
                "box": 115204,
            }
        },
    )

    assert migration_state == "box_shadow"
    assert delivery_targets[0]["contour"] == "cloud"
    assert delivery_targets[0]["owner_overrides"] == {"finance": 10105}
    assert delivery_targets[0]["watcher_overrides"] == {"hr": 777}
    assert delivery_targets[0]["default_responsible_id"] == 10105
    assert delivery_targets[0]["default_observer_ids"] == [1, 2]
    assert delivery_targets[0]["default_created_by_id"] == 55

    assert delivery_targets[1]["contour"] == "box"
    assert delivery_targets[1]["mode"] == "shadow"
    assert delivery_targets[1]["owner_overrides"] == {}
    assert delivery_targets[1]["watcher_overrides"] == {}
    assert delivery_targets[1]["default_responsible_id"] == 115204
    assert delivery_targets[1]["default_observer_ids"] == []
    assert delivery_targets[1]["default_created_by_id"] is None


def test_sync_management_tasks_box_shadow_uses_separate_state_keys(tmp_path: Path) -> None:
    created: list[tuple[str, int, str]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [_task_payload(dedupe_key="finance|2026-03-20|cp-1")]}

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append((webhook_url, assignee_id, str(payload["title"])))
        return 700 + len(created)

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url=None,
        anchor_date=date(2026, 3, 20),
        state_path=tmp_path / "state.json",
        owner_overrides={},
        watcher_overrides={},
        team_roles={
            "cfo": {
                "legacy": 10105,
                "cloud": 10105,
                "box": 115204,
            }
        },
        default_responsible_id=None,
        default_observer_ids=[],
        default_created_by_id=None,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        dry_run=False,
        create_task=create_task,
        delivery_targets=[
            {
                "contour": "cloud",
                "mode": "primary",
                "webhook_url": "https://bitrix-cloud.example/rest/1/token",
                "owner_overrides": {},
                "watcher_overrides": {},
                "default_responsible_id": 10105,
                "default_observer_ids": [],
                "default_created_by_id": None,
                "disk_folder_id": None,
            },
            {
                "contour": "box",
                "mode": "shadow",
                "webhook_url": "https://bitrix-box.example/rest/115204/token",
                "owner_overrides": {},
                "watcher_overrides": {},
                "default_responsible_id": 115204,
                "default_observer_ids": [],
                "default_created_by_id": None,
                "disk_folder_id": None,
            },
        ],
        migration_state="box_shadow",
    )

    assert summary["created"] == 2
    assert summary["by_contour"] == {
        "box": {"created": 1, "updated": 0, "noop": 0},
        "cloud": {"created": 1, "updated": 0, "noop": 0},
    }
    assert created == [
        ("https://bitrix-cloud.example/rest/1/token", 10105, "Task title"),
        ("https://bitrix-box.example/rest/115204/token", 115204, "Task title"),
    ]

    state = json.loads((tmp_path / "state.json").read_text(encoding="utf-8"))
    assert "finance|2026-03-20|cp-1" in state["tasks"]
    assert "task|box|finance|2026-03-20|cp-1" in state["tasks"]
    assert state["tasks"]["finance|2026-03-20|cp-1"]["bitrix_contour"] == "cloud"
    assert state["tasks"]["task|box|finance|2026-03-20|cp-1"]["bitrix_contour"] == "box"


def test_sync_management_tasks_box_requires_explicit_box_user_id(tmp_path: Path) -> None:
    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [_task_payload(dedupe_key="finance|2026-03-20|cp-1")]}

    with pytest.raises(RuntimeError, match="owner_code=finance.*contour=box"):
        sync_management_tasks(
            fetch_json=fetch_json,
            webhook_url=None,
            anchor_date=date(2026, 3, 20),
            state_path=tmp_path / "state.json",
            owner_overrides={},
            watcher_overrides={},
            team_roles={
                "cfo": {
                    "legacy": 10105,
                    "cloud": 10105,
                    "box": None,
                }
            },
            default_responsible_id=None,
            default_observer_ids=[],
            default_created_by_id=None,
            report_dir=tmp_path / "reports",
            disk_folder_id=None,
            dry_run=False,
            delivery_targets=[
                {
                    "contour": "box",
                    "mode": "primary",
                    "webhook_url": "https://bitrix-box.example/rest/115204/token",
                    "owner_overrides": {},
                    "watcher_overrides": {},
                    "default_responsible_id": None,
                    "default_observer_ids": [],
                    "default_created_by_id": None,
                    "disk_folder_id": None,
                }
            ],
            migration_state="box_primary",
        )


def test_sync_management_tasks_box_does_not_use_default_structured_override(tmp_path: Path) -> None:
    created: list[str] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [_task_payload(dedupe_key="finance|2026-03-20|cp-1")]}

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append(webhook_url)
        return 901

    with pytest.raises(RuntimeError, match="owner_code=finance.*contour=box"):
        sync_management_tasks(
            fetch_json=fetch_json,
            webhook_url=None,
            anchor_date=date(2026, 3, 20),
            state_path=tmp_path / "state.json",
            owner_overrides={"finance": {"default": 10105}},
            watcher_overrides={},
            team_roles={"cfo": {"legacy": 10105, "cloud": 10105, "box": None}},
            default_responsible_id=None,
            default_observer_ids=[],
            default_created_by_id=None,
            report_dir=tmp_path / "reports",
            disk_folder_id=None,
            dry_run=False,
            create_task=create_task,
            delivery_targets=[
                {
                    "contour": "box",
                    "mode": "primary",
                    "webhook_url": "https://bitrix-box.example/rest/115204/token",
                    "owner_overrides": {"finance": {"default": 10105}},
                    "watcher_overrides": {},
                    "default_responsible_id": None,
                    "default_observer_ids": [],
                    "default_created_by_id": None,
                    "disk_folder_id": None,
                }
            ],
            migration_state="box_primary",
        )

    assert created == []


def test_sync_management_tasks_box_does_not_use_legacy_scalar_team_role(tmp_path: Path) -> None:
    created: list[str] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [_task_payload(dedupe_key="finance|2026-03-20|cp-1")]}

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append(webhook_url)
        return 902

    with pytest.raises(RuntimeError, match="owner_code=finance.*contour=box"):
        sync_management_tasks(
            fetch_json=fetch_json,
            webhook_url=None,
            anchor_date=date(2026, 3, 20),
            state_path=tmp_path / "state.json",
            owner_overrides={},
            watcher_overrides={},
            team_roles={"cfo": 10105},
            default_responsible_id=None,
            default_observer_ids=[],
            default_created_by_id=None,
            report_dir=tmp_path / "reports",
            disk_folder_id=None,
            dry_run=False,
            create_task=create_task,
            delivery_targets=[
                {
                    "contour": "box",
                    "mode": "primary",
                    "webhook_url": "https://bitrix-box.example/rest/115204/token",
                    "owner_overrides": {},
                    "watcher_overrides": {},
                    "default_responsible_id": None,
                    "default_observer_ids": [],
                    "default_created_by_id": None,
                    "disk_folder_id": None,
                }
            ],
            migration_state="box_primary",
        )

    assert created == []


def test_sync_management_tasks_falls_back_to_cloud_when_box_fails(tmp_path: Path) -> None:
    created: list[str] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [_task_payload(dedupe_key="finance|2026-03-20|cp-1")]}

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        if "bitrix-box" in webhook_url:
            raise FallbackableTaskSyncError("box unavailable before durable side effect")
        created.append(webhook_url)
        return 900

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url=None,
        anchor_date=date(2026, 3, 20),
        state_path=tmp_path / "state.json",
        owner_overrides={},
        watcher_overrides={},
        team_roles={
            "cfo": {
                "legacy": 10105,
                "cloud": 10105,
                "box": 115204,
            }
        },
        default_responsible_id=None,
        default_observer_ids=[],
        default_created_by_id=None,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        dry_run=False,
        create_task=create_task,
        delivery_targets=[
            {
                "contour": "box",
                "mode": "primary",
                "webhook_url": "https://bitrix-box.example/rest/115204/token",
                "owner_overrides": {},
                "watcher_overrides": {},
                "default_responsible_id": 115204,
                "default_observer_ids": [],
                "default_created_by_id": None,
                "disk_folder_id": None,
                "fallback": {
                    "contour": "cloud",
                    "mode": "fallback",
                    "webhook_url": "https://bitrix-cloud.example/rest/1/token",
                    "owner_overrides": {},
                    "watcher_overrides": {},
                    "default_responsible_id": 10105,
                    "default_observer_ids": [],
                    "default_created_by_id": None,
                    "disk_folder_id": None,
                },
            }
        ],
        migration_state="cloud_fallback",
    )

    assert created == ["https://bitrix-cloud.example/rest/1/token"]
    assert summary["created"] == 1
    assert summary["actions"][0]["action"] == "fallback_create"
    assert summary["actions"][0]["fallback_from_contour"] == "box"
    assert summary["actions"][0]["bitrix_contour"] == "cloud"

    state = json.loads((tmp_path / "state.json").read_text(encoding="utf-8"))
    assert "finance|2026-03-20|cp-1" in state["tasks"]
    assert state["tasks"]["finance|2026-03-20|cp-1"]["bitrix_contour"] == "cloud"


def test_sync_management_tasks_falls_back_to_cloud_with_default_create_task(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    calls: list[tuple[str, str]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [_task_payload(dedupe_key="finance|2026-03-20|cp-1")]}

    def fake_b24_call(
        base_url: str, method: str, params: list[tuple[str, str]]
    ) -> dict[str, object]:
        calls.append((base_url, method))
        if "bitrix-box" in base_url:
            raise urllib.error.URLError(ConnectionRefusedError("connection refused"))
        assert method == "tasks.task.add"
        return {"result": 901}

    monkeypatch.setattr(management_tasks_module, "_b24_call", fake_b24_call)

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url=None,
        anchor_date=date(2026, 3, 20),
        state_path=tmp_path / "state.json",
        owner_overrides={},
        watcher_overrides={},
        team_roles={
            "cfo": {
                "legacy": 10105,
                "cloud": 10105,
                "box": 115204,
            }
        },
        default_responsible_id=None,
        default_observer_ids=[],
        default_created_by_id=None,
        report_dir=tmp_path / "reports",
        disk_folder_id=None,
        dry_run=False,
        delivery_targets=[
            {
                "contour": "box",
                "mode": "primary",
                "webhook_url": "https://bitrix-box.example/rest/115204/token",
                "owner_overrides": {},
                "watcher_overrides": {},
                "default_responsible_id": 115204,
                "default_observer_ids": [],
                "default_created_by_id": None,
                "disk_folder_id": None,
                "fallback": {
                    "contour": "cloud",
                    "mode": "fallback",
                    "webhook_url": "https://bitrix-cloud.example/rest/1/token",
                    "owner_overrides": {},
                    "watcher_overrides": {},
                    "default_responsible_id": 10105,
                    "default_observer_ids": [],
                    "default_created_by_id": None,
                    "disk_folder_id": None,
                },
            }
        ],
        migration_state="cloud_fallback",
    )

    assert calls == [
        ("https://bitrix-box.example/rest/115204/token", "tasks.task.add"),
        ("https://bitrix-cloud.example/rest/1/token", "tasks.task.add"),
    ]
    assert summary["created"] == 1
    assert summary["actions"][0]["action"] == "fallback_create"
    assert summary["actions"][0]["bitrix_contour"] == "cloud"


def test_sync_management_tasks_does_not_fallback_on_box_mapping_error(tmp_path: Path) -> None:
    created: list[str] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [_task_payload(dedupe_key="finance|2026-03-20|cp-1")]}

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append(webhook_url)
        return 901

    with pytest.raises(RuntimeError, match="owner_code=finance.*contour=box"):
        sync_management_tasks(
            fetch_json=fetch_json,
            webhook_url=None,
            anchor_date=date(2026, 3, 20),
            state_path=tmp_path / "state.json",
            owner_overrides={},
            watcher_overrides={},
            team_roles={
                "cfo": {
                    "legacy": 10105,
                    "cloud": 10105,
                    "box": None,
                }
            },
            default_responsible_id=None,
            default_observer_ids=[],
            default_created_by_id=None,
            report_dir=tmp_path / "reports",
            disk_folder_id=None,
            dry_run=False,
            create_task=create_task,
            delivery_targets=[
                {
                    "contour": "box",
                    "mode": "primary",
                    "webhook_url": "https://bitrix-box.example/rest/115204/token",
                    "owner_overrides": {},
                    "watcher_overrides": {},
                    "default_responsible_id": None,
                    "default_observer_ids": [],
                    "default_created_by_id": None,
                    "disk_folder_id": None,
                    "fallback": {
                        "contour": "cloud",
                        "mode": "fallback",
                        "webhook_url": "https://bitrix-cloud.example/rest/1/token",
                        "owner_overrides": {},
                        "watcher_overrides": {},
                        "default_responsible_id": 10105,
                        "default_observer_ids": [],
                        "default_created_by_id": None,
                        "disk_folder_id": None,
                    },
                }
            ],
            migration_state="cloud_fallback",
        )

    assert created == []


def test_sync_management_tasks_does_not_fallback_after_uncertain_box_create(
    tmp_path: Path,
) -> None:
    created: list[str] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [_task_payload(dedupe_key="finance|2026-03-20|cp-1")]}

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append(webhook_url)
        raise RuntimeError("timeout during create")

    with pytest.raises(RuntimeError, match="create result is uncertain"):
        sync_management_tasks(
            fetch_json=fetch_json,
            webhook_url=None,
            anchor_date=date(2026, 3, 20),
            state_path=tmp_path / "state.json",
            owner_overrides={},
            watcher_overrides={},
            team_roles={
                "cfo": {
                    "legacy": 10105,
                    "cloud": 10105,
                    "box": 115204,
                }
            },
            default_responsible_id=None,
            default_observer_ids=[],
            default_created_by_id=None,
            report_dir=tmp_path / "reports",
            disk_folder_id=None,
            dry_run=False,
            create_task=create_task,
            delivery_targets=[
                {
                    "contour": "box",
                    "mode": "primary",
                    "webhook_url": "https://bitrix-box.example/rest/115204/token",
                    "owner_overrides": {},
                    "watcher_overrides": {},
                    "default_responsible_id": 115204,
                    "default_observer_ids": [],
                    "default_created_by_id": None,
                    "disk_folder_id": None,
                    "fallback": {
                        "contour": "cloud",
                        "mode": "fallback",
                        "webhook_url": "https://bitrix-cloud.example/rest/1/token",
                        "owner_overrides": {},
                        "watcher_overrides": {},
                        "default_responsible_id": 10105,
                        "default_observer_ids": [],
                        "default_created_by_id": None,
                        "disk_folder_id": None,
                    },
                }
            ],
            migration_state="cloud_fallback",
        )

    assert created == ["https://bitrix-box.example/rest/115204/token"]


def test_sync_management_tasks_does_not_recreate_after_update_error(tmp_path: Path) -> None:
    payload = _task_payload(
        dedupe_key="finance|2026-03-20|cp-1",
        title="Updated title",
        summary="Updated summary",
    )
    state_path = tmp_path / "state.json"
    state_path.write_text(
        json.dumps(
            {
                "tasks": {
                    "finance|2026-03-20|cp-1": {
                        "task_id": 501,
                        "fingerprint": "old",
                        "rule_code": "custom_rule",
                    }
                }
            }
        ),
        encoding="utf-8",
    )
    created: list[str] = []
    updated: list[int] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [payload]}

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append(webhook_url)
        return 902

    def update_task(
        *,
        webhook_url: str,
        task_id: int,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> None:
        updated.append(task_id)
        raise RuntimeError("timeout after update")

    with pytest.raises(RuntimeError, match="Bitrix task update failed"):
        sync_management_tasks(
            fetch_json=fetch_json,
            webhook_url="https://bitrix.example/rest/1/token",
            anchor_date=date(2026, 3, 20),
            state_path=state_path,
            owner_overrides={},
            watcher_overrides={},
            team_roles={"cfo": 10105},
            default_responsible_id=10105,
            default_observer_ids=[],
            default_created_by_id=None,
            report_dir=tmp_path / "reports",
            disk_folder_id=None,
            dry_run=False,
            create_task=create_task,
            update_task=update_task,
        )

    assert updated == [501]
    assert created == []


def test_sync_management_tasks_attach_report_uploads_missing_batch_attachment(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = {
        **_task_payload(
            dedupe_key="receivable_finance_daily_batch|2026-03-20",
            owner_code="finance",
            title="Дебиторка: единый финансовый пакет за 2026-03-20",
            summary="Единый daily-пакет по дебиторке.",
        ),
        "rule_code": "receivable_finance_daily_batch",
        "metrics": {
            "counterparty_count": 1,
            "current_balance_total": "12000.00",
            "sections": {
                "receivable_new_daily_batch": {
                    "counterparty_count": 1,
                    "current_balance_total": "12000.00",
                }
            },
        },
        "references": [
            {
                "batch_rule_code": "receivable_new_daily_batch",
                "counterparty_name": "Контрагент 1",
                "counterparty_ref": "cp-1",
                "current_balance": "12000",
                "original_task_key": "receivable_new_daily|2026-03-20|cp-1",
            }
        ],
    }
    state_path = tmp_path / "state.json"
    state_path.write_text(
        json.dumps(
            {
                "tasks": {
                    "receivable_finance_daily_batch|2026-03-20": {
                        "task_id": 501,
                        "fingerprint": _fingerprint_payload(payload),
                        "rule_code": "receivable_finance_daily_batch",
                    }
                }
            }
        ),
        encoding="utf-8",
    )

    uploaded: list[tuple[str, int, str]] = []
    attached: list[tuple[str, int, int]] = []

    monkeypatch.setattr(
        management_tasks_module,
        "_upload_b24_disk_file",
        lambda *, webhook_url, folder_id, file_path: uploaded.append(
            (webhook_url, folder_id, str(file_path))
        )
        or 901,
    )
    monkeypatch.setattr(
        management_tasks_module,
        "_attach_b24_file_to_task",
        lambda *, webhook_url, task_id, file_object_id: attached.append(
            (webhook_url, task_id, file_object_id)
        )
        or 902,
    )

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [payload]}

    summary = sync_management_tasks(
        fetch_json=fetch_json,
        webhook_url="https://bitrix.example/rest/1/token",
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        owner_overrides={},
        watcher_overrides={},
        team_roles={"cfo": 10105},
        default_responsible_id=10105,
        default_observer_ids=[],
        default_created_by_id=None,
        report_dir=tmp_path / "reports",
        disk_folder_id=3,
        dry_run=False,
    )

    assert summary["created"] == 0
    assert summary["updated"] == 1
    assert summary["actions"][0]["action"] == "attach_report"
    assert uploaded == [
        (
            "https://bitrix.example/rest/1/token",
            3,
            str(
                tmp_path
                / "reports"
                / "cloud"
                / "2026-03-20"
                / "receivable-finance-daily-2026-03-20.xlsx"
            ),
        )
    ]
    assert attached == [("https://bitrix.example/rest/1/token", 501, 901)]

    state = json.loads(state_path.read_text(encoding="utf-8"))
    task_state = state["tasks"]["receivable_finance_daily_batch|2026-03-20"]
    assert task_state["attachment_id"] == 902
    assert task_state["disk_object_id"] == 901


def test_sync_management_tasks_does_not_fallback_after_partial_box_success(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    created: list[str] = []

    payload = {
        **_task_payload(
            dedupe_key="receivable_finance_daily_batch|2026-03-20",
            owner_code="finance",
            title="Дебиторка: единый финансовый пакет за 2026-03-20",
            summary="Единый daily-пакет по дебиторке.",
        ),
        "rule_code": "receivable_finance_daily_batch",
        "metrics": {
            "counterparty_count": 1,
            "current_balance_total": "12000.00",
            "sections": {
                "receivable_new_daily_batch": {
                    "counterparty_count": 1,
                    "current_balance_total": "12000.00",
                }
            },
        },
        "references": [
            {
                "batch_rule_code": "receivable_new_daily_batch",
                "counterparty_name": "Контрагент 1",
                "counterparty_ref": "cp-1",
                "current_balance": "12000",
                "original_task_key": "receivable_new_daily|2026-03-20|cp-1",
            }
        ],
    }
    state_path = tmp_path / "state.json"

    monkeypatch.setattr(
        management_tasks_module,
        "_upload_b24_disk_file",
        lambda *, webhook_url, folder_id, file_path: (_ for _ in ()).throw(
            RuntimeError("disk unavailable")
        ),
    )

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        return {"payload": [payload]}

    def create_task(
        *,
        webhook_url: str,
        payload: dict[str, object],
        assignee_id: int,
        observer_ids: list[int],
        created_by_id: int | None = None,
    ) -> int:
        created.append(webhook_url)
        return 800 + len(created)

    with pytest.raises(RuntimeError, match="Batch attachment failed after task sync"):
        sync_management_tasks(
            fetch_json=fetch_json,
            webhook_url=None,
            anchor_date=date(2026, 3, 20),
            state_path=state_path,
            owner_overrides={},
            watcher_overrides={},
            team_roles={
                "cfo": {
                    "legacy": 10105,
                    "cloud": 10105,
                    "box": 115204,
                }
            },
            default_responsible_id=None,
            default_observer_ids=[],
            default_created_by_id=None,
            report_dir=tmp_path / "reports",
            disk_folder_id=None,
            dry_run=False,
            create_task=create_task,
            delivery_targets=[
                {
                    "contour": "box",
                    "mode": "primary",
                    "webhook_url": "https://bitrix-box.example/rest/115204/token",
                    "owner_overrides": {},
                    "watcher_overrides": {},
                    "default_responsible_id": 115204,
                    "default_observer_ids": [],
                    "default_created_by_id": None,
                    "disk_folder_id": 3,
                    "fallback": {
                        "contour": "cloud",
                        "mode": "fallback",
                        "webhook_url": "https://bitrix-cloud.example/rest/1/token",
                        "owner_overrides": {},
                        "watcher_overrides": {},
                        "default_responsible_id": 10105,
                        "default_observer_ids": [],
                        "default_created_by_id": None,
                        "disk_folder_id": 3,
                    },
                }
            ],
            migration_state="cloud_fallback",
        )

    assert created == ["https://bitrix-box.example/rest/115204/token"]
    state = json.loads(state_path.read_text(encoding="utf-8"))
    task_state = state["tasks"]["task|box|receivable_finance_daily_batch|2026-03-20"]
    assert task_state["task_id"] == 801
    assert task_state["bitrix_contour"] == "box"
    assert "attachment_id" not in task_state
