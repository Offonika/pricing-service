from __future__ import annotations

import json
import urllib.error
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from infra.cron.new_daily_receivables_from_a import (
    _collect_recent_anchor_dates,
    _resolve_chat_ids,
    render_summary,
    sync_new_daily_receivables_report,
)


def _payload(
    *, counterparty_name: str = "Контрагент 1", balance: str = "12000"
) -> dict[str, object]:
    return {
        "freshness_status": "fresh",
        "source_status": "ready",
        "payload": [
            {
                "counterparty_ref": "cp-1",
                "counterparty_name": counterparty_name,
                "current_balance": balance,
                "origin_document_ref": "doc-1",
                "origin_document_number": "РТ-1",
                "origin_document_date": "2026-03-16T09:30:00",
                "current_manager_name": "Менеджер 1",
                "planned_payment_date": "2026-03-21T00:00:00",
                "due_date": "2026-03-21T00:00:00",
                "overdue_days": 0,
            }
        ],
    }


def test_sync_new_daily_receivables_report_noops_when_payload_is_empty(tmp_path: Path) -> None:
    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        assert path == "/api/receivables/new-daily"
        assert params == {"date": "2026-03-20"}
        return {"freshness_status": "fresh", "source_status": "ready", "payload": []}

    summary = sync_new_daily_receivables_report(
        fetch_json=fetch_json,
        deliver_report=lambda **kwargs: {"sent_count": 1},
        anchor_date=date(2026, 3, 20),
        state_path=tmp_path / "state.json",
        artifact_dir=tmp_path / "artifacts",
    )

    assert summary["status"] == "ok"
    assert summary["noop"] == 1
    assert summary["actions"][0]["action"] == "noop_empty"


def test_sync_new_daily_receivables_report_marks_pending_when_source_not_ready(
    tmp_path: Path,
) -> None:
    calls: list[tuple[str, dict[str, str]]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        calls.append((path, params))
        if path == "/api/receivables/new-daily":
            return {"freshness_status": "missing", "source_status": "empty", "payload": []}
        assert path == "/api/management/health"
        return {
            "components": [
                {
                    "component": "receivables",
                    "source_status": "loading",
                    "latest_snapshot_date": "2026-03-19",
                }
            ]
        }

    state_path = tmp_path / "state.json"
    summary = sync_new_daily_receivables_report(
        fetch_json=fetch_json,
        deliver_report=lambda **kwargs: {"sent_count": 1},
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        artifact_dir=tmp_path / "artifacts",
        pending_retry_attempts=0,
    )

    assert summary["status"] == "pending"
    assert summary["actions"][0]["action"] == "pending_source"
    state = json.loads(state_path.read_text(encoding="utf-8"))
    pending = state["reports"]["receivable-new-daily|2026-03-20|pending"]
    assert pending["delivery_status"] == "pending_source"
    assert calls == [
        ("/api/receivables/new-daily", {"date": "2026-03-20"}),
        ("/api/management/health", {"date": "2026-03-20"}),
    ]


def test_sync_new_daily_receivables_report_deduplicates_by_revision(tmp_path: Path) -> None:
    state_path = tmp_path / "state.json"
    delivered: list[dict[str, object]] = []

    initial = sync_new_daily_receivables_report(
        fetch_json=lambda path, params: _payload(),
        deliver_report=lambda **kwargs: delivered.append(kwargs) or {"sent_count": 2},
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        artifact_dir=tmp_path / "artifacts",
    )
    repeated = sync_new_daily_receivables_report(
        fetch_json=lambda path, params: _payload(),
        deliver_report=lambda **kwargs: delivered.append(kwargs) or {"sent_count": 2},
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        artifact_dir=tmp_path / "artifacts",
    )

    assert initial["status"] == "ok"
    assert initial["delivered"] == 1
    assert repeated["status"] == "ok"
    assert repeated["noop"] == 1
    assert len(delivered) == 1


def test_sync_new_daily_receivables_report_exports_formatted_workbook(tmp_path: Path) -> None:
    delivered: list[dict[str, object]] = []

    summary = sync_new_daily_receivables_report(
        fetch_json=lambda path, params: _payload(balance="125000"),
        deliver_report=lambda **kwargs: delivered.append(kwargs) or {"sent_count": 1},
        anchor_date=date(2026, 3, 20),
        state_path=tmp_path / "state.json",
        artifact_dir=tmp_path / "artifacts",
    )

    assert summary["status"] == "ok"
    assert len(delivered) == 1
    assert delivered[0]["report_path"].name == "Новая_дебиторка_2026-03-20.xlsx"

    workbook = load_workbook(delivered[0]["report_path"])
    assert workbook.sheetnames == ["Сводка", "Новая дебиторка"]

    summary_sheet = workbook["Сводка"]
    assert summary_sheet["A1"].value == "Утренняя новая дебиторка"
    assert summary_sheet.freeze_panes == "A2"
    assert summary_sheet["A2"].value == "Дата отчета"
    assert summary_sheet["B5"].number_format == "#,##0.00"
    assert summary_sheet["A12"].value == "Топ контрагентов по сумме"

    details_sheet = workbook["Новая дебиторка"]
    assert details_sheet["A1"].value == "Новая дебиторка на 20.03.2026"
    assert details_sheet["A3"].value == "№"
    assert [details_sheet.cell(row=3, column=index).value for index in range(1, 11)] == [
        "№",
        "Контрагент",
        "Сумма долга, ₽",
        "Возраст долга, дн",
        "Исходный документ",
        "Дата возникновения",
        "Текущий менеджер",
        "План оплаты",
        "Due date",
        "Просрочка, дн",
    ]
    assert details_sheet.freeze_panes == "A4"
    assert details_sheet.auto_filter.ref == "A3:J4"
    assert details_sheet["C4"].number_format == "#,##0.00"
    assert details_sheet["A5"].value == "Итого"
    assert details_sheet["C5"].value == "=SUM(C4:C4)"


def test_sync_new_daily_receivables_report_delivers_correction_once(tmp_path: Path) -> None:
    state_path = tmp_path / "state.json"
    delivered: list[dict[str, object]] = []

    first = sync_new_daily_receivables_report(
        fetch_json=lambda path, params: _payload(counterparty_name="Контрагент 1", balance="12000"),
        deliver_report=lambda **kwargs: delivered.append(kwargs) or {"sent_count": 2},
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        artifact_dir=tmp_path / "artifacts",
    )
    second = sync_new_daily_receivables_report(
        fetch_json=lambda path, params: _payload(counterparty_name="Контрагент 2", balance="15000"),
        deliver_report=lambda **kwargs: delivered.append(kwargs) or {"sent_count": 2},
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        artifact_dir=tmp_path / "artifacts",
    )

    assert first["status"] == "ok"
    assert second["status"] == "ok"
    assert second["delivered"] == 1
    assert len(delivered) == 2
    assert delivered[1]["is_correction"] is True

    state = json.loads(state_path.read_text(encoding="utf-8"))
    delivered_reports = [
        item for item in state["reports"].values() if item["delivery_status"] == "delivered"
    ]
    assert len(delivered_reports) == 2


def test_sync_new_daily_receivables_report_delivers_after_pending_as_correction(
    tmp_path: Path,
) -> None:
    state_path = tmp_path / "state.json"
    state_path.write_text(
        json.dumps(
            {
                "reports": {
                    "receivable-new-daily|2026-03-20|pending": {
                        "report_key": "receivable-new-daily|2026-03-20",
                        "anchor_date": "2026-03-20",
                        "delivery_status": "pending_source",
                    }
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    delivered: list[dict[str, object]] = []

    summary = sync_new_daily_receivables_report(
        fetch_json=lambda path, params: _payload(counterparty_name="Контрагент 2", balance="15000"),
        deliver_report=lambda **kwargs: delivered.append(kwargs) or {"sent_count": 2},
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        artifact_dir=tmp_path / "artifacts",
    )

    assert summary["status"] == "ok"
    assert summary["delivered"] == 1
    assert delivered[0]["is_correction"] is True
    state = json.loads(state_path.read_text(encoding="utf-8"))
    assert "receivable-new-daily|2026-03-20|pending" not in state["reports"]


def test_sync_new_daily_receivables_report_returns_error_when_source_unavailable(
    tmp_path: Path,
) -> None:
    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        raise urllib.error.HTTPError(
            url="http://example.test",
            code=401,
            msg="Unauthorized",
            hdrs=None,
            fp=None,
        )

    summary = sync_new_daily_receivables_report(
        fetch_json=fetch_json,
        deliver_report=lambda **kwargs: {"sent_count": 1},
        anchor_date=date(2026, 3, 20),
        state_path=tmp_path / "state.json",
        artifact_dir=tmp_path / "artifacts",
    )

    assert summary["status"] == "error"
    assert summary["failed"] == 1


def test_sync_new_daily_receivables_report_marks_pending_when_source_error_is_retryable(
    tmp_path: Path,
) -> None:
    state_path = tmp_path / "state.json"

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        raise urllib.error.URLError("server-a down")

    summary = sync_new_daily_receivables_report(
        fetch_json=fetch_json,
        deliver_report=lambda **kwargs: {"sent_count": 1},
        anchor_date=date(2026, 3, 20),
        state_path=state_path,
        artifact_dir=tmp_path / "artifacts",
        pending_retry_attempts=0,
    )

    assert summary["status"] == "pending"
    assert summary["actions"][0]["action"] == "pending_source_error"
    state = json.loads(state_path.read_text(encoding="utf-8"))
    pending = state["reports"]["receivable-new-daily|2026-03-20|pending"]
    assert pending["delivery_status"] == "pending_source"


def test_resolve_chat_ids_prefers_dedicated_value() -> None:
    env = {
        "MANAGEMENT_NEW_DAILY_TELEGRAM_CHAT_ID": "1001,1002",
        "WEEKLY_BUYER_DIGEST_ALERT_TELEGRAM_CHAT_ID": "2001",
    }

    assert _resolve_chat_ids(env) == ["1001", "1002"]


def test_collect_recent_anchor_dates_includes_today_and_lookback() -> None:
    assert _collect_recent_anchor_dates(today=date(2026, 3, 20), lookback_days=3) == [
        date(2026, 3, 17),
        date(2026, 3, 18),
        date(2026, 3, 19),
        date(2026, 3, 20),
    ]


def test_render_summary_includes_status_line() -> None:
    summary = {
        "status": "ok",
        "anchor_date": "2026-03-20",
        "freshness_status": "fresh",
        "source_status": "ready",
        "fetched": 1,
        "delivered": 1,
        "noop": 0,
        "failed": 0,
        "sent_documents": 2,
        "actions": [
            {
                "action": "deliver",
                "report_key": "receivable-new-daily|2026-03-20",
                "revision": "abc123",
                "is_correction": False,
            }
        ],
    }

    rendered = render_summary(summary)

    assert "new_daily_receivables_from_a: ok" in rendered
    assert "sent_documents: 2" in rendered
    assert "revision=abc123" in rendered
