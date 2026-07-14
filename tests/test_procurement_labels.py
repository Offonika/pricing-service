from __future__ import annotations

import zipfile
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.core.config import Settings
from app.schemas.procurement_labels import ProcurementLabelOrderPreview, ProcurementLabelRow
from app.services import procurement_labels


def _mapping() -> dict:
    return {
        "process": {"entity_type_id": 1056},
        "category_map": {"ved_import": {"id": 12}, "ordinary": {"id": 52}},
        "field_map": {
            "procurement_contour": "UF_CRM_8_PROCUREMENTCONTOUR",
            "onec_source_number": "UF_CRM_8_ONECSOURCENUMBER",
        },
        "enum_map": {"procurement_contour": {"ved_import": "369", "ordinary": "367"}},
    }


def _line(**overrides) -> dict:
    row = {
        "line_no": 1,
        "onec_item_code": "РБ000046282",
        "item_name": "Аккумулятор для Apple iPhone 11 F5ENERGY",
        "article_1c": "052841",
        "sku": "OEM-BAT-IPH11-3470-HC-SP-F5",
        "unit": "шт",
        "quantity": Decimal("160"),
        "price": Decimal("38"),
        "amount": Decimal("6080"),
    }
    row.update(overrides)
    return row


def _cert(**overrides) -> dict:
    row = {
        "certificate_id": "ДС-TEST-1",
        "number": "ЕАЭС N RU Д-CN.TEST/26",
        "status": "covered",
        "valid_to": "2031-04-23",
        "file_id": "22316",
        "eac": True,
    }
    row.update(overrides)
    return row


def _ready_preview() -> ProcurementLabelOrderPreview:
    return ProcurementLabelOrderPreview(
        item_id="134",
        entity_type_id=1056,
        onec_number="РБГУ0000377",
        title="ВЭД импорт РБГУ0000377",
        contour="ved_import",
        status="draft",
        ready=True,
        blocked=False,
        rows=[
            ProcurementLabelRow(
                line_no=1,
                onec_item_code="РБ000046282",
                item_name="Аккумулятор для Apple iPhone 11 F5ENERGY",
                article_1c="052841",
                sku="OEM-BAT-IPH11-3470-HC-SP-F5",
                barcode="4601234567892",
                unit="шт",
                quantity=Decimal("160"),
                price=Decimal("38"),
                amount=Decimal("6080"),
                certificate_id="ДС-TEST-1",
                certificate_status="covered",
                eac_allowed=True,
                status="ready",
                blockers=[],
            )
        ],
    )


def test_non_ved_order_is_blocked() -> None:
    preview = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 52,
            "title": "Закупка РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[_line()],
        mapping=_mapping(),
        barcode_catalog={"РБ000046282": "4601234567892"},
        certificate_catalog={"РБ000046282": _cert()},
    )

    assert preview.blocked is True
    assert procurement_labels.BLOCK_NON_VED in preview.blockers


def test_missing_barcode_blocks_zip_generation() -> None:
    preview = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 12,
            "title": "ВЭД импорт РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[_line()],
        mapping=_mapping(),
        certificate_catalog={"РБ000046282": _cert()},
    )

    assert preview.blocked is True
    assert procurement_labels.BLOCK_MISSING_BARCODE in preview.rows[0].blockers
    with pytest.raises(ValueError):
        procurement_labels.build_artifacts(preview, artifact_dir=Path("/tmp/noop"))


def test_onec_barcode_from_order_line_is_used_without_catalog() -> None:
    preview = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 12,
            "title": "ВЭД импорт РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[_line(barcode="2900000536753")],
        mapping=_mapping(),
        certificate_catalog={"РБ000046282": _cert()},
    )

    assert preview.blocked is False
    assert preview.rows[0].barcode == "2900000536753"
    assert preview.rows[0].barcode_source == "1c_internal"


def test_article_and_sku_are_kept_separate() -> None:
    preview = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 12,
            "title": "ВЭД импорт РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[
            _line(
                article_1c="052841",
                sku="OEM-BAT-IPH11-3470-HC-SP-F5",
                barcode="2900000536753",
            )
        ],
        mapping=_mapping(),
        certificate_catalog={"OEM-BAT-IPH11-3470-HC-SP-F5": _cert()},
    )

    assert preview.blocked is False
    assert preview.rows[0].article_1c == "052841"
    assert preview.rows[0].sku == "OEM-BAT-IPH11-3470-HC-SP-F5"
    assert preview.rows[0].certificate_status == "covered"


def test_eac_is_allowed_only_with_confirmed_certificate() -> None:
    covered = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 12,
            "title": "ВЭД импорт РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[_line()],
        mapping=_mapping(),
        barcode_catalog={"РБ000046282": "4601234567892"},
        certificate_catalog={"РБ000046282": _cert(eac=True)},
    )
    missing = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 12,
            "title": "ВЭД импорт РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[_line()],
        mapping=_mapping(),
        barcode_catalog={"РБ000046282": "4601234567892"},
        certificate_catalog={},
    )

    assert covered.rows[0].eac_allowed is True
    assert covered.blocked is False
    assert missing.rows[0].eac_allowed is False
    assert procurement_labels.BLOCK_MISSING_CERTIFICATE in missing.rows[0].blockers


def test_expired_certificate_blocks_zip() -> None:
    preview = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 12,
            "title": "ВЭД импорт РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[_line(barcode="2900000536753")],
        mapping=_mapping(),
        certificate_catalog={"OEM-BAT-IPH11-3470-HC-SP-F5": _cert(valid_to="2020-01-01")},
    )

    assert preview.blocked is True
    assert procurement_labels.BLOCK_EXPIRED_CERTIFICATE in preview.rows[0].blockers


def test_missing_certificate_file_blocks_zip() -> None:
    preview = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 12,
            "title": "ВЭД импорт РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[_line(barcode="2900000536753")],
        mapping=_mapping(),
        certificate_catalog={"OEM-BAT-IPH11-3470-HC-SP-F5": _cert(file_id="")},
    )

    assert preview.blocked is True
    assert procurement_labels.BLOCK_MISSING_CERTIFICATE_FILE in preview.rows[0].blockers


def test_multiple_active_certificates_block_zip() -> None:
    preview = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 12,
            "title": "ВЭД импорт РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[_line(barcode="2900000536753")],
        mapping=_mapping(),
        certificate_catalog={
            "OEM-BAT-IPH11-3470-HC-SP-F5": [
                _cert(certificate_id="ДС-1"),
                _cert(certificate_id="ДС-2"),
            ]
        },
    )

    assert preview.blocked is True
    assert procurement_labels.BLOCK_AMBIGUOUS_CERTIFICATE in preview.rows[0].blockers


def test_repeated_generation_creates_new_zip_version(tmp_path: Path) -> None:
    preview = _ready_preview()

    v1, zip1, _ = procurement_labels.build_artifacts(preview, artifact_dir=tmp_path)
    v2, zip2, _ = procurement_labels.build_artifacts(preview, artifact_dir=tmp_path)

    assert v1 == 1
    assert v2 == 2
    assert zip1.exists()
    assert zip2.exists()
    assert zip1 != zip2
    with zipfile.ZipFile(zip2) as archive:
        names = set(archive.namelist())
        register_payload = archive.read("label-register.xlsx")
    assert "label-register.xlsx" in names
    assert "factory-labels-readme.txt" in names
    assert any(name.startswith("labels-preview/") and name.endswith(".png") for name in names)
    workbook = load_workbook(BytesIO(register_payload), data_only=True)
    register_sheet = workbook["label-register"]
    headers = [cell.value for cell in register_sheet[1]]
    values = dict(
        zip(headers, next(register_sheet.iter_rows(min_row=2, values_only=True)), strict=False)
    )
    assert values["article_1c"] == "052841"
    assert values["sku"] == "OEM-BAT-IPH11-3470-HC-SP-F5"


def test_certification_docs_package_contains_master_gtin_and_checklist(tmp_path: Path) -> None:
    preview = _ready_preview()

    version, zip_path, zip_url, gtin_rows, missing_rows = (
        procurement_labels.build_certification_docs_artifacts(
            preview,
            artifact_dir=tmp_path,
            base_url="https://pricing.example",
        )
    )

    assert version == 1
    assert zip_path.name == "ved-certification-docs-РБГУ0000377-v1.zip"
    assert zip_url.endswith(
        "/api/procurement-labels/artifacts/ved-certification-docs-%D0%A0%D0%91%D0%93%D0%A30000377-v1.zip"
    )
    assert gtin_rows == 1
    assert missing_rows == 1
    with zipfile.ZipFile(zip_path) as archive:
        names = set(archive.namelist())
        master_payload = archive.read("мастер-таблица-для-декларации.xlsx")
        gtin_payload = archive.read("gtin-order-list.xlsx")
        checklist = archive.read("certification-documents-checklist.md").decode("utf-8")
    assert {
        "мастер-таблица-для-декларации.xlsx",
        "gtin-order-list.xlsx",
        "certification-documents-checklist.md",
        "broker-request.md",
        "package-readme.txt",
    } <= names
    master = load_workbook(BytesIO(master_payload), data_only=True)["Мастер-реестр"]
    master_values = dict(
        zip(
            [cell.value for cell in master[1]],
            next(master.iter_rows(min_row=2, values_only=True)),
            strict=False,
        )
    )
    assert master_values["Артикул 1С (числовой)"] == "052841"
    assert master_values["SKU 1С (отдельное поле)"] == "OEM-BAT-IPH11-3470-HC-SP-F5"
    assert master_values["UN Code"] == "UN3480"
    assert master_values["Что не заполнено / ошибки"] == (
        "TradeName / семейная формула для ДС; Совместимость; Ёмкость, mAh; "
        "Напряжение, V; Энергия, Wh; Размеры, DIM; BMS / плата защиты; "
        "Коннектор; GTIN/EAN-13; Новая ДС ГОСТ Р; Покрытие новой ДС по SKU"
    )
    assert master_values["Статус проверки"] == "Нет GTIN и новой ДС ГОСТ Р"
    gtin = load_workbook(BytesIO(gtin_payload), data_only=True)["Заказ GTIN"]
    gtin_values = dict(
        zip(
            [cell.value for cell in gtin[1]],
            next(gtin.iter_rows(min_row=2, values_only=True)),
            strict=False,
        )
    )
    assert gtin_values["Действие по GTIN"] == "заказать GTIN/EAN-13"
    assert "GTIN/EAN-13" in checklist


def test_certification_docs_package_is_not_blocked_by_missing_declaration(tmp_path: Path) -> None:
    preview = procurement_labels.build_preview_from_sources(
        item_id="134",
        entity_type_id=1056,
        bitrix_item={
            "categoryId": 12,
            "title": "ВЭД импорт РБГУ0000377",
            "ufCrm8Onecsourcenumber": "РБГУ0000377",
        },
        onec_lines=[_line(barcode="2900000536753")],
        mapping=_mapping(),
        certificate_catalog={},
    )

    assert preview.blocked is True
    assert procurement_labels.BLOCK_MISSING_CERTIFICATE in preview.blockers[0]
    assert procurement_labels.certification_package_blockers(preview) == []
    version, zip_path, _zip_url, _gtin_rows, _missing_rows = (
        procurement_labels.build_certification_docs_artifacts(preview, artifact_dir=tmp_path)
    )
    assert version == 1
    assert zip_path.exists()


def test_certification_docs_status_fields_update_separate_bitrix_fields() -> None:
    mapping = {
        "field_map": {
            "certification_docs_status": "UF_CRM_8_CERTIFICATIONDOCSSTATUS",
            "certification_docs_version": "UF_CRM_8_CERTIFICATIONDOCSVERSION",
            "certification_docs_zip_url": "UF_CRM_8_CERTIFICATIONDOCSZIPURL",
            "certification_docs_disk_file_id": "UF_CRM_8_CERTIFICATIONDOCSDISKFILEID",
            "certification_docs_errors": "UF_CRM_8_CERTIFICATIONDOCSERRORS",
            "certification_docs_generated_at": "UF_CRM_8_CERTIFICATIONDOCSGENERATEDAT",
        },
        "enum_map": {"certification_docs_status": {"needs_data": "501"}},
    }

    fields = procurement_labels.certification_docs_status_fields(
        mapping,
        status="needs_data",
        version=2,
        zip_url="https://example.test/file.zip",
        disk_file_id="777",
        errors=[],
    )

    assert fields["ufCrm8Certificationdocsstatus"] == "501"
    assert fields["ufCrm8Certificationdocsversion"] == 2
    assert fields["ufCrm8Certificationdocszipurl"] == "https://example.test/file.zip"
    assert fields["ufCrm8Certificationdocsdiskfileid"] == "777"
    assert fields["ufCrm8Certificationdocserrors"] == ""
    assert "ufCrm8Certificationdocsgeneratedat" in fields


def test_approve_without_persisted_disk_version_is_blocked(tmp_path: Path) -> None:
    mapping_path = tmp_path / "mapping.json"
    mapping_path.write_text(
        """
        {
          "process": {"entity_type_id": 1056},
          "field_map": {
            "label_generation_version": "UF_CRM_VERSION",
            "label_generation_zip_url": "UF_CRM_ZIP",
            "label_generation_disk_file_id": "UF_CRM_DISK"
          }
        }
        """,
        encoding="utf-8",
    )

    class FakeBitrixClient:
        def get_item(self, *, entity_type_id: int, item_id: str) -> dict:
            return {"id": item_id, "title": "ВЭД импорт"}

    settings = Settings(
        procurement_labels_mapping_path=str(mapping_path),
        onec_database_url="mssql+pyodbc://example",
    )
    with pytest.raises(Exception) as exc:
        procurement_labels.approve_zip("134", settings=settings, bitrix_client=FakeBitrixClient())

    assert "Сначала сформируйте ZIP" in str(exc.value)
