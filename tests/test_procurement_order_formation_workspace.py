from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

import app.services.bitrix_order_formation as bitrix_order_service
from app.core.config import Settings
from app.models.procurement_order_formation import (
    ProcurementLifecycleTransitionProposal,
    ProcurementOrderFormation,
    ProcurementOrderFormationEvent,
    ProcurementOrderFormationLine,
)
from app.services.assortment_lifecycle_classification_store import (
    ASSORTMENT_LIFECYCLE_METADATA,
    build_classification_rows,
    persist_classification_rows,
)
from app.services.bitrix_order_formation import BitrixCatalogProduct
from app.services.bitrix_procurement_order_formation_auth import (
    ProcurementOrderFormationSession,
)
from app.services.procurement_order_formation_workspace import (
    LIFECYCLE_ORDER,
    approve_lifecycle_transitions,
    build_dashboard,
    build_order_calculation_excel,
    list_lifecycle_transitions,
    sync_lifecycle_transition_proposals,
)

DISPLAY_GUID = "2685293e-967c-11e1-bdb9-0025901e48ef"


@pytest.fixture()
def lifecycle_db(sqlite_engine, db_session):
    ASSORTMENT_LIFECYCLE_METADATA.create_all(sqlite_engine)
    try:
        yield db_session
    finally:
        db_session.rollback()
        ASSORTMENT_LIFECYCLE_METADATA.drop_all(sqlite_engine)


def _session(user_id: str, name: str) -> ProcurementOrderFormationSession:
    return ProcurementOrderFormationSession(
        actor=f"bitrix:member:{user_id}",
        domain="crm.example.test",
        member_id="member",
        user_id=user_id,
        expires_at=datetime.now(UTC),
        user_name=name,
    )


def _settings() -> Settings:
    return Settings(
        procurement_order_formation_lifecycle_approver_user_ids=["130757", "4241"],
        procurement_order_formation_display_responsible_user_id="130757",
        procurement_order_formation_property_apply_enabled=False,
    )


def _seed_lifecycle(sqlite_engine) -> int:
    now = datetime(2026, 7, 10, 9, 20)
    records = [
        {
            "nomenclature_code": "FRUIT-1",
            "product_ref": DISPLAY_GUID.replace("-", ""),
            "manual_status": "fruit",
            "created_at": date(2026, 7, 1).isoformat(),
            "first_supplier_order_at": date(2026, 7, 10).isoformat(),
        },
        {
            "nomenclature_code": "NEWBORN-NEED-1",
            "product_ref": DISPLAY_GUID,
            "created_at": date(2026, 7, 1).isoformat(),
            "first_supplier_order_at": date(2026, 7, 8).isoformat(),
            "has_need_signal": True,
        },
        {
            "nomenclature_code": "NEWBORN-1",
            "product_ref": DISPLAY_GUID,
            "manual_status": "newborn",
        },
        {
            "nomenclature_code": "WORKING-1",
            "product_ref": DISPLAY_GUID,
            "manual_status": "working",
        },
        {
            "nomenclature_code": "MATRIX-1",
            "product_ref": DISPLAY_GUID,
            "manual_status": "matrix",
        },
    ]
    summaries = [
        {
            "nomenclature_code": "FRUIT-1",
            "name": "Дисплей Fruit",
            "folder": "дисплеи",
            "status": "newborn",
            "status_label": "Новорожденный",
            "recommended_status": "newborn",
            "reason_text": "Создан первый заказ поставщику",
            "export_blockers": ["fact_status_decision_requires_1c_approval"],
        },
        {
            "nomenclature_code": "NEWBORN-NEED-1",
            "name": "Дисплей ДН",
            "folder": "дисплеи",
            "status": "newborn_need",
            "status_label": "ДН / Добор новорождённого",
            "reason_text": "Есть явная потребность",
            "manual_review_required": True,
        },
        {
            "nomenclature_code": "NEWBORN-1",
            "name": "Дисплей к Новинке",
            "folder": "дисплеи",
            "status": "new_item",
            "status_label": "Новинка",
            "recommended_status": "new_item",
            "reason_text": "Первый груз передан перевозчику",
        },
        {
            "nomenclature_code": "WORKING-1",
            "name": "Дисплей рабочий",
            "folder": "дисплеи",
            "status": "working",
            "status_label": "Рабочий",
            "reason_text": "Нужен пересмотр",
            "manual_review_required": True,
        },
        {
            "nomenclature_code": "MATRIX-1",
            "name": "Дисплей матричный",
            "folder": "дисплеи",
            "status": "matrix",
            "status_label": "Матричный",
            "reason_text": "Основной товар группы",
        },
    ]
    rows = build_classification_rows(
        records=records,
        summaries=summaries,
        source="test",
        classified_at=now,
    )
    result = persist_classification_rows(
        sqlite_engine,
        rows=rows,
        run_key="display-test-run-1",
        folder="дисплеи",
        source="test",
        started_at=now,
        finished_at=now,
    )
    assert result.run_id is not None
    return result.run_id


def _approval_item(proposal: ProcurementLifecycleTransitionProposal) -> dict[str, object]:
    return {
        "proposal_id": proposal.id,
        "expected_run_id": proposal.run_id,
        "expected_current_status": proposal.current_status,
        "facts_hash": proposal.facts_hash,
    }


def test_order_calculation_excel_contains_classification_and_active_filtered_lines(
    lifecycle_db,
    sqlite_engine,
) -> None:
    now = datetime(2026, 8, 1, 9, 30)
    rows = build_classification_rows(
        records=[
            {
                "nomenclature_code": "DISPLAY-1",
                "article": "ART-001",
                "subject_1c": "Запчасти для телефонов",
                "category_1c": "Дисплеи",
            }
        ],
        summaries=[
            {
                "nomenclature_code": "DISPLAY-1",
                "name": "Дисплей тестовый",
                "folder": "Дисплеи Apple",
                "status": "working",
                "status_label": "Рабочий",
            }
        ],
        source="test",
        classified_at=now,
    )
    persist_classification_rows(
        sqlite_engine,
        rows=rows,
        run_key="excel-export-test",
        folder="дисплеи",
        source="test",
        started_at=now,
        finished_at=now,
    )

    active_order = ProcurementOrderFormation(
        stable_key="excel:active",
        status="draft",
        supplier_name="Поставщик Excel",
        contract_name="Основной договор",
        warehouse_name="Центральный склад",
        currency="RUB",
        route="ordinary",
        batch_id="2026-08-01",
        order_date=date(2026, 8, 1),
        calculation_id="excel-calculation",
    )
    active_order.lines = [
        ProcurementOrderFormationLine(
            stable_key="excel:line:active",
            line_number=1,
            bitrix_product_xml_id="display-guid-1",
            nomenclature_ref="display-ref-1",
            nomenclature_code="DISPLAY-1",
            nomenclature_name="Дисплей iPhone тестовый",
            recommended_quantity=Decimal("3"),
            final_quantity=Decimal("3"),
            purchase_price=Decimal("1250.50"),
            amount=Decimal("3751.50"),
            currency="RUB",
        ),
        ProcurementOrderFormationLine(
            stable_key="excel:line:removed",
            line_number=2,
            bitrix_product_xml_id="display-guid-2",
            nomenclature_ref="display-ref-2",
            nomenclature_code="DISPLAY-2",
            nomenclature_name="Удалённая строка",
            final_quantity=Decimal("1"),
            purchase_price=Decimal("10"),
            amount=Decimal("10"),
            currency="RUB",
            removed=True,
        ),
    ]
    superseded_order = ProcurementOrderFormation(
        stable_key="excel:superseded",
        status="superseded",
        supplier_name="Старый поставщик",
        contract_name="Старый договор",
        warehouse_name="Старый склад",
        currency="RUB",
        route="ordinary",
        batch_id="2026-07-31",
        order_date=date(2026, 7, 31),
        calculation_id="old-calculation",
    )
    superseded_order.lines = [
        ProcurementOrderFormationLine(
            stable_key="excel:line:superseded",
            line_number=1,
            bitrix_product_xml_id="old-guid",
            nomenclature_ref="old-ref",
            nomenclature_code="OLD-1",
            nomenclature_name="Строка старого расчёта",
            final_quantity=Decimal("2"),
            purchase_price=Decimal("100"),
            amount=Decimal("200"),
            currency="RUB",
        )
    ]
    lifecycle_db.add_all([active_order, superseded_order])
    lifecycle_db.commit()

    workbook = load_workbook(BytesIO(build_order_calculation_excel(lifecycle_db)))
    worksheet = workbook["Расчёт заказа"]
    exported_rows = list(worksheet.values)

    assert exported_rows[0][:5] == (
        "Предмет",
        "Категория",
        "Группа",
        "Номенклатура",
        "Артикул",
    )
    assert len(exported_rows) == 2
    assert exported_rows[1][:5] == (
        "Запчасти для телефонов",
        "Дисплеи",
        "Дисплеи Apple",
        "Дисплей iPhone тестовый",
        "ART-001",
    )
    assert exported_rows[1][5:15] == (
        "Поставщик Excel",
        "Основной договор",
        "Центральный склад",
        3,
        1250.5,
        3751.5,
        "RUB",
        "На подтверждении",
        "2026-08-01",
        datetime(2026, 8, 1),
    )
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == worksheet.dimensions

    superseded_workbook = load_workbook(
        BytesIO(build_order_calculation_excel(lifecycle_db, status="superseded"))
    )
    superseded_rows = list(superseded_workbook["Расчёт заказа"].values)
    assert len(superseded_rows) == 2
    assert superseded_rows[1][3] == "Строка старого расчёта"


def test_dashboard_keeps_lifecycle_order_and_nests_newborn_need(
    lifecycle_db,
    sqlite_engine,
) -> None:
    run_id = _seed_lifecycle(sqlite_engine)
    summary = sync_lifecycle_transition_proposals(
        lifecycle_db,
        run_id=run_id,
        settings=_settings(),
    )

    dashboard = build_dashboard(lifecycle_db, settings=_settings())

    assert [card["status"] for card in dashboard["cards"]] == list(LIFECYCLE_ORDER)
    newborn = next(card for card in dashboard["cards"] if card["status"] == "newborn")
    fruit = next(card for card in dashboard["cards"] if card["status"] == "fruit")
    working = next(card for card in dashboard["cards"] if card["status"] == "working")
    assert newborn["total_count"] == 2
    assert newborn["action_count"] == 2
    assert newborn["action_breakdown"] == {"new_item": 1, "review": 1}
    assert fruit["action_count"] == 0
    assert working["action_label"] == "На пересмотр"
    assert working["action_count"] == 1
    assert dashboard["manual_status_counts"]["matrix"] == 1
    matrix_attention = next(
        item for item in dashboard["manual_attention"] if item["filter_status"] == "matrix"
    )
    assert matrix_attention == {
        "proposal_id": None,
        "nomenclature_code": "MATRIX-1",
        "product_name": "Дисплей матричный",
        "current_status": "matrix",
        "current_status_label": "Матричный",
        "kind": "manual",
        "filter_status": "matrix",
        "action_label": "Проверить матрицу и минимальный запас",
        "fact_summary": "Основной товар группы",
        "decision_state": "control",
        "decision_state_label": "Контроль",
        "reason": "Основной товар группы",
        "recommendation": "Проверить матрицу и минимальный запас",
        "deadline_label": "Контроль",
        "urgency": "warning",
    }
    assert dashboard["decision_summary"] == {
        "ready_count": 1,
        "review_count": 2,
        "blocked_count": 0,
    }
    assert all(item["deadline_label"] != "сегодня" for item in dashboard["attention"])
    assert all("pricing-service" not in item["fact_summary"] for item in dashboard["attention"])
    assert {item["filter_status"] for item in dashboard["manual_attention"]} == {
        "matrix",
        "review",
    }
    assert summary == {
        "created": 4,
        "updated": 0,
        "automatic": 1,
        "stale": 0,
        "run_id": run_id,
    }


def test_queue_starts_unselected_and_only_ready_rows_are_selectable(
    lifecycle_db,
    sqlite_engine,
) -> None:
    run_id = _seed_lifecycle(sqlite_engine)
    sync_lifecycle_transition_proposals(
        lifecycle_db,
        run_id=run_id,
        settings=_settings(),
    )

    queue = list_lifecycle_transitions(lifecycle_db, status="fruit", scope="action")
    newborn_queue = list_lifecycle_transitions(
        lifecycle_db,
        status="newborn",
        scope="action",
    )

    assert queue["total"] == 0
    assert newborn_queue["total"] == 2
    assert newborn_queue["ready_count"] == 1
    assert sum(1 for item in newborn_queue["items"] if item["selectable"]) == 1
    assert all("selected" not in item for item in newborn_queue["items"])
    review = next(item for item in newborn_queue["items"] if item["action_kind"] == "review")
    transition = next(
        item for item in newborn_queue["items"] if item["action_kind"] == "transition"
    )
    assert review["current_status"] == "newborn"
    assert review["target_status"] is None
    assert transition["reason"].startswith("Рекомендуется переход Заказали → Завезли.")


def test_queue_filters_manual_reviews_and_opens_exact_proposal(
    lifecycle_db,
    sqlite_engine,
) -> None:
    run_id = _seed_lifecycle(sqlite_engine)
    sync_lifecycle_transition_proposals(
        lifecycle_db,
        run_id=run_id,
        settings=_settings(),
    )

    reviews = list_lifecycle_transitions(
        lifecycle_db,
        status="all",
        scope="action",
        readiness="review",
    )

    assert reviews["total"] == 2
    assert reviews["review_count"] == 2
    assert {item["decision_state"] for item in reviews["items"]} == {"review"}

    exact = list_lifecycle_transitions(
        lifecycle_db,
        status="all",
        scope="action",
        readiness="review",
        proposal_id=reviews["items"][0]["proposal_id"],
    )

    assert exact["total"] == 1
    assert exact["items"][0]["proposal_id"] == reviews["items"][0]["proposal_id"]


def test_action_queue_hides_stale_runs_until_archive_filter(
    lifecycle_db,
    sqlite_engine,
) -> None:
    run_id = _seed_lifecycle(sqlite_engine)
    sync_lifecycle_transition_proposals(
        lifecycle_db,
        run_id=run_id,
        settings=_settings(),
    )
    stale = ProcurementLifecycleTransitionProposal(
        nomenclature_code="FRUIT-STALE",
        nomenclature_ref=DISPLAY_GUID,
        product_guid=DISPLAY_GUID,
        product_name="Дисплей из прошлого расчёта",
        folder="дисплеи",
        action_kind="transition",
        current_status="fruit",
        target_status="sales_start",
        status="stale",
        reason="Старое предложение",
        facts={},
        blockers=[],
        risk_codes=[],
        run_id=run_id - 1,
        run_key="display-old-run",
        facts_hash="b" * 64,
        idempotency_key="fruit-stale-test",
        responsible_bitrix_user_id="130757",
        responsible_name="Омар",
    )
    lifecycle_db.add(stale)
    lifecycle_db.commit()

    current_queue = list_lifecycle_transitions(
        lifecycle_db,
        status="fruit",
        scope="action",
    )
    archive_queue = list_lifecycle_transitions(
        lifecycle_db,
        status="fruit",
        scope="action",
        readiness="stale",
    )

    assert current_queue["total"] == 0
    assert current_queue["stale_count"] == 0
    assert archive_queue["total"] == 1
    assert archive_queue["stale_count"] == 1
    assert archive_queue["items"][0]["nomenclature_code"] == "FRUIT-STALE"


def test_first_supplier_order_moves_fruit_to_newborn_without_approval(
    lifecycle_db,
    sqlite_engine,
) -> None:
    run_id = _seed_lifecycle(sqlite_engine)

    sync_lifecycle_transition_proposals(
        lifecycle_db,
        run_id=run_id,
        settings=_settings(),
    )

    proposal = lifecycle_db.scalar(
        select(ProcurementLifecycleTransitionProposal).where(
            ProcurementLifecycleTransitionProposal.nomenclature_code == "FRUIT-1"
        )
    )
    assert proposal is not None
    assert proposal.current_status == "fruit"
    assert proposal.target_status == "newborn"
    assert proposal.status == "auto_applied"
    assert proposal.approved_by_actor == "system:onec-facts"
    event = lifecycle_db.scalar(
        select(ProcurementOrderFormationEvent).where(
            ProcurementOrderFormationEvent.event_type == "lifecycle_transition_auto_applied"
        )
    )
    assert event is not None
    assert event.entity_id == "FRUIT-1"
    assert event.before == {"status": "fruit"}
    assert event.after == {"status": "newborn"}
    sync_lifecycle_transition_proposals(
        lifecycle_db,
        run_id=run_id,
        settings=_settings(),
    )
    assert (
        lifecycle_db.scalar(
            select(func.count(ProcurementOrderFormationEvent.id)).where(
                ProcurementOrderFormationEvent.event_type == "lifecycle_transition_auto_applied"
            )
        )
        == 1
    )


def test_v2_live_auto_applies_computed_transition_internally_and_protects_manual_status(
    lifecycle_db,
    sqlite_engine,
) -> None:
    now = datetime(2026, 8, 12, 10, 0)
    rows = build_classification_rows(
        records=[
            {
                "nomenclature_code": "V2-AUTO",
                "previous_status": "working",
            },
            {
                "nomenclature_code": "V2-MANUAL",
                "previous_status": "working",
                "manual_status": "pension",
            },
        ],
        summaries=[
            {
                "nomenclature_code": "V2-AUTO",
                "name": "Дисплей растущий",
                "folder": "дисплеи",
                "status": "sale",
                "status_label": "Растим",
                "classification_model": "v2-live",
                "legacy_status": "working",
                "demand_state": "growing",
                "reason_text": "Устойчивый рост подтверждён.",
            },
            {
                "nomenclature_code": "V2-MANUAL",
                "name": "Дисплей ручной",
                "folder": "дисплеи",
                "status": "pension",
                "status_label": "Допродаём",
                "classification_model": "v2-live",
                "legacy_status": "working",
                "demand_state": "growing",
                "reason_text": "Ручное решение.",
            },
        ],
        source="test",
        classified_at=now,
    )
    result = persist_classification_rows(
        sqlite_engine,
        rows=rows,
        run_key="v2-live-auto-test",
        folder="дисплеи",
        source="test",
        started_at=now,
        finished_at=now,
    )
    summary = sync_lifecycle_transition_proposals(
        lifecycle_db,
        run_id=result.run_id,
        settings=_settings(),
    )
    proposals = lifecycle_db.scalars(
        select(ProcurementLifecycleTransitionProposal).where(
            ProcurementLifecycleTransitionProposal.run_id == result.run_id
        )
    ).all()
    assert summary["automatic"] == 1
    assert len(proposals) == 1
    assert proposals[0].nomenclature_code == "V2-AUTO"
    assert proposals[0].current_status == "working"
    assert proposals[0].target_status == "sale"
    assert proposals[0].status == "auto_applied"
    assert proposals[0].onec_status == "not_required"


def test_v2_live_blocker_creates_review_and_never_applies_target_stage(
    lifecycle_db,
    sqlite_engine,
) -> None:
    now = datetime(2026, 8, 12, 11, 0)
    rows = build_classification_rows(
        records=[
            {
                "nomenclature_code": "V2-BLOCKED",
                "previous_status": "working",
            }
        ],
        summaries=[
            {
                "nomenclature_code": "V2-BLOCKED",
                "name": "Дисплей с неполными данными",
                "folder": "дисплеи",
                "status": "sale",
                "status_label": "Растим",
                "classification_model": "v2-live",
                "legacy_status": "working",
                "demand_state": "no_data",
                "manual_review_required": True,
                "blockers": ["demand_data_missing"],
                "reason_text": "Переход заблокирован до проверки данных.",
            }
        ],
        source="test",
        classified_at=now,
    )
    result = persist_classification_rows(
        sqlite_engine,
        rows=rows,
        run_key="v2-live-blocker-test",
        folder="дисплеи",
        source="test",
        started_at=now,
        finished_at=now,
    )

    summary = sync_lifecycle_transition_proposals(
        lifecycle_db,
        run_id=result.run_id,
        settings=_settings(),
    )
    proposal = lifecycle_db.scalar(
        select(ProcurementLifecycleTransitionProposal).where(
            ProcurementLifecycleTransitionProposal.run_id == result.run_id
        )
    )

    assert summary["automatic"] == 0
    assert proposal is not None
    assert proposal.action_kind == "review"
    assert proposal.current_status == "working"
    assert proposal.target_status is None
    assert proposal.status == "pending"
    assert proposal.blockers == ["demand_data_missing"]


def test_batch_approval_returns_partial_result_and_is_idempotent(
    lifecycle_db,
    sqlite_engine,
) -> None:
    run_id = _seed_lifecycle(sqlite_engine)
    sync_lifecycle_transition_proposals(
        lifecycle_db,
        run_id=run_id,
        settings=_settings(),
    )
    proposals = lifecycle_db.scalars(
        select(ProcurementLifecycleTransitionProposal).order_by(
            ProcurementLifecycleTransitionProposal.id
        )
    ).all()
    ready = next(
        item for item in proposals if item.action_kind == "transition" and item.status == "pending"
    )
    review = next(
        item for item in proposals if item.action_kind == "review" and item.status == "pending"
    )
    request_items = [_approval_item(ready), _approval_item(review)]

    first = approve_lifecycle_transitions(
        lifecycle_db,
        items=request_items,
        idempotency_key="test-lifecycle-batch-1",
        session=_session("4241", "Эльдар"),
        settings=_settings(),
    )
    second = approve_lifecycle_transitions(
        lifecycle_db,
        items=request_items,
        idempotency_key="test-lifecycle-batch-1",
        session=_session("4241", "Эльдар"),
        settings=_settings(),
    )

    assert first["mode"] == "internal"
    assert first["xml_preview"] == ""
    assert first["written_path"] is None
    assert first["summary"]["approved"] == 1
    assert first["summary"]["blocked"] == 1
    assert second == first
    assert (
        lifecycle_db.scalar(
            select(func.count(ProcurementOrderFormationEvent.id)).where(
                ProcurementOrderFormationEvent.event_type == "lifecycle_transitions_approved"
            )
        )
        == 1
    )


def test_sale_to_working_is_omar_only(lifecycle_db, sqlite_engine) -> None:
    run_id = _seed_lifecycle(sqlite_engine)
    proposal = ProcurementLifecycleTransitionProposal(
        nomenclature_code="SALE-1",
        nomenclature_ref=DISPLAY_GUID,
        product_guid=DISPLAY_GUID,
        product_name="Дисплей к рабочему",
        folder="Дисплеи",
        action_kind="transition",
        current_status="sale",
        target_status="working",
        status="pending",
        reason="Подтверждение ответственного",
        facts={},
        blockers=[],
        risk_codes=[],
        run_id=run_id,
        run_key="display-test-run-1",
        facts_hash="a" * 64,
        idempotency_key="sale-to-working-test",
        responsible_bitrix_user_id="130757",
        responsible_name="Омар",
    )
    lifecycle_db.add(proposal)
    lifecycle_db.commit()

    eldar_result = approve_lifecycle_transitions(
        lifecycle_db,
        items=[_approval_item(proposal)],
        idempotency_key="sale-to-working-eldar",
        session=_session("4241", "Эльдар"),
        settings=_settings(),
    )
    assert eldar_result["summary"]["blocked"] == 1

    omar_result = approve_lifecycle_transitions(
        lifecycle_db,
        items=[_approval_item(proposal)],
        idempotency_key="sale-to-working-omar",
        session=_session("130757", "Омар"),
        settings=_settings(),
    )
    assert omar_result["summary"]["approved"] == 1


def test_batch_limit_is_100(lifecycle_db) -> None:
    with pytest.raises(ValueError, match="1..100"):
        approve_lifecycle_transitions(
            lifecycle_db,
            items=[{"proposal_id": index} for index in range(101)],
            idempotency_key="too-large-batch",
            session=_session("130757", "Омар"),
            settings=_settings(),
        )


def test_commerceml_readback_reflects_lifecycle_transition(
    lifecycle_db,
    monkeypatch,
) -> None:
    proposal = ProcurementLifecycleTransitionProposal(
        nomenclature_code="SALE-READBACK",
        nomenclature_ref=DISPLAY_GUID,
        product_guid=DISPLAY_GUID,
        product_name="Дисплей readback",
        folder="дисплеи",
        action_kind="transition",
        current_status="sale",
        target_status="working",
        status="applied",
        reason="Подтверждено",
        facts={},
        blockers=[],
        risk_codes=[],
        run_id=1,
        run_key="readback-run",
        facts_hash="b" * 64,
        idempotency_key="readback-transition-test",
    )
    lifecycle_db.add(proposal)
    lifecycle_db.commit()
    monkeypatch.setattr(bitrix_order_service, "load_order_formation_mapping", lambda _settings: {})
    monkeypatch.setattr(
        bitrix_order_service,
        "resolve_catalog_product_by_xml_id",
        lambda *_args, **_kwargs: BitrixCatalogProduct(
            product_id="1646",
            name="Дисплей readback",
            xml_id=DISPLAY_GUID,
            assortment_status="Рабочий",
        ),
    )

    summary = bitrix_order_service.reflect_classifications_from_bitrix(
        lifecycle_db,
        settings=Settings(),
    )

    lifecycle_db.refresh(proposal)
    assert summary == {"reflected": 1, "pending": 0, "missing": 0, "unrecognized": 0}
    assert proposal.status == "reflected"
    assert proposal.bitrix_readback_value == "Рабочий"
