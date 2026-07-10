from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from typing import Any

from sqlalchemy import create_engine, select, text
from sqlalchemy.orm import Session

from app.core.config import Settings, get_settings
from app.models import (
    ReceivableCase,
    ReceivableOpenDebtCache,
    ReceivableSmsLog,
    ReceivableWorkEvent,
    ReceivableWorkItem,
)
from app.services.receivable_workflow import (
    EVENT_DATA_QUALITY,
    SMS_DRY_RUN,
    SMS_SKIPPED_NO_PHONE,
    STATUS_CALLING,
    STATUS_CLOSED,
    STATUS_DATA_QUALITY,
    STATUS_ESCALATED,
    debt_key_for_case,
    format_chain_documents_for_bitrix,
    stable_key_for_counterparty,
    sync_receivable_workflow,
)
from app.services.receivables import CASE_BUYERS, CASE_OVERDUE
from tasks.export_receivable_work_report import export_receivable_work_report


class FakeBitrixClient:
    def __init__(self) -> None:
        self.added: list[dict[str, Any]] = []
        self.updated: list[tuple[str, dict[str, Any]]] = []
        self.matches: list[dict[str, Any]] = []
        self.lookup_calls: list[dict[str, Any]] = []
        self.next_id = 100

    def add_smart_process_item(
        self,
        *,
        entity_type_id: int,
        fields: dict[str, Any],
    ) -> tuple[str, str | None]:
        self.added.append({"entity_type_id": entity_type_id, "fields": fields})
        self.next_id += 1
        return str(self.next_id), f"/crm/type/{entity_type_id}/details/{self.next_id}/"

    def update_smart_process_item(
        self,
        *,
        entity_type_id: int,
        item_id: str,
        fields: dict[str, Any],
    ) -> None:
        self.updated.append((item_id, {"entity_type_id": entity_type_id, "fields": fields}))

    def list_items_by_ref(
        self,
        *,
        entity_type_id: int,
        ref_field: str,
        ref_value: str,
    ) -> list[dict[str, Any]]:
        self.lookup_calls.append(
            {
                "entity_type_id": entity_type_id,
                "ref_field": ref_field,
                "ref_value": ref_value,
            }
        )
        return list(self.matches)


def _settings(**overrides: Any) -> Settings:
    data = {
        "receivable_sms_mode": "dry_run",
        "receivable_bitrix_entity_type_id": 187,
        "receivable_bitrix_field_map": {
            "title": "TITLE",
            "stable_key": "UF_CRM_RECEIVABLE_STABLE_KEY",
            "counterparty_ref": "UF_CRM_RECEIVABLE_COUNTERPARTY_REF",
            "current_balance": "UF_CRM_RECEIVABLE_CURRENT_BALANCE",
            "assigned_by": "ASSIGNED_BY_ID",
        },
        "receivable_bitrix_stage_map": {
            "calling": "DT187_1:CALLING",
            "escalated": "DT187_1:ESCALATED",
            "closed": "DT187_1:CLOSED",
            "no_phone": "DT187_1:NO_PHONE",
            "new_debt": "DT187_1:NEW",
        },
        "receivable_workflow_department_refs": [],
        "receivable_workflow_department_names": [],
    }
    data.update(overrides)
    return Settings(**data)


def _case(
    *,
    snapshot_date: date,
    segment: str,
    counterparty_ref: str = "cp-a",
    counterparty_name: str = "Ромашка ООО",
    balance: Decimal = Decimal("15000"),
    origin_date: datetime = datetime(2026, 3, 14, 10, 0),
    due_date: datetime = datetime(2026, 3, 19),
    overdue_days: int = 1,
    department_ref: str | None = "dep-1",
    department_name: str | None = "Продажи",
) -> ReceivableCase:
    return ReceivableCase(
        snapshot_date=snapshot_date,
        segment=segment,
        owner_type="sales_manager",
        recommendation="Проверить просрочку.",
        counterparty_ref=counterparty_ref,
        counterparty_name=counterparty_name,
        current_balance=balance,
        aged_bucket="1-7",
        activity_segment="active",
        origin_document_ref="sale-a",
        origin_document_number="S-001",
        origin_document_date=origin_date,
        origin_manager_ref="mgr-1",
        origin_manager_name="Менеджер 1",
        current_manager_ref="mgr-1",
        current_manager_name="Менеджер 1",
        department_ref=department_ref,
        department_name=department_name,
        planned_payment_date=due_date,
        credit_depth_days=None,
        shipment_ban=False,
        payment_term_source="planned_payment_date",
        due_date=due_date,
        overdue_days=overdue_days,
        is_overdue=segment == CASE_OVERDUE,
        chain_documents=[
            {
                "document_ref": "sale-a",
                "document_number": "S-001",
                "amount_delta": "15000",
            },
            {
                "document_ref": "sale-b",
                "document_number": "S-002",
                "amount_delta": "2500",
            },
        ],
    )


def test_chain_documents_are_formatted_for_bitrix_without_technical_values() -> None:
    formatted = format_chain_documents_for_bitrix(
        [
            {
                "event_type": "sale",
                "document_ref": "0xbbdd002590803daf11f139862db68db7",
                "document_number": "РБГУ0172002",
                "document_date": "2026-04-16T14:22:59",
                "amount_delta": 6436.0,
            },
            {
                "event_type": "sale",
                "document_ref": "0xbbdd002590803daf11f1398cead5f437",
                "document_number": "РБГУ0172201",
                "document_date": "2026-04-16T15:23:06",
                "amount_delta": 518.0,
            },
        ]
    )

    assert formatted.splitlines() == [
        "1. Реализация РБГУ0172002 от 16.04.2026 14:22 на 6 436 руб.",
        "2. Реализация РБГУ0172201 от 16.04.2026 15:23 на 518 руб.",
    ]
    assert "document_ref" not in formatted
    assert "event_type" not in formatted
    assert "0xbbdd" not in formatted
    assert "{" not in formatted


def test_open_debt_documents_are_formatted_for_bitrix_with_source_details() -> None:
    formatted = format_chain_documents_for_bitrix(
        [
            {
                "document_ref": "sale-open",
                "document_number": "РТУ-1",
                "document_date": "2026-06-10T09:00:00",
                "open_amount": "12000.00",
                "sale_amount": "15000.00",
                "closing_amount": "-3000.00",
                "statement_selection_rule": "statement_bottom_up_balance_cutoff",
            }
        ]
    )

    assert formatted == (
        "1. Открытый долг РТУ-1 от 10.06.2026 09:00 на 12 000 руб. "
        "(исходно 15 000 руб.; закрытия -3 000 руб.; правило: подбор от текущего остатка)"
    )


def test_sms_outbox_is_deduplicated_per_debt_day(db_session: Session) -> None:
    as_of = date(2026, 3, 20)
    buyer_case = _case(
        snapshot_date=as_of,
        segment=CASE_BUYERS,
        due_date=datetime(2026, 3, 21),
    )
    db_session.add(buyer_case)

    first = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=_settings(),
        dry_run_bitrix=True,
    )
    second = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=_settings(),
        dry_run_bitrix=True,
    )

    logs = db_session.scalars(select(ReceivableSmsLog)).all()
    assert len(logs) == 1
    assert logs[0].status == SMS_DRY_RUN
    assert logs[0].debt_key == debt_key_for_case(buyer_case)
    assert first.sms_created == 1
    assert second.sms_reused == 1


def test_sms_without_phone_is_logged_as_skipped(db_session: Session) -> None:
    as_of = date(2026, 3, 20)
    db_session.add(
        _case(
            snapshot_date=as_of,
            segment=CASE_BUYERS,
            due_date=datetime(2026, 3, 21),
        )
    )

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        settings=_settings(),
        dry_run_bitrix=True,
    )

    log = db_session.scalar(select(ReceivableSmsLog))
    assert log is not None
    assert log.status == SMS_SKIPPED_NO_PHONE
    assert log.error == "Нет телефона"
    assert summary.sms_skipped_no_phone == 1


def test_work_item_is_not_created_before_payment_due_date(db_session: Session) -> None:
    as_of = date(2026, 3, 18)
    db_session.add_all(
        [
            _case(
                snapshot_date=as_of,
                segment=CASE_BUYERS,
                due_date=datetime(2026, 3, 19),
            ),
            _case(
                snapshot_date=as_of,
                segment=CASE_OVERDUE,
                due_date=datetime(2026, 3, 19),
            ),
        ]
    )

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=_settings(),
        dry_run_bitrix=True,
    )

    assert db_session.scalar(select(ReceivableWorkItem)) is None
    assert summary.work_items_created == 0


def test_worker_enriches_phone_from_onec_before_sync(
    monkeypatch,
    sqlite_engine,
    db_session: Session,
) -> None:
    from app.workers import receivable_workflow as worker

    as_of = date(2026, 3, 21)
    db_session.add_all(
        [
            _case(snapshot_date=as_of, segment=CASE_BUYERS, origin_date=datetime(2026, 3, 14)),
            _case(snapshot_date=as_of, segment=CASE_OVERDUE, origin_date=datetime(2026, 3, 14)),
        ]
    )
    db_session.commit()

    onec_engine = create_engine("sqlite:///:memory:")
    with onec_engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE _Reference25 (
                _IDRRef TEXT,
                _Description TEXT
            )
        """))
        conn.execute(text("""
            CREATE TABLE _InfoRg6402 (
                _Fld6403_RTRef TEXT,
                _Fld6403_RRRef TEXT,
                _Fld6405_RRRef TEXT,
                _Fld6406 TEXT
            )
        """))
        conn.execute(text("""
            INSERT INTO _Reference25 (_IDRRef, _Description)
            VALUES ('kind-main', 'Телефон контрагента')
        """))
        conn.execute(text("""
            INSERT INTO _InfoRg6402 (
                _Fld6403_RTRef, _Fld6403_RRRef, _Fld6405_RRRef, _Fld6406
            )
            VALUES ('0x00000036', 'cp-a', 'kind-main', '8 (999) 111-22-33')
        """))

    settings = _settings(
        receivable_workflow_enabled=True,
        onec_database_url="sqlite:///:memory:",
    )
    monkeypatch.setattr(worker, "get_settings", lambda: settings)
    monkeypatch.setattr(worker, "_get_app_engine", lambda: sqlite_engine)
    monkeypatch.setattr(worker, "_get_onec_engine", lambda: onec_engine)

    try:
        result = worker.run_receivable_workflow_sync(as_of=as_of, dry_run_bitrix=True)
    finally:
        onec_engine.dispose()

    item = db_session.scalar(select(ReceivableWorkItem))
    assert result["status"] == "ok"
    assert item is not None
    assert item.phone == "+79991112233"
    assert item.phone_status == "present"


def test_overdue_buyer_gets_one_work_item_and_bitrix_item(db_session: Session) -> None:
    as_of = date(2026, 3, 21)
    db_session.add_all(
        [
            _case(snapshot_date=as_of, segment=CASE_BUYERS, origin_date=datetime(2026, 3, 14)),
            _case(snapshot_date=as_of, segment=CASE_OVERDUE, origin_date=datetime(2026, 3, 14)),
        ]
    )
    bitrix = FakeBitrixClient()

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=_settings(),
        bitrix_client=bitrix,
    )

    item = db_session.scalar(select(ReceivableWorkItem))
    assert item is not None
    assert item.stable_key == stable_key_for_counterparty("cp-a")
    assert item.status == STATUS_CALLING
    assert item.needs_call_today is True
    assert item.current_balance == Decimal("15000")
    assert item.department_ref == "dep-1"
    assert item.department_name == "Продажи"
    assert len(item.chain_documents or []) == 2
    assert item.bitrix_item_id == 101
    assert summary.work_items_created == 1
    assert summary.bitrix_created == 1
    assert len(bitrix.added) == 1
    assert bitrix.lookup_calls == [
        {
            "entity_type_id": 187,
            "ref_field": "UF_CRM_RECEIVABLE_STABLE_KEY",
            "ref_value": stable_key_for_counterparty("cp-a"),
        }
    ]


def test_bitrix_sync_uses_open_debt_cache_for_documents(db_session: Session) -> None:
    as_of = date(2026, 3, 21)
    db_session.add_all(
        [
            _case(snapshot_date=as_of, segment=CASE_BUYERS, origin_date=datetime(2026, 3, 14)),
            _case(snapshot_date=as_of, segment=CASE_OVERDUE, origin_date=datetime(2026, 3, 14)),
            ReceivableOpenDebtCache(
                snapshot_date=as_of,
                counterparty_ref="cp-a",
                department_ref="dep-1",
                documents=[
                    {
                        "document_ref": "sale-open",
                        "document_number": "РТУ-1",
                        "document_date": "2026-03-14T10:00:00",
                        "open_amount": "12000.00",
                        "sale_amount": "15000.00",
                        "closing_amount": "-3000.00",
                        "statement_selection_rule": "statement_bottom_up_balance_cutoff",
                    }
                ],
            ),
        ]
    )
    bitrix = FakeBitrixClient()
    settings = _settings(
        receivable_bitrix_field_map={
            **_settings().receivable_bitrix_field_map,
            "chain_documents": "UF_CRM_RECEIVABLE_CHAIN_DOCUMENTS",
        }
    )

    sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=settings,
        bitrix_client=bitrix,
    )

    docs_field = bitrix.added[0]["fields"]["UF_CRM_RECEIVABLE_CHAIN_DOCUMENTS"]
    assert "Открытый долг РТУ-1" in docs_field
    assert "правило: подбор от текущего остатка" in docs_field
    assert "Реализация S-001" not in docs_field


def test_bitrix_sync_falls_back_to_chain_documents_without_open_debt_cache(
    db_session: Session,
) -> None:
    as_of = date(2026, 3, 21)
    db_session.add_all(
        [
            _case(snapshot_date=as_of, segment=CASE_BUYERS, origin_date=datetime(2026, 3, 14)),
            _case(snapshot_date=as_of, segment=CASE_OVERDUE, origin_date=datetime(2026, 3, 14)),
        ]
    )
    bitrix = FakeBitrixClient()
    settings = _settings(
        receivable_bitrix_field_map={
            **_settings().receivable_bitrix_field_map,
            "chain_documents": "UF_CRM_RECEIVABLE_CHAIN_DOCUMENTS",
        }
    )

    sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=settings,
        bitrix_client=bitrix,
    )

    docs_field = bitrix.added[0]["fields"]["UF_CRM_RECEIVABLE_CHAIN_DOCUMENTS"]
    assert "Документ S-001" in docs_field
    assert "Открытый долг" not in docs_field


def test_bitrix_sync_does_not_fall_back_when_open_debt_cache_row_is_empty(
    db_session: Session,
) -> None:
    as_of = date(2026, 3, 21)
    db_session.add_all(
        [
            _case(snapshot_date=as_of, segment=CASE_BUYERS, origin_date=datetime(2026, 3, 14)),
            _case(snapshot_date=as_of, segment=CASE_OVERDUE, origin_date=datetime(2026, 3, 14)),
            ReceivableOpenDebtCache(
                snapshot_date=as_of,
                counterparty_ref="cp-a",
                department_ref="dep-1",
                documents=[],
            ),
        ]
    )
    bitrix = FakeBitrixClient()
    settings = _settings(
        receivable_bitrix_field_map={
            **_settings().receivable_bitrix_field_map,
            "chain_documents": "UF_CRM_RECEIVABLE_CHAIN_DOCUMENTS",
        }
    )

    sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=settings,
        bitrix_client=bitrix,
    )

    assert bitrix.added[0]["fields"]["UF_CRM_RECEIVABLE_CHAIN_DOCUMENTS"] == ""


def test_existing_bitrix_item_is_reused_by_stable_key(db_session: Session) -> None:
    as_of = date(2026, 3, 21)
    db_session.add_all(
        [
            _case(snapshot_date=as_of, segment=CASE_BUYERS, origin_date=datetime(2026, 3, 14)),
            _case(snapshot_date=as_of, segment=CASE_OVERDUE, origin_date=datetime(2026, 3, 14)),
        ]
    )
    bitrix = FakeBitrixClient()
    bitrix.matches = [{"id": 555, "detailUrl": "/crm/type/187/details/555/"}]

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=_settings(),
        bitrix_client=bitrix,
    )

    item = db_session.scalar(select(ReceivableWorkItem))
    assert item is not None
    assert item.bitrix_item_id == 555
    assert item.bitrix_detail_url == "/crm/type/187/details/555/"
    assert summary.bitrix_created == 0
    assert summary.bitrix_updated == 1
    assert bitrix.added == []
    assert bitrix.updated[0][0] == "555"


def test_department_scope_limits_workflow_to_one_department(db_session: Session) -> None:
    as_of = date(2026, 3, 21)
    db_session.add_all(
        [
            _case(
                snapshot_date=as_of,
                segment=CASE_BUYERS,
                counterparty_ref="cp-a",
                department_ref="dep-1",
                department_name="Пилот",
            ),
            _case(
                snapshot_date=as_of,
                segment=CASE_OVERDUE,
                counterparty_ref="cp-a",
                department_ref="dep-1",
                department_name="Пилот",
            ),
            _case(
                snapshot_date=as_of,
                segment=CASE_BUYERS,
                counterparty_ref="cp-b",
                department_ref="dep-2",
                department_name="Не пилот",
            ),
            _case(
                snapshot_date=as_of,
                segment=CASE_OVERDUE,
                counterparty_ref="cp-b",
                department_ref="dep-2",
                department_name="Не пилот",
            ),
        ]
    )
    bitrix = FakeBitrixClient()

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        settings=_settings(receivable_workflow_department_refs=["dep-1"]),
        bitrix_client=bitrix,
    )

    items = db_session.scalars(select(ReceivableWorkItem)).all()
    assert [item.counterparty_ref for item in items] == ["cp-a"]
    assert summary.work_items_created == 1
    assert summary.bitrix_created == 1


def test_department_scope_does_not_close_out_of_scope_items(db_session: Session) -> None:
    as_of = date(2026, 3, 21)
    db_session.add(
        ReceivableWorkItem(
            stable_key=stable_key_for_counterparty("cp-b"),
            counterparty_ref="cp-b",
            status=STATUS_CALLING,
            current_balance=Decimal("1000"),
            department_ref="dep-2",
            department_name="Не пилот",
        )
    )
    db_session.add_all(
        [
            _case(
                snapshot_date=as_of,
                segment=CASE_BUYERS,
                counterparty_ref="cp-a",
                department_ref="dep-1",
                department_name="Пилот",
            ),
            _case(
                snapshot_date=as_of,
                segment=CASE_OVERDUE,
                counterparty_ref="cp-a",
                department_ref="dep-1",
                department_name="Пилот",
            ),
        ]
    )

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        settings=_settings(receivable_workflow_department_names=["пилот"]),
        dry_run_bitrix=True,
    )

    out_of_scope = db_session.scalar(
        select(ReceivableWorkItem).where(ReceivableWorkItem.counterparty_ref == "cp-b")
    )
    assert out_of_scope is not None
    assert out_of_scope.status == STATUS_CALLING
    assert summary.work_items_closed == 0


def test_bitrix_enum_fields_are_sent_as_enum_ids_when_mapping_exists(
    db_session: Session,
) -> None:
    sms_as_of = date(2026, 3, 18)
    as_of = date(2026, 3, 20)
    db_session.add_all(
        [
            _case(
                snapshot_date=sms_as_of,
                segment=CASE_BUYERS,
                origin_date=datetime(2026, 3, 15),
                due_date=datetime(2026, 3, 19),
            ),
            _case(
                snapshot_date=as_of,
                segment=CASE_BUYERS,
                origin_date=datetime(2026, 3, 15),
                due_date=datetime(2026, 3, 19),
            ),
            _case(
                snapshot_date=as_of,
                segment=CASE_OVERDUE,
                origin_date=datetime(2026, 3, 15),
                due_date=datetime(2026, 3, 19),
            ),
        ]
    )
    bitrix = FakeBitrixClient()
    settings = _settings(
        receivable_bitrix_field_map={
            "title": "TITLE",
            "stable_key": "UF_CRM_RECEIVABLE_STABLE_KEY",
            "counterparty_ref": "UF_CRM_RECEIVABLE_COUNTERPARTY_REF",
            "phone_status": "UF_CRM_RECEIVABLE_PHONE_STATUS",
            "sms_status": "UF_CRM_RECEIVABLE_SMS_STATUS",
            "escalation_level": "UF_CRM_RECEIVABLE_ESCALATION_LEVEL",
        },
        receivable_bitrix_enum_map={
            "phone_status": {"present": "11", "missing": "12"},
            "sms_status": {"dry_run": "21", "sent": "22"},
            "escalation_level": {"retail_network_head": "31"},
        },
    )

    sync_receivable_workflow(
        db_session,
        as_of=sms_as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=settings,
        dry_run_bitrix=True,
    )
    sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=settings,
        bitrix_client=bitrix,
    )

    fields = bitrix.added[0]["fields"]
    assert fields["UF_CRM_RECEIVABLE_PHONE_STATUS"] == "11"
    assert fields["UF_CRM_RECEIVABLE_SMS_STATUS"] == "21"
    assert (
        "UF_CRM_RECEIVABLE_ESCALATION_LEVEL" not in fields
        or fields["UF_CRM_RECEIVABLE_ESCALATION_LEVEL"] is None
    )


def test_duplicate_bitrix_items_by_stable_key_are_reported(db_session: Session) -> None:
    as_of = date(2026, 3, 21)
    db_session.add_all(
        [
            _case(snapshot_date=as_of, segment=CASE_BUYERS, origin_date=datetime(2026, 3, 14)),
            _case(snapshot_date=as_of, segment=CASE_OVERDUE, origin_date=datetime(2026, 3, 14)),
        ]
    )
    bitrix = FakeBitrixClient()
    bitrix.matches = [{"id": 555}, {"id": 556}]

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=_settings(),
        bitrix_client=bitrix,
    )

    item = db_session.scalar(select(ReceivableWorkItem))
    assert item is not None
    assert item.bitrix_item_id is None
    assert summary.bitrix_created == 0
    assert summary.bitrix_errors == 1
    assert "несколько карточек" in (item.bitrix_last_error or "")
    assert bitrix.added == []


def test_overdue_buyer_without_origin_document_is_skipped(db_session: Session) -> None:
    as_of = date(2026, 3, 21)
    buyer_case = _case(snapshot_date=as_of, segment=CASE_BUYERS, overdue_days=10000)
    overdue_case = _case(snapshot_date=as_of, segment=CASE_OVERDUE, overdue_days=10000)
    for case in (buyer_case, overdue_case):
        case.origin_document_ref = None
        case.origin_document_number = None
        case.origin_document_date = None
        case.chain_documents = []
    db_session.add_all([buyer_case, overdue_case])

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        settings=_settings(),
        dry_run_bitrix=True,
    )

    assert summary.work_items_created == 0
    assert summary.data_quality_skipped == 1
    assert db_session.scalar(select(ReceivableWorkItem)) is None


def test_existing_work_item_with_no_doc_debt_gets_data_quality_event(
    db_session: Session,
) -> None:
    as_of = date(2026, 3, 21)
    item = ReceivableWorkItem(
        stable_key=stable_key_for_counterparty("cp-a"),
        counterparty_ref="cp-a",
        counterparty_name="Ромашка ООО",
        status=STATUS_CALLING,
        current_balance=Decimal("1000"),
    )
    buyer_case = _case(snapshot_date=as_of, segment=CASE_BUYERS, overdue_days=10000)
    overdue_case = _case(snapshot_date=as_of, segment=CASE_OVERDUE, overdue_days=10000)
    for case in (buyer_case, overdue_case):
        case.origin_document_ref = None
        case.origin_document_number = None
        case.origin_document_date = None
        case.chain_documents = []
    db_session.add_all([item, buyer_case, overdue_case])

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        settings=_settings(),
        dry_run_bitrix=True,
    )

    assert item.status == STATUS_DATA_QUALITY
    assert item.current_balance == Decimal("1000")
    assert item.closed_at is None
    assert summary.work_items_closed == 0
    assert summary.data_quality_skipped == 1
    events = db_session.scalars(select(ReceivableWorkEvent)).all()
    assert [event.event_type for event in events] == [EVENT_DATA_QUALITY]


def test_closed_work_item_with_no_doc_debt_stays_closed(db_session: Session) -> None:
    as_of = date(2026, 3, 21)
    item = ReceivableWorkItem(
        stable_key=stable_key_for_counterparty("cp-a"),
        counterparty_ref="cp-a",
        counterparty_name="Ромашка ООО",
        status=STATUS_CLOSED,
        current_balance=Decimal("0"),
        closed_at=datetime(2026, 3, 20),
    )
    buyer_case = _case(snapshot_date=as_of, segment=CASE_BUYERS, overdue_days=10000)
    overdue_case = _case(snapshot_date=as_of, segment=CASE_OVERDUE, overdue_days=10000)
    for case in (buyer_case, overdue_case):
        case.origin_document_ref = None
        case.origin_document_number = None
        case.origin_document_date = None
        case.chain_documents = []
    db_session.add_all([item, buyer_case, overdue_case])

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        settings=_settings(),
        dry_run_bitrix=True,
    )

    assert item.status == STATUS_CLOSED
    assert item.current_balance == Decimal("0")
    assert item.closed_at == datetime(2026, 3, 20)
    assert summary.work_items_closed == 0
    assert summary.data_quality_skipped == 1
    assert db_session.scalars(select(ReceivableWorkEvent)).all() == []


def test_previously_closed_no_doc_item_is_not_reopened_as_data_quality(
    db_session: Session,
) -> None:
    as_of = date(2026, 3, 21)
    item = ReceivableWorkItem(
        stable_key=stable_key_for_counterparty("cp-a"),
        counterparty_ref="cp-a",
        counterparty_name="Ромашка ООО",
        status=STATUS_DATA_QUALITY,
        current_balance=Decimal("1000"),
        closed_at=datetime(2026, 3, 20),
    )
    buyer_case = _case(snapshot_date=as_of, segment=CASE_BUYERS, overdue_days=10000)
    overdue_case = _case(snapshot_date=as_of, segment=CASE_OVERDUE, overdue_days=10000)
    for case in (buyer_case, overdue_case):
        case.origin_document_ref = None
        case.origin_document_number = None
        case.origin_document_date = None
        case.chain_documents = []
    db_session.add_all([item, buyer_case, overdue_case])

    sync_receivable_workflow(
        db_session,
        as_of=as_of,
        settings=_settings(),
        dry_run_bitrix=True,
    )

    assert item.status == STATUS_CLOSED
    assert item.current_balance == Decimal("0")
    assert db_session.scalars(select(ReceivableWorkEvent)).all() == []


def test_existing_work_item_is_closed_when_debt_disappears(db_session: Session) -> None:
    as_of = date(2026, 3, 22)
    item = ReceivableWorkItem(
        stable_key=stable_key_for_counterparty("cp-a"),
        counterparty_ref="cp-a",
        counterparty_name="Ромашка ООО",
        status=STATUS_CALLING,
        current_balance=Decimal("1000"),
    )
    db_session.add_all(
        [
            item,
            _case(
                snapshot_date=as_of,
                segment=CASE_BUYERS,
                counterparty_ref="cp-b",
                origin_date=datetime(2026, 3, 15),
            ),
        ]
    )

    summary = sync_receivable_workflow(
        db_session,
        as_of=as_of,
        settings=_settings(),
        dry_run_bitrix=True,
    )

    assert item.status == STATUS_CLOSED
    assert item.current_balance == Decimal("0")
    assert item.closed_at is not None
    assert summary.work_items_closed == 1
    events = db_session.scalars(select(ReceivableWorkEvent)).all()
    assert {event.event_type for event in events} == {"closed_by_onec"}


def test_empty_case_snapshot_does_not_close_existing_work_items(db_session: Session) -> None:
    item = ReceivableWorkItem(
        stable_key=stable_key_for_counterparty("cp-a"),
        counterparty_ref="cp-a",
        counterparty_name="Ромашка ООО",
        status=STATUS_CALLING,
        current_balance=Decimal("1000"),
    )
    db_session.add(item)

    summary = sync_receivable_workflow(
        db_session,
        as_of=date(2026, 3, 22),
        settings=_settings(),
        dry_run_bitrix=True,
    )

    assert item.status == STATUS_CALLING
    assert item.current_balance == Decimal("1000")
    assert item.closed_at is None
    assert summary.work_items_closed == 0
    assert db_session.scalars(select(ReceivableWorkEvent)).all() == []


def test_overdue_15_days_escalates_to_retail_network_head(db_session: Session) -> None:
    as_of = date(2026, 4, 3)
    db_session.add_all(
        [
            _case(
                snapshot_date=as_of,
                segment=CASE_BUYERS,
                origin_date=datetime(2026, 3, 10),
                overdue_days=15,
            ),
            _case(
                snapshot_date=as_of,
                segment=CASE_OVERDUE,
                origin_date=datetime(2026, 3, 10),
                overdue_days=15,
            ),
        ]
    )

    sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=_settings(receivable_retail_network_head_user_id=777),
        dry_run_bitrix=True,
    )

    item = db_session.scalar(select(ReceivableWorkItem))
    assert item is not None
    assert item.status == STATUS_ESCALATED
    assert item.assigned_bitrix_user_id == 777
    assert item.assigned_source == "retail_network_head"
    assert item.escalated_at is not None


def test_regular_work_item_is_not_assigned_from_legacy_department_map(
    db_session: Session,
) -> None:
    as_of = date(2026, 3, 21)
    db_session.add_all(
        [
            _case(
                snapshot_date=as_of,
                segment=CASE_BUYERS,
                origin_date=datetime(2026, 3, 14),
                overdue_days=7,
            ),
            _case(
                snapshot_date=as_of,
                segment=CASE_OVERDUE,
                origin_date=datetime(2026, 3, 14),
                overdue_days=7,
            ),
        ]
    )

    sync_receivable_workflow(
        db_session,
        as_of=as_of,
        phone_by_counterparty={"cp-a": "+79990000000"},
        settings=_settings(receivable_department_manager_map={"mgr-1": 555}),
        dry_run_bitrix=True,
    )

    item = db_session.scalar(select(ReceivableWorkItem))
    assert item is not None
    assert item.status == STATUS_CALLING
    assert item.assigned_bitrix_user_id is None
    assert item.assigned_source is None


def test_management_receivable_task_payloads_can_be_disabled(
    monkeypatch, db_session: Session
) -> None:
    from app.models import ReceivableBalanceSnapshot, StaffingSnapshot
    from app.services.management_rules import build_management_task_payloads

    as_of = date(2026, 3, 20)
    db_session.add(
        ReceivableBalanceSnapshot(
            snapshot_date=as_of,
            counterparty_ref="cp-a",
            counterparty_name="Ромашка ООО",
            current_balance=Decimal("1000"),
            due_date=datetime(2026, 3, 19),
            overdue_days=1,
            is_overdue=True,
            aged_bucket="1-7",
            activity_segment="active",
        )
    )
    db_session.add(
        StaffingSnapshot(
            snapshot_date=as_of,
            store_ref="store-1",
            store_name="Магазин 1",
            shift_code="day",
            planned_count=2,
            assigned_count=1,
            confirmed_count=1,
            no_show_count=0,
            deficit_count=1,
            fill_rate=0.5,
            criticality="warning",
        )
    )
    monkeypatch.setenv("RECEIVABLE_TASK_PAYLOADS_ENABLED", "false")
    get_settings.cache_clear()

    try:
        payloads = build_management_task_payloads(db_session, as_of=as_of)
    finally:
        get_settings.cache_clear()

    rule_codes = {item["rule_code"] for item in payloads}
    assert "receivable_overdue" not in rule_codes
    assert "staffing_shift_deficit" in rule_codes


def test_receivable_work_report_first_sheet_has_four_control_columns(tmp_path) -> None:
    from openpyxl import load_workbook

    item = ReceivableWorkItem(
        stable_key=stable_key_for_counterparty("cp-a"),
        counterparty_ref="cp-a",
        counterparty_name="Ромашка ООО",
        status=STATUS_CALLING,
        current_balance=Decimal("1234.56"),
        overdue_days=3,
        current_manager_name="Менеджер 1",
    )
    output_path = tmp_path / "receivable.xlsx"

    export_receivable_work_report([item], output_path=output_path)

    workbook = load_workbook(output_path)
    ws = workbook["Контроль"]
    assert [cell.value for cell in ws[1]] == [
        "Клиент",
        "Сумма просрочки",
        "Дней просрочки",
        "Ответственный",
    ]
    assert ws.max_column == 4
    assert ws["A2"].value == "Ромашка ООО"
    assert ws["B2"].value == 1234.56
