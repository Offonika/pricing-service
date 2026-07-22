from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from infra.cron.retail_price_type_recommendations_from_a import (
    REPORT_ENDPOINT,
    export_recommendations_xlsx,
    render_telegram_message,
    sync_retail_price_type_recommendations,
)


def _report() -> dict[str, object]:
    return {
        "month": "2026-03",
        "freshness_status": "fresh",
        "source_status": "ready",
        "summary": {
            "actionable_count": 2,
            "set_silver_count": 0,
            "set_gold_count": 0,
            "downgrade_to_silver_count": 0,
            "downgrade_to_bronze_count": 1,
            "downgrade_to_gold_count": 0,
            "downgrade_to_retail_count": 0,
            "manager_work_count": 0,
            "isolate_count": 1,
            "recovery_count": 0,
            "data_check_count": 1,
            "special_review_count": 0,
            "rules": {
                "silver": "300 000 ₽ <= чистые продажи < 1 200 000 ₽",
                "gold": "1 200 000 ₽ <= чистые продажи",
            },
        },
        "payload": [
            {
                "action_label": "Сверка данных",
                "counterparty_name": "Клиент 1",
                "current_price_type": "2.Бронзовый",
                "current_level_label": "Бронза",
                "recommended_price_type": "3.Серебряный",
                "purchase_amount": "500000.00",
                "net_sales_amount": "500000.00",
                "previous_purchase_amount": "200000.00",
                "previous_net_sales_amount": "200000.00",
                "purchase_delta_amount": "300000.00",
                "net_sales_delta_amount": "300000.00",
                "purchase_delta_pct": "1.5000",
                "net_sales_delta_pct": "1.5000",
                "sales_amount": "500000.00",
                "return_amount": "0.00",
                "document_count": 1,
                "last_sale_at": "2026-03-10T12:00:00",
                "current_price_seen_at": "2026-03-10T12:00:00",
                "rule_note": "Чистые продажи от 300 000 ₽ до 1 200 000 ₽.",
                "counterparty_ref": "cp-1",
                "counterparty_code": "РБ000001",
            },
            {
                "action_label": "Перевести на бронзу",
                "counterparty_name": "Клиент 2",
                "current_price_type": "4.Золотой",
                "current_level_label": "Золото",
                "recommended_price_type": "2.Бронзовый",
                "purchase_amount": "0.00",
                "net_sales_amount": "0.00",
                "previous_purchase_amount": "100000.00",
                "previous_net_sales_amount": "100000.00",
                "purchase_delta_amount": "-100000.00",
                "net_sales_delta_amount": "-100000.00",
                "purchase_delta_pct": "-1.0000",
                "net_sales_delta_pct": "-1.0000",
                "sales_amount": "0.00",
                "return_amount": "0.00",
                "document_count": 0,
                "last_sale_at": None,
                "current_price_seen_at": "2026-02-10T12:00:00",
                "rule_note": "Чистые продажи ниже 5 000 ₽.",
                "counterparty_ref": "cp-2",
                "counterparty_code": "РБ000002",
            },
        ],
    }


def test_export_recommendations_xlsx(tmp_path: Path) -> None:
    path = export_recommendations_xlsx(_report(), tmp_path / "report.xlsx")

    wb = load_workbook(path)
    ws = wb["Типы цен"]

    assert ws["B1"].value == "Рекомендации и ручные проверки типов цен клиентов"
    assert ws["B3"].value == 2
    assert ws["E15"].value == "Продажи (чистые)"
    assert ws["F15"].value == "Продажи прошлый месяц"
    assert ws["G15"].value == "Изменение продаж"
    assert ws["H15"].value == "Изменение, %"
    assert ws["I15"].value == "Возвраты"
    assert ws["N15"].value == "Код 1С"
    assert ws["A16"].value == "Сверка данных"
    assert ws["F16"].value == 200000
    assert ws["G16"].value == 300000
    assert ws["H16"].value == 1.5
    assert ws["N16"].value == "РБ000001"
    assert ws["B17"].value == "Клиент 2"
    assert len(ws.conditional_formatting) >= 3


def test_render_telegram_message_includes_action_counts() -> None:
    message = render_telegram_message(_report())

    assert "Ежемесячный отчет по типам цен клиентов за 2026-03." in message
    assert "К ручной работе: 2" in message
    assert "изолятор 1" in message
    assert "сверка данных 1" in message
    assert "не меняют тип цены автоматически" in message


def test_sync_retail_price_type_recommendations_delivers_once(tmp_path: Path) -> None:
    fetched: list[tuple[str, dict[str, str]]] = []
    delivered: list[dict[str, object]] = []

    def fetch_json(path: str, params: dict[str, str]) -> dict[str, object]:
        fetched.append((path, params))
        return _report()

    def deliver_report(**kwargs):
        delivered.append(kwargs)
        return {"sent_count": 1, "chat_ids": ["-1001"]}

    summary = sync_retail_price_type_recommendations(
        fetch_json=fetch_json,
        deliver_report=deliver_report,
        month="2026-03",
        state_path=tmp_path / "state.json",
        artifact_dir=tmp_path / "artifacts",
        delivery_target="-1001",
    )

    assert summary["delivered"] == 1
    assert fetched == [
        (
            REPORT_ENDPOINT,
            {
                "month": "2026-03",
                "actionable_only": "true",
                "buyers_group_only": "true",
                "buyer_group_name": "ПОКУПАТЕЛИ",
            },
        )
    ]
    assert delivered[0]["report_key"] == "retail-price-types|2026-03"
    assert Path(str(delivered[0]["artifact_path"])).exists()

    second = sync_retail_price_type_recommendations(
        fetch_json=fetch_json,
        deliver_report=deliver_report,
        month="2026-03",
        state_path=tmp_path / "state.json",
        artifact_dir=tmp_path / "artifacts",
        delivery_target="-1001",
    )

    assert second["action"] == "noop"
    state = json.loads((tmp_path / "state.json").read_text(encoding="utf-8"))
    assert state["months"]["2026-03"]["delivery_status"] == "delivered"
    assert state["months"]["2026-03"]["delivery_channel"] == "telegram"
    assert state["months"]["2026-03"]["telegram_chat_ids"] == ["-1001"]
