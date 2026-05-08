from __future__ import annotations

from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path

import httpx
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.models import Base, ReturnSchemeIncident
from app.services.return_scheme import (
    DetectedReturnSchemeIncident,
    OperationEvent,
    acknowledge_return_scheme_alert_batch,
    create_return_scheme_alert_batch,
    detect_return_scheme_incidents,
    export_return_scheme_report_xlsx,
    mark_return_scheme_incidents_notified,
    parse_retail_price_types,
    send_return_scheme_telegram_report,
    serialize_return_scheme_alert_batch,
    upsert_return_scheme_incidents,
)


def _event(
    *,
    event_type: str,
    hours: int,
    product_ref: str = "prod-1",
    store_ref: str = "store-1",
    price_type: str | None = None,
    quantity: str = "1",
    amount: str = "100",
    suffix: str = "1",
) -> OperationEvent:
    base = datetime(2026, 3, 1, 10, 0, 0)
    return OperationEvent(
        event_type=event_type,
        doc_ref=f"doc-{suffix}",
        doc_number=f"N-{suffix}",
        doc_datetime=base + timedelta(hours=hours),
        product_ref=product_ref,
        product_name=f"Товар {product_ref}",
        store_ref=store_ref,
        store_name=f"Магазин {store_ref}",
        employee_ref=f"emp-{suffix}",
        employee_name=f"Менеджер {suffix}",
        price_type=price_type,
        quantity=Decimal(quantity),
        amount=Decimal(amount),
    )


def test_detect_return_scheme_incidents_positive_case() -> None:
    events = [
        _event(event_type="sale", hours=0, price_type="Розница", amount="120", suffix="sale1"),
        _event(event_type="return", hours=12, amount="120", suffix="return1"),
        _event(event_type="sale", hours=18, price_type="Опт2", amount="90", suffix="sale2"),
    ]

    incidents = detect_return_scheme_incidents(
        events,
        retail_price_types=parse_retail_price_types("Розница"),
        window_days=7,
    )

    assert len(incidents) == 1
    assert incidents[0].second_price_type == "Опт2"
    assert incidents[0].manager_name == "Менеджер sale2"
    assert incidents[0].matched_qty == Decimal("1.000")
    assert incidents[0].amount == Decimal("90.00")


def test_detect_return_scheme_ignores_sequence_outside_window() -> None:
    events = [
        _event(event_type="sale", hours=0, price_type="Розница", suffix="sale1"),
        _event(event_type="return", hours=12, suffix="return1"),
        _event(
            event_type="sale",
            hours=24 * 8,
            price_type="Опт",
            amount="90",
            suffix="sale2",
        ),
    ]

    incidents = detect_return_scheme_incidents(
        events,
        retail_price_types=parse_retail_price_types("Розница"),
        window_days=7,
    )

    assert incidents == []


def test_detect_return_scheme_ignores_other_store_and_second_retail() -> None:
    events = [
        _event(event_type="sale", hours=0, price_type="Розница", suffix="sale1"),
        _event(event_type="return", hours=1, suffix="return1"),
        _event(
            event_type="sale",
            hours=2,
            store_ref="store-2",
            price_type="Опт",
            suffix="sale2",
        ),
        _event(
            event_type="sale",
            hours=3,
            price_type="Розница",
            suffix="sale3",
        ),
    ]

    incidents = detect_return_scheme_incidents(
        events,
        retail_price_types=parse_retail_price_types("Розница"),
        window_days=7,
    )

    assert incidents == []


def test_detect_return_scheme_fifo_partial_quantity() -> None:
    events = [
        _event(
            event_type="sale",
            hours=0,
            price_type="Розница",
            quantity="3",
            amount="300",
            suffix="sale1",
        ),
        _event(event_type="return", hours=1, quantity="2", amount="200", suffix="return1"),
        _event(
            event_type="sale",
            hours=2,
            price_type="Опт",
            quantity="1.5",
            amount="135",
            suffix="sale2",
        ),
        _event(
            event_type="sale",
            hours=3,
            price_type="Опт",
            quantity="1",
            amount="90",
            suffix="sale3",
        ),
    ]

    incidents = detect_return_scheme_incidents(
        events,
        retail_price_types=parse_retail_price_types("Розница"),
        window_days=7,
    )

    assert len(incidents) == 2
    assert incidents[0].matched_qty == Decimal("1.500")
    assert incidents[0].amount == Decimal("135.00")
    assert incidents[1].matched_qty == Decimal("0.500")
    assert incidents[1].amount == Decimal("45.00")


def test_upsert_return_scheme_incidents_deduplicates_and_keeps_unsent() -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    events = [
        _event(event_type="sale", hours=0, price_type="Розница", suffix="sale1"),
        _event(event_type="return", hours=1, suffix="return1"),
        _event(event_type="sale", hours=2, price_type="Опт", amount="90", suffix="sale2"),
    ]
    incidents = detect_return_scheme_incidents(
        events,
        retail_price_types=parse_retail_price_types("Розница"),
        window_days=7,
    )

    with Session(engine) as session:
        result1 = upsert_return_scheme_incidents(session, incidents, detected_at=datetime.now())
        session.commit()

        assert len(result1["new"]) == 1
        assert len(result1["pending_notification"]) == 1

        stored = session.query(ReturnSchemeIncident).one()
        assert stored.notified_at is None

        result2 = upsert_return_scheme_incidents(session, incidents, detected_at=datetime.now())
        session.commit()
        assert len(result2["new"]) == 0
        assert len(result2["pending_notification"]) == 1

        mark_return_scheme_incidents_notified(session, [stored.id], notified_at=datetime.now())
        session.commit()

        result3 = upsert_return_scheme_incidents(session, incidents, detected_at=datetime.now())
        session.commit()
        assert len(result3["new"]) == 0
        assert len(result3["pending_notification"]) == 0


def test_export_return_scheme_report_xlsx(tmp_path: Path) -> None:
    incident = DetectedReturnSchemeIncident(
        product_ref="prod-1",
        product_name="Дисплей iPhone",
        store_ref="store-1",
        store_name="Магазин 1",
        manager_ref="emp-2",
        manager_name="Иван",
        first_sale_event=_event(
            event_type="sale", hours=0, price_type="Розница", suffix="sale1", amount="120"
        ),
        return_event=_event(event_type="return", hours=1, suffix="return1", amount="120"),
        second_sale_event=_event(
            event_type="sale", hours=2, price_type="Опт", suffix="sale2", amount="90"
        ),
        second_price_type="Опт",
        matched_qty=Decimal("1.000"),
        amount=Decimal("90.00"),
    )

    output_path = tmp_path / "return-scheme.xlsx"
    export_return_scheme_report_xlsx([incident], output_path)

    workbook = load_workbook(output_path)
    sheet = workbook.active
    assert sheet.cell(1, 1).value == "Магазин"
    assert sheet.cell(2, 1).value == "Магазин 1"
    assert sheet.cell(2, 2).value == "Дисплей iPhone"
    assert sheet.cell(2, 7).value == "Опт"
    assert sheet.cell(2, 8).value == 1
    assert sheet.cell(2, 9).value == 90


def test_create_alert_batch_and_acknowledge_marks_incidents_notified(tmp_path: Path) -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    events = [
        _event(event_type="sale", hours=0, price_type="Розница", suffix="sale1"),
        _event(event_type="return", hours=1, suffix="return1"),
        _event(event_type="sale", hours=2, price_type="Опт", amount="90", suffix="sale2"),
    ]
    incidents = detect_return_scheme_incidents(
        events,
        retail_price_types=parse_retail_price_types("Розница"),
        window_days=7,
    )

    report_path = tmp_path / "batch.xlsx"
    with Session(engine) as session:
        persisted = upsert_return_scheme_incidents(session, incidents, detected_at=datetime.now())
        export_return_scheme_report_xlsx(persisted["pending_notification"], report_path)
        batch = create_return_scheme_alert_batch(
            session,
            incidents=persisted["pending_notification"],
            generated_at=datetime(2026, 3, 2, 9, 0, 0),
            window_start=datetime(2026, 2, 24, 9, 0, 0),
            window_end=datetime(2026, 3, 2, 9, 0, 0),
            report_path=report_path,
            new_incidents_count=len(persisted["new"]),
        )
        assert batch is not None
        assert batch.status == "pending"
        payload = serialize_return_scheme_alert_batch(session, batch)
        assert payload["incident_ids"]
        assert payload["incidents"][0]["repeat_store_product_7d_count"] == 1

        batch = acknowledge_return_scheme_alert_batch(
            session, batch.id, delivered_at=datetime.now()
        )
        session.commit()

        stored = session.query(ReturnSchemeIncident).one()
        assert batch.status == "delivered"
        assert stored.notified_at is not None
        assert stored.alert_batch_id == batch.id


def test_send_return_scheme_telegram_report_uploads_document(tmp_path: Path) -> None:
    sent: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        sent.append(request)
        return httpx.Response(200, json={"ok": True})

    report_path = tmp_path / "report.xlsx"
    report_path.write_bytes(b"test")

    send_return_scheme_telegram_report(
        token="token",
        chat_id="123",
        message="test-message",
        report_path=report_path,
        transport=httpx.MockTransport(handler),
    )

    assert len(sent) == 1
    assert str(sent[0].url) == "https://api.telegram.org/bottoken/sendDocument"
