from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

from app.models import Base, ReceivableCase
from tasks.send_employee_receivable_report import (
    CHANGES_SHEET_TITLE,
    CHART_DATA_SHEET_TITLE,
    CONTROL_SHEET_TITLE,
    CURRENT_SHEET_TITLE,
    RELATED_DOCS_SHEET_TITLE,
    SUMMARY_SHEET_TITLE,
    EmployeeReceivableItem,
    EmployeeReceivableRelatedDocument,
    build_employee_receivable_changes,
    export_employee_receivable_report,
    load_employee_related_documents,
    resolve_employee_snapshot_dates,
)


def _item(
    *,
    snapshot_date: date,
    counterparty_ref: str,
    counterparty_name: str,
    current_balance: str,
    aged_bucket: str = "unknown",
    activity_segment: str = "inactive",
    counterparty_code: str | None = None,
) -> EmployeeReceivableItem:
    return EmployeeReceivableItem(
        snapshot_date=snapshot_date,
        counterparty_ref=counterparty_ref,
        counterparty_name=counterparty_name,
        current_balance=Decimal(current_balance),
        aged_bucket=aged_bucket,
        activity_segment=activity_segment,
        current_manager_ref=None,
        current_manager_name=None,
        origin_document_ref=None,
        origin_document_number=None,
        counterparty_code=counterparty_code,
    )


def test_build_employee_receivable_changes_marks_new_increased_decreased_and_closed() -> None:
    previous_items = [
        _item(
            snapshot_date=date(2026, 4, 4),
            counterparty_ref="cp-same",
            counterparty_name="Без изменений",
            current_balance="100.00",
        ),
        _item(
            snapshot_date=date(2026, 4, 4),
            counterparty_ref="cp-down",
            counterparty_name="Снижение",
            current_balance="300.00",
        ),
        _item(
            snapshot_date=date(2026, 4, 4),
            counterparty_ref="cp-closed",
            counterparty_name="Закрыт",
            current_balance="250.00",
        ),
    ]
    current_items = [
        _item(
            snapshot_date=date(2026, 4, 5),
            counterparty_ref="cp-same",
            counterparty_name="Без изменений",
            current_balance="100.00",
        ),
        _item(
            snapshot_date=date(2026, 4, 5),
            counterparty_ref="cp-up",
            counterparty_name="Новый",
            current_balance="500.00",
        ),
        _item(
            snapshot_date=date(2026, 4, 5),
            counterparty_ref="cp-down",
            counterparty_name="Снижение",
            current_balance="200.00",
        ),
    ]

    changes = build_employee_receivable_changes(current_items, previous_items)
    by_ref = {item.counterparty_ref: item for item in changes}

    assert by_ref["cp-same"].status == "unchanged"
    assert by_ref["cp-same"].delta_balance == Decimal("0.00")
    assert by_ref["cp-up"].status == "new"
    assert by_ref["cp-up"].delta_balance == Decimal("500.00")
    assert by_ref["cp-down"].status == "decreased"
    assert by_ref["cp-down"].delta_balance == Decimal("-100.00")
    assert by_ref["cp-closed"].status == "closed"
    assert by_ref["cp-closed"].current_balance == Decimal("0.00")
    assert by_ref["cp-closed"].delta_balance == Decimal("-250.00")


def test_export_employee_receivable_report_creates_formatted_workbook(tmp_path) -> None:
    current_items = [
        _item(
            snapshot_date=date(2026, 4, 5),
            counterparty_ref="cp-1",
            counterparty_name="Платонов Андрей",
            current_balance="2742646.65",
            aged_bucket="unknown",
            counterparty_code="РБ010101",
        ),
        _item(
            snapshot_date=date(2026, 4, 5),
            counterparty_ref="cp-2",
            counterparty_name="Байрамов Эльвин",
            current_balance="674126.00",
            aged_bucket="unknown",
            counterparty_code="РБ010102",
        ),
    ]
    previous_items = [
        _item(
            snapshot_date=date(2026, 4, 4),
            counterparty_ref="cp-1",
            counterparty_name="Платонов Андрей",
            current_balance="842646.65",
            counterparty_code="РБ010101",
        )
    ]
    changes = build_employee_receivable_changes(current_items, previous_items)
    related_documents = [
        EmployeeReceivableRelatedDocument(
            counterparty_name="Платонов Андрей",
            document_kind="Реализация товаров и услуг",
            document_date=datetime(2026, 4, 5, 14, 30),
            document_number="РБГУ000111",
            document_ref="0xabc",
            responsible_name="Иванов Иван",
            department_name="ТК Савеловский",
            organization_name="MASTER MOBILE",
            item_name="Дисплей iPhone 11",
            quantity=Decimal("2"),
            amount_delta=Decimal("145000.00"),
        ),
        EmployeeReceivableRelatedDocument(
            counterparty_name="Списание товара",
            document_kind="Реализация товаров и услуг",
            document_date=datetime(2026, 4, 4, 10, 15),
            document_number="РБГУ000222",
            document_ref="0xdef",
            responsible_name="Петров Петр",
            department_name="Мега",
            organization_name="MASTER MOBILE",
            item_name="Аккумулятор Xiaomi",
            quantity=Decimal("1"),
            amount_delta=Decimal("8200.00"),
        ),
    ]
    output_path = tmp_path / "employee-report.xlsx"

    export_employee_receivable_report(
        snapshot_date=date(2026, 4, 5),
        previous_date=date(2026, 4, 4),
        current_items=current_items,
        changes=changes,
        output_path=output_path,
        snapshot_history=[
            (date(2026, 4, 1), Decimal("900000.00")),
            (date(2026, 4, 2), Decimal("1100000.00")),
            (date(2026, 4, 5), Decimal("3416772.65")),
        ],
        related_documents=related_documents,
    )

    workbook = load_workbook(output_path)
    assert [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"] == [
        CONTROL_SHEET_TITLE,
        SUMMARY_SHEET_TITLE,
        CURRENT_SHEET_TITLE,
        CHANGES_SHEET_TITLE,
        RELATED_DOCS_SHEET_TITLE,
    ]
    assert CHART_DATA_SHEET_TITLE in workbook.sheetnames
    assert workbook[CHART_DATA_SHEET_TITLE].sheet_state == "hidden"

    control = workbook[CONTROL_SHEET_TITLE]
    assert control["A1"].value == "Контроль долгов сотрудников"
    assert len(control._charts) >= 3

    summary = workbook[SUMMARY_SHEET_TITLE]
    assert summary.freeze_panes == "A2"
    assert summary.auto_filter.ref == "A1:B11"

    current_sheet = workbook[CURRENT_SHEET_TITLE]
    assert current_sheet.freeze_panes == "A2"
    assert current_sheet.auto_filter.ref == "A1:J3"
    assert current_sheet["C2"].number_format == "#,##0.00"
    assert current_sheet["B2"].value == "Платонов Андрей"
    assert current_sheet["J1"].value == "Код 1С контрагента"
    assert current_sheet["J2"].value == "РБ010101"
    assert len(current_sheet.conditional_formatting) > 0

    changes_sheet = workbook[CHANGES_SHEET_TITLE]
    assert changes_sheet.freeze_panes == "A2"
    assert changes_sheet["A2"].value in {"Новый", "Рост"}
    assert changes_sheet["I1"].value == "Код 1С контрагента"
    assert changes_sheet["I2"].value in {"РБ010101", "РБ010102"}

    related_sheet = workbook[RELATED_DOCS_SHEET_TITLE]
    assert related_sheet.freeze_panes == "A2"
    assert related_sheet["B1"].value == "Документ"
    assert related_sheet["B2"].value == "Реализация товаров и услуг"
    assert related_sheet["J2"].number_format == "#,##0.00"
    assert related_sheet["I3"].value == "Аккумулятор Xiaomi"
    assert related_sheet["J3"].value == 1


def test_load_employee_related_documents_limits_rows_to_previous_week_and_supported_types() -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        session.execute(
            text(
                "CREATE TABLE _Document203 (_IDRRef TEXT, _Date_Time TEXT, _Number TEXT, _Posted INTEGER, _Marked INTEGER, _Fld4932RRef TEXT, _Fld4940RRef TEXT, _Fld4942RRef TEXT, _Fld4950RRef TEXT)"
            )
        )
        session.execute(
            text(
                "CREATE TABLE _Document203_VT4966 (_Document203_IDRRef TEXT, _LineNo4967 INTEGER, _Fld4974RRef TEXT, _Fld4971 NUMERIC, _Fld4982 NUMERIC)"
            )
        )
        session.execute(text("CREATE TABLE _Reference66 (_IDRRef TEXT, _Description TEXT)"))
        session.execute(text("CREATE TABLE _Reference80 (_IDRRef TEXT, _Description TEXT)"))
        session.execute(text("CREATE TABLE _Reference54 (_IDRRef TEXT, _Description TEXT)"))
        session.execute(text("CREATE TABLE _Reference69 (_IDRRef TEXT, _Description TEXT)"))
        session.execute(text("CREATE TABLE _Reference62 (_IDRRef TEXT, _Description TEXT)"))

        session.execute(text("INSERT INTO _Reference66 VALUES ('org1', 'MASTER MOBILE')"))
        session.execute(text("INSERT INTO _Reference80 VALUES ('dep1', 'ТК Савеловский')"))
        session.execute(text("INSERT INTO _Reference54 VALUES ('cp_writeoff', 'Списание товара')"))
        session.execute(text("INSERT INTO _Reference54 VALUES ('cp_other', 'Другой контрагент')"))
        session.execute(text("INSERT INTO _Reference69 VALUES ('resp1', 'Иванов Иван')"))
        session.execute(text("INSERT INTO _Reference62 VALUES ('item1', 'Дисплей iPhone 11')"))
        session.execute(text("INSERT INTO _Reference62 VALUES ('item2', 'Аккумулятор Xiaomi')"))

        session.execute(
            text(
                "INSERT INTO _Document203 VALUES "
                "('doc1', '2026-04-10 14:00:00', 'РБГУ000001', 1, 0, 'org1', 'dep1', 'cp_writeoff', 'resp1'),"
                "('doc2', '2026-04-12 10:30:00', 'РБГУ000002', 1, 0, 'org1', 'dep1', 'cp_writeoff', 'resp1'),"
                "('doc3', '2026-04-15 09:00:00', 'РБГУ000003', 1, 0, 'org1', 'dep1', 'cp_writeoff', 'resp1'),"
                "('doc4', '2026-04-09 12:00:00', 'РБГУ000004', 1, 0, 'org1', 'dep1', 'cp_other', 'resp1')"
            )
        )
        session.execute(
            text(
                "INSERT INTO _Document203_VT4966 VALUES "
                "('doc1', 1, 'item1', 2, 1500.00),"
                "('doc2', 1, 'item2', 1, 820.00),"
                "('doc3', 1, 'item1', 1, 900.00),"
                "('doc4', 1, 'item2', 1, 500.00)"
            )
        )
        session.commit()

        related_documents = load_employee_related_documents(
            onec_engine=engine,
            snapshot_date=date(2026, 4, 17),
        )

    assert [item.document_number for item in related_documents] == ["РБГУ000002", "РБГУ000001"]
    assert [item.counterparty_name for item in related_documents] == [
        "Списание товара",
        "Списание товара",
    ]
    assert [item.responsible_name for item in related_documents] == ["Иванов Иван", "Иванов Иван"]
    assert [item.item_name for item in related_documents] == [
        "Аккумулятор Xiaomi",
        "Дисплей iPhone 11",
    ]


def test_resolve_employee_snapshot_dates_supports_latest_not_after() -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        for snapshot_date in (date(2026, 4, 6), date(2026, 4, 5), date(2026, 4, 3)):
            session.add(
                ReceivableCase(
                    snapshot_date=snapshot_date,
                    segment="employee",
                    owner_type="finance_hr",
                    recommendation="test",
                    counterparty_ref=f"cp-{snapshot_date.isoformat()}",
                    counterparty_name=f"Контрагент {snapshot_date.isoformat()}",
                    current_balance=Decimal("100.00"),
                    aged_bucket="unknown",
                    activity_segment="inactive",
                    origin_document_ref=None,
                    origin_document_number=None,
                    origin_document_date=None,
                    origin_manager_ref=None,
                    origin_manager_name=None,
                    current_manager_ref=None,
                    current_manager_name=None,
                    planned_payment_date=None,
                    credit_depth_days=None,
                    shipment_ban=None,
                    payment_term_source=None,
                    due_date=None,
                    overdue_days=None,
                    is_overdue=False,
                    chain_documents=None,
                )
            )
        session.commit()

        snapshot_date, previous_date = resolve_employee_snapshot_dates(
            session,
            requested_date=None,
            latest_not_after=date(2026, 4, 5),
        )

    assert snapshot_date == date(2026, 4, 5)
    assert previous_date == date(2026, 4, 3)
