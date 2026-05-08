from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook

from tasks.send_weekly_manager_sales_report import (
    ATTENTION_SHEET_TITLE,
    CASH_ORDERS_SHEET_TITLE,
    CHART_DATA_SHEET_TITLE,
    DASHBOARD_SHEET_TITLE,
    MANAGERS_SHEET_TITLE,
    STORE_DETAIL_SHEET_TITLE,
    SUMMARY_SHEET_TITLE,
    SalesKpiRecord,
    ShortageCashOrderItem,
    WeeklySalesWindow,
    build_attention_manager_sales_items,
    build_weekly_manager_sales_items,
    build_weekly_manager_store_sales_items,
    export_weekly_manager_sales_report,
    send_weekly_reports_to_telegram,
)


def _record(
    *,
    sales_date: date,
    manager_ref: str,
    manager_name: str,
    manager_code: str,
    store_ref: str,
    store_name: str,
    store_code: str,
    revenue: str,
    sales_count: str,
) -> SalesKpiRecord:
    return SalesKpiRecord(
        sales_date=sales_date,
        manager_ref=manager_ref,
        manager_name=manager_name,
        manager_code=manager_code,
        store_ref=store_ref,
        store_name=store_name,
        store_code=store_code,
        revenue=Decimal(revenue),
        sales_count=Decimal(sales_count),
    )


def _cash_order(
    *,
    document_number: str,
    document_date: datetime,
    counterparty_name: str,
    counterparty_code: str,
    amount: str,
    employee_name: str,
    cashbox_name: str,
    store_name: str,
    base_sale_number: str,
) -> ShortageCashOrderItem:
    return ShortageCashOrderItem(
        document_number=document_number,
        document_date=document_date,
        counterparty_name=counterparty_name,
        counterparty_code=counterparty_code,
        amount=Decimal(amount),
        employee_name=employee_name,
        cashbox_name=cashbox_name,
        store_name=store_name,
        base_sale_number=base_sale_number,
    )


def test_build_weekly_manager_sales_items_marks_attention_and_ranks() -> None:
    current_records = [
        _record(
            sales_date=date(2026, 4, 1),
            manager_ref="mgr-top",
            manager_name="Лидер",
            manager_code="РБ000101",
            store_ref="store-1",
            store_name="Лира",
            store_code="РБ000201",
            revenue="150000.00",
            sales_count="10.000",
        ),
        _record(
            sales_date=date(2026, 4, 2),
            manager_ref="mgr-drop",
            manager_name="Просадка",
            manager_code="РБ000102",
            store_ref="store-2",
            store_name="Мега",
            store_code="РБ000202",
            revenue="40000.00",
            sales_count="4.000",
        ),
        _record(
            sales_date=date(2026, 4, 3),
            manager_ref="mgr-low",
            manager_name="Слабый",
            manager_code="РБ000103",
            store_ref="store-3",
            store_name="Парнас",
            store_code="РБ000203",
            revenue="15000.00",
            sales_count="2.000",
        ),
        _record(
            sales_date=date(2026, 4, 4),
            manager_ref="mgr-zero",
            manager_name="Без продаж",
            manager_code="РБ000104",
            store_ref="store-4",
            store_name="Невский",
            store_code="РБ000204",
            revenue="0.00",
            sales_count="0.000",
        ),
    ]
    previous_records = [
        _record(
            sales_date=date(2026, 3, 25),
            manager_ref="mgr-top",
            manager_name="Лидер",
            manager_code="РБ000101",
            store_ref="store-1",
            store_name="Лира",
            store_code="РБ000201",
            revenue="100000.00",
            sales_count="8.000",
        ),
        _record(
            sales_date=date(2026, 3, 26),
            manager_ref="mgr-drop",
            manager_name="Просадка",
            manager_code="РБ000102",
            store_ref="store-2",
            store_name="Мега",
            store_code="РБ000202",
            revenue="90000.00",
            sales_count="8.000",
        ),
        _record(
            sales_date=date(2026, 3, 27),
            manager_ref="mgr-low",
            manager_name="Слабый",
            manager_code="РБ000103",
            store_ref="store-3",
            store_name="Парнас",
            store_code="РБ000203",
            revenue="12000.00",
            sales_count="1.000",
        ),
        _record(
            sales_date=date(2026, 3, 28),
            manager_ref="mgr-zero",
            manager_name="Без продаж",
            manager_code="РБ000104",
            store_ref="store-4",
            store_name="Невский",
            store_code="РБ000204",
            revenue="5000.00",
            sales_count="1.000",
        ),
    ]

    items = build_weekly_manager_sales_items(current_records, previous_records)
    by_ref = {item.manager_ref: item for item in items}

    assert items[0].manager_ref == "mgr-top"
    assert items[0].revenue_rank == 1
    assert by_ref["mgr-top"].signal == "Норма"
    assert by_ref["mgr-drop"].signal == "Просадка"
    assert by_ref["mgr-low"].signal == "Низкий объем"
    assert by_ref["mgr-zero"].signal == "Нет продаж"
    assert by_ref["mgr-drop"].revenue_delta == Decimal("-50000.00")
    assert by_ref["mgr-top"].current_avg_ticket == Decimal("15000.00")


def test_export_weekly_manager_sales_report_creates_formatted_workbook(tmp_path) -> None:
    current_records = [
        _record(
            sales_date=date(2026, 4, 1),
            manager_ref="mgr-1",
            manager_name="Артем",
            manager_code="РБ010001",
            store_ref="store-1",
            store_name="Лира",
            store_code="РБ020001",
            revenue="250000.00",
            sales_count="20.000",
        ),
        _record(
            sales_date=date(2026, 4, 2),
            manager_ref="mgr-2",
            manager_name="Борис",
            manager_code="РБ010002",
            store_ref="store-2",
            store_name="Мега",
            store_code="РБ020002",
            revenue="30000.00",
            sales_count="3.000",
        ),
    ]
    previous_records = [
        _record(
            sales_date=date(2026, 3, 25),
            manager_ref="mgr-1",
            manager_name="Артем",
            manager_code="РБ010001",
            store_ref="store-1",
            store_name="Лира",
            store_code="РБ020001",
            revenue="200000.00",
            sales_count="16.000",
        ),
        _record(
            sales_date=date(2026, 3, 26),
            manager_ref="mgr-2",
            manager_name="Борис",
            manager_code="РБ010002",
            store_ref="store-2",
            store_name="Мега",
            store_code="РБ020002",
            revenue="60000.00",
            sales_count="5.000",
        ),
    ]

    manager_items = build_weekly_manager_sales_items(current_records, previous_records)
    attention_items = build_attention_manager_sales_items(manager_items)
    manager_store_items = build_weekly_manager_store_sales_items(current_records, previous_records)
    cash_order_items = [
        _cash_order(
            document_number="РКО-002",
            document_date=datetime(2026, 4, 4, 18, 30),
            counterparty_name="Излишек Савелово",
            counterparty_code="РБ051790",
            amount="2500.00",
            employee_name="Садыков Энвер",
            cashbox_name="Савелово касса 4",
            store_name="ТК Савеловский",
            base_sale_number="000123",
        ),
        _cash_order(
            document_number="РКО-001",
            document_date=datetime(2026, 4, 1, 12, 15),
            counterparty_name="Недостача Мега",
            counterparty_code="РБ051791",
            amount="1100.00",
            employee_name="Шевцов Вячеслав",
            cashbox_name="Мега касса 1",
            store_name="Мега",
            base_sale_number="000122",
        ),
    ]
    output_path = tmp_path / "weekly-manager-sales.xlsx"

    export_weekly_manager_sales_report(
        window=WeeklySalesWindow(
            week_start=date(2026, 3, 30),
            week_end=date(2026, 4, 5),
            compare_week_start=date(2026, 3, 23),
            compare_week_end=date(2026, 3, 29),
        ),
        manager_items=manager_items,
        attention_items=attention_items,
        manager_store_items=manager_store_items,
        cash_order_items=cash_order_items,
        output_path=output_path,
        weekly_history=[
            (date(2026, 3, 16), Decimal("150000.00"), Decimal("12.000")),
            (date(2026, 3, 23), Decimal("260000.00"), Decimal("21.000")),
            (date(2026, 3, 30), Decimal("280000.00"), Decimal("23.000")),
        ],
    )

    workbook = load_workbook(output_path)
    assert [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"] == [
        DASHBOARD_SHEET_TITLE,
        SUMMARY_SHEET_TITLE,
        MANAGERS_SHEET_TITLE,
        ATTENTION_SHEET_TITLE,
        STORE_DETAIL_SHEET_TITLE,
        CASH_ORDERS_SHEET_TITLE,
    ]
    assert CHART_DATA_SHEET_TITLE in workbook.sheetnames
    assert workbook[CHART_DATA_SHEET_TITLE].sheet_state == "hidden"

    dashboard = workbook[DASHBOARD_SHEET_TITLE]
    assert dashboard["A1"].value == "Управленческий weekly по личным продажам"
    assert len(dashboard._charts) >= 3

    summary = workbook[SUMMARY_SHEET_TITLE]
    assert summary.freeze_panes == "A2"
    assert summary.auto_filter.ref == "A1:B19"
    summary_values = {
        summary.cell(row=row_index, column=1).value: summary.cell(row=row_index, column=2).value
        for row_index in range(2, summary.max_row + 1)
    }
    assert summary_values["РКО Излишек/недостача, шт"] == 2
    assert summary_values["РКО Излишек/недостача, сумма, ₽"] == 3600

    managers = workbook[MANAGERS_SHEET_TITLE]
    assert managers.freeze_panes == "A2"
    assert managers.auto_filter.ref == "A1:O3"
    assert managers["E2"].number_format == "#,##0.00"
    assert managers["F2"].number_format == "#,##0.000"
    assert managers["B3"].value in {"Просадка", "Низкий объем"}
    assert managers["O1"].value == "Код 1С менеджера"
    assert managers["O2"].value == "РБ010001"
    assert len(managers.conditional_formatting) > 0

    attention = workbook[ATTENTION_SHEET_TITLE]
    assert attention.freeze_panes == "A2"
    assert attention["B2"].value in {"Просадка", "Низкий объем", "Нет продаж"}
    assert attention["K1"].value == "Код 1С менеджера"

    detail = workbook[STORE_DETAIL_SHEET_TITLE]
    assert detail.freeze_panes == "A2"
    assert detail.auto_filter.ref == "A1:O3"
    assert detail["D2"].value == "Лира"
    assert detail["N1"].value == "Код 1С менеджера"
    assert detail["O1"].value == "Код 1С магазина"
    assert detail["N2"].value == "РБ010001"
    assert detail["O2"].value == "РБ020001"

    cash_orders = workbook[CASH_ORDERS_SHEET_TITLE]
    assert cash_orders.freeze_panes == "A2"
    assert cash_orders.auto_filter.ref == "A1:J3"
    assert cash_orders["B2"].value == "04.04.2026 18:30"
    assert cash_orders["F2"].number_format == "#,##0.00"
    assert cash_orders["E2"].value == "РБ051790"
    assert cash_orders["G2"].value == "Садыков Энвер"


def test_send_weekly_reports_to_telegram_sends_main_and_employee_attachments(tmp_path) -> None:
    sent: list[dict[str, object]] = []
    weekly_report = tmp_path / "weekly.xlsx"
    weekly_report.write_bytes(b"weekly")
    employee_report = tmp_path / "employee.xlsx"
    employee_report.write_bytes(b"employee")

    def fake_sender(*, token: str, chat_id: str, message: str, report_path):
        sent.append(
            {
                "token": token,
                "chat_id": chat_id,
                "message": message,
                "report_path": str(report_path),
            }
        )

    sent_count = send_weekly_reports_to_telegram(
        token="token",
        chat_ids="chat",
        weekly_message="weekly-message",
        weekly_report_path=weekly_report,
        employee_message="employee-message",
        employee_report_path=employee_report,
        sender=fake_sender,
    )

    assert sent_count == 2
    assert sent == [
        {
            "token": "token",
            "chat_id": "chat",
            "message": "weekly-message",
            "report_path": str(weekly_report),
        },
        {
            "token": "token",
            "chat_id": "chat",
            "message": "employee-message",
            "report_path": str(employee_report),
        },
    ]


def test_send_weekly_reports_to_telegram_supports_multiple_chat_ids(tmp_path) -> None:
    sent: list[dict[str, object]] = []
    weekly_report = tmp_path / "weekly.xlsx"
    weekly_report.write_bytes(b"weekly")
    employee_report = tmp_path / "employee.xlsx"
    employee_report.write_bytes(b"employee")

    def fake_sender(*, token: str, chat_id: str, message: str, report_path):
        sent.append(
            {
                "token": token,
                "chat_id": chat_id,
                "message": message,
                "report_path": str(report_path),
            }
        )

    sent_count = send_weekly_reports_to_telegram(
        token="token",
        chat_ids="karina-chat, arsen-chat",
        weekly_message="weekly-message",
        weekly_report_path=weekly_report,
        employee_message="employee-message",
        employee_report_path=employee_report,
        sender=fake_sender,
    )

    assert sent_count == 4
    assert sent == [
        {
            "token": "token",
            "chat_id": "karina-chat",
            "message": "weekly-message",
            "report_path": str(weekly_report),
        },
        {
            "token": "token",
            "chat_id": "karina-chat",
            "message": "employee-message",
            "report_path": str(employee_report),
        },
        {
            "token": "token",
            "chat_id": "arsen-chat",
            "message": "weekly-message",
            "report_path": str(weekly_report),
        },
        {
            "token": "token",
            "chat_id": "arsen-chat",
            "message": "employee-message",
            "report_path": str(employee_report),
        },
    ]
